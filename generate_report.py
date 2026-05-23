#!/usr/bin/env python3
"""Generate Al Matar Restaurant Monthly Business Report (Sample)"""

import sys, os
sys.path.insert(0, '/home/z/my-project/skills/xlsx/templates')
sys.path.insert(0, '/home/z/my-project/skills/xlsx')
from base import *

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
from openpyxl.chart.marker import Marker

# ── Apply warm palette BEFORE creating any styles ──
use_palette_explicit("warm")

wb = Workbook()

# ============================================================
# HELPER: Add footer to a sheet
# ============================================================
def add_footer(ws, col_end, row=None):
    """Add a footer row with analyst credit."""
    if row is None:
        row = ws.max_row + 2
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=col_end)
    fc = ws.cell(row=row, column=2, value="Prepared by Ahmed Ali | Data Analysis Services | ahmed-ali-ops.vercel.app")
    fc.font = font_caption()
    fc.alignment = Alignment(horizontal="center", vertical="center")


# ============================================================
# SHEET 1: Dashboard
# ============================================================
ws1 = wb.active
ws1.title = "Dashboard"

setup_sheet(ws1, title="Al Matar Restaurant — Monthly Business Report", last_col=6)

# Subtitle row 3
ws1.merge_cells(start_row=3, start_column=2, end_row=3, end_column=6)
subtitle_cell = ws1.cell(row=3, column=2, value="January 2026 | Sharjah, UAE")
subtitle_cell.font = font_subheader()
subtitle_cell.alignment = align_title()

# ── KPI Row (row 5-6) ──
kpi_data = [
    ("Total Revenue", "AED 187,450"),
    ("Total Expenses", "AED 142,800"),
    ("Net Profit", "AED 44,650"),
    ("Profit Margin", "23.8%"),
]

kpi_start_row = 5
kpi_start_col = 2
for i, (label, value) in enumerate(kpi_data):
    col = kpi_start_col + i
    val_cell = ws1.cell(row=kpi_start_row, column=col, value=value)
    val_cell.font = font_kpi()
    val_cell.alignment = Alignment(horizontal="center", vertical="center")
    lbl_cell = ws1.cell(row=kpi_start_row + 1, column=col, value=label)
    lbl_cell.font = font_kpi_label()
    lbl_cell.alignment = Alignment(horizontal="center", vertical="center")

ws1.row_dimensions[5].height = 36
ws1.row_dimensions[6].height = 18

# ── Spacer row 7 ──
ws1.row_dimensions[7].height = 12

# ── Monthly Data Table ──
table_start = 8
headers = ["Month", "Revenue (AED)", "Expenses (AED)", "Net Profit (AED)", "Profit Margin"]
for i, h in enumerate(headers):
    ws1.cell(row=table_start, column=2 + i, value=h)
style_header_row(ws1, table_start, 2, 6)

monthly_data = [
    ("Oct 2025", 165200, 138500, 26700, 0.162),
    ("Nov 2025", 172800, 140200, 32600, 0.189),
    ("Dec 2025", 198500, 151300, 47200, 0.238),
    ("Jan 2026", 187450, 142800, 44650, 0.238),
]

for idx, (month, rev, exp, profit, margin) in enumerate(monthly_data):
    r = table_start + 1 + idx
    ws1.cell(row=r, column=2, value=month).alignment = align_text()
    ws1.cell(row=r, column=3, value=rev).number_format = '#,##0'
    ws1.cell(row=r, column=3).alignment = align_number()
    ws1.cell(row=r, column=4, value=exp).number_format = '#,##0'
    ws1.cell(row=r, column=4).alignment = align_number()
    ws1.cell(row=r, column=5, value=profit).number_format = '#,##0'
    ws1.cell(row=r, column=5).alignment = align_number()
    ws1.cell(row=r, column=6, value=margin).number_format = '0.0%'
    ws1.cell(row=r, column=6).alignment = align_number()
    style_data_row(ws1, r, 2, 6, idx)

# Auto-fit columns BEFORE adding footer merged cells
auto_fit_columns(ws1, header_row=table_start, data_start_row=table_start + 1)

# ── Bar Chart: Revenue vs Expenses ──
chart1 = create_bar_chart(chart_type="col", grouping="clustered", gap_width=80, overlap=-20, width=18, height=10)
data_ref = Reference(ws1, min_col=3, min_row=table_start, max_col=4, max_row=table_start + 4)
cats_ref = Reference(ws1, min_col=2, min_row=table_start + 1, max_row=table_start + 4)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.shape = 4
setup_chart_titles(chart1, title="Revenue vs Expenses by Month", y_title="AED", x_title="Month")
apply_chart_colors(chart1)
ws1.add_chart(chart1, "B14")

# ── Line Chart: Profit Margin Trend ──
chart2 = create_line_chart(width=18, height=10)
margin_ref = Reference(ws1, min_col=6, min_row=table_start, max_row=table_start + 4)
cats_ref2 = Reference(ws1, min_col=2, min_row=table_start + 1, max_row=table_start + 4)
chart2.add_data(margin_ref, titles_from_data=True)
chart2.set_categories(cats_ref2)
setup_chart_titles(chart2, title="Profit Margin Trend", y_title="Margin", x_title="Month")
apply_chart_colors(chart2)
s = chart2.series[0]
s.graphicalProperties.line.width = 28000
s.marker = Marker(symbol="circle", size=7)
ws1.add_chart(chart2, "B30")

# ── Footer ──
add_footer(ws1, 6)


# ============================================================
# SHEET 2: Sales Breakdown
# ============================================================
ws2 = wb.create_sheet("Sales Breakdown")
setup_sheet(ws2, title="Sales by Category — January 2026", last_col=6)

table_start2 = 4
headers2 = ["Category", "Orders", "Revenue (AED)", "Avg Order (AED)", "% of Revenue"]
for i, h in enumerate(headers2):
    ws2.cell(row=table_start2, column=2 + i, value=h)
style_header_row(ws2, table_start2, 2, 6)

sales_data = [
    ("Grills & BBQ",  1240, 52080, 42.0, 0.278),
    ("Rice & Biryani", 980, 33320, 34.0, 0.178),
    ("Shawarma",      1560, 37440, 24.0, 0.200),
    ("Beverages",     2100, 25200, 12.0, 0.134),
    ("Desserts",       640, 17280, 27.0, 0.092),
    ("Appetizers",     890, 22130, 24.9, 0.118),
]

for idx, (cat, orders, rev, avg, pct) in enumerate(sales_data):
    r = table_start2 + 1 + idx
    ws2.cell(row=r, column=2, value=cat).alignment = align_text()
    ws2.cell(row=r, column=3, value=orders).number_format = '#,##0'
    ws2.cell(row=r, column=3).alignment = align_number()
    ws2.cell(row=r, column=4, value=rev).number_format = '#,##0'
    ws2.cell(row=r, column=4).alignment = align_number()
    ws2.cell(row=r, column=5, value=avg).number_format = '#,##0.0'
    ws2.cell(row=r, column=5).alignment = align_number()
    ws2.cell(row=r, column=6, value=pct).number_format = '0.0%'
    ws2.cell(row=r, column=6).alignment = align_number()
    style_data_row(ws2, r, 2, 6, idx)

# Total row
total_row2 = table_start2 + 1 + len(sales_data)
ws2.cell(row=total_row2, column=2, value="TOTAL").alignment = align_text()
ws2.cell(row=total_row2, column=3, value=7410).number_format = '#,##0'
ws2.cell(row=total_row2, column=3).alignment = align_number()
ws2.cell(row=total_row2, column=4, value=187450).number_format = '#,##0'
ws2.cell(row=total_row2, column=4).alignment = align_number()
ws2.cell(row=total_row2, column=5, value=25.3).number_format = '#,##0.0'
ws2.cell(row=total_row2, column=5).alignment = align_number()
ws2.cell(row=total_row2, column=6, value=1.0).number_format = '0.0%'
ws2.cell(row=total_row2, column=6).alignment = align_number()
style_total_row(ws2, total_row2, 2, 6)

# Auto-fit before footer
auto_fit_columns(ws2, header_row=table_start2, data_start_row=table_start2 + 1)

# ── Pie Chart: Revenue Share ──
pie1 = create_pie_chart(width=14, height=10)
pie_data = Reference(ws2, min_col=4, min_row=table_start2, max_row=table_start2 + len(sales_data))
pie_cats = Reference(ws2, min_col=2, min_row=table_start2 + 1, max_row=table_start2 + len(sales_data))
pie1.add_data(pie_data, titles_from_data=True)
pie1.set_categories(pie_cats)
setup_chart_titles(pie1, title="Revenue Share by Category")
apply_pie_colors(pie1, count=len(sales_data))
ws2.add_chart(pie1, "B13")

# ── Horizontal Bar Chart: Orders by Category ──
bar2 = create_bar_chart(chart_type="bar", grouping="clustered", gap_width=80, width=18, height=10)
orders_ref = Reference(ws2, min_col=3, min_row=table_start2, max_row=table_start2 + len(sales_data))
bar_cats = Reference(ws2, min_col=2, min_row=table_start2 + 1, max_row=table_start2 + len(sales_data))
bar2.add_data(orders_ref, titles_from_data=True)
bar2.set_categories(bar_cats)
setup_chart_titles(bar2, title="Orders by Category", x_title="Orders")
apply_chart_colors(bar2)
ws2.add_chart(bar2, "B28")

# ── Footer ──
add_footer(ws2, 6)


# ============================================================
# SHEET 3: Expense Analysis
# ============================================================
ws3 = wb.create_sheet("Expense Analysis")
setup_sheet(ws3, title="Expense Breakdown — January 2026", last_col=6)

table_start3 = 4
headers3 = ["Category", "Budget (AED)", "Actual (AED)", "Variance (AED)", "Status"]
for i, h in enumerate(headers3):
    ws3.cell(row=table_start3, column=2 + i, value=h)
style_header_row(ws3, table_start3, 2, 6)

expense_data = [
    ("Rent",                 35000, 35000,  0,     "On Track"),
    ("Staff Salaries",       42000, 43500, -1500,  "Over Budget"),
    ("Food Supplies",        28000, 26800,  1200,  "Under Budget"),
    ("Utilities",             8000,  9200, -1200,  "Over Budget"),
    ("Marketing",             5000,  4800,   200,  "Under Budget"),
    ("Maintenance",           3500,  3200,   300,  "Under Budget"),
    ("Delivery Commissions", 12000, 13100, -1100,  "Over Budget"),
    ("Miscellaneous",         6000,  7200, -1200,  "Over Budget"),
]

for idx, (cat, budget, actual, var, status) in enumerate(expense_data):
    r = table_start3 + 1 + idx
    ws3.cell(row=r, column=2, value=cat).alignment = align_text()
    ws3.cell(row=r, column=3, value=budget).number_format = '#,##0'
    ws3.cell(row=r, column=3).alignment = align_number()
    ws3.cell(row=r, column=4, value=actual).number_format = '#,##0'
    ws3.cell(row=r, column=4).alignment = align_number()
    ws3.cell(row=r, column=5, value=var).number_format = '#,##0'
    ws3.cell(row=r, column=5).alignment = align_number()
    ws3.cell(row=r, column=6, value=status).alignment = align_text()
    style_data_row(ws3, r, 2, 6, idx)

    # Conditional formatting for variance
    var_cell = ws3.cell(row=r, column=5)
    if var > 0:
        var_cell.fill = CF_POSITIVE_FILL
        var_cell.font = CF_POSITIVE_FONT
    elif var < 0:
        var_cell.fill = CF_NEGATIVE_FILL
        var_cell.font = CF_NEGATIVE_FONT

    # Conditional formatting for status
    status_cell = ws3.cell(row=r, column=6)
    if status == "Over Budget":
        status_cell.font = Font(name=FONT_NAME, size=11, color=ACCENT_NEGATIVE, bold=HEADER_BOLD)
    elif status == "Under Budget":
        status_cell.font = Font(name=FONT_NAME, size=11, color=ACCENT_POSITIVE, bold=HEADER_BOLD)
    else:
        status_cell.font = Font(name=FONT_NAME, size=11, color=NEUTRAL_600)

# Total row
total_row3 = table_start3 + 1 + len(expense_data)
ws3.cell(row=total_row3, column=2, value="TOTAL").alignment = align_text()
ws3.cell(row=total_row3, column=3, value=139500).number_format = '#,##0'
ws3.cell(row=total_row3, column=3).alignment = align_number()
ws3.cell(row=total_row3, column=4, value=142800).number_format = '#,##0'
ws3.cell(row=total_row3, column=4).alignment = align_number()
ws3.cell(row=total_row3, column=5, value=-3300).number_format = '#,##0'
ws3.cell(row=total_row3, column=5).alignment = align_number()
ws3.cell(row=total_row3, column=6, value="—").alignment = align_text()
style_total_row(ws3, total_row3, 2, 6)

# Conditional formatting for total variance
total_var_cell = ws3.cell(row=total_row3, column=5)
total_var_cell.fill = CF_NEGATIVE_FILL
total_var_cell.font = CF_NEGATIVE_FONT

# Auto-fit before footer
auto_fit_columns(ws3, header_row=table_start3, data_start_row=table_start3 + 1)

# ── Clustered Bar Chart: Budget vs Actual ──
chart3 = create_bar_chart(chart_type="col", grouping="clustered", gap_width=80, overlap=-20, width=18, height=10)
budget_actual_ref = Reference(ws3, min_col=3, min_row=table_start3, max_col=4, max_row=table_start3 + len(expense_data))
exp_cats = Reference(ws3, min_col=2, min_row=table_start3 + 1, max_row=table_start3 + len(expense_data))
chart3.add_data(budget_actual_ref, titles_from_data=True)
chart3.set_categories(exp_cats)
setup_chart_titles(chart3, title="Budget vs Actual Expenses", y_title="AED", x_title="Category")
apply_chart_colors(chart3)
ws3.add_chart(chart3, "B14")

# ── Footer ──
add_footer(ws3, 6)


# ============================================================
# SHEET 4: Recommendations
# ============================================================
ws4 = wb.create_sheet("Recommendations")
setup_sheet(ws4, title="Key Insights & Recommendations", last_col=4)

table_start4 = 4
headers4 = ["#", "Insight", "Priority"]
for i, h in enumerate(headers4):
    ws4.cell(row=table_start4, column=2 + i, value=h)
style_header_row(ws4, table_start4, 2, 4)

insights = [
    (1, "Profit margin improved from 16.2% to 23.8% over 4 months — strong upward trend", "High"),
    (2, "Shawarma has the highest order volume (1,560) but below-average ticket size (AED 24). Consider combo upselling.", "Medium"),
    (3, "Staff salaries exceeded budget by AED 1,500. Review overtime hours.", "High"),
    (4, "Delivery commissions are 9.2% over budget. Negotiate better rates with Talabat/Noon Food.", "High"),
    (5, "December peak revenue (AED 198,500) was seasonal. Plan promotions for slower months.", "Low"),
]

for idx, (num, insight, priority) in enumerate(insights):
    r = table_start4 + 1 + idx
    ws4.cell(row=r, column=2, value=num).alignment = Alignment(horizontal="center", vertical="center")
    ws4.cell(row=r, column=3, value=insight).alignment = align_text()
    ws4.cell(row=r, column=4, value=priority).alignment = Alignment(horizontal="center", vertical="center")
    style_data_row(ws4, r, 2, 4, idx)

    # Conditional formatting for Priority
    prio_cell = ws4.cell(row=r, column=4)
    if priority == "High":
        prio_cell.fill = PatternFill("solid", fgColor="FDEDEC")
        prio_cell.font = Font(name=FONT_NAME, size=11, color=ACCENT_NEGATIVE, bold=HEADER_BOLD)
    elif priority == "Medium":
        prio_cell.fill = PatternFill("solid", fgColor="FEF9E7")
        prio_cell.font = Font(name=FONT_NAME, size=11, color=ACCENT_WARNING, bold=HEADER_BOLD)
    elif priority == "Low":
        prio_cell.fill = PatternFill("solid", fgColor="E8F5E9")
        prio_cell.font = Font(name=FONT_NAME, size=11, color=ACCENT_POSITIVE, bold=HEADER_BOLD)

# Row heights for wrapped text
for idx in range(len(insights)):
    r = table_start4 + 1 + idx
    ws4.row_dimensions[r].height = 36

# Auto-fit before footer
auto_fit_columns(ws4, header_row=table_start4, data_start_row=table_start4 + 1)

# Override insight column width for readability
ws4.column_dimensions['C'].width = 60
ws4.column_dimensions['D'].width = 14
ws4.column_dimensions['B'].width = 8

# ── Footer ──
add_footer(ws4, 4, row=table_start4 + 1 + len(insights) + 2)


# ============================================================
# Save
# ============================================================
wb.properties.creator = "Z.ai"
output_path = "/home/z/my-project/download/sample_report_restaurant.xlsx"
wb.save(output_path)
print(f"Report saved to: {output_path}")
