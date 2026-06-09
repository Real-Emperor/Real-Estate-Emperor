"""
Real Estate Emperor - Data Import Template Generator
=====================================================
Creates a professional XLSX template with:
  Sheet 1: Instructions (English + Bengali)
  Sheet 2: Properties
  Sheet 3: Tenants
  Sheet 4: Payments
"""

import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from templates.base import (
    setup_sheet, style_header_row, style_data_row, style_total_row,
    font_title, font_header, font_subheader, font_body, font_caption,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_text, align_number, align_date,
    ROW_HEIGHTS, COLUMN_WIDTHS, FORMATS,
    PRIMARY, PRIMARY_LIGHT, SECONDARY, HEADER_TEXT,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_0,
)

# Override with Real Estate Emperor brand colors (green/gold Islamic theme)
BRAND_GREEN     = "1B6B3A"    # Deep Islamic green
BRAND_GREEN_LT  = "D4EDDA"    # Light green
BRAND_GOLD      = "C5960C"    # Rich gold
BRAND_GOLD_LT   = "FFF3CD"    # Light gold
BRAND_DARK      = "1A3C2A"    # Very dark green for headers
BRAND_WHITE     = "FFFFFF"
BRAND_LIGHT_BG  = "F8F9FA"    # Off-white alternating row
BRAND_WARM_BG   = "F0F7F1"    # Warm green-tinted alternating row

# Custom fills
fill_brand_header = PatternFill("solid", fgColor=BRAND_DARK)
fill_brand_gold   = PatternFill("solid", fgColor=BRAND_GOLD_LT)
fill_brand_green  = PatternFill("solid", fgColor=BRAND_GREEN_LT)
fill_alt_row      = PatternFill("solid", fgColor=BRAND_WARM_BG)
fill_white        = PatternFill("solid", fgColor=BRAND_WHITE)
fill_section_en   = PatternFill("solid", fgColor="E8F5E9")   # Light green for EN sections
fill_section_bn   = PatternFill("solid", fgColor="FFF8E1")   # Light gold for BN sections

# Custom fonts
font_brand_title   = Font(name="Calibri", size=18, bold=True, color=BRAND_GREEN)
font_brand_header  = Font(name="Calibri", size=11, bold=True, color=BRAND_WHITE)
font_brand_body    = Font(name="Calibri", size=11, color=NEUTRAL_900)
font_brand_caption = Font(name="Calibri", size=9, color=NEUTRAL_600)
font_brand_section = Font(name="Calibri", size=13, bold=True, color=BRAND_GREEN)
font_brand_note    = Font(name="Calibri", size=10, italic=True, color=BRAND_GOLD)
font_brand_bold    = Font(name="Calibri", size=11, bold=True, color=NEUTRAL_900)
font_brand_gold_h  = Font(name="Calibri", size=11, bold=True, color=BRAND_GOLD)

# Bengali font (will use system fallback on recipient's machine)
font_bn_body = Font(name="Noto Sans Bengali, Calibri", size=11, color=NEUTRAL_900)
font_bn_section = Font(name="Noto Sans Bengali, Calibri", size=13, bold=True, color=BRAND_GREEN)
font_bn_note = Font(name="Noto Sans Bengali, Calibri", size=10, italic=True, color=BRAND_GOLD)
font_bn_bold = Font(name="Noto Sans Bengali, Calibri", size=11, bold=True, color=NEUTRAL_900)
font_bn_caption = Font(name="Noto Sans Bengali, Calibri", size=9, color=NEUTRAL_600)

# Borders
thin_bottom = Border(bottom=Side(style="thin", color=NEUTRAL_200))
medium_top  = Border(top=Side(style="medium", color=BRAND_GREEN))
gold_bottom = Border(bottom=Side(style="medium", color=BRAND_GOLD))

# Alignments
align_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_right_wrap  = Alignment(horizontal="right", vertical="center", wrap_text=True)

wb = Workbook()

# ============================================================
# SHEET 1: INSTRUCTIONS
# ============================================================
ws1 = wb.active
ws1.title = "Instructions"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = BRAND_GREEN

# Column widths
ws1.column_dimensions["A"].width = 3    # margin
ws1.column_dimensions["B"].width = 5    # #
ws1.column_dimensions["C"].width = 35   # English
ws1.column_dimensions["D"].width = 35   # Bengali
ws1.column_dimensions["E"].width = 5    # #
ws1.column_dimensions["F"].width = 35   # English
ws1.column_dimensions["G"].width = 35   # Bengali

# ---- HEADER BANNER ----
row = 1
ws1.row_dimensions[row].height = 15  # margin

row = 2
ws1.merge_cells("B2:G2")
cell = ws1.cell(row=2, column=2, value="Real Estate Emperor - Data Import Template")
cell.font = Font(name="Calibri", size=20, bold=True, color=BRAND_WHITE)
cell.fill = PatternFill("solid", fgColor=BRAND_GREEN)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 45

row = 3
ws1.merge_cells("B3:G3")
cell = ws1.cell(row=3, column=2, value="الإمبراطور العقاري لإدارة الممتلكات ذ.م.م")
cell.font = Font(name="Calibri", size=14, bold=False, color=BRAND_GOLD)
cell.fill = PatternFill("solid", fgColor=BRAND_GREEN)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[3].height = 30

row = 4
ws1.merge_cells("B4:G4")
cell = ws1.cell(row=4, column=2, value="Data Import Instructions | ডাটা ইমপোর্ট নির্দেশনা")
cell.font = Font(name="Calibri", size=12, bold=True, color=BRAND_WHITE)
cell.fill = PatternFill("solid", fgColor=BRAND_DARK)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[4].height = 28

row = 5
ws1.row_dimensions[5].height = 12  # spacer

# ---- SECTION: GENERAL OVERVIEW ----
row = 6
ws1.merge_cells("B6:G6")
cell = ws1.cell(row=6, column=2, value="SECTION 1: GENERAL OVERVIEW | বিভাগ ১: সাধারণ ভূমিকা")
cell.font = font_brand_section
cell.fill = fill_brand_green
cell.alignment = align_center_wrap
cell.border = gold_bottom
ws1.row_dimensions[6].height = 28

# EN/BN header
row = 7
ws1.merge_cells("B7:C7")
cell = ws1.cell(row=7, column=2, value="English")
cell.font = font_brand_header
cell.fill = PatternFill("solid", fgColor=BRAND_GREEN)
cell.alignment = align_center_wrap

ws1.merge_cells("D7:D7")
cell = ws1.cell(row=7, column=4, value="")
cell.fill = PatternFill("solid", fgColor=BRAND_GREEN)

ws1.merge_cells("E7:F7")
cell = ws1.cell(row=7, column=5, value="বাংলা (Bengali)")
cell.font = font_brand_header
cell.fill = PatternFill("solid", fgColor=BRAND_GOLD)
cell.alignment = align_center_wrap

ws1.cell(row=7, column=7).fill = PatternFill("solid", fgColor=BRAND_GOLD)
ws1.row_dimensions[7].height = 25

overview_data = [
    (
        "This Excel file is a template for importing real estate data into the Real Estate Emperor Dashboard. "
        "Please fill in your actual property, tenant, and payment information in the sheets provided. "
        "The dashboard will then be populated with your real data replacing the fictional examples currently shown.",
        (
            "এই এক্সেল ফাইলটি রিয়েল এস্টেট এম্পেরর প্রপার্টি ড্যাশবোর্ডে রিয়েল এস্টেট ডাটা আমদানি করার একটি টেমপ্লেট। "
            "অনুগ্রহ করে প্রদত্ত শীটগুলিতে আপনার প্রকৃত সম্পত্তি, ভাড়াটে এবং পেমেন্ট তথ্য পূরণ করুন। "
            "ড্যাশবোর্ডে তখন বর্তমানে প্রদর্শিত কাল্পনিক উদাহরণগুলির পরিবর্তে আপনার প্রকৃত ডাটা দেখানো হবে।"
        )
    ),
    (
        "There are 3 data sheets in this file: Properties, Tenants, and Payments. "
        "Each sheet has column headers explaining what information is needed. "
        "Columns marked with a * (asterisk) are mandatory and must be filled. "
        "Columns without * are optional but recommended for a complete dashboard.",
        (
            "এই ফাইলে ৩টি ডাটা শীট আছে: Properties (সম্পত্তি), Tenants (ভাড়াটে), এবং Payments (পেমেন্ট)। "
            "প্রতিটি শীটে কলাম হেডার রয়েছে যা ব্যাখ্যা করে কোন তথ্য প্রয়োজন। "
            "* (তারকাচিহ্ন) দিয়ে চিহ্নিত কলামগুলি বাধ্যতামূলক এবং অবশ্যই পূরণ করতে হবে। "
            "* ছাড়া কলামগুলি ঐচ্ছিক কিন্তু সম্পূর্ণ ড্যাশবোর্ডের জন্য সুপারিশ করা হয়।"
        )
    ),
    (
        "Do NOT rename the sheet tabs or column headers, as the import system relies on these exact names. "
        "Simply fill in the data rows below the headers. You may add as many rows as needed. "
        "Do NOT change the order of columns.",
        (
            "শীট ট্যাব বা কলাম হেডারের নাম পরিবর্তন করবেন না, কারণ আমদানি সিস্টেম এই সঠিক নামগুলির উপর নির্ভর করে। "
            "হেডারের নিচের ডাটা সারিগুলিতে সহজেই পূরণ করুন। আপনার প্রয়োজন অনুযায়ী যত সারি যোগ করতে পারেন। "
            "কলামের ক্রম পরিবর্তন করবেন না।"
        )
    ),
]

for i, (en_text, bn_text) in enumerate(overview_data):
    r = 8 + i
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    cell = ws1.cell(row=r, column=2, value=en_text)
    cell.font = font_brand_body
    cell.fill = fill_section_en if i % 2 == 0 else fill_white
    cell.alignment = align_wrap

    ws1.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    cell = ws1.cell(row=r, column=5, value=bn_text)
    cell.font = font_bn_body
    cell.fill = fill_section_bn if i % 2 == 0 else fill_white
    cell.alignment = align_wrap

    # Column G (spacer between EN/BN) - style only
    ws1.cell(row=r, column=4).fill = fill_white
    ws1.cell(row=r, column=7).fill = fill_white

    ws1.row_dimensions[r].height = 55

row = 11
ws1.row_dimensions[11].height = 12  # spacer

# ---- SECTION: DATA FORMATS ----
row = 12
ws1.merge_cells("B12:G12")
cell = ws1.cell(row=12, column=2, value="SECTION 2: DATA FORMATS & RULES | বিভাগ ২: ডাটা ফরম্যাট ও নিয়ম")
cell.font = font_brand_section
cell.fill = fill_brand_green
cell.alignment = align_center_wrap
cell.border = gold_bottom
ws1.row_dimensions[12].height = 28

# Format rules - EN/BN
row = 13
ws1.merge_cells("B13:C13")
cell = ws1.cell(row=13, column=2, value="Field Rule (English)")
cell.font = font_brand_header
cell.fill = PatternFill("solid", fgColor=BRAND_GREEN)
cell.alignment = align_center_wrap

ws1.cell(row=13, column=4).fill = PatternFill("solid", fgColor=BRAND_GREEN)

ws1.merge_cells("E13:F13")
cell = ws1.cell(row=13, column=5, value="নিয়ম (বাংলা)")
cell.font = font_brand_header
cell.fill = PatternFill("solid", fgColor=BRAND_GOLD)
cell.alignment = align_center_wrap
ws1.cell(row=13, column=7).fill = PatternFill("solid", fgColor=BRAND_GOLD)
ws1.row_dimensions[13].height = 25

format_rules = [
    (
        "Dates: Use YYYY-MM-DD format only. Example: 2025-06-15",
        "তারিখ: শুধুমাত্র YYYY-MM-DD ফরম্যাট ব্যবহার করুন। উদাহরণ: 2025-06-15"
    ),
    (
        "Amounts: Numbers only, no commas or currency symbols. Example: 45000 (not 45,000 or 45,000 AED)",
        "পরিমাণ: শুধুমাত্র সংখ্যা, কোনো কমা বা মুদ্রা প্রতীক নয়। উদাহরণ: 45000 (45000 AED বা 45,000 নয়)"
    ),
    (
        "Phone Numbers: UAE format starting with 05. Example: 0501234567 (the system will add +971 automatically)",
        "ফোন নম্বর: 05 দিয়ে শুরু UAE ফরম্যাট। উদাহরণ: 0501234567 (সিস্টেম স্বয়ংক্রিয়ভাবে +971 যোগ করবে)"
    ),
    (
        "Tenant Names: English name is MANDATORY. Arabic name is OPTIONAL but recommended for official records.",
        "ভাড়াটে নাম: ইংরেজি নাম বাধ্যতামূলক। আরবি নাম ঐচ্ছিক কিন্তু অফিসিয়াল রেকর্ডের জন্য সুপারিশ করা হয়।"
    ),
    (
        "Status Values: Use only the allowed values shown in dropdown lists. Do not type custom statuses.",
        "স্ট্যাটাস মান: ড্রপডাউন তালিকায় দেখানো অনুমোদিত মানগুলি ব্যবহার করুন। কাস্টম স্ট্যাটাস টাইপ করবেন না।"
    ),
    (
        "Property Types: Residential, Commercial, Industrial, Mixed Use",
        "সম্পত্তির ধরন: Residential (আবাসিক), Commercial (বাণিজ্যিক), Industrial (শিল্প), Mixed Use (মিশ্র ব্যবহার)"
    ),
    (
        "Payment Methods: Cash, Bank Transfer, Cheque, Online",
        "পেমেন্ট পদ্ধতি: Cash (নগদ), Bank Transfer (ব্যাংক ট্রান্সফার), Cheque (চেক), Online (অনলাইন)"
    ),
    (
        "Payment Status: Paid, Partial, Pending, Overdue",
        "পেমেন্ট স্ট্যাটাস: Paid (পরিশোধিত), Partial (আংশিক), Pending (বাকি), Overdue (মেয়াদোত্তীর্ণ)"
    ),
]

for i, (en_text, bn_text) in enumerate(format_rules):
    r = 14 + i
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    cell = ws1.cell(row=r, column=2, value=en_text)
    cell.font = font_brand_body
    cell.fill = fill_section_en if i % 2 == 0 else fill_white
    cell.alignment = align_wrap

    ws1.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    cell = ws1.cell(row=r, column=5, value=bn_text)
    cell.font = font_bn_body
    cell.fill = fill_section_bn if i % 2 == 0 else fill_white
    cell.alignment = align_wrap

    ws1.cell(row=r, column=4).fill = fill_white
    ws1.cell(row=r, column=7).fill = fill_white
    ws1.row_dimensions[r].height = 35

row = 22
ws1.row_dimensions[22].height = 12  # spacer

# ---- SECTION: SHEET DESCRIPTIONS ----
row = 23
ws1.merge_cells("B23:G23")
cell = ws1.cell(row=23, column=2, value="SECTION 3: SHEET DESCRIPTIONS | বিভাগ ৩: শীট বিবরণ")
cell.font = font_brand_section
cell.fill = fill_brand_green
cell.alignment = align_center_wrap
cell.border = gold_bottom
ws1.row_dimensions[23].height = 28

# Sheet descriptions table
row = 24
headers_sheet_desc = ["Sheet Name", "Description (English)", "বিবরণ (বাংলা)"]
cols_sd = [(2, 2), (3, 5), (6, 7)]
for (cs, ce), hdr in zip(cols_sd, headers_sheet_desc):
    if cs != ce:
        ws1.merge_cells(start_row=24, start_column=cs, end_row=24, end_column=ce)
    cell = ws1.cell(row=24, column=cs, value=hdr)
    cell.font = font_brand_header
    cell.fill = PatternFill("solid", fgColor=BRAND_DARK)
    cell.alignment = align_center_wrap
    # Fill all cells in merge
    for c in range(cs, ce + 1):
        ws1.cell(row=24, column=c).fill = PatternFill("solid", fgColor=BRAND_DARK)
ws1.row_dimensions[24].height = 25

sheet_descs = [
    (
        "Properties",
        "List all buildings, villas, or units you own/manage. Include name, type, location, total units, and status.",
        "আপনার মালিকানাধীন/পরিচালিত সকল ভবন, ভিলা বা ইউনিটের তালিকা করুন। নাম, ধরন, অবস্থান, মোট ইউনিট এবং স্ট্যাটাস অন্তর্ভুক্ত করুন।"
    ),
    (
        "Tenants",
        "List all tenants with their details. English name is mandatory, Arabic name optional. Link each tenant to a property.",
        "সকল ভাড়াটেদের বিস্তারিত সহ তালিকা করুন। ইংরেজি নাম বাধ্যতামূলক, আরবি নাম ঐচ্ছিক। প্রতিটি ভাড়াটেকে একটি সম্পত্তির সাথে লিঙ্ক করুন।"
    ),
    (
        "Payments",
        "Record rent payment history. Link each payment to a tenant and property. Include amount, date, method, and status.",
        "ভাড়া পেমেন্টের ইতিহাস রেকর্ড করুন। প্রতিটি পেমেন্টকে একটি ভাড়াটে এবং সম্পত্তির সাথে লিঙ্ক করুন। পরিমাণ, তারিখ, পদ্ধতি এবং স্ট্যাটাস অন্তর্ভুক্ত করুন।"
    ),
]

for i, (sheet_name, en_desc, bn_desc) in enumerate(sheet_descs):
    r = 25 + i
    cell = ws1.cell(row=r, column=2, value=sheet_name)
    cell.font = font_brand_bold
    cell.fill = fill_section_en if i % 2 == 0 else fill_white
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    cell = ws1.cell(row=r, column=3, value=en_desc)
    cell.font = font_brand_body
    cell.fill = fill_section_en if i % 2 == 0 else fill_white
    cell.alignment = align_wrap

    ws1.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    cell = ws1.cell(row=r, column=6, value=bn_desc)
    cell.font = font_bn_body
    cell.fill = fill_section_bn if i % 2 == 0 else fill_white
    cell.alignment = align_wrap

    ws1.row_dimensions[r].height = 40

row = 28
ws1.row_dimensions[28].height = 12  # spacer

# ---- SECTION: IMPORTANT NOTES ----
row = 29
ws1.merge_cells("B29:G29")
cell = ws1.cell(row=29, column=2, value="SECTION 4: IMPORTANT NOTES | বিভাগ ৪: গুরুত্বপূর্ণ নোট")
cell.font = font_brand_section
cell.fill = fill_brand_green
cell.alignment = align_center_wrap
cell.border = gold_bottom
ws1.row_dimensions[29].height = 28

notes_data = [
    (
        "If a tenant does not have an Arabic name, leave the Arabic name column blank. English name is always required.",
        "কোনো ভাড়াটের আরবি নাম না থাকলে, আরবি নাম কলামটি খালি রাখুন। ইংরেজি নাম সবসময় প্রয়োজন।"
    ),
    (
        "Property names in the Tenants sheet must exactly match the property names in the Properties sheet.",
        "ভাড়াটে শীটের সম্পত্তির নাম অবশ্যই Properties শীটের সম্পত্তির নামের সাথে হুবহু মিলতে হবে।"
    ),
    (
        "For tenants who pay rent in multiple installments, create a separate payment row for each installment.",
        "যে ভাড়াটেরা একাধিক কিস্তিতে ভাড়া দেয়, তাদের জন্য প্রতিটি কিস্তির জন্য আলাদা পেমেন্ট সারি তৈরি করুন।"
    ),
    (
        "If you have existing data in another spreadsheet, send it as-is. We can map your columns automatically.",
        "আপনার অন্য স্প্রেডশীটে বিদ্যমান ডাটা থাকলে, সেটি যেমন আছে পাঠান। আমরা আপনার কলামগুলি স্বয়ংক্রিয়ভাবে ম্যাপ করতে পারি।"
    ),
    (
        "Do NOT merge cells, use formulas, or add images in the data sheets. Keep data clean and simple.",
        "ডাটা শীটে সেল মার্জ করবেন না, সূত্র ব্যবহার করবেন না, বা ছবি যোগ করবেন না। ডাটা পরিষ্কার ও সহজ রাখুন।"
    ),
    (
        "Once completed, save the file and send it back. The dashboard will be updated with your real data within 24 hours.",
        "সম্পন্ন হলে ফাইলটি সেভ করে ফেরত পাঠান। ২৪ ঘন্টার মধ্যে আপনার প্রকৃত ডাটা দিয়ে ড্যাশবোর্ড আপডেট করা হবে।"
    ),
]

row = 30
ws1.merge_cells("B30:C30")
cell = ws1.cell(row=30, column=2, value="Important Note (English)")
cell.font = font_brand_header
cell.fill = PatternFill("solid", fgColor=BRAND_GREEN)
cell.alignment = align_center_wrap
ws1.cell(row=30, column=4).fill = PatternFill("solid", fgColor=BRAND_GREEN)

ws1.merge_cells("E30:F30")
cell = ws1.cell(row=30, column=5, value="গুরুত্বপূর্ণ নোট (বাংলা)")
cell.font = font_brand_header
cell.fill = PatternFill("solid", fgColor=BRAND_GOLD)
cell.alignment = align_center_wrap
ws1.cell(row=30, column=7).fill = PatternFill("solid", fgColor=BRAND_GOLD)
ws1.row_dimensions[30].height = 25

for i, (en_text, bn_text) in enumerate(notes_data):
    r = 31 + i
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    cell = ws1.cell(row=r, column=2, value=f"{i+1}. {en_text}")
    cell.font = font_brand_body
    cell.fill = fill_section_en if i % 2 == 0 else fill_white
    cell.alignment = align_wrap

    ws1.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    cell = ws1.cell(row=r, column=5, value=f"{i+1}. {bn_text}")
    cell.font = font_bn_body
    cell.fill = fill_section_bn if i % 2 == 0 else fill_white
    cell.alignment = align_wrap

    ws1.cell(row=r, column=4).fill = fill_white
    ws1.cell(row=r, column=7).fill = fill_white
    ws1.row_dimensions[r].height = 40

# Footer
row = 37
ws1.row_dimensions[37].height = 12
row = 38
ws1.merge_cells("B38:G38")
cell = ws1.cell(row=38, column=2, value="Real Estate Emperor Property Management L.L.C. | Prepared for Shafiul Azam")
cell.font = Font(name="Calibri", size=10, italic=True, color=BRAND_GOLD)
cell.fill = PatternFill("solid", fgColor=BRAND_DARK)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[38].height = 25

# Print setup
ws1.print_area = "A1:G38"
ws1.page_setup.orientation = 'landscape'
ws1.page_setup.fitToWidth = 1

# ============================================================
# SHEET 2: PROPERTIES
# ============================================================
ws2 = wb.create_sheet("Properties")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = BRAND_GREEN

# Column definitions
prop_cols = [
    ("A", 3, None),    # margin
    ("B", 8, "#"),     # ID
    ("C", 25, "Property Name (EN) *"),   # mandatory
    ("D", 25, "Property Name (AR)"),      # optional
    ("E", 16, "Type *"),                  # mandatory
    ("F", 10, "Units *"),                 # mandatory
    ("G", 20, "Location *"),              # mandatory
    ("H", 14, "Status *"),                # mandatory
]

for col_letter, width, _ in prop_cols:
    ws2.column_dimensions[col_letter].width = width

# Row 1: margin
ws2.row_dimensions[1].height = ROW_HEIGHTS["margin"]

# Row 2: Title
ws2.merge_cells("B2:H2")
cell = ws2.cell(row=2, column=2, value="Properties | সম্পত্তি")
cell.font = font_brand_title
cell.alignment = align_title()
ws2.row_dimensions[2].height = ROW_HEIGHTS["title"]

# Row 3: spacer
ws2.row_dimensions[3].height = ROW_HEIGHTS["spacer"]

# Row 4: Headers
for col_idx, (_, _, header) in enumerate(prop_cols):
    if header is None:
        continue
    cell = ws2.cell(row=4, column=col_idx + 1, value=header)
    cell.font = font_brand_header
    cell.fill = PatternFill("solid", fgColor=BRAND_GREEN)
    cell.alignment = align_center_wrap
    cell.border = Border(bottom=Side(style="thin", color=BRAND_GOLD))
ws2.row_dimensions[4].height = ROW_HEIGHTS["header"]

# Sample data rows
prop_samples = [
    [1, "Emperor Tower 1", "برج الإمبراطور 1", "Residential", 24, "Abu Dhabi", "Active"],
    [2, "Souk Commercial Center", "سوق المركز التجاري", "Commercial", 12, "Dubai", "Active"],
    [3, "Al Fahad Residence", "مسكن الفهد", "Residential", 16, "Sharjah", "Active"],
    [4, "Industrial Warehouse Zone", "منطقة المستودعات الصناعية", "Industrial", 8, "Ajman", "Maintenance"],
    [5, "Marina Mixed Use Complex", "مجمع مارينا متعدد الاستخدامات", "Mixed Use", 30, "Abu Dhabi", "Active"],
]

for i, row_data in enumerate(prop_samples):
    r = 5 + i
    for j, val in enumerate(row_data):
        cell = ws2.cell(row=r, column=j + 2, value=val)
        cell.font = font_brand_body
        cell.fill = fill_alt_row if i % 2 == 0 else fill_white
        cell.alignment = align_text() if j > 0 else Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[r].height = ROW_HEIGHTS["data"]

# Add a few empty rows with styling for user input
for i in range(5):
    r = 10 + i
    for j in range(len(prop_cols) - 1):
        cell = ws2.cell(row=r, column=j + 2)
        cell.fill = fill_alt_row if i % 2 == 0 else fill_white
        cell.font = font_brand_body
        cell.border = Border(bottom=Side(style="hair", color=NEUTRAL_200))
    ws2.row_dimensions[r].height = ROW_HEIGHTS["data"]

# Data validations
dv_type = DataValidation(type="list", formula1='"Residential,Commercial,Industrial,Mixed Use"', allow_blank=False)
dv_type.error = "Please select a valid property type"
dv_type.errorTitle = "Invalid Type"
ws2.add_data_validation(dv_type)
dv_type.add(f"E5:E14")

dv_status = DataValidation(type="list", formula1='"Active,Inactive,Maintenance,Vacant"', allow_blank=False)
dv_status.error = "Please select a valid status"
dv_status.errorTitle = "Invalid Status"
ws2.add_data_validation(dv_status)
dv_status.add(f"H5:H14")

# Freeze panes
ws2.freeze_panes = "B5"

# ============================================================
# SHEET 3: TENANTS
# ============================================================
ws3 = wb.create_sheet("Tenants")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = BRAND_GOLD

tenant_cols = [
    ("A", 3, None),    # margin
    ("B", 8, "#"),
    ("C", 22, "Tenant Name (EN) *"),    # mandatory
    ("D", 22, "Tenant Name (AR)"),       # optional
    ("E", 16, "Phone *"),                # mandatory
    ("F", 25, "Email"),                   # optional
    ("G", 22, "Property *"),             # mandatory - must match Properties sheet
    ("H", 10, "Unit No. *"),             # mandatory
    ("I", 14, "Rent (AED) *"),           # mandatory
    ("J", 14, "Lease Start *"),          # mandatory
    ("K", 14, "Lease End *"),            # mandatory
    ("L", 12, "Status *"),               # mandatory
]

for col_letter, width, _ in tenant_cols:
    ws3.column_dimensions[col_letter].width = width

# Row 1: margin
ws3.row_dimensions[1].height = ROW_HEIGHTS["margin"]

# Row 2: Title
ws3.merge_cells("B2:L2")
cell = ws3.cell(row=2, column=2, value="Tenants | ভাড়াটে")
cell.font = font_brand_title
cell.alignment = align_title()
ws3.row_dimensions[2].height = ROW_HEIGHTS["title"]

# Row 3: spacer
ws3.row_dimensions[3].height = ROW_HEIGHTS["spacer"]

# Row 4: Headers
for col_idx, (_, _, header) in enumerate(tenant_cols):
    if header is None:
        continue
    cell = ws3.cell(row=4, column=col_idx + 1, value=header)
    cell.font = font_brand_header
    cell.fill = PatternFill("solid", fgColor=BRAND_GOLD)
    cell.font = Font(name="Calibri", size=11, bold=True, color=BRAND_WHITE)
    cell.alignment = align_center_wrap
    cell.border = Border(bottom=Side(style="thin", color=BRAND_GREEN))
ws3.row_dimensions[4].height = ROW_HEIGHTS["header"]

# Sample data
tenant_samples = [
    [1, "Mohammed Al Rashid", "محمد الراشد", "0501234567", "mohammed@email.com", "Emperor Tower 1", "101", 45000, "2025-01-01", "2025-12-31", "Active"],
    [2, "Ahmad Trading LLC", "شركة أحمد التجارية", "0559876543", "info@ahmadtrading.ae", "Souk Commercial Center", "S-5", 120000, "2025-03-01", "2026-02-28", "Active"],
    [3, "Fatima Al Zaabi", "فاطمة الزعابي", "0521112233", "", "Al Fahad Residence", "3A", 38000, "2025-06-01", "2026-05-31", "Active"],
    [4, "Khalid Construction", "مؤسسة خالد للمقاولات", "0543334455", "khalid@construction.ae", "Industrial Warehouse Zone", "W-2", 65000, "2025-01-15", "2025-07-14", "Overdue"],
    [5, "Al Noor Electronics", "النور للإلكترونيات", "0567778899", "alnoor@electronics.ae", "Marina Mixed Use Complex", "M-12", 95000, "2025-04-01", "2026-03-31", "Active"],
]

for i, row_data in enumerate(tenant_samples):
    r = 5 + i
    for j, val in enumerate(row_data):
        cell = ws3.cell(row=r, column=j + 2, value=val)
        cell.font = font_brand_body
        cell.fill = fill_alt_row if i % 2 == 0 else fill_white
        if j == 0:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif j == 8:  # Rent amount
            cell.number_format = '#,##0'
            cell.alignment = align_number()
        elif j in (9, 10):  # Dates
            cell.alignment = align_date()
        else:
            cell.alignment = align_text()
    ws3.row_dimensions[r].height = ROW_HEIGHTS["data"]

# Empty rows for input
for i in range(5):
    r = 10 + i
    for j in range(len(tenant_cols) - 1):
        cell = ws3.cell(row=r, column=j + 2)
        cell.fill = fill_alt_row if i % 2 == 0 else fill_white
        cell.font = font_brand_body
        cell.border = Border(bottom=Side(style="hair", color=NEUTRAL_200))
    ws3.row_dimensions[r].height = ROW_HEIGHTS["data"]

# Data validations
dv_tenant_status = DataValidation(type="list", formula1='"Active,Inactive,Vacant,Overdue"', allow_blank=False)
dv_tenant_status.error = "Please select a valid status"
ws3.add_data_validation(dv_tenant_status)
dv_tenant_status.add("L5:L14")

# Phone validation (text starting with 05)
dv_phone = DataValidation(type="textLength", operator="greaterThanOrEqual", formula1="10")
dv_phone.error = "UAE phone number should start with 05 and be at least 10 digits"
dv_phone.errorTitle = "Invalid Phone"
ws3.add_data_validation(dv_phone)
dv_phone.add("E5:E14")

ws3.freeze_panes = "B5"

# ============================================================
# SHEET 4: PAYMENTS
# ============================================================
ws4 = wb.create_sheet("Payments")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = ACCENT_POSITIVE

pay_cols = [
    ("A", 3, None),    # margin
    ("B", 8, "#"),
    ("C", 22, "Tenant Name (EN) *"),    # mandatory
    ("D", 22, "Property *"),             # mandatory
    ("E", 10, "Unit *"),                 # mandatory
    ("F", 14, "Amount (AED) *"),         # mandatory
    ("G", 14, "Date *"),                 # mandatory
    ("H", 16, "Method *"),               # mandatory
    ("I", 12, "Status *"),               # mandatory
    ("J", 25, "Notes"),                   # optional
]

for col_letter, width, _ in pay_cols:
    ws4.column_dimensions[col_letter].width = width

# Row 1: margin
ws4.row_dimensions[1].height = ROW_HEIGHTS["margin"]

# Row 2: Title
ws4.merge_cells("B2:J2")
cell = ws4.cell(row=2, column=2, value="Payments | পেমেন্ট")
cell.font = font_brand_title
cell.alignment = align_title()
ws4.row_dimensions[2].height = ROW_HEIGHTS["title"]

# Row 3: spacer
ws4.row_dimensions[3].height = ROW_HEIGHTS["spacer"]

# Row 4: Headers
for col_idx, (_, _, header) in enumerate(pay_cols):
    if header is None:
        continue
    cell = ws4.cell(row=4, column=col_idx + 1, value=header)
    cell.font = font_brand_header
    cell.fill = PatternFill("solid", fgColor=ACCENT_POSITIVE)
    cell.font = Font(name="Calibri", size=11, bold=True, color=BRAND_WHITE)
    cell.alignment = align_center_wrap
    cell.border = Border(bottom=Side(style="thin", color=BRAND_GOLD))
ws4.row_dimensions[4].height = ROW_HEIGHTS["header"]

# Sample data
payment_samples = [
    [1, "Mohammed Al Rashid", "Emperor Tower 1", "101", 45000, "2025-01-05", "Bank Transfer", "Paid", "Annual rent - January"],
    [2, "Ahmad Trading LLC", "Souk Commercial Center", "S-5", 60000, "2025-03-10", "Cheque", "Partial", "First half payment"],
    [3, "Fatima Al Zaabi", "Al Fahad Residence", "3A", 38000, "2025-06-01", "Cash", "Paid", "Full payment received"],
    [4, "Khalid Construction", "Industrial Warehouse Zone", "W-2", 32500, "2025-01-20", "Bank Transfer", "Overdue", "Half payment - remainder overdue"],
    [5, "Al Noor Electronics", "Marina Mixed Use Complex", "M-12", 95000, "2025-04-03", "Online", "Paid", "Full annual payment"],
]

for i, row_data in enumerate(payment_samples):
    r = 5 + i
    for j, val in enumerate(row_data):
        cell = ws4.cell(row=r, column=j + 2, value=val)
        cell.font = font_brand_body
        cell.fill = fill_alt_row if i % 2 == 0 else fill_white
        if j == 0:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif j == 5:  # Amount
            cell.number_format = '#,##0'
            cell.alignment = align_number()
        elif j == 6:  # Date
            cell.alignment = align_date()
        else:
            cell.alignment = align_text()
    ws4.row_dimensions[r].height = ROW_HEIGHTS["data"]

# Empty rows
for i in range(5):
    r = 10 + i
    for j in range(len(pay_cols) - 1):
        cell = ws4.cell(row=r, column=j + 2)
        cell.fill = fill_alt_row if i % 2 == 0 else fill_white
        cell.font = font_brand_body
        cell.border = Border(bottom=Side(style="hair", color=NEUTRAL_200))
    ws4.row_dimensions[r].height = ROW_HEIGHTS["data"]

# Data validations
dv_method = DataValidation(type="list", formula1='"Cash,Bank Transfer,Cheque,Online"', allow_blank=False)
dv_method.error = "Please select a valid payment method"
ws4.add_data_validation(dv_method)
dv_method.add("H5:H14")

dv_pay_status = DataValidation(type="list", formula1='"Paid,Partial,Pending,Overdue"', allow_blank=False)
dv_pay_status.error = "Please select a valid payment status"
ws4.add_data_validation(dv_pay_status)
dv_pay_status.add("I5:I14")

ws4.freeze_panes = "B5"

# ============================================================
# CONDITIONAL FORMATTING for Payment Status
# ============================================================
from openpyxl.formatting.rule import CellIsRule

# Green for Paid
ws4.conditional_formatting.add('I5:I14',
    CellIsRule(operator='equal', formula=['"Paid"'],
              fill=PatternFill(bgColor="C6EFCE"),
              font=Font(color="006100")))

# Red for Overdue
ws4.conditional_formatting.add('I5:I14',
    CellIsRule(operator='equal', formula=['"Overdue"'],
              fill=PatternFill(bgColor="FFC7CE"),
              font=Font(color="9C0006")))

# Yellow for Partial/Pending
ws4.conditional_formatting.add('I5:I14',
    CellIsRule(operator='equal', formula=['"Partial"'],
              fill=PatternFill(bgColor="FFEB9C"),
              font=Font(color="9C6500")))
ws4.conditional_formatting.add('I5:I14',
    CellIsRule(operator='equal', formula=['"Pending"'],
              fill=PatternFill(bgColor="FFEB9C"),
              font=Font(color="9C6500")))

# Same conditional formatting for Tenant status
ws3.conditional_formatting.add('L5:L14',
    CellIsRule(operator='equal', formula=['"Active"'],
              fill=PatternFill(bgColor="C6EFCE"),
              font=Font(color="006100")))
ws3.conditional_formatting.add('L5:L14',
    CellIsRule(operator='equal', formula=['"Overdue"'],
              fill=PatternFill(bgColor="FFC7CE"),
              font=Font(color="9C0006")))
ws3.conditional_formatting.add('L5:L14',
    CellIsRule(operator='equal', formula=['"Vacant"'],
              fill=PatternFill(bgColor="FFEB9C"),
              font=Font(color="9C6500")))

# Property status conditional formatting
ws2.conditional_formatting.add('H5:H14',
    CellIsRule(operator='equal', formula=['"Active"'],
              fill=PatternFill(bgColor="C6EFCE"),
              font=Font(color="006100")))
ws2.conditional_formatting.add('H5:H14',
    CellIsRule(operator='equal', formula=['"Maintenance"'],
              fill=PatternFill(bgColor="FFEB9C"),
              font=Font(color="9C6500")))
ws2.conditional_formatting.add('H5:H14',
    CellIsRule(operator='equal', formula=['"Inactive"'],
              fill=PatternFill(bgColor="FFC7CE"),
              font=Font(color="9C0006")))

# ============================================================
# SAVE
# ============================================================
output_path = "/home/z/my-project/download/Real_Estate_Emperor_Data_Template.xlsx"
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Template saved to: {output_path}")
