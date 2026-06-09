#!/usr/bin/env python3
"""
Generate POS_Basic_Sample.xlsx — Professional POS Sample with Demo Data
Uses warm palette and base.py styling system.
Restaurant/Cafeteria demo data for client presentations.
"""

import sys
import random
from datetime import date, timedelta

sys.path.insert(0, "/home/z/my-project/skills/xlsx")

from openpyxl import Workbook
from openpyxl.styles import Border as OpenpyxlBorder, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule

from templates.base import (
    use_palette_explicit,
    setup_sheet, style_header_row, style_data_row, style_total_row,
    font_title, font_header, font_body, font_caption, font_subheader,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    COLUMN_WIDTHS, ROW_HEIGHTS, FORMATS,
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    create_bar_chart, setup_chart_titles, apply_chart_colors,
    CHART_COLORS,
)

# ─── Activate warm palette BEFORE creating any styles ───
use_palette_explicit("warm")

wb = Workbook()
wb.properties.creator = "Z.ai"

CURRENCY_FMT = "#,##0.00"
DATE_FMT = "YYYY-MM-DD"
INTEGER_FMT = "#,##0"

# ─── Helper: set column widths from a list of (width_key_or_value, ...) ───
def set_col_widths(ws, widths):
    """widths: list of int or COLUMN_WIDTHS key strings, starting from col B."""
    for i, w in enumerate(widths, start=2):
        if isinstance(w, str):
            ws.column_dimensions[get_column_letter(i)].width = COLUMN_WIDTHS.get(w, w)
        else:
            ws.column_dimensions[get_column_letter(i)].width = w

# ─── Helper: apply print setup ───
def apply_print_setup(ws, orientation="landscape", fit_to_width=1, fit_to_height=0,
                      print_area=None, print_title_rows=None):
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = fit_to_width
    ws.page_setup.fitToHeight = fit_to_height
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    if print_area:
        ws.print_area = print_area
    if print_title_rows:
        ws.print_title_rows = print_title_rows

# ─── Helper: set warm tab color ───
def set_tab_color(ws):
    ws.sheet_properties.tabColor = "B85C1E"

# ─── Demo Data Definitions ───
PRODUCTS = [
    ("P001", "Coffee", "Beverages", 15.00, "Freshly brewed Arabic coffee"),
    ("P002", "Tea", "Beverages", 10.00, "Karak chai with spices"),
    ("P003", "Fresh Juice", "Beverages", 18.00, "Freshly squeezed orange juice"),
    ("P004", "Smoothie", "Beverages", 22.00, "Mixed berry smoothie"),
    ("P005", "Water", "Beverages", 5.00, "Mineral water 500ml"),
    ("P006", "Soft Drink", "Beverages", 8.00, "Carbonated soft drink"),
    ("P007", "Sandwich", "Main Course", 25.00, "Grilled chicken sandwich"),
    ("P008", "Burger", "Main Course", 35.00, "Classic beef burger with fries"),
    ("P009", "Pasta", "Main Course", 45.00, "Creamy Alfredo pasta"),
    ("P010", "Grilled Chicken", "Main Course", 55.00, "Herb-marinated grilled chicken"),
    ("P011", "Shawarma", "Main Course", 30.00, "Chicken shawarma wrap"),
    ("P012", "Biryani", "Main Course", 40.00, "Hyderabadi chicken biryani"),
    ("P013", "Salad", "Appetizers", 30.00, "Caesar salad with croutons"),
    ("P014", "Soup", "Appetizers", 20.00, "Lentil soup with lemon"),
    ("P015", "Fries", "Appetizers", 15.00, "Crispy seasoned fries"),
    ("P016", "Hummus", "Appetizers", 18.00, "Classic hummus with pita"),
    ("P017", "Cake", "Desserts", 22.00, "Chocolate lava cake"),
    ("P018", "Cheesecake", "Desserts", 28.00, "New York style cheesecake"),
    ("P019", "Ice Cream", "Desserts", 16.00, "Three scoops choice of flavour"),
    ("P020", "Kunafa", "Desserts", 25.00, "Traditional Palestinian kunafa"),
    ("P021", "Latte", "Beverages", 20.00, "Caffe latte with steamed milk"),
    ("P022", "Mojito Mocktail", "Beverages", 24.00, "Fresh lime and mint mocktail"),
    ("P023", "Mandi", "Main Course", 50.00, "Lamb mandi with saffron rice"),
    ("P024", "Manakish", "Appetizers", 16.00, "Zaatar and cheese manakish"),
]

CUSTOMERS = [
    "Walk-in", "Walk-in", "Walk-in", "Walk-in", "Walk-in",  # Walk-in is most common
    "Ahmed Al Maktoum", "Fatima Hassan", "Mohammed Al Rashid",
    "Sarah Johnson", "Khalid Ibrahim", "Aisha Al Sayed",
    "Omar Farooq", "Noor Abdullah", "Yusuf Ali", "Layla Khalil",
]

PAYMENT_METHODS = ["Cash", "Cash", "Cash", "Card", "Card", "Online", "Online", "Cheque"]

# Generate 55 sales transactions
random.seed(42)  # Reproducible demo data

sales_data = []
invoice_counter = 1
start_date = date(2025, 1, 1)

for i in range(55):
    tx_date = start_date + timedelta(days=random.randint(0, 30))
    inv_num = f"INV-{invoice_counter:04d}"
    invoice_counter += 1

    product = random.choice(PRODUCTS)
    item_name = product[1]
    category = product[2]
    unit_price = product[3]

    qty = random.choices([1, 1, 1, 2, 2, 3, 4], k=1)[0]
    payment = random.choice(PAYMENT_METHODS)
    customer = random.choice(CUSTOMERS)

    notes = ""
    if qty >= 3:
        notes = "Bulk order"
    elif payment == "Online":
        notes = "Online order"
    elif customer != "Walk-in":
        notes = "Regular customer"

    sales_data.append((tx_date, inv_num, item_name, category, qty, unit_price, payment, customer, notes))

# Sort by date then invoice number
sales_data.sort(key=lambda x: (x[0], x[1]))


# ═══════════════════════════════════════════════════════════
# SHEET 1: Sales Entry
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Sales Entry"

setup_sheet(ws1, title="Sales Entry", last_col=11)
set_tab_color(ws1)

# Column widths (B–K = cols 2–11)
col_widths_1 = ["date", "id_short", "name_en", "status", "id_short", "number", "number", "status", "name_en", "description"]
set_col_widths(ws1, col_widths_1)

# Header row (row 4)
headers_1 = ["Date", "Invoice #", "Item / Service", "Category", "Qty", "Unit Price", "Total", "Payment Method", "Customer Name", "Notes"]
for i, h in enumerate(headers_1, start=2):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, 2, 11)

# Data rows - fill with demo data
DATA_START_1 = 5
DATA_END_1 = DATA_START_1 + len(sales_data) - 1  # Row 59 for 55 items

for idx, (tx_date, inv_num, item_name, category, qty, unit_price, payment, customer, notes) in enumerate(sales_data):
    r = DATA_START_1 + idx
    row_idx = idx
    style_data_row(ws1, r, 2, 11, row_idx)

    # Date
    c = ws1.cell(row=r, column=2, value=tx_date)
    c.number_format = DATE_FMT
    c.alignment = align_date()

    # Invoice #
    c = ws1.cell(row=r, column=3, value=inv_num)
    c.alignment = align_text()

    # Item / Service
    c = ws1.cell(row=r, column=4, value=item_name)
    c.alignment = align_text()

    # Category
    c = ws1.cell(row=r, column=5, value=category)
    c.alignment = align_text()

    # Qty
    c = ws1.cell(row=r, column=6, value=qty)
    c.alignment = align_number()

    # Unit Price
    c = ws1.cell(row=r, column=7, value=unit_price)
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

    # Total = Qty × Unit Price (formula)
    c = ws1.cell(row=r, column=8)
    c.value = f'=IFERROR(F{r}*G{r},"")'
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

    # Payment Method
    c = ws1.cell(row=r, column=9, value=payment)
    c.alignment = align_text()

    # Customer Name
    c = ws1.cell(row=r, column=10, value=customer)
    c.alignment = align_text()

    # Notes
    c = ws1.cell(row=r, column=11, value=notes)
    c.alignment = align_text()

# Data validation: Payment Method dropdown (col I = col 9)
dv_payment = DataValidation(type="list", formula1='"Cash,Card,Online,Cheque"', allow_blank=True)
dv_payment.error = "Please select a valid payment method"
dv_payment.errorTitle = "Invalid Payment Method"
dv_payment.prompt = "Select payment method"
dv_payment.promptTitle = "Payment Method"
ws1.add_data_validation(dv_payment)
dv_payment.add(f"I{DATA_START_1}:I{DATA_END_1 + 50}")  # Extra rows for future entries

# Conditional formatting: Total column > 100 = green highlight
from templates.base import ACCENT_POSITIVE, CF_POSITIVE_FILL
green_fill = PatternFill(bgColor="E8F5E9")
green_font = Font(color=ACCENT_POSITIVE)
ws1.conditional_formatting.add(
    f"H{DATA_START_1}:H{DATA_END_1}",
    CellIsRule(operator="greaterThan", formula=["100"], fill=green_fill, font=green_font)
)

# Totals row
total_row_1 = DATA_END_1 + 2
ws1.cell(row=total_row_1, column=2, value="TOTALS")
ws1.cell(row=total_row_1, column=2).alignment = align_text()
ws1.cell(row=total_row_1, column=6).value = f'=SUMPRODUCT((F{DATA_START_1}:F{DATA_END_1}<>"")*1)'
ws1.cell(row=total_row_1, column=6).alignment = align_number()
ws1.cell(row=total_row_1, column=7).value = ""
ws1.cell(row=total_row_1, column=8).value = f'=IFERROR(SUM(H{DATA_START_1}:H{DATA_END_1}),"")'
ws1.cell(row=total_row_1, column=8).number_format = CURRENCY_FMT
ws1.cell(row=total_row_1, column=8).alignment = align_number()
style_total_row(ws1, total_row_1, 2, 11)

# Freeze panes at C5
ws1.freeze_panes = "C5"

# Print setup
apply_print_setup(ws1, orientation="landscape", fit_to_width=1,
                  print_area=f"B1:K{total_row_1}",
                  print_title_rows="1:4")


# ═══════════════════════════════════════════════════════════
# SHEET 2: Daily Summary
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Daily Summary")

setup_sheet(ws2, title="Daily Summary", last_col=8)
set_tab_color(ws2)

# Columns: Date, Total Sales, # Transactions, Avg Transaction, Cash Total, Card Total, Online Total
col_widths_2 = ["date", "number", "id_short", "number", "number", "number", "number"]
set_col_widths(ws2, col_widths_2)

headers_2 = ["Date", "Total Sales", "# Transactions", "Avg Transaction", "Cash Total", "Card Total", "Online Total"]
for i, h in enumerate(headers_2, start=2):
    ws2.cell(row=4, column=i, value=h)
style_header_row(ws2, 4, 2, 8)

# 31 days of January 2025 (rows 5–35)
DAILY_START = 5
DAILY_END = 35

for day in range(1, 32):
    r = DAILY_START + day - 1
    row_idx = r - DAILY_START
    style_data_row(ws2, r, 2, 8, row_idx)

    tx_date = date(2025, 1, day)

    # Date
    c = ws2.cell(row=r, column=2, value=tx_date)
    c.number_format = DATE_FMT
    c.alignment = align_date()

    # Total Sales = SUMPRODUCT matching date in Sales Entry
    c = ws2.cell(row=r, column=3)
    c.value = (
        f'=IFERROR(SUMPRODUCT((\'Sales Entry\'!B${DATA_START_1}:B${DATA_END_1}=B{r})*'
        f'(\'Sales Entry\'!H${DATA_START_1}:H${DATA_END_1})),"")'
    )
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

    # # Transactions
    c = ws2.cell(row=r, column=4)
    c.value = (
        f'=IFERROR(SUMPRODUCT((\'Sales Entry\'!B${DATA_START_1}:B${DATA_END_1}=B{r})*1),"")'
    )
    c.alignment = align_number()

    # Avg Transaction
    c = ws2.cell(row=r, column=5)
    c.value = f'=IFERROR(C{r}/D{r},"")'
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

    # Cash Total
    c = ws2.cell(row=r, column=6)
    c.value = (
        f'=IFERROR(SUMPRODUCT((\'Sales Entry\'!B${DATA_START_1}:B${DATA_END_1}=B{r})*'
        f'(\'Sales Entry\'!I${DATA_START_1}:I${DATA_END_1}="Cash")*'
        f'(\'Sales Entry\'!H${DATA_START_1}:H${DATA_END_1})),"")'
    )
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

    # Card Total
    c = ws2.cell(row=r, column=7)
    c.value = (
        f'=IFERROR(SUMPRODUCT((\'Sales Entry\'!B${DATA_START_1}:B${DATA_END_1}=B{r})*'
        f'(\'Sales Entry\'!I${DATA_START_1}:I${DATA_END_1}="Card")*'
        f'(\'Sales Entry\'!H${DATA_START_1}:H${DATA_END_1})),"")'
    )
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

    # Online Total
    c = ws2.cell(row=r, column=8)
    c.value = (
        f'=IFERROR(SUMPRODUCT((\'Sales Entry\'!B${DATA_START_1}:B${DATA_END_1}=B{r})*'
        f'(\'Sales Entry\'!I${DATA_START_1}:I${DATA_END_1}="Online")*'
        f'(\'Sales Entry\'!H${DATA_START_1}:H${DATA_END_1})),"")'
    )
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

# Totals row (row 36)
total_row_2 = DAILY_END + 1
ws2.cell(row=total_row_2, column=2, value="TOTALS")
ws2.cell(row=total_row_2, column=2).alignment = align_text()
ws2.cell(row=total_row_2, column=3).value = f'=IFERROR(SUM(C{DAILY_START}:C{DAILY_END}),"")'
ws2.cell(row=total_row_2, column=3).number_format = CURRENCY_FMT
ws2.cell(row=total_row_2, column=3).alignment = align_number()
ws2.cell(row=total_row_2, column=4).value = f'=IFERROR(SUM(D{DAILY_START}:D{DAILY_END}),"")'
ws2.cell(row=total_row_2, column=4).alignment = align_number()
ws2.cell(row=total_row_2, column=5).value = f'=IFERROR(C{total_row_2}/D{total_row_2},"")'
ws2.cell(row=total_row_2, column=5).number_format = CURRENCY_FMT
ws2.cell(row=total_row_2, column=5).alignment = align_number()
ws2.cell(row=total_row_2, column=6).value = f'=IFERROR(SUM(F{DAILY_START}:F{DAILY_END}),"")'
ws2.cell(row=total_row_2, column=6).number_format = CURRENCY_FMT
ws2.cell(row=total_row_2, column=6).alignment = align_number()
ws2.cell(row=total_row_2, column=7).value = f'=IFERROR(SUM(G{DAILY_START}:G{DAILY_END}),"")'
ws2.cell(row=total_row_2, column=7).number_format = CURRENCY_FMT
ws2.cell(row=total_row_2, column=7).alignment = align_number()
ws2.cell(row=total_row_2, column=8).value = f'=IFERROR(SUM(H{DAILY_START}:H{DAILY_END}),"")'
ws2.cell(row=total_row_2, column=8).number_format = CURRENCY_FMT
ws2.cell(row=total_row_2, column=8).alignment = align_number()
style_total_row(ws2, total_row_2, 2, 8)

apply_print_setup(ws2, orientation="landscape", fit_to_width=1,
                  print_area=f"B1:H{total_row_2}",
                  print_title_rows="1:4")


# ═══════════════════════════════════════════════════════════
# SHEET 3: Weekly Summary
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Weekly Summary")

setup_sheet(ws3, title="Weekly Summary", last_col=7)
set_tab_color(ws3)

# Columns: Week #, Week Start Date, Week End Date, Total Sales, # Transactions, Avg Transaction
col_widths_3 = ["id_short", "date", "date", "number", "id_short", "number"]
set_col_widths(ws3, col_widths_3)

headers_3 = ["Week #", "Week Start Date", "Week End Date", "Total Sales", "# Transactions", "Avg Transaction"]
for i, h in enumerate(headers_3, start=2):
    ws3.cell(row=4, column=i, value=h)
style_header_row(ws3, 4, 2, 7)

# 4 weeks of January 2025
weeks = [
    (1, date(2025, 1, 1), date(2025, 1, 5)),
    (2, date(2025, 1, 6), date(2025, 1, 12)),
    (3, date(2025, 1, 13), date(2025, 1, 19)),
    (4, date(2025, 1, 20), date(2025, 1, 26)),
    (5, date(2025, 1, 27), date(2025, 1, 31)),
]

WEEKLY_START = 5
WEEKLY_END = WEEKLY_START + len(weeks) - 1

for idx, (week_num, week_start, week_end) in enumerate(weeks):
    r = WEEKLY_START + idx
    row_idx = idx
    style_data_row(ws3, r, 2, 7, row_idx)

    # Week #
    c = ws3.cell(row=r, column=2, value=week_num)
    c.alignment = align_number()

    # Week Start Date
    c = ws3.cell(row=r, column=3, value=week_start)
    c.number_format = DATE_FMT
    c.alignment = align_date()

    # Week End Date
    c = ws3.cell(row=r, column=4, value=week_end)
    c.number_format = DATE_FMT
    c.alignment = align_date()

    # Total Sales = SUMPRODUCT matching dates in Daily Summary
    c = ws3.cell(row=r, column=5)
    c.value = (
        f'=IFERROR(SUMPRODUCT((\'Daily Summary\'!B${DAILY_START}:B${DAILY_END}>=C{r})*'
        f'(\'Daily Summary\'!B${DAILY_START}:B${DAILY_END}<=D{r})*'
        f'(\'Daily Summary\'!C${DAILY_START}:C${DAILY_END})),"")'
    )
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

    # # Transactions
    c = ws3.cell(row=r, column=6)
    c.value = (
        f'=IFERROR(SUMPRODUCT((\'Daily Summary\'!B${DAILY_START}:B${DAILY_END}>=C{r})*'
        f'(\'Daily Summary\'!B${DAILY_START}:B${DAILY_END}<=D{r})*'
        f'(\'Daily Summary\'!D${DAILY_START}:D${DAILY_END})),"")'
    )
    c.alignment = align_number()

    # Avg Transaction
    c = ws3.cell(row=r, column=7)
    c.value = f'=IFERROR(E{r}/F{r},"")'
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

# Totals row
total_row_3 = WEEKLY_END + 1
ws3.cell(row=total_row_3, column=2, value="TOTALS")
ws3.cell(row=total_row_3, column=2).alignment = align_text()
ws3.cell(row=total_row_3, column=5).value = f'=IFERROR(SUM(E{WEEKLY_START}:E{WEEKLY_END}),"")'
ws3.cell(row=total_row_3, column=5).number_format = CURRENCY_FMT
ws3.cell(row=total_row_3, column=5).alignment = align_number()
ws3.cell(row=total_row_3, column=6).value = f'=IFERROR(SUM(F{WEEKLY_START}:F{WEEKLY_END}),"")'
ws3.cell(row=total_row_3, column=6).alignment = align_number()
ws3.cell(row=total_row_3, column=7).value = f'=IFERROR(E{total_row_3}/F{total_row_3},"")'
ws3.cell(row=total_row_3, column=7).number_format = CURRENCY_FMT
ws3.cell(row=total_row_3, column=7).alignment = align_number()
style_total_row(ws3, total_row_3, 2, 7)

apply_print_setup(ws3, orientation="landscape", fit_to_width=1,
                  print_area=f"B1:G{total_row_3}",
                  print_title_rows="1:4")


# ═══════════════════════════════════════════════════════════
# SHEET 4: Monthly Summary
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Monthly Summary")

setup_sheet(ws4, title="Monthly Summary", last_col=7)
set_tab_color(ws4)

# Columns: Month, Total Sales, # Transactions, Avg Transaction, Best Day Sales, Best Day Date
col_widths_4 = ["name_en", "number", "id_short", "number", "number", "date"]
set_col_widths(ws4, col_widths_4)

headers_4 = ["Month", "Total Sales", "# Transactions", "Avg Transaction", "Best Day Sales", "Best Day Date"]
for i, h in enumerate(headers_4, start=2):
    ws4.cell(row=4, column=i, value=h)
style_header_row(ws4, 4, 2, 7)

# Just January 2025
MONTHLY_START = 5
MONTHLY_END = 5

r = MONTHLY_START
style_data_row(ws4, r, 2, 7, 0)

# Month
c = ws4.cell(row=r, column=2, value="January 2025")
c.alignment = align_text()

# Total Sales
c = ws4.cell(row=r, column=3)
c.value = f'=IFERROR(SUMPRODUCT((\'Daily Summary\'!C${DAILY_START}:C${DAILY_END})),"")'
c.number_format = CURRENCY_FMT
c.alignment = align_number()

# # Transactions
c = ws4.cell(row=r, column=4)
c.value = f'=IFERROR(SUMPRODUCT((\'Daily Summary\'!D${DAILY_START}:D${DAILY_END})),"")'
c.alignment = align_number()

# Avg Transaction
c = ws4.cell(row=r, column=5)
c.value = f'=IFERROR(C{r}/D{r},"")'
c.number_format = CURRENCY_FMT
c.alignment = align_number()

# Best Day Sales
c = ws4.cell(row=r, column=6)
c.value = f'=IFERROR(MAX(\'Daily Summary\'!C${DAILY_START}:C${DAILY_END}),"")'
c.number_format = CURRENCY_FMT
c.alignment = align_number()

# Best Day Date
c = ws4.cell(row=r, column=7)
c.value = (
    f'=IFERROR(SUMPRODUCT((\'Daily Summary\'!C${DAILY_START}:C${DAILY_END}=F{r})*'
    f'(\'Daily Summary\'!B${DAILY_START}:B${DAILY_END})),"")'
)
c.number_format = DATE_FMT
c.alignment = align_date()

# Totals row
total_row_4 = MONTHLY_END + 1
ws4.cell(row=total_row_4, column=2, value="TOTALS")
ws4.cell(row=total_row_4, column=2).alignment = align_text()
ws4.cell(row=total_row_4, column=3).value = f'=IFERROR(SUM(C{MONTHLY_START}:C{MONTHLY_END}),"")'
ws4.cell(row=total_row_4, column=3).number_format = CURRENCY_FMT
ws4.cell(row=total_row_4, column=3).alignment = align_number()
ws4.cell(row=total_row_4, column=4).value = f'=IFERROR(SUM(D{MONTHLY_START}:D{MONTHLY_END}),"")'
ws4.cell(row=total_row_4, column=4).alignment = align_number()
ws4.cell(row=total_row_4, column=5).value = f'=IFERROR(C{total_row_4}/D{total_row_4},"")'
ws4.cell(row=total_row_4, column=5).number_format = CURRENCY_FMT
ws4.cell(row=total_row_4, column=5).alignment = align_number()
ws4.cell(row=total_row_4, column=6).value = ""
ws4.cell(row=total_row_4, column=7).value = ""
style_total_row(ws4, total_row_4, 2, 7)

# Bar chart showing daily sales for January
chart = create_bar_chart(chart_type="col", grouping="clustered", gap_width=80, width=22, height=12)
data_ref = Reference(ws2, min_col=3, min_row=4, max_row=DAILY_END)
cats_ref = Reference(ws2, min_col=2, min_row=DAILY_START, max_row=DAILY_END)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
setup_chart_titles(chart, title="January 2025 — Daily Sales", y_title="Sales (AED)", x_title="Day")
apply_chart_colors(chart)
ws4.add_chart(chart, "B8")

apply_print_setup(ws4, orientation="landscape", fit_to_width=1,
                  print_area=f"B1:G{total_row_4}",
                  print_title_rows="1:4")


# ═══════════════════════════════════════════════════════════
# SHEET 5: Invoice
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Invoice")

setup_sheet(ws5, title="", last_col=8)
set_tab_color(ws5)

# Column widths for invoice layout
ws5.column_dimensions["A"].width = 3
ws5.column_dimensions["B"].width = 18
ws5.column_dimensions["C"].width = 22
ws5.column_dimensions["D"].width = 3
ws5.column_dimensions["E"].width = 28
ws5.column_dimensions["F"].width = 10
ws5.column_dimensions["G"].width = 14
ws5.column_dimensions["H"].width = 14

# ── Company Header ──
ws5.merge_cells("B2:H2")
c = ws5.cell(row=2, column=2, value="Al Madina Restaurant")
c.font = font_title()
c.alignment = align_title()

ws5.merge_cells("B3:H3")
c = ws5.cell(row=3, column=2, value="Business Bay, Dubai, UAE  |  Tel: +971 4 555 0100  |  info@realestateemperor.ae")
c.font = font_caption()
c.alignment = align_text()

# Spacer
ws5.row_dimensions[4].height = ROW_HEIGHTS["spacer"]

# ── INVOICE Title ──
ws5.merge_cells("B5:H5")
c = ws5.cell(row=5, column=2, value="INVOICE")
c.font = Font(name=FONT_NAME if 'FONT_NAME' in dir() else "Calibri", size=20, bold=True, color=PRIMARY)
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[5].height = 36

# Divider line
ws5.row_dimensions[6].height = 4

# ── Invoice Details ──
# Row 7: Invoice #
ws5.cell(row=7, column=2, value="Invoice #").font = font_subheader()
ws5.cell(row=7, column=2).alignment = align_text()
ws5.cell(row=7, column=3, value="INV-0025").font = font_body()
ws5.cell(row=7, column=3).alignment = align_text()

# Row 8: Date
ws5.cell(row=8, column=2, value="Date").font = font_subheader()
ws5.cell(row=8, column=2).alignment = align_text()
c = ws5.cell(row=8, column=3, value=date(2025, 1, 15))
c.font = font_body()
c.number_format = "DD-MMM-YYYY"
c.alignment = align_date()

# Row 9: Customer
ws5.cell(row=9, column=2, value="Customer").font = font_subheader()
ws5.cell(row=9, column=2).alignment = align_text()
ws5.cell(row=9, column=3, value="Walk-in").font = font_body()
ws5.cell(row=9, column=3).alignment = align_text()

# Row 10: Payment
ws5.cell(row=10, column=2, value="Payment").font = font_subheader()
ws5.cell(row=10, column=2).alignment = align_text()
ws5.cell(row=10, column=3, value="Cash").font = font_body()
ws5.cell(row=10, column=3).alignment = align_text()

# Spacer
ws5.row_dimensions[11].height = ROW_HEIGHTS["spacer"]

# ── Items Table Header (row 12) ──
item_headers = ["Item", "Qty", "Price (AED)", "Total (AED)"]
item_header_cols = [5, 6, 7, 8]
for i, h in enumerate(item_headers):
    col = item_header_cols[i]
    ws5.cell(row=12, column=col, value=h)
style_header_row(ws5, 12, 5, 8)

# ── Invoice Items ──
invoice_items = [
    ("Coffee", 2, 15.00),
    ("Sandwich", 1, 25.00),
    ("Salad", 1, 30.00),
    ("Fresh Juice", 2, 18.00),
    ("Cake", 1, 22.00),
]

INV_ITEM_START = 13
INV_ITEM_END = INV_ITEM_START + len(invoice_items) - 1

for idx, (item_name, qty, price) in enumerate(invoice_items):
    r = INV_ITEM_START + idx
    row_idx = idx
    style_data_row(ws5, r, 5, 8, row_idx)

    c = ws5.cell(row=r, column=5, value=item_name)
    c.alignment = align_text()

    c = ws5.cell(row=r, column=6, value=qty)
    c.alignment = align_number()

    c = ws5.cell(row=r, column=7, value=price)
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

    c = ws5.cell(row=r, column=8)
    c.value = f'=IFERROR(F{r}*G{r},"")'
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

# Spacer after items
spacer_row = INV_ITEM_END + 1
ws5.row_dimensions[spacer_row].height = ROW_HEIGHTS["spacer"]

# ── Summary Section ──
subtotal_row = INV_ITEM_END + 2
ws5.cell(row=subtotal_row, column=7, value="Subtotal").font = font_subheader()
ws5.cell(row=subtotal_row, column=7).alignment = Alignment(horizontal="right", vertical="center")
c = ws5.cell(row=subtotal_row, column=8)
c.value = f'=IFERROR(SUM(H{INV_ITEM_START}:H{INV_ITEM_END}),"")'
c.number_format = CURRENCY_FMT
c.font = font_body()
c.alignment = align_number()

vat_row = subtotal_row + 1
ws5.cell(row=vat_row, column=7, value="VAT 5%").font = font_subheader()
ws5.cell(row=vat_row, column=7).alignment = Alignment(horizontal="right", vertical="center")
c = ws5.cell(row=vat_row, column=8)
c.value = f'=IFERROR(H{subtotal_row}*0.05,"")'
c.number_format = CURRENCY_FMT
c.font = font_body()
c.alignment = align_number()

grand_row = vat_row + 1
ws5.cell(row=grand_row, column=7, value="Grand Total").font = font_subheader()
ws5.cell(row=grand_row, column=7).alignment = Alignment(horizontal="right", vertical="center")
c = ws5.cell(row=grand_row, column=8)
c.value = f'=IFERROR(H{subtotal_row}+H{vat_row},"")'
c.number_format = CURRENCY_FMT
c.font = font_subheader()
c.alignment = align_number()
# Style grand total with total row styling
for col in [7, 8]:
    ws5.cell(row=grand_row, column=col).fill = fill_total()
    ws5.cell(row=grand_row, column=col).border = border_total()

# Spacer
ws5.row_dimensions[grand_row + 1].height = ROW_HEIGHTS["spacer"]

# ── Footer ──
footer_row = grand_row + 2
ws5.merge_cells(f"B{footer_row}:H{footer_row}")
c = ws5.cell(row=footer_row, column=2, value="Thank you for dining with us!")
c.font = font_caption()
c.alignment = Alignment(horizontal="center", vertical="center")

# Second footer line
ws5.merge_cells(f"B{footer_row + 1}:H{footer_row + 1}")
c = ws5.cell(row=footer_row + 1, column=2, value="TRN: 10001 23456 7890 1234  |  VAT Registration No: AE123456789012345")
c.font = font_caption()
c.alignment = Alignment(horizontal="center", vertical="center")

apply_print_setup(ws5, orientation="portrait", fit_to_width=1, fit_to_height=1,
                  print_area=f"B1:H{footer_row + 1}")


# ═══════════════════════════════════════════════════════════
# SHEET 6: Product List
# ═══════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Product List")

setup_sheet(ws6, title="Product List", last_col=6)
set_tab_color(ws6)

# Columns: Product ID, Item Name, Category, Unit Price, Description
col_widths_6 = ["id_short", "name_en", "status", "number", "description"]
set_col_widths(ws6, col_widths_6)

headers_6 = ["Product ID", "Item Name", "Category", "Unit Price", "Description"]
for i, h in enumerate(headers_6, start=2):
    ws6.cell(row=4, column=i, value=h)
style_header_row(ws6, 4, 2, 6)

# Fill with product data
PROD_START = 5
PROD_END = PROD_START + len(PRODUCTS) - 1

for idx, (pid, item_name, category, price, desc) in enumerate(PRODUCTS):
    r = PROD_START + idx
    row_idx = idx
    style_data_row(ws6, r, 2, 6, row_idx)

    c = ws6.cell(row=r, column=2, value=pid)
    c.alignment = align_text()

    c = ws6.cell(row=r, column=3, value=item_name)
    c.alignment = align_text()

    c = ws6.cell(row=r, column=4, value=category)
    c.alignment = align_text()

    c = ws6.cell(row=r, column=5, value=price)
    c.number_format = CURRENCY_FMT
    c.alignment = align_number()

    c = ws6.cell(row=r, column=6, value=desc)
    c.alignment = align_text()

# Data validation: Category dropdown (col D = col 4)
dv_category = DataValidation(
    type="list",
    formula1='"Beverages,Main Course,Appetizers,Desserts"',
    allow_blank=True
)
dv_category.error = "Please select a valid category"
dv_category.errorTitle = "Invalid Category"
dv_category.prompt = "Select product category"
dv_category.promptTitle = "Category"
ws6.add_data_validation(dv_category)
dv_category.add(f"D{PROD_START}:D{PROD_END + 30}")  # Extra rows for new items

# Totals row
total_row_6 = PROD_END + 1
ws6.cell(row=total_row_6, column=2, value="TOTALS")
ws6.cell(row=total_row_6, column=2).alignment = align_text()
ws6.cell(row=total_row_6, column=3).value = f'=COUNTA(C{PROD_START}:C{PROD_END})&" items"'
ws6.cell(row=total_row_6, column=3).alignment = align_text()
ws6.cell(row=total_row_6, column=5).value = f'=IFERROR(SUM(E{PROD_START}:E{PROD_END}),"")'
ws6.cell(row=total_row_6, column=5).number_format = CURRENCY_FMT
ws6.cell(row=total_row_6, column=5).alignment = align_number()
style_total_row(ws6, total_row_6, 2, 6)

apply_print_setup(ws6, orientation="landscape", fit_to_width=1,
                  print_area=f"B1:F{total_row_6}",
                  print_title_rows="1:4")


# ═══════════════════════════════════════════════════════════
# Save workbook
# ═══════════════════════════════════════════════════════════
output_path = "/home/z/my-project/download/POS_Basic_Sample.xlsx"
wb.save(output_path)
print(f"✅ Saved: {output_path}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Sales Entry: {len(sales_data)} transactions")
print(f"   Products: {len(PRODUCTS)} items")
