"""
Al Madina Grocery — Monthly Business Report Generator
Generates a professional Excel report at /home/z/my-project/download/sample_report_grocery.xlsx
"""

import sys, os
sys.path.insert(0, '/home/z/my-project/skills/xlsx')
sys.path.insert(0, '/home/z/my-project/skills/xlsx/templates')
from base import *

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# ── Activate professional palette ──
use_palette_explicit("professional")

wb = Workbook()
FMT_CURRENCY = '#,##0'
FMT_PERCENT  = '0.0%'

FOOTER_TEXT = "Real Estate Emperor Property Management L.L.C."


# ============================================================
# Helper: write footer row
# ============================================================
def write_footer(ws, row, col_start, col_end):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=FOOTER_TEXT)
    cell.font = font_caption()
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20


# ============================================================
# SHEET 1: Dashboard
# ============================================================
ws1 = wb.active
ws1.title = "Dashboard"

setup_sheet(ws1, title="Al Madina Grocery \u2014 Monthly Business Report", last_col=8)

# Subtitle row
ws1.merge_cells('B3:H3')
sub_cell = ws1.cell(row=3, column=2, value="January 2026 | Ajman, UAE")
sub_cell.font = font_caption()
sub_cell.alignment = align_title()
ws1.row_dimensions[3].height = 18

# ── KPI Row ──
kpi_row = 5
ws1.row_dimensions[kpi_row].height = 48
ws1.row_dimensions[kpi_row + 1].height = 18

kpis = [
    ("Total Revenue", 312800, "AED"),
    ("Total Expenses", 268400, "AED"),
    ("Net Profit", 44400, "AED"),
    ("Profit Margin", 0.142, "%"),
]

kpi_cols = [2, 4, 6, 8]
for i, (label, value, fmt) in enumerate(kpis):
    col = kpi_cols[i]
    # KPI number
    cell = ws1.cell(row=kpi_row, column=col, value=value)
    cell.font = font_kpi()
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    if fmt == "AED":
        cell.number_format = FMT_CURRENCY
    else:
        cell.number_format = FMT_PERCENT
    # KPI label
    lbl = ws1.cell(row=kpi_row + 1, column=col, value=label)
    lbl.font = font_kpi_label()
    lbl.alignment = Alignment(horizontal="center", vertical="top")

# ── Spacer ──
data_start = kpi_row + 3  # row 8

# ── Monthly data table ──
headers = ["Month", "Revenue (AED)", "Expenses (AED)", "Net Profit (AED)", "Profit Margin"]
header_row = data_start
col_start = 2
col_end = col_start + len(headers) - 1  # col 6

for i, h in enumerate(headers):
    ws1.cell(row=header_row, column=col_start + i, value=h)
style_header_row(ws1, header_row, col_start, col_end)

monthly_data = [
    ("Oct 2025", 285600, 252100, 33500, 0.117),
    ("Nov 2025", 298400, 259800, 38600, 0.129),
    ("Dec 2025", 342100, 278500, 63600, 0.186),
    ("Jan 2026", 312800, 268400, 44400, 0.142),
]

for idx, (month, rev, exp, profit, margin) in enumerate(monthly_data):
    r = header_row + 1 + idx
    ws1.cell(row=r, column=2, value=month)
    ws1.cell(row=r, column=3, value=rev).number_format = FMT_CURRENCY
    ws1.cell(row=r, column=4, value=exp).number_format = FMT_CURRENCY
    ws1.cell(row=r, column=5, value=profit).number_format = FMT_CURRENCY
    ws1.cell(row=r, column=6, value=margin).number_format = FMT_PERCENT
    style_data_row(ws1, r, col_start, col_end, idx)
    # Alignment: text left, numbers right
    ws1.cell(row=r, column=2).alignment = align_text()
    for c in range(3, 7):
        ws1.cell(row=r, column=c).alignment = align_number()

# ── Bar Chart: Revenue vs Expenses ──
chart1 = create_bar_chart(chart_type="col", grouping="clustered", gap_width=100, overlap=0, width=16, height=10)

data_ref = Reference(ws1, min_col=3, max_col=4, min_row=header_row, max_row=header_row + 4)
cats_ref = Reference(ws1, min_col=2, min_row=header_row + 1, max_row=header_row + 4)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
setup_chart_titles(chart1, title="Revenue vs Expenses by Month", y_title="Amount (AED)")
apply_chart_colors(chart1, [CHART_COLORS[0], CHART_COLORS[2]])
chart1.legend.position = 'b'

ws1.add_chart(chart1, "B15")

# ── Line Chart: Profit Margin trend ──
chart2 = create_line_chart(width=16, height=10)

margin_ref = Reference(ws1, min_col=6, min_row=header_row, max_row=header_row + 4)
cats_ref2 = Reference(ws1, min_col=2, min_row=header_row + 1, max_row=header_row + 4)
chart2.add_data(margin_ref, titles_from_data=True)
chart2.set_categories(cats_ref2)
setup_chart_titles(chart2, title="Profit Margin Trend", y_title="Margin")
apply_chart_colors(chart2, [CHART_COLORS[1]])

# Style the line
s = chart2.series[0]
s.graphicalProperties.line.width = 28000  # ~2pt
s.smooth = True

# Set y-axis number format
chart2.y_axis.numFmt = '0.0%'
chart2.y_axis.scaling.min = 0
chart2.y_axis.scaling.max = 0.25
chart2.legend = None

ws1.add_chart(chart2, "B31")

# Footer
write_footer(ws1, 47, 2, 8)

auto_fit_columns(ws1, header_row=header_row)


# ============================================================
# SHEET 2: Product Analysis
# ============================================================
ws2 = wb.create_sheet("Product Analysis")

setup_sheet(ws2, title="Top Products \u2014 January 2026", last_col=8)

headers2 = ["Product", "Category", "Units Sold", "Revenue (AED)", "Unit Price (AED)", "Profit Margin"]
header_row2 = 5
col_start2 = 2
col_end2 = col_start2 + len(headers2) - 1  # col 7

for i, h in enumerate(headers2):
    ws2.cell(row=header_row2, column=col_start2 + i, value=h)
style_header_row(ws2, header_row2, col_start2, col_end2)

product_data = [
    ("Basmati Rice (25kg)", "Grains",       420, 33600, 80.0, 0.220),
    ("Fresh Chicken (kg)",  "Meat",         680, 27200, 40.0, 0.150),
    ("Cooking Oil (5L)",    "Oil & Ghee",   540, 24300, 45.0, 0.185),
    ("Fresh Milk (1L)",     "Dairy",       1200,  8400,  7.0, 0.280),
    ("Bread (pack)",        "Bakery",       950,  4750,  5.0, 0.350),
    ("Dates (1kg)",         "Dry Fruits",   310, 12400, 40.0, 0.420),
    ("Spices Mix",          "Spices",       480,  7200, 15.0, 0.550),
    ("Water (24-pack)",     "Beverages",    890,  5340,  6.0, 0.300),
    ("Canned Beans",        "Canned Food",  620,  4960,  8.0, 0.380),
    ("Laundry Detergent",   "Household",    280,  8400, 30.0, 0.250),
]

total_units = sum(r[2] for r in product_data)
total_revenue = sum(r[3] for r in product_data)

for idx, (product, cat, units, rev, price, margin) in enumerate(product_data):
    r = header_row2 + 1 + idx
    ws2.cell(row=r, column=2, value=product)
    ws2.cell(row=r, column=3, value=cat)
    ws2.cell(row=r, column=4, value=units).number_format = FMT_CURRENCY
    ws2.cell(row=r, column=5, value=rev).number_format = FMT_CURRENCY
    ws2.cell(row=r, column=6, value=price).number_format = '#,##0.0'
    ws2.cell(row=r, column=7, value=margin).number_format = FMT_PERCENT
    style_data_row(ws2, r, col_start2, col_end2, idx)
    # Alignment
    ws2.cell(row=r, column=2).alignment = align_text()
    ws2.cell(row=r, column=3).alignment = align_text()
    for c in range(4, 8):
        ws2.cell(row=r, column=c).alignment = align_number()

# Total row
total_r = header_row2 + 1 + len(product_data)
ws2.cell(row=total_r, column=2, value="TOTAL")
ws2.cell(row=total_r, column=3, value="\u2014")
ws2.cell(row=total_r, column=4, value=total_units).number_format = FMT_CURRENCY
ws2.cell(row=total_r, column=5, value=total_revenue).number_format = FMT_CURRENCY
ws2.cell(row=total_r, column=6, value="\u2014")
ws2.cell(row=total_r, column=7, value="\u2014")
style_total_row(ws2, total_r, col_start2, col_end2)
ws2.cell(row=total_r, column=2).alignment = align_text()
ws2.cell(row=total_r, column=3).alignment = align_text()
for c in range(4, 8):
    ws2.cell(row=total_r, column=c).alignment = align_number()

# ── Horizontal Bar Chart: Revenue by product ──
chart3 = create_bar_chart(chart_type="bar", grouping="clustered", gap_width=80, overlap=100, width=18, height=11)

rev_ref = Reference(ws2, min_col=5, min_row=header_row2, max_row=header_row2 + len(product_data))
prod_cats = Reference(ws2, min_col=2, min_row=header_row2 + 1, max_row=header_row2 + len(product_data))
chart3.add_data(rev_ref, titles_from_data=True)
chart3.set_categories(prod_cats)
setup_chart_titles(chart3, title="Revenue by Product (AED)", x_title="Revenue (AED)")
apply_chart_colors(chart3, [CHART_COLORS[0]])
chart3.legend = None

ws2.add_chart(chart3, "B18")

# ── Pie Chart: Revenue by category ──
# We need category-aggregated data for the pie chart. Let's create a small helper area.
# Place category summary starting at column 10 (J)
cat_rev = {}
for product, cat, units, rev, price, margin in product_data:
    cat_rev[cat] = cat_rev.get(cat, 0) + rev

cat_col = 10
ws2.cell(row=header_row2, column=cat_col, value="Category")
ws2.cell(row=header_row2, column=cat_col + 1, value="Revenue (AED)")

sorted_cats = sorted(cat_rev.items(), key=lambda x: -x[1])
for i, (cat, rev) in enumerate(sorted_cats):
    ws2.cell(row=header_row2 + 1 + i, column=cat_col, value=cat)
    ws2.cell(row=header_row2 + 1 + i, column=cat_col + 1, value=rev).number_format = FMT_CURRENCY

chart4 = create_pie_chart(width=16, height=11)
pie_data = Reference(ws2, min_col=cat_col + 1, min_row=header_row2, max_row=header_row2 + len(sorted_cats))
pie_cats = Reference(ws2, min_col=cat_col, min_row=header_row2 + 1, max_row=header_row2 + len(sorted_cats))
chart4.add_data(pie_data, titles_from_data=True)
chart4.set_categories(pie_cats)
setup_chart_titles(chart4, title="Revenue by Category")
apply_pie_colors(chart4, len(sorted_cats))
chart4.dataLabels = None

# Add data labels with percentages
from openpyxl.chart.label import DataLabelList
chart4.dataLabels = DataLabelList()
chart4.dataLabels.showPercent = True
chart4.dataLabels.showCatName = True
chart4.dataLabels.showVal = False

ws2.add_chart(chart4, "B35")

# Footer
write_footer(ws2, 52, 2, 8)

auto_fit_columns(ws2, header_row=header_row2)

# Hide helper columns
ws2.column_dimensions[get_column_letter(cat_col)].width = 0
ws2.column_dimensions[get_column_letter(cat_col + 1)].width = 0


# ============================================================
# SHEET 3: Inventory Alerts
# ============================================================
ws3 = wb.create_sheet("Inventory Alerts")

setup_sheet(ws3, title="Inventory Status & Reorder Alerts", last_col=8)

headers3 = ["Product", "Current Stock", "Reorder Level", "Status", "Days of Supply", "Action Required"]
header_row3 = 5
col_start3 = 2
col_end3 = col_start3 + len(headers3) - 1  # col 7

for i, h in enumerate(headers3):
    ws3.cell(row=header_row3, column=col_start3 + i, value=h)
style_header_row(ws3, header_row3, col_start3, col_end3)

inventory_data = [
    ("Basmati Rice (25kg)", "85 bags",   "100 bags",  "Low Stock",  8, "Reorder Now"),
    ("Fresh Chicken (kg)",  "45 kg",     "50 kg",     "Low Stock",  3, "Reorder Urgent"),
    ("Cooking Oil (5L)",    "180 bottles","100 bottles","Adequate",  22, "No Action"),
    ("Fresh Milk (1L)",     "320 units", "200 units", "Adequate",  15, "No Action"),
    ("Bread (pack)",        "60 packs",  "100 packs", "Low Stock",  4, "Reorder Urgent"),
    ("Dates (1kg)",         "150 boxes", "80 boxes",  "Surplus",   35, "Reduce Order"),
    ("Spices Mix",          "200 packs", "100 packs", "Adequate",  28, "No Action"),
    ("Water (24-pack)",     "40 packs",  "150 packs", "Critical",   2, "Reorder Urgent"),
    ("Canned Beans",        "300 cans",  "150 cans",  "Adequate",  30, "No Action"),
    ("Laundry Detergent",   "25 units",  "50 units",  "Low Stock",  5, "Reorder Now"),
]

for idx, (product, stock, reorder, status, days, action) in enumerate(inventory_data):
    r = header_row3 + 1 + idx
    ws3.cell(row=r, column=2, value=product)
    ws3.cell(row=r, column=3, value=stock)
    ws3.cell(row=r, column=4, value=reorder)
    ws3.cell(row=r, column=5, value=status)
    ws3.cell(row=r, column=6, value=days)
    ws3.cell(row=r, column=7, value=action)
    style_data_row(ws3, r, col_start3, col_end3, idx)
    # Alignment
    ws3.cell(row=r, column=2).alignment = align_text()
    ws3.cell(row=r, column=3).alignment = align_number()
    ws3.cell(row=r, column=4).alignment = align_number()
    ws3.cell(row=r, column=5).alignment = Alignment(horizontal="center", vertical="center")
    ws3.cell(row=r, column=6).alignment = align_number()
    ws3.cell(row=r, column=7).alignment = align_text()

# ── Conditional formatting on Status column (E) ──
status_col_letter = get_column_letter(5)  # column E (col 5 in 1-based)
status_range = f"{status_col_letter}{header_row3+1}:{status_col_letter}{header_row3 + len(inventory_data)}"

# Critical = red fill + red font
ws3.conditional_formatting.add(status_range,
    CellIsRule(operator='equal', formula=['"Critical"'],
              fill=PatternFill(bgColor="FDEDEC"),
              font=Font(color=ACCENT_NEGATIVE, bold=True)))

# Low Stock = amber fill + amber font
ws3.conditional_formatting.add(status_range,
    CellIsRule(operator='equal', formula=['"Low Stock"'],
              fill=PatternFill(bgColor="FEF9E7"),
              font=Font(color=ACCENT_WARNING, bold=True)))

# Adequate = green fill + green font
ws3.conditional_formatting.add(status_range,
    CellIsRule(operator='equal', formula=['"Adequate"'],
              fill=PatternFill(bgColor="E8F5E9"),
              font=Font(color=ACCENT_POSITIVE, bold=True)))

# Surplus = blue fill + blue font
ws3.conditional_formatting.add(status_range,
    CellIsRule(operator='equal', formula=['"Surplus"'],
              fill=PatternFill(bgColor="D6E4F0"),
              font=Font(color=PRIMARY, bold=True)))

# Footer
write_footer(ws3, header_row3 + len(inventory_data) + 2, 2, 7)

auto_fit_columns(ws3, header_row=header_row3)


# ============================================================
# SHEET 4: Recommendations
# ============================================================
ws4 = wb.create_sheet("Recommendations")

setup_sheet(ws4, title="Key Insights & Recommendations", last_col=5)

headers4 = ["#", "Insight", "Priority"]
header_row4 = 5
col_start4 = 2
col_end4 = col_start4 + len(headers4) - 1  # col 4

for i, h in enumerate(headers4):
    ws4.cell(row=header_row4, column=col_start4 + i, value=h)
style_header_row(ws4, header_row4, col_start4, col_end4)

insights_data = [
    (1, "Water inventory is at critical level (2 days supply). Immediate reorder needed.", "High"),
    (2, "Fresh bread and chicken are below reorder levels. Daily deliveries need adjustment.", "High"),
    (3, "Dates have 35 days of supply \u2014 surplus. Reduce next order by 40%.", "Medium"),
    (4, "Spices have highest margin (55%) but only 4.4% of revenue. Consider promotional display.", "Medium"),
    (5, "December revenue spike (AED 342,100) was seasonal. Stock planning needed for Q4 2026.", "Low"),
]

for idx, (num, insight, priority) in enumerate(insights_data):
    r = header_row4 + 1 + idx
    ws4.cell(row=r, column=2, value=num)
    ws4.cell(row=r, column=3, value=insight)
    ws4.cell(row=r, column=4, value=priority)
    style_data_row(ws4, r, col_start4, col_end4, idx)
    # Alignment
    ws4.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws4.cell(row=r, column=3).alignment = align_text()
    ws4.cell(row=r, column=4).alignment = Alignment(horizontal="center", vertical="center")

# ── Conditional formatting on Priority column (D) ──
priority_col_letter = get_column_letter(4)  # column D
priority_range = f"{priority_col_letter}{header_row4+1}:{priority_col_letter}{header_row4 + len(insights_data)}"

# High = red
ws4.conditional_formatting.add(priority_range,
    CellIsRule(operator='equal', formula=['"High"'],
              fill=PatternFill(bgColor="FDEDEC"),
              font=Font(color=ACCENT_NEGATIVE, bold=True)))

# Medium = amber
ws4.conditional_formatting.add(priority_range,
    CellIsRule(operator='equal', formula=['"Medium"'],
              fill=PatternFill(bgColor="FEF9E7"),
              font=Font(color=ACCENT_WARNING, bold=True)))

# Low = green
ws4.conditional_formatting.add(priority_range,
    CellIsRule(operator='equal', formula=['"Low"'],
              fill=PatternFill(bgColor="E8F5E9"),
              font=Font(color=ACCENT_POSITIVE, bold=True)))

# Footer
write_footer(ws4, header_row4 + len(insights_data) + 2, 2, 4)

auto_fit_columns(ws4, header_row=header_row4)

# Widen insight column specifically
ws4.column_dimensions['C'].width = 60


# ============================================================
# Finalize
# ============================================================
wb.properties.creator = "Z.ai"
output_path = "/home/z/my-project/download/sample_report_grocery.xlsx"
wb.save(output_path)
print(f"Report saved to: {output_path}")
