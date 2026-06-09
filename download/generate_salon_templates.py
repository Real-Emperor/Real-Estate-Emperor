#!/usr/bin/env python3
"""
Generate 2 UAE Salon Monthly Sales Report Excel files:
  1. EMPTY TEMPLATE with formulas (yellow input cells)
  2. FILLED SAMPLE with realistic data for January 2026

Both use the xlsx skill base template with elegant palette.
"""

import sys, os
sys.path.insert(0, '/home/z/my-project/skills/xlsx/templates')
sys.path.insert(0, '/home/z/my-project/skills/xlsx')
from base import *

from openpyxl import Workbook
from openpyxl.chart import Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList
import random

# ── Activate elegant palette ──
use_palette_explicit("elegant")

YELLOW_FILL = PatternFill("solid", fgColor="FFFDE7")
YELLOW_INPUT_FONT = Font(name=FONT_NAME, size=11, color=NEUTRAL_900)

FOOTER_TEXT = "Real Estate Emperor Property Management L.L.C."
AED_FMT = "#,##0"
AED_FMT2 = "#,##0.00"
PCT_FMT = "0.0%"


def write_footer(ws, row, col_start, col_end):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=FOOTER_TEXT)
    cell.font = font_caption()
    cell.alignment = Alignment(horizontal="left", vertical="center")


def apply_yellow_input(cell):
    """Mark a cell as yellow user-input."""
    cell.fill = YELLOW_FILL
    cell.font = YELLOW_INPUT_FONT


def add_conditional_variance(ws, col_letter, start_row, end_row):
    """Add red/green conditional formatting on a variance column."""
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT)
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="greaterThanOrEqual", formula=["0"],
                   fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT)
    )


def add_conditional_status(ws, col_letter, start_row, end_row):
    """Add red/green conditional formatting on a Status text column."""
    rng = f"{col_letter}{start_row}:{col_letter}{end_row}"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Over Budget"'],
                   fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT)
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="equal", formula=['"Under Budget"'],
                   fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT)
    )


# ════════════════════════════════════════════════════════════════
#  SHEET BUILDERS (used by both template and sample)
# ════════════════════════════════════════════════════════════════

def build_daily_sales(wb, title_suffix, fill_data=False, daily_data=None):
    """Build the Daily Sales sheet. Returns the worksheet."""
    ws = wb.active
    ws.title = "Daily Sales"

    LAST_COL = 14  # B through N (13 data cols + margin A)
    setup_sheet(ws, title=f"Daily Sales Record — {title_suffix}", last_col=LAST_COL)

    # Instruction row (row 3)
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=LAST_COL)
    instr = ws.cell(row=3, column=2,
                    value="Enter daily data in the yellow-highlighted columns. All other columns calculate automatically.")
    instr.font = font_caption()
    instr.alignment = align_text()
    ws.row_dimensions[3].height = 18

    # Header row (row 5, with row 4 spacer)
    HDR_ROW = 5
    headers = ["Day", "Date", "Walk-Ins", "Appointments", "Total Clients",
               "Service Revenue", "Product Revenue", "Total Revenue",
               "Cash", "Card", "Online", "Total Payments", "Payment Variance"]
    for i, h in enumerate(headers):
        ws.cell(row=HDR_ROW, column=2 + i, value=h)
    style_header_row(ws, HDR_ROW, 2, 2 + len(headers) - 1)
    ws.row_dimensions[4].height = ROW_HEIGHTS["spacer"]

    DATA_START = 6  # row 6 = day 1
    DATA_END = 36   # row 36 = day 31
    # Column mapping (1-indexed from B):
    # B=Day, C=Date, D=Walk-Ins, E=Appointments, F=TotalClients,
    # G=ServiceRev, H=ProductRev, I=TotalRev,
    # J=Cash, K=Card, L=Online, M=TotalPayments, N=PaymentVariance

    col_day = 2   # B
    col_date = 3  # C
    col_walk = 4  # D
    col_appt = 5  # E
    col_tcli = 6  # F
    col_srev = 7  # G
    col_prev = 8  # H
    col_trev = 9  # I
    col_cash = 10 # J
    col_card = 11 # K
    col_onln = 12 # L
    col_tpay = 13 # M
    col_pvar = 14 # N

    for day_num in range(1, 32):
        r = DATA_START + day_num - 1
        # Day
        ws.cell(row=r, column=col_day, value=day_num)
        ws.cell(row=r, column=col_day).alignment = align_date()

        # Date - blank or filled
        if fill_data and daily_data:
            ws.cell(row=r, column=col_date, value=daily_data[day_num-1]["date"])
            ws.cell(row=r, column=col_date).number_format = "DD-MMM-YYYY"
        else:
            apply_yellow_input(ws.cell(row=r, column=col_date))

        # Walk-Ins (yellow input)
        if fill_data and daily_data:
            ws.cell(row=r, column=col_walk, value=daily_data[day_num-1]["walk_ins"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_walk))
        ws.cell(row=r, column=col_walk).number_format = AED_FMT
        ws.cell(row=r, column=col_walk).alignment = align_number()

        # Appointments (yellow input)
        if fill_data and daily_data:
            ws.cell(row=r, column=col_appt, value=daily_data[day_num-1]["appointments"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_appt))
        ws.cell(row=r, column=col_appt).number_format = AED_FMT
        ws.cell(row=r, column=col_appt).alignment = align_number()

        # Total Clients = Walk-Ins + Appointments (FORMULA)
        walk_ref = f"{get_column_letter(col_walk)}{r}"
        appt_ref = f"{get_column_letter(col_appt)}{r}"
        ws.cell(row=r, column=col_tcli).value = f"={walk_ref}+{appt_ref}"
        ws.cell(row=r, column=col_tcli).number_format = AED_FMT
        ws.cell(row=r, column=col_tcli).alignment = align_number()

        # Service Revenue (yellow input)
        if fill_data and daily_data:
            ws.cell(row=r, column=col_srev, value=daily_data[day_num-1]["service_rev"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_srev))
        ws.cell(row=r, column=col_srev).number_format = AED_FMT
        ws.cell(row=r, column=col_srev).alignment = align_number()

        # Product Revenue (yellow input)
        if fill_data and daily_data:
            ws.cell(row=r, column=col_prev, value=daily_data[day_num-1]["product_rev"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_prev))
        ws.cell(row=r, column=col_prev).number_format = AED_FMT
        ws.cell(row=r, column=col_prev).alignment = align_number()

        # Total Revenue = Service + Product (FORMULA)
        srev_ref = f"{get_column_letter(col_srev)}{r}"
        prev_ref = f"{get_column_letter(col_prev)}{r}"
        ws.cell(row=r, column=col_trev).value = f"={srev_ref}+{prev_ref}"
        ws.cell(row=r, column=col_trev).number_format = AED_FMT
        ws.cell(row=r, column=col_trev).alignment = align_number()

        # Cash (yellow input)
        if fill_data and daily_data:
            ws.cell(row=r, column=col_cash, value=daily_data[day_num-1]["cash"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_cash))
        ws.cell(row=r, column=col_cash).number_format = AED_FMT
        ws.cell(row=r, column=col_cash).alignment = align_number()

        # Card (yellow input)
        if fill_data and daily_data:
            ws.cell(row=r, column=col_card, value=daily_data[day_num-1]["card"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_card))
        ws.cell(row=r, column=col_card).number_format = AED_FMT
        ws.cell(row=r, column=col_card).alignment = align_number()

        # Online (yellow input)
        if fill_data and daily_data:
            ws.cell(row=r, column=col_onln, value=daily_data[day_num-1]["online"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_onln))
        ws.cell(row=r, column=col_onln).number_format = AED_FMT
        ws.cell(row=r, column=col_onln).alignment = align_number()

        # Total Payments = Cash + Card + Online (FORMULA)
        cash_ref = f"{get_column_letter(col_cash)}{r}"
        card_ref = f"{get_column_letter(col_card)}{r}"
        onln_ref = f"{get_column_letter(col_onln)}{r}"
        ws.cell(row=r, column=col_tpay).value = f"={cash_ref}+{card_ref}+{onln_ref}"
        ws.cell(row=r, column=col_tpay).number_format = AED_FMT
        ws.cell(row=r, column=col_tpay).alignment = align_number()

        # Payment Variance = Total Revenue - Total Payments (FORMULA)
        trev_ref = f"{get_column_letter(col_trev)}{r}"
        tpay_ref = f"{get_column_letter(col_tpay)}{r}"
        ws.cell(row=r, column=col_pvar).value = f"={trev_ref}-{tpay_ref}"
        ws.cell(row=r, column=col_pvar).number_format = AED_FMT
        ws.cell(row=r, column=col_pvar).alignment = align_number()

        # Style the data row (alternating)
        style_data_row(ws, r, 2, LAST_COL, day_num - 1)

        # Re-apply alignments and yellow fills that style_data_row may override
        ws.cell(row=r, column=col_day).alignment = align_date()
        for c in [col_walk, col_appt, col_srev, col_prev, col_cash, col_card, col_onln, col_date]:
            if ws.cell(row=r, column=c).fill == PatternFill() or (not fill_data):
                apply_yellow_input(ws.cell(row=r, column=c))
            ws.cell(row=r, column=c).alignment = align_number()
        for c in [col_tcli, col_trev, col_tpay, col_pvar]:
            ws.cell(row=r, column=c).alignment = align_number()
        ws.cell(row=r, column=col_date).alignment = align_date()
        if fill_data and daily_data:
            ws.cell(row=r, column=col_date).number_format = "DD-MMM-YYYY"

    # ── TOTALS row (row 37) ──
    TOTAL_ROW = DATA_END + 1  # 37
    ws.cell(row=TOTAL_ROW, column=col_day, value="TOTAL")
    ws.cell(row=TOTAL_ROW, column=col_day).alignment = align_text()
    for c in [col_walk, col_appt, col_tcli, col_srev, col_prev, col_trev,
              col_cash, col_card, col_onln, col_tpay, col_pvar]:
        cl = get_column_letter(c)
        ws.cell(row=TOTAL_ROW, column=c).value = f"=SUM({cl}{DATA_START}:{cl}{DATA_END})"
        ws.cell(row=TOTAL_ROW, column=c).number_format = AED_FMT
        ws.cell(row=TOTAL_ROW, column=c).alignment = align_number()
    style_total_row(ws, TOTAL_ROW, 2, LAST_COL)
    # Re-apply number formats and alignment
    for c in [col_walk, col_appt, col_tcli, col_srev, col_prev, col_trev,
              col_cash, col_card, col_onln, col_tpay, col_pvar]:
        ws.cell(row=TOTAL_ROW, column=c).number_format = AED_FMT
        ws.cell(row=TOTAL_ROW, column=c).alignment = align_number()
    ws.cell(row=TOTAL_ROW, column=col_day).alignment = align_text()

    # ── AVERAGES row (row 38) ──
    AVG_ROW = TOTAL_ROW + 1  # 38
    ws.cell(row=AVG_ROW, column=col_day, value="AVERAGE")
    ws.cell(row=AVG_ROW, column=col_day).alignment = align_text()
    for c in [col_walk, col_appt, col_tcli, col_srev, col_prev, col_trev,
              col_cash, col_card, col_onln, col_tpay, col_pvar]:
        cl = get_column_letter(c)
        ws.cell(row=AVG_ROW, column=c).value = f"=AVERAGE({cl}{DATA_START}:{cl}{DATA_END})"
        ws.cell(row=AVG_ROW, column=c).number_format = AED_FMT
        ws.cell(row=AVG_ROW, column=c).alignment = align_number()
    style_total_row(ws, AVG_ROW, 2, LAST_COL)
    for c in [col_walk, col_appt, col_tcli, col_srev, col_prev, col_trev,
              col_cash, col_card, col_onln, col_tpay, col_pvar]:
        ws.cell(row=AVG_ROW, column=c).number_format = AED_FMT
        ws.cell(row=AVG_ROW, column=c).alignment = align_number()
    ws.cell(row=AVG_ROW, column=col_day).alignment = align_text()

    # ── BEST DAY row (row 39) ──
    BEST_ROW = AVG_ROW + 1  # 39
    ws.cell(row=BEST_ROW, column=col_day, value="BEST DAY")
    ws.cell(row=BEST_ROW, column=col_day).alignment = align_text()
    for c in [col_walk, col_appt, col_tcli, col_srev, col_prev, col_trev,
              col_cash, col_card, col_onln, col_tpay, col_pvar]:
        cl = get_column_letter(c)
        ws.cell(row=BEST_ROW, column=c).value = f"=MAX({cl}{DATA_START}:{cl}{DATA_END})"
        ws.cell(row=BEST_ROW, column=c).number_format = AED_FMT
        ws.cell(row=BEST_ROW, column=c).alignment = align_number()
    style_total_row(ws, BEST_ROW, 2, LAST_COL)
    for c in [col_walk, col_appt, col_tcli, col_srev, col_prev, col_trev,
              col_cash, col_card, col_onln, col_tpay, col_pvar]:
        ws.cell(row=BEST_ROW, column=c).number_format = AED_FMT
        ws.cell(row=BEST_ROW, column=c).alignment = align_number()
    ws.cell(row=BEST_ROW, column=col_day).alignment = align_text()

    # ── Conditional formatting: Payment Variance ──
    add_conditional_variance(ws, get_column_letter(col_pvar), DATA_START, DATA_END)

    # ── Charts ──
    # Bar chart: Daily Total Revenue
    chart_bar = create_bar_chart(chart_type="col", grouping="clustered", width=22, height=12)
    cats = Reference(ws, min_col=col_day, min_row=DATA_START, max_row=DATA_END)
    vals = Reference(ws, min_col=col_trev, min_row=HDR_ROW, max_row=DATA_END)
    chart_bar.add_data(vals, titles_from_data=True)
    chart_bar.set_categories(cats)
    setup_chart_titles(chart_bar, title="Daily Total Revenue", y_title="AED", x_title="Day")
    apply_chart_colors(chart_bar, [CHART_COLORS[0]])

    # Pie chart: Payment Method Split (from totals row)
    chart_pie = create_pie_chart(width=14, height=12)
    pie_cats = Reference(ws, min_col=col_cash, max_col=col_onln, min_row=HDR_ROW, max_row=HDR_ROW)
    pie_vals = Reference(ws, min_col=col_cash, max_col=col_onln, min_row=TOTAL_ROW, max_row=TOTAL_ROW)
    chart_pie.add_data(pie_vals, titles_from_data=True, from_rows=True)
    chart_pie.set_categories(pie_cats)
    setup_chart_titles(chart_pie, title="Payment Method Split")
    apply_pie_colors(chart_pie, 3)
    chart_pie.dataLabels = DataLabelList(showPercent=True, showCatName=True, showVal=False)

    chart_anchor = BEST_ROW + 2
    ws.add_chart(chart_bar, f"B{chart_anchor}")
    ws.add_chart(chart_pie, f"H{chart_anchor}")

    # Footer
    footer_row = chart_anchor + 18
    write_footer(ws, footer_row, 2, LAST_COL)

    auto_fit_columns(ws, header_row=HDR_ROW, data_start_row=DATA_START)

    return ws


def build_service_breakdown(wb, title_suffix, fill_data=False, svc_data=None):
    """Build the Service Breakdown sheet."""
    ws = wb.create_sheet("Service Breakdown")

    LAST_COL = 8  # B through H
    setup_sheet(ws, title=f"Service Revenue Breakdown — {title_suffix}", last_col=LAST_COL)

    HDR_ROW = 4
    headers = ["Service", "Bookings", "Avg Price (AED)", "Revenue", "% of Total", "Target", "vs Target"]
    for i, h in enumerate(headers):
        ws.cell(row=HDR_ROW, column=2 + i, value=h)
    style_header_row(ws, HDR_ROW, 2, 2 + len(headers) - 1)

    SERVICES = [
        "Haircut & Styling", "Hair Coloring", "Highlights & Balayage",
        "Keratin Treatment", "Bridal Package", "Facial Treatment",
        "Manicure & Pedicure", "Henna Design", "Hair Extensions",
        "Scalp Treatment", "Makeup", "Other"
    ]

    # Column mapping
    col_svc = 2    # B
    col_bk = 3     # C
    col_ap = 4     # D
    col_rev = 5    # E
    col_pct = 6    # F
    col_tgt = 7    # G
    col_vs = 8     # H

    DATA_START = HDR_ROW + 1
    DATA_END = DATA_START + len(SERVICES) - 1

    for idx, svc_name in enumerate(SERVICES):
        r = DATA_START + idx
        ws.cell(row=r, column=col_svc, value=svc_name)
        ws.cell(row=r, column=col_svc).alignment = align_text()

        # Bookings (yellow)
        if fill_data and svc_data and idx < len(svc_data):
            ws.cell(row=r, column=col_bk, value=svc_data[idx]["bookings"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_bk))
        ws.cell(row=r, column=col_bk).number_format = AED_FMT
        ws.cell(row=r, column=col_bk).alignment = align_number()

        # Avg Price (yellow)
        if fill_data and svc_data and idx < len(svc_data):
            ws.cell(row=r, column=col_ap, value=svc_data[idx]["avg_price"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_ap))
        ws.cell(row=r, column=col_ap).number_format = AED_FMT
        ws.cell(row=r, column=col_ap).alignment = align_number()

        # Revenue = Bookings * Avg Price (FORMULA)
        bk_ref = f"{get_column_letter(col_bk)}{r}"
        ap_ref = f"{get_column_letter(col_ap)}{r}"
        ws.cell(row=r, column=col_rev).value = f"={bk_ref}*{ap_ref}"
        ws.cell(row=r, column=col_rev).number_format = AED_FMT
        ws.cell(row=r, column=col_rev).alignment = align_number()

        # % of Total = IFERROR(Revenue / SUM(Revenue), 0) (FORMULA)
        rev_range = f"{get_column_letter(col_rev)}${DATA_START}:{get_column_letter(col_rev)}${DATA_END}"
        rev_ref = f"{get_column_letter(col_rev)}{r}"
        ws.cell(row=r, column=col_pct).value = f"=IFERROR({rev_ref}/SUM({rev_range}),0)"
        ws.cell(row=r, column=col_pct).number_format = PCT_FMT
        ws.cell(row=r, column=col_pct).alignment = align_number()

        # Target (yellow)
        if fill_data and svc_data and idx < len(svc_data):
            ws.cell(row=r, column=col_tgt, value=svc_data[idx]["target"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_tgt))
        ws.cell(row=r, column=col_tgt).number_format = AED_FMT
        ws.cell(row=r, column=col_tgt).alignment = align_number()

        # vs Target = Revenue - Target (FORMULA)
        tgt_ref = f"{get_column_letter(col_tgt)}{r}"
        ws.cell(row=r, column=col_vs).value = f"={rev_ref}-{tgt_ref}"
        ws.cell(row=r, column=col_vs).number_format = AED_FMT
        ws.cell(row=r, column=col_vs).alignment = align_number()

        style_data_row(ws, r, 2, LAST_COL, idx)
        # Re-apply yellow fills and alignments
        ws.cell(row=r, column=col_svc).alignment = align_text()
        for c in [col_bk, col_ap, col_tgt]:
            if not fill_data:
                apply_yellow_input(ws.cell(row=r, column=c))
            ws.cell(row=r, column=c).alignment = align_number()
        for c in [col_rev, col_pct, col_vs]:
            ws.cell(row=r, column=c).alignment = align_number()

    # TOTALS row
    TOTAL_ROW = DATA_END + 1
    ws.cell(row=TOTAL_ROW, column=col_svc, value="TOTAL")
    ws.cell(row=TOTAL_ROW, column=col_svc).alignment = align_text()
    for c in [col_bk, col_rev, col_tgt, col_vs]:
        cl = get_column_letter(c)
        ws.cell(row=TOTAL_ROW, column=c).value = f"=SUM({cl}{DATA_START}:{cl}{DATA_END})"
        ws.cell(row=TOTAL_ROW, column=c).number_format = AED_FMT
        ws.cell(row=TOTAL_ROW, column=c).alignment = align_number()
    # % of Total sum should be 100%
    cl_pct = get_column_letter(col_pct)
    ws.cell(row=TOTAL_ROW, column=col_pct).value = f"=SUM({cl_pct}{DATA_START}:{cl_pct}{DATA_END})"
    ws.cell(row=TOTAL_ROW, column=col_pct).number_format = PCT_FMT
    ws.cell(row=TOTAL_ROW, column=col_pct).alignment = align_number()
    ws.cell(row=TOTAL_ROW, column=col_ap, value="—")
    ws.cell(row=TOTAL_ROW, column=col_ap).alignment = align_text()
    style_total_row(ws, TOTAL_ROW, 2, LAST_COL)
    ws.cell(row=TOTAL_ROW, column=col_svc).alignment = align_text()
    for c in [col_bk, col_rev, col_tgt, col_vs]:
        ws.cell(row=TOTAL_ROW, column=c).number_format = AED_FMT
        ws.cell(row=TOTAL_ROW, column=c).alignment = align_number()
    ws.cell(row=TOTAL_ROW, column=col_pct).number_format = PCT_FMT
    ws.cell(row=TOTAL_ROW, column=col_pct).alignment = align_number()

    # Conditional formatting: vs Target
    add_conditional_variance(ws, get_column_letter(col_vs), DATA_START, DATA_END)

    # ── Charts ──
    # Pie chart: Revenue by Service
    chart_pie = create_pie_chart(width=16, height=12)
    pie_cats = Reference(ws, min_col=col_svc, min_row=DATA_START, max_row=DATA_END)
    pie_vals = Reference(ws, min_col=col_rev, min_row=DATA_START, max_row=DATA_END)
    chart_pie.add_data(pie_vals, titles_from_data=False)
    chart_pie.set_categories(pie_cats)
    setup_chart_titles(chart_pie, title="Revenue by Service")
    apply_pie_colors(chart_pie, len(SERVICES))
    chart_pie.dataLabels = DataLabelList(showPercent=True, showCatName=False, showVal=False)

    # Bar chart: Bookings by Service
    chart_bar = create_bar_chart(chart_type="col", grouping="clustered", width=20, height=12)
    bar_cats = Reference(ws, min_col=col_svc, min_row=DATA_START, max_row=DATA_END)
    bar_vals = Reference(ws, min_col=col_bk, min_row=HDR_ROW, max_row=DATA_END)
    chart_bar.add_data(bar_vals, titles_from_data=True)
    chart_bar.set_categories(bar_cats)
    setup_chart_titles(chart_bar, title="Bookings by Service", y_title="Bookings")
    apply_chart_colors(chart_bar, [CHART_COLORS[0]])

    chart_anchor = TOTAL_ROW + 2
    ws.add_chart(chart_pie, f"B{chart_anchor}")
    ws.add_chart(chart_bar, f"B{chart_anchor + 16}")

    # Footer
    write_footer(ws, chart_anchor + 32, 2, LAST_COL)

    auto_fit_columns(ws, header_row=HDR_ROW, data_start_row=DATA_START)

    return ws


def build_staff_performance(wb, title_suffix, fill_data=False, staff_data=None):
    """Build the Staff Performance sheet."""
    ws = wb.create_sheet("Staff Performance")

    LAST_COL = 7  # B through G
    setup_sheet(ws, title=f"Staff Performance — {title_suffix}", last_col=LAST_COL)

    HDR_ROW = 4
    headers = ["Staff Name", "Clients Served", "Revenue Generated", "Avg Ticket", "Tips", "Rating (1-5)"]
    for i, h in enumerate(headers):
        ws.cell(row=HDR_ROW, column=2 + i, value=h)
    style_header_row(ws, HDR_ROW, 2, 2 + len(headers) - 1)

    col_name = 2   # B
    col_cli = 3    # C
    col_rev = 4    # D
    col_avg = 5    # E
    col_tips = 6   # F
    col_rat = 7    # G

    NUM_STAFF = 8
    DATA_START = HDR_ROW + 1
    DATA_END = DATA_START + NUM_STAFF - 1

    for idx in range(NUM_STAFF):
        r = DATA_START + idx

        # Staff Name (yellow)
        if fill_data and staff_data and idx < len(staff_data):
            ws.cell(row=r, column=col_name, value=staff_data[idx]["name"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_name))
        ws.cell(row=r, column=col_name).alignment = align_text()

        # Clients Served (yellow)
        if fill_data and staff_data and idx < len(staff_data):
            ws.cell(row=r, column=col_cli, value=staff_data[idx]["clients"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_cli))
        ws.cell(row=r, column=col_cli).number_format = AED_FMT
        ws.cell(row=r, column=col_cli).alignment = align_number()

        # Revenue Generated (yellow)
        if fill_data and staff_data and idx < len(staff_data):
            ws.cell(row=r, column=col_rev, value=staff_data[idx]["revenue"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_rev))
        ws.cell(row=r, column=col_rev).number_format = AED_FMT
        ws.cell(row=r, column=col_rev).alignment = align_number()

        # Avg Ticket = IFERROR(Revenue/Clients, 0) (FORMULA)
        rev_ref = f"{get_column_letter(col_rev)}{r}"
        cli_ref = f"{get_column_letter(col_cli)}{r}"
        ws.cell(row=r, column=col_avg).value = f"=IFERROR({rev_ref}/{cli_ref},0)"
        ws.cell(row=r, column=col_avg).number_format = AED_FMT2
        ws.cell(row=r, column=col_avg).alignment = align_number()

        # Tips (yellow)
        if fill_data and staff_data and idx < len(staff_data):
            ws.cell(row=r, column=col_tips, value=staff_data[idx]["tips"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_tips))
        ws.cell(row=r, column=col_tips).number_format = AED_FMT
        ws.cell(row=r, column=col_tips).alignment = align_number()

        # Rating (yellow)
        if fill_data and staff_data and idx < len(staff_data):
            ws.cell(row=r, column=col_rat, value=staff_data[idx]["rating"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_rat))
        ws.cell(row=r, column=col_rat).number_format = "0.0"
        ws.cell(row=r, column=col_rat).alignment = align_number()

        style_data_row(ws, r, 2, LAST_COL, idx)
        # Re-apply yellow fills and alignments
        ws.cell(row=r, column=col_name).alignment = align_text()
        for c in [col_cli, col_rev, col_tips, col_rat]:
            if not fill_data:
                apply_yellow_input(ws.cell(row=r, column=c))
            ws.cell(row=r, column=c).alignment = align_number()
        ws.cell(row=r, column=col_avg).alignment = align_number()

    # TOTALS row
    TOTAL_ROW = DATA_END + 1
    ws.cell(row=TOTAL_ROW, column=col_name, value="TOTAL")
    ws.cell(row=TOTAL_ROW, column=col_name).alignment = align_text()
    for c in [col_cli, col_rev, col_tips]:
        cl = get_column_letter(c)
        ws.cell(row=TOTAL_ROW, column=c).value = f"=SUM({cl}{DATA_START}:{cl}{DATA_END})"
        ws.cell(row=TOTAL_ROW, column=c).number_format = AED_FMT
        ws.cell(row=TOTAL_ROW, column=c).alignment = align_number()
    # Avg Ticket for total = IFERROR(total rev / total clients, 0)
    total_rev_ref = f"{get_column_letter(col_rev)}{TOTAL_ROW}"
    total_cli_ref = f"{get_column_letter(col_cli)}{TOTAL_ROW}"
    ws.cell(row=TOTAL_ROW, column=col_avg).value = f"=IFERROR({total_rev_ref}/{total_cli_ref},0)"
    ws.cell(row=TOTAL_ROW, column=col_avg).number_format = AED_FMT2
    ws.cell(row=TOTAL_ROW, column=col_avg).alignment = align_number()
    # Average rating
    cl_rat = get_column_letter(col_rat)
    ws.cell(row=TOTAL_ROW, column=col_rat).value = f"=AVERAGE({cl_rat}{DATA_START}:{cl_rat}{DATA_END})"
    ws.cell(row=TOTAL_ROW, column=col_rat).number_format = "0.0"
    ws.cell(row=TOTAL_ROW, column=col_rat).alignment = align_number()
    style_total_row(ws, TOTAL_ROW, 2, LAST_COL)
    ws.cell(row=TOTAL_ROW, column=col_name).alignment = align_text()
    for c in [col_cli, col_rev, col_tips]:
        ws.cell(row=TOTAL_ROW, column=c).number_format = AED_FMT
        ws.cell(row=TOTAL_ROW, column=c).alignment = align_number()
    ws.cell(row=TOTAL_ROW, column=col_avg).number_format = AED_FMT2
    ws.cell(row=TOTAL_ROW, column=col_avg).alignment = align_number()
    ws.cell(row=TOTAL_ROW, column=col_rat).number_format = "0.0"
    ws.cell(row=TOTAL_ROW, column=col_rat).alignment = align_number()

    # AVERAGES row
    AVG_ROW = TOTAL_ROW + 1
    ws.cell(row=AVG_ROW, column=col_name, value="AVERAGE")
    ws.cell(row=AVG_ROW, column=col_name).alignment = align_text()
    for c in [col_cli, col_rev, col_tips]:
        cl = get_column_letter(c)
        ws.cell(row=AVG_ROW, column=c).value = f"=AVERAGE({cl}{DATA_START}:{cl}{DATA_END})"
        ws.cell(row=AVG_ROW, column=c).number_format = AED_FMT
        ws.cell(row=AVG_ROW, column=c).alignment = align_number()
    ws.cell(row=AVG_ROW, column=col_avg).value = f"=AVERAGE({get_column_letter(col_avg)}{DATA_START}:{get_column_letter(col_avg)}{DATA_END})"
    ws.cell(row=AVG_ROW, column=col_avg).number_format = AED_FMT2
    ws.cell(row=AVG_ROW, column=col_avg).alignment = align_number()
    ws.cell(row=AVG_ROW, column=col_rat).value = f"=AVERAGE({cl_rat}{DATA_START}:{cl_rat}{DATA_END})"
    ws.cell(row=AVG_ROW, column=col_rat).number_format = "0.0"
    ws.cell(row=AVG_ROW, column=col_rat).alignment = align_number()
    style_total_row(ws, AVG_ROW, 2, LAST_COL)
    ws.cell(row=AVG_ROW, column=col_name).alignment = align_text()
    for c in [col_cli, col_rev, col_tips]:
        ws.cell(row=AVG_ROW, column=c).number_format = AED_FMT
        ws.cell(row=AVG_ROW, column=c).alignment = align_number()
    ws.cell(row=AVG_ROW, column=col_avg).number_format = AED_FMT2
    ws.cell(row=AVG_ROW, column=col_avg).alignment = align_number()
    ws.cell(row=AVG_ROW, column=col_rat).number_format = "0.0"
    ws.cell(row=AVG_ROW, column=col_rat).alignment = align_number()

    # ── Bar chart: Revenue by Staff ──
    chart_hbar = create_bar_chart(chart_type="bar", grouping="clustered", width=18, height=10)
    hbar_cats = Reference(ws, min_col=col_name, min_row=DATA_START, max_row=DATA_END)
    hbar_vals = Reference(ws, min_col=col_rev, min_row=HDR_ROW, max_row=DATA_END)
    chart_hbar.add_data(hbar_vals, titles_from_data=True)
    chart_hbar.set_categories(hbar_cats)
    setup_chart_titles(chart_hbar, title="Revenue by Staff Member", x_title="AED")
    apply_chart_colors(chart_hbar, [CHART_COLORS[0]])

    chart_anchor = AVG_ROW + 2
    ws.add_chart(chart_hbar, f"B{chart_anchor}")

    # Footer
    write_footer(ws, chart_anchor + 14, 2, LAST_COL)

    auto_fit_columns(ws, header_row=HDR_ROW, data_start_row=DATA_START)

    return ws


def build_expense_tracker(wb, title_suffix, fill_data=False, exp_data=None):
    """Build the Expense Tracker sheet."""
    ws = wb.create_sheet("Expense Tracker")

    LAST_COL = 7  # B through G
    setup_sheet(ws, title=f"Monthly Expenses — {title_suffix}", last_col=LAST_COL)

    HDR_ROW = 4
    headers = ["Category", "Budget (AED)", "Actual (AED)", "Variance", "Variance %", "Status"]
    for i, h in enumerate(headers):
        ws.cell(row=HDR_ROW, column=2 + i, value=h)
    style_header_row(ws, HDR_ROW, 2, 2 + len(headers) - 1)

    CATEGORIES = [
        "Rent", "Staff Salaries", "Product Supplies", "Utilities",
        "Marketing", "Insurance", "Maintenance", "POS/Banking Fees",
        "Delivery Commissions", "Licenses & Permits", "Miscellaneous"
    ]

    col_cat = 2   # B
    col_bud = 3   # C
    col_act = 4   # D
    col_var = 5   # E
    col_vpct = 6  # F
    col_stat = 7  # G

    DATA_START = HDR_ROW + 1
    DATA_END = DATA_START + len(CATEGORIES) - 1

    for idx, cat_name in enumerate(CATEGORIES):
        r = DATA_START + idx

        ws.cell(row=r, column=col_cat, value=cat_name)
        ws.cell(row=r, column=col_cat).alignment = align_text()

        # Budget (yellow)
        if fill_data and exp_data and idx < len(exp_data):
            ws.cell(row=r, column=col_bud, value=exp_data[idx]["budget"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_bud))
        ws.cell(row=r, column=col_bud).number_format = AED_FMT
        ws.cell(row=r, column=col_bud).alignment = align_number()

        # Actual (yellow)
        if fill_data and exp_data and idx < len(exp_data):
            ws.cell(row=r, column=col_act, value=exp_data[idx]["actual"])
        else:
            apply_yellow_input(ws.cell(row=r, column=col_act))
        ws.cell(row=r, column=col_act).number_format = AED_FMT
        ws.cell(row=r, column=col_act).alignment = align_number()

        # Variance = Budget - Actual (FORMULA)
        bud_ref = f"{get_column_letter(col_bud)}{r}"
        act_ref = f"{get_column_letter(col_act)}{r}"
        ws.cell(row=r, column=col_var).value = f"={bud_ref}-{act_ref}"
        ws.cell(row=r, column=col_var).number_format = AED_FMT
        ws.cell(row=r, column=col_var).alignment = align_number()

        # Variance % = IFERROR(Variance/Budget, 0) (FORMULA)
        var_ref = f"{get_column_letter(col_var)}{r}"
        ws.cell(row=r, column=col_vpct).value = f"=IFERROR({var_ref}/{bud_ref},0)"
        ws.cell(row=r, column=col_vpct).number_format = PCT_FMT
        ws.cell(row=r, column=col_vpct).alignment = align_number()

        # Status = IF(Variance<0,"Over Budget",IF(Variance=0,"On Budget","Under Budget")) (FORMULA)
        ws.cell(row=r, column=col_stat).value = f'=IF({var_ref}<0,"Over Budget",IF({var_ref}=0,"On Budget","Under Budget"))'
        ws.cell(row=r, column=col_stat).alignment = align_date()

        style_data_row(ws, r, 2, LAST_COL, idx)
        # Re-apply yellow fills and alignments
        ws.cell(row=r, column=col_cat).alignment = align_text()
        for c in [col_bud, col_act]:
            if not fill_data:
                apply_yellow_input(ws.cell(row=r, column=c))
            ws.cell(row=r, column=c).alignment = align_number()
        ws.cell(row=r, column=col_var).alignment = align_number()
        ws.cell(row=r, column=col_vpct).alignment = align_number()
        ws.cell(row=r, column=col_stat).alignment = align_date()

    # TOTALS row
    TOTAL_ROW = DATA_END + 1
    ws.cell(row=TOTAL_ROW, column=col_cat, value="TOTAL")
    ws.cell(row=TOTAL_ROW, column=col_cat).alignment = align_text()
    for c in [col_bud, col_act, col_var]:
        cl = get_column_letter(c)
        ws.cell(row=TOTAL_ROW, column=c).value = f"=SUM({cl}{DATA_START}:{cl}{DATA_END})"
        ws.cell(row=TOTAL_ROW, column=c).number_format = AED_FMT
        ws.cell(row=TOTAL_ROW, column=c).alignment = align_number()
    # Variance % total
    total_bud_ref = f"{get_column_letter(col_bud)}{TOTAL_ROW}"
    total_var_ref = f"{get_column_letter(col_var)}{TOTAL_ROW}"
    ws.cell(row=TOTAL_ROW, column=col_vpct).value = f"=IFERROR({total_var_ref}/{total_bud_ref},0)"
    ws.cell(row=TOTAL_ROW, column=col_vpct).number_format = PCT_FMT
    ws.cell(row=TOTAL_ROW, column=col_vpct).alignment = align_number()
    ws.cell(row=TOTAL_ROW, column=col_stat, value="—")
    ws.cell(row=TOTAL_ROW, column=col_stat).alignment = align_text()
    style_total_row(ws, TOTAL_ROW, 2, LAST_COL)
    ws.cell(row=TOTAL_ROW, column=col_cat).alignment = align_text()
    for c in [col_bud, col_act, col_var]:
        ws.cell(row=TOTAL_ROW, column=c).number_format = AED_FMT
        ws.cell(row=TOTAL_ROW, column=c).alignment = align_number()
    ws.cell(row=TOTAL_ROW, column=col_vpct).number_format = PCT_FMT
    ws.cell(row=TOTAL_ROW, column=col_vpct).alignment = align_number()

    # Conditional formatting: Variance
    add_conditional_variance(ws, get_column_letter(col_var), DATA_START, DATA_END)
    # Conditional formatting: Status
    add_conditional_status(ws, get_column_letter(col_stat), DATA_START, DATA_END)

    # ── Bar chart: Budget vs Actual ──
    chart_bar = create_bar_chart(chart_type="col", grouping="clustered", width=20, height=12)
    bar_cats = Reference(ws, min_col=col_cat, min_row=DATA_START, max_row=DATA_END)
    bud_vals = Reference(ws, min_col=col_bud, min_row=HDR_ROW, max_row=DATA_END)
    act_vals = Reference(ws, min_col=col_act, min_row=HDR_ROW, max_row=DATA_END)
    chart_bar.add_data(bud_vals, titles_from_data=True)
    chart_bar.add_data(act_vals, titles_from_data=True)
    chart_bar.set_categories(bar_cats)
    setup_chart_titles(chart_bar, title="Budget vs Actual", y_title="AED")
    apply_chart_colors(chart_bar, [CHART_COLORS[0], CHART_COLORS[2]])

    chart_anchor = TOTAL_ROW + 2
    ws.add_chart(chart_bar, f"B{chart_anchor}")

    # Footer
    write_footer(ws, chart_anchor + 16, 2, LAST_COL)

    auto_fit_columns(ws, header_row=HDR_ROW, data_start_row=DATA_START)

    return ws


def build_monthly_summary(wb, title_suffix, fill_data=False, six_month_data=None):
    """Build the Monthly Summary sheet with cross-sheet formulas."""
    ws = wb.create_sheet("Monthly Summary")

    LAST_COL = 3  # B and C
    setup_sheet(ws, title=f"Monthly Summary — {title_suffix}", last_col=LAST_COL)

    HDR_ROW = 4
    headers = ["Metric", "Value"]
    for i, h in enumerate(headers):
        ws.cell(row=HDR_ROW, column=2 + i, value=h)
    style_header_row(ws, HDR_ROW, 2, 3)

    col_metric = 2  # B
    col_value = 3   # C

    # Metrics with formulas referencing other sheets
    metrics = [
        ("Total Revenue", f"='Daily Sales'!I37"),
        ("Total Expenses", f"='Expense Tracker'!D{5 + 11}"),  # TOTAL row = DATA_START(5) + 11 = 16
        ("Net Profit", f"=C5-C6"),
        ("Profit Margin", f"=IFERROR(C7/C5,0)"),
        ("Total Clients", f"='Daily Sales'!F37"),
        ("Avg Revenue per Client", f"=IFERROR(C5/C9,0)"),
        ("Best Day Revenue", f"=MAX('Daily Sales'!I6:I36)"),
        ("Payment Variance", f"='Daily Sales'!N37"),
        ("Top Service", f'="See Service Breakdown"'),
        ("Staff Count", f"=COUNTA('Staff Performance'!B5:B12)"),
        ("Avg Staff Revenue", f"=IFERROR(C5/C14,0)"),
    ]

    DATA_START = HDR_ROW + 1  # row 5
    for idx, (metric_name, formula) in enumerate(metrics):
        r = DATA_START + idx
        ws.cell(row=r, column=col_metric, value=metric_name)
        ws.cell(row=r, column=col_metric).alignment = align_text()
        ws.cell(row=r, column=col_value).value = formula
        ws.cell(row=r, column=col_value).alignment = align_number()

        # Apply number formats
        if metric_name in ("Profit Margin",):
            ws.cell(row=r, column=col_value).number_format = PCT_FMT
        elif metric_name in ("Total Revenue", "Total Expenses", "Net Profit",
                             "Avg Revenue per Client", "Best Day Revenue",
                             "Payment Variance", "Avg Staff Revenue"):
            ws.cell(row=r, column=col_value).number_format = AED_FMT
        elif metric_name == "Staff Count":
            ws.cell(row=r, column=col_value).number_format = AED_FMT
        elif metric_name == "Top Service":
            ws.cell(row=r, column=col_value).alignment = align_text()

        style_data_row(ws, r, 2, 3, idx)
        ws.cell(row=r, column=col_metric).alignment = align_text()
        ws.cell(row=r, column=col_value).alignment = align_number()
        if metric_name == "Top Service":
            ws.cell(row=r, column=col_value).alignment = align_text()

    # ── 6-Month Comparison Table ──
    SIX_MO_START = DATA_START + len(metrics) + 2  # leave a gap
    ws.merge_cells(start_row=SIX_MO_START - 1, start_column=2, end_row=SIX_MO_START - 1, end_column=6)
    sec_title = ws.cell(row=SIX_MO_START - 1, column=2, value="6-Month Comparison")
    sec_title.font = font_subheader()
    sec_title.alignment = align_text()

    six_headers = ["Month", "Revenue", "Expenses", "Profit", "Margin"]
    for i, h in enumerate(six_headers):
        ws.cell(row=SIX_MO_START, column=2 + i, value=h)
    style_header_row(ws, SIX_MO_START, 2, 6)

    SIX_DATA_START = SIX_MO_START + 1
    SIX_ROWS = 6

    for idx in range(SIX_ROWS):
        r = SIX_DATA_START + idx
        for ci in range(5):
            c = 2 + ci
            if fill_data and six_month_data and idx < len(six_month_data):
                ws.cell(row=r, column=c, value=six_month_data[idx][ci])
            else:
                apply_yellow_input(ws.cell(row=r, column=c))

            if ci == 0:
                ws.cell(row=r, column=c).alignment = align_text()
            elif ci == 4:
                ws.cell(row=r, column=c).number_format = PCT_FMT
                ws.cell(row=r, column=c).alignment = align_number()
            else:
                ws.cell(row=r, column=c).number_format = AED_FMT
                ws.cell(row=r, column=c).alignment = align_number()

        style_data_row(ws, r, 2, 6, idx)
        # Re-apply
        ws.cell(row=r, column=2).alignment = align_text()
        for ci in range(1, 5):
            c = 2 + ci
            if ci == 4:
                ws.cell(row=r, column=c).number_format = PCT_FMT
            else:
                ws.cell(row=r, column=c).number_format = AED_FMT
            ws.cell(row=r, column=c).alignment = align_number()
        if not fill_data:
            for ci in range(5):
                apply_yellow_input(ws.cell(row=r, column=2 + ci))

    SIX_DATA_END = SIX_DATA_START + SIX_ROWS - 1

    # ── Bar chart: 6-month Revenue vs Expenses ──
    chart_bar = create_bar_chart(chart_type="col", grouping="clustered", width=18, height=10)
    bar_cats = Reference(ws, min_col=2, min_row=SIX_DATA_START, max_row=SIX_DATA_END)
    rev_vals = Reference(ws, min_col=3, min_row=SIX_MO_START, max_row=SIX_DATA_END)
    exp_vals = Reference(ws, min_col=4, min_row=SIX_MO_START, max_row=SIX_DATA_END)
    chart_bar.add_data(rev_vals, titles_from_data=True)
    chart_bar.add_data(exp_vals, titles_from_data=True)
    chart_bar.set_categories(bar_cats)
    setup_chart_titles(chart_bar, title="Revenue vs Expenses (6 Months)", y_title="AED")
    apply_chart_colors(chart_bar, [CHART_COLORS[0], CHART_COLORS[2]])

    chart_anchor = SIX_DATA_END + 2
    ws.add_chart(chart_bar, f"B{chart_anchor}")

    # Footer
    write_footer(ws, chart_anchor + 14, 2, 6)

    auto_fit_columns(ws, header_row=HDR_ROW, data_start_row=DATA_START)
    # Make metric column wider
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20

    return ws


def build_instructions(wb):
    """Build the Instructions sheet."""
    ws = wb.create_sheet("Instructions")

    LAST_COL = 4  # B-D
    setup_sheet(ws, title="How to Use This Template", last_col=LAST_COL)

    instructions = [
        ('1.', 'Go to "Daily Sales" sheet — enter your daily numbers in the yellow cells'),
        ('2.', 'Go to "Service Breakdown" — enter bookings and average prices for each service'),
        ('3.', 'Go to "Staff Performance" — enter each staff member\'s data'),
        ('4.', 'Go to "Expense Tracker" — enter your budget and actual expenses'),
        ('5.', 'The "Monthly Summary" updates automatically!'),
        ('6.', 'Yellow cells = YOU fill in. White cells = AUTO-CALCULATED.'),
    ]

    r = 4
    for step_num, step_text in instructions:
        ws.cell(row=r, column=2, value=step_num)
        ws.cell(row=r, column=2).font = font_subheader()
        ws.cell(row=r, column=2).alignment = align_date()
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=LAST_COL)
        ws.cell(row=r, column=3, value=step_text)
        ws.cell(row=r, column=3).font = font_body()
        ws.cell(row=r, column=3).alignment = align_text()
        ws.row_dimensions[r].height = 28
        r += 1

    # Spacer
    r += 1

    # Contact info
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=LAST_COL)
    contact = ws.cell(row=r, column=2,
                      value="Need help? Contact Ahmed Ali | Data Analysis Services | ahmed-ali-ops.vercel.app")
    contact.font = font_subheader()
    contact.alignment = align_text()
    ws.row_dimensions[r].height = 30

    # Footer
    write_footer(ws, r + 2, 2, LAST_COL)

    # Manual column widths (auto_fit_columns fails with merged cells)
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 10

    return ws


# ════════════════════════════════════════════════════════════════
#  SAMPLE DATA GENERATION
# ════════════════════════════════════════════════════════════════

def generate_daily_data():
    """Generate realistic January 2026 daily sales data for a Dubai salon."""
    from datetime import date, timedelta
    data = []
    base_date = date(2026, 1, 1)

    # Weekend days (Fri/Sat in UAE) get higher traffic
    for day_num in range(1, 32):
        d = date(2026, 1, day_num)
        dow = d.weekday()  # 0=Mon, 4=Fri, 5=Sat

        # Base traffic varies by day of week
        if dow in (4, 5):  # Fri, Sat - busy
            walk_base = random.randint(8, 16)
            appt_base = random.randint(12, 22)
            svc_base = random.randint(2800, 5200)
            prod_base = random.randint(300, 700)
        elif dow == 6:  # Sun - moderate
            walk_base = random.randint(5, 10)
            appt_base = random.randint(8, 15)
            svc_base = random.randint(1800, 3500)
            prod_base = random.randint(200, 500)
        else:  # Mon-Thu - normal
            walk_base = random.randint(4, 9)
            appt_base = random.randint(6, 14)
            svc_base = random.randint(1500, 3000)
            prod_base = random.randint(150, 400)

        # Round to nice numbers
        svc_rev = round(svc_base / 50) * 50
        prod_rev = round(prod_base / 25) * 25
        total_rev = svc_rev + prod_rev

        # Payment split: ~40% cash, ~45% card, ~15% online
        cash = round(total_rev * random.uniform(0.35, 0.45))
        card = round(total_rev * random.uniform(0.40, 0.50))
        online = total_rev - cash - card  # remainder
        # Add small variance for realism
        online += random.randint(-50, 50)
        if online < 0:
            online = 50

        data.append({
            "date": d,
            "walk_ins": walk_base,
            "appointments": appt_base,
            "service_rev": svc_rev,
            "product_rev": prod_rev,
            "cash": cash,
            "card": card,
            "online": online,
        })
    return data


def generate_service_data():
    """Generate realistic service breakdown data."""
    return [
        {"bookings": 320, "avg_price": 65, "target": 20000},   # Haircut & Styling
        {"bookings": 180, "avg_price": 150, "target": 28000},  # Hair Coloring
        {"bookings": 85,  "avg_price": 280, "target": 25000},  # Highlights & Balayage
        {"bookings": 45,  "avg_price": 190, "target": 9000},   # Keratin Treatment
        {"bookings": 12,  "avg_price": 1400, "target": 15000}, # Bridal Package
        {"bookings": 210, "avg_price": 60, "target": 13000},   # Facial Treatment
        {"bookings": 280, "avg_price": 40, "target": 12000},   # Manicure & Pedicure
        {"bookings": 60,  "avg_price": 55, "target": 3500},    # Henna Design
        {"bookings": 18,  "avg_price": 450, "target": 8000},   # Hair Extensions
        {"bookings": 35,  "avg_price": 120, "target": 4500},   # Scalp Treatment
        {"bookings": 90,  "avg_price": 75, "target": 7000},    # Makeup
        {"bookings": 0,   "avg_price": 0, "target": 0},        # Other
    ]


def generate_staff_data():
    """Generate realistic staff performance data for 7 staff members."""
    return [
        {"name": "Fatima K.", "clients": 95,  "revenue": 18500, "tips": 1850, "rating": 4.9},
        {"name": "Noor A.",   "clients": 78,  "revenue": 15200, "tips": 1400, "rating": 4.8},
        {"name": "Sara M.",   "clients": 22,  "revenue": 16800, "tips": 2200, "rating": 5.0},
        {"name": "Maryam H.", "clients": 120, "revenue": 8400,  "tips": 750,  "rating": 4.7},
        {"name": "Aisha R.",  "clients": 88,  "revenue": 7920,  "tips": 680,  "rating": 4.6},
        {"name": "Layla S.",  "clients": 65,  "revenue": 5850,  "tips": 520,  "rating": 4.5},
        {"name": "Huda B.",   "clients": 50,  "revenue": 2700,  "tips": 380,  "rating": 4.8},
        # Row 8 left blank (template has 8 rows, sample fills 7)
    ]


def generate_expense_data():
    """Generate realistic UAE salon expense data."""
    return [
        {"budget": 25000, "actual": 25000},   # Rent
        {"budget": 28000, "actual": 27500},   # Staff Salaries
        {"budget": 8000,  "actual": 7200},    # Product Supplies
        {"budget": 4500,  "actual": 4800},    # Utilities
        {"budget": 3000,  "actual": 2500},    # Marketing
        {"budget": 2000,  "actual": 2000},    # Insurance
        {"budget": 1500,  "actual": 1200},    # Maintenance
        {"budget": 800,   "actual": 750},     # POS/Banking Fees
        {"budget": 600,   "actual": 550},     # Delivery Commissions
        {"budget": 1200,  "actual": 1100},    # Licenses & Permits
        {"budget": 1500,  "actual": 1800},    # Miscellaneous
    ]


def generate_six_month_data():
    """Generate 6-month comparison data (Aug 2025 - Jan 2026)."""
    return [
        ("Aug 2025",  82300,  68500,  13800, 0.168),
        ("Sep 2025",  85600,  69200,  16400, 0.192),
        ("Oct 2025",  88900,  70200,  18700, 0.210),
        ("Nov 2025",  91200,  71000,  20200, 0.222),
        ("Dec 2025", 112400,  76800,  35600, 0.317),
        ("Jan 2026",  98600,  72400,  26200, 0.266),
    ]


# ════════════════════════════════════════════════════════════════
#  MAIN: Generate both files
# ════════════════════════════════════════════════════════════════

def create_template():
    """Create the EMPTY template file."""
    wb = Workbook()
    wb.properties.creator = "Z.ai"

    title_suffix = "[Month/Year]"

    build_daily_sales(wb, title_suffix, fill_data=False)
    build_service_breakdown(wb, title_suffix, fill_data=False)
    build_staff_performance(wb, title_suffix, fill_data=False)
    build_expense_tracker(wb, title_suffix, fill_data=False)
    build_monthly_summary(wb, title_suffix, fill_data=False)
    build_instructions(wb)

    # Move Instructions to last position
    wb.move_sheet("Instructions", offset=0)

    path = "/home/z/my-project/download/Salon_Monthly_Sales_Report_TEMPLATE.xlsx"
    wb.save(path)
    print(f"✅ Template saved to {path}")
    return path


def create_sample():
    """Create the FILLED sample file."""
    wb = Workbook()
    wb.properties.creator = "Z.ai"

    title_suffix = "January 2026"
    daily_data = generate_daily_data()
    svc_data = generate_service_data()
    staff_data = generate_staff_data()
    exp_data = generate_expense_data()
    six_mo_data = generate_six_month_data()

    build_daily_sales(wb, title_suffix, fill_data=True, daily_data=daily_data)
    build_service_breakdown(wb, title_suffix, fill_data=True, svc_data=svc_data)
    build_staff_performance(wb, title_suffix, fill_data=True, staff_data=staff_data)
    build_expense_tracker(wb, title_suffix, fill_data=True, exp_data=exp_data)
    build_monthly_summary(wb, title_suffix, fill_data=True, six_month_data=six_mo_data)
    build_instructions(wb)

    path = "/home/z/my-project/download/Salon_Monthly_Sales_Report_SAMPLE.xlsx"
    wb.save(path)
    print(f"✅ Sample saved to {path}")
    return path


if __name__ == "__main__":
    random.seed(42)  # Reproducible data
    t1 = create_template()
    t2 = create_sample()
    print(f"\n🎯 Both files generated successfully!")
    print(f"   Template: {t1}")
    print(f"   Sample:   {t2}")
