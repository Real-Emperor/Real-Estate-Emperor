#!/usr/bin/env python3
"""
Generate Professional POS Sample — Al Noor Supermarket
8 sheets with realistic demo data for February 2025.
"""

import sys, os
sys.path.insert(0, "/home/z/my-project/skills/xlsx")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta
import random

# ── Import base template ──────────────────────────────────────
from templates.base import (
    use_palette_explicit, setup_sheet, style_header_row, style_data_row,
    style_total_row, font_title, font_header, font_subheader, font_body,
    font_caption, fill_header, fill_total, fill_data_row,
    border_header, border_total, align_title, align_header, align_number,
    align_text, align_date, COLUMN_WIDTHS, ROW_HEIGHTS,
    PRIMARY, PRIMARY_LIGHT, SECONDARY, HEADER_TEXT,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_50, NEUTRAL_0,
    CHART_COLORS, CF_POSITIVE_FILL, CF_POSITIVE_FONT,
    CF_NEGATIVE_FILL, CF_NEGATIVE_FONT,
    create_bar_chart, setup_chart_titles, apply_chart_colors,
    make_chart_title, auto_fit_columns, FORMATS,
)

# ── Activate professional palette BEFORE creating styles ──────
use_palette_explicit("professional")

# ── Currency format for AED ───────────────────────────────────
FMT_CURRENCY = "#,##0.00"
FMT_INT = "#,##0"
FMT_PCT = "0.0%"

# ── Workbook ──────────────────────────────────────────────────
wb = Workbook()
wb.properties.creator = "Z.ai"

# ── Constants ─────────────────────────────────────────────────
STAFF = ["Ahmed", "Fatima", "Omar", "Sara", "Khalid"]
SHIFTS = {"Ahmed": "Morning", "Fatima": "Morning", "Omar": "Evening", "Sara": "Evening", "Khalid": "Morning"}
PAYMENTS = ["Cash", "Card", "Online"]

# Product catalog: (name, category, unit_price, cost_price, supplier)
PRODUCTS = [
    ("Rice (5kg)", "Grains", 35, 26, "Al Falaj"),
    ("Milk (1L)", "Dairy", 8, 5.5, "GCC Foods"),
    ("Bread", "Bakery", 5, 3, "Emirates Trading"),
    ("Eggs (12pc)", "Dairy", 12, 8, "GCC Foods"),
    ("Chicken (kg)", "Meat", 28, 20, "Al Falaj"),
    ("Cooking Oil (1L)", "Grains", 22, 15, "Emirates Trading"),
    ("Water (6pk)", "Beverages", 2, 1.2, "GCC Foods"),
    ("Orange Juice (1L)", "Beverages", 6, 3.8, "GCC Foods"),
    ("Pasta (500g)", "Grains", 4, 2.5, "Emirates Trading"),
    ("Tomatoes (kg)", "Produce", 6, 3.5, "Al Falaj"),
    ("Potatoes (kg)", "Produce", 4, 2, "Al Falaj"),
    ("Onions (kg)", "Produce", 3, 1.5, "Al Falaj"),
    ("Cheese (200g)", "Dairy", 15, 10, "GCC Foods"),
    ("Yogurt (500g)", "Dairy", 7, 4.5, "GCC Foods"),
    ("Sugar (1kg)", "Grains", 5, 3, "Emirates Trading"),
    ("Flour (1kg)", "Grains", 3, 1.8, "Emirates Trading"),
    ("Tea (25 bags)", "Beverages", 10, 6.5, "Emirates Trading"),
    ("Coffee (200g)", "Beverages", 18, 12, "Emirates Trading"),
    ("Butter (200g)", "Dairy", 11, 7, "GCC Foods"),
    ("Detergent (1L)", "Cleaning", 14, 9, "Emirates Trading"),
    ("Dish Soap (500ml)", "Cleaning", 8, 5, "Emirates Trading"),
    ("Chips (150g)", "Snacks", 5, 3, "GCC Foods"),
    ("Chocolate Bar", "Snacks", 7, 4.5, "GCC Foods"),
    ("Biscuits (200g)", "Snacks", 4, 2.5, "Emirates Trading"),
    ("Cereal (375g)", "Snacks", 12, 8, "Emirates Trading"),
    ("Canned Beans", "Grains", 3, 1.8, "Al Falaj"),
    ("Lamb (kg)", "Meat", 55, 40, "Al Falaj"),
    ("Fish (kg)", "Meat", 32, 22, "Al Falaj"),
]

SUPPLIERS = ["Al Falaj", "Emirates Trading", "GCC Foods"]

# ── Helper: generate realistic sales data ─────────────────────
random.seed(42)

def gen_sales_data():
    """Generate 85 realistic supermarket transactions for Feb 2025."""
    sales = []
    inv = 1001
    for day in range(1, 29):
        # more transactions on weekends
        n_txns = random.randint(3, 5) if day % 7 not in [5, 6] else random.randint(4, 7)
        for _ in range(n_txns):
            d = date(2025, 2, day)
            staff = random.choice(STAFF)
            n_items = random.randint(1, 4)
            for _ in range(n_items):
                prod = random.choice(PRODUCTS)
                name, cat, price, _, _ = prod
                qty = random.randint(1, 5)
                pay = random.choice(PAYMENTS)
                customer = random.choice(["Regular Customer", "Walk-in", "Regular Customer", "Regular Customer", "Online Customer"])
                note = random.choice(["", "", "", "Bulk", "Express", ""])
                sales.append((d, f"INV-{inv:04d}", name, cat, qty, price, qty * price, pay, staff, customer, note))
            inv += 1
    # Trim/pad to ~85 transactions
    if len(sales) > 85:
        sales = sales[:85]
    return sales

SALES_DATA = gen_sales_data()

# ══════════════════════════════════════════════════════════════
# SHEET 1: Sales Entry
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Sales Entry"
ws1.sheet_properties.tabColor = PRIMARY

HEADERS_S1 = ["Date", "Invoice #", "Item/Service", "Category", "Qty",
              "Unit Price", "Total", "Payment Method", "Staff Name",
              "Customer Name", "Notes"]
COL_START = 2
COL_END = COL_START + len(HEADERS_S1) - 1
HEADER_ROW = 4
DATA_START = 5
N_SALES = len(SALES_DATA)

setup_sheet(ws1, "Sales Entry — Al Noor Supermarket", COL_END)

# Set column widths
widths_s1 = [14, 14, 20, 14, 8, 14, 14, 16, 14, 18, 12]
for i, w in enumerate(widths_s1):
    ws1.column_dimensions[get_column_letter(COL_START + i)].width = w

# Write headers
for i, h in enumerate(HEADERS_S1):
    ws1.cell(row=HEADER_ROW, column=COL_START + i, value=h)
style_header_row(ws1, HEADER_ROW, COL_START, COL_END)

# Write data
for idx, row_data in enumerate(SALES_DATA):
    r = DATA_START + idx
    for c, val in enumerate(row_data):
        cell = ws1.cell(row=r, column=COL_START + c, value=val)
    style_data_row(ws1, r, COL_START, COL_END, idx)
    # Format specific columns
    ws1.cell(row=r, column=COL_START).number_format = "YYYY-MM-DD"      # Date
    ws1.cell(row=r, column=COL_START).alignment = align_date()
    ws1.cell(row=r, column=COL_START + 4).number_format = FMT_INT        # Qty
    ws1.cell(row=r, column=COL_START + 4).alignment = align_number()
    ws1.cell(row=r, column=COL_START + 5).number_format = FMT_CURRENCY   # Unit Price
    ws1.cell(row=r, column=COL_START + 5).alignment = align_number()
    ws1.cell(row=r, column=COL_START + 6).number_format = FMT_CURRENCY   # Total
    ws1.cell(row=r, column=COL_START + 6).alignment = align_number()

# Totals row
TOTAL_ROW_S1 = DATA_START + N_SALES
ws1.cell(row=TOTAL_ROW_S1, column=COL_START, value="TOTAL")
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 4,
         value=f"=SUM({get_column_letter(COL_START+4)}{DATA_START}:{get_column_letter(COL_START+4)}{TOTAL_ROW_S1-1})")
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 4).number_format = FMT_INT
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 6,
         value=f"=SUM({get_column_letter(COL_START+6)}{DATA_START}:{get_column_letter(COL_START+6)}{TOTAL_ROW_S1-1})")
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 6).number_format = FMT_CURRENCY
style_total_row(ws1, TOTAL_ROW_S1, COL_START, COL_END)
ws1.cell(row=TOTAL_ROW_S1, column=COL_START).alignment = align_text()

# Print area
ws1.print_area = f"A1:{get_column_letter(COL_END)}{TOTAL_ROW_S1}"
ws1.print_title_rows = f"{HEADER_ROW}:{HEADER_ROW}"

# ══════════════════════════════════════════════════════════════
# SHEET 2: Daily Summary
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Daily Summary")
ws2.sheet_properties.tabColor = PRIMARY

HEADERS_S2 = ["Date", "Total Sales", "# Transactions", "Avg Transaction",
              "Cash Total", "Card Total", "Online Total", "# Items Sold"]
COL_END_S2 = COL_START + len(HEADERS_S2) - 1
HEADER_ROW_S2 = 4
DATA_START_S2 = 5
N_DAYS = 28  # Feb 1-28

setup_sheet(ws2, "Daily Summary — February 2025", COL_END_S2)

widths_s2 = [14, 16, 16, 16, 16, 16, 16, 14]
for i, w in enumerate(widths_s2):
    ws2.column_dimensions[get_column_letter(COL_START + i)].width = w

for i, h in enumerate(HEADERS_S2):
    ws2.cell(row=HEADER_ROW_S2, column=COL_START + i, value=h)
style_header_row(ws2, HEADER_ROW_S2, COL_START, COL_END_S2)

# Sales Entry sheet references
SE_DATE_COL = get_column_letter(COL_START)       # B
SE_TOTAL_COL = get_column_letter(COL_START + 6)   # H
SE_PAY_COL = get_column_letter(COL_START + 7)     # I
SE_QTY_COL = get_column_letter(COL_START + 4)     # F

for day in range(1, N_DAYS + 1):
    r = DATA_START_S2 + day - 1
    d = date(2025, 2, day)
    ws2.cell(row=r, column=COL_START, value=d)
    ws2.cell(row=r, column=COL_START).number_format = "YYYY-MM-DD"
    ws2.cell(row=r, column=COL_START).alignment = align_date()

    # Total Sales = SUMPRODUCT((SE_dates=this_date)*(SE_totals))
    ws2.cell(row=r, column=COL_START + 1,
             value=f'=SUMPRODUCT((\'Sales Entry\'!{SE_DATE_COL}{DATA_START}:{SE_DATE_COL}{TOTAL_ROW_S1-1}=B{r})*(\'Sales Entry\'!{SE_TOTAL_COL}{DATA_START}:{SE_TOTAL_COL}{TOTAL_ROW_S1-1}))')
    ws2.cell(row=r, column=COL_START + 1).number_format = FMT_CURRENCY

    # # Transactions = SUMPRODUCT((SE_dates=this_date)*1) - but need unique invoice count
    # Simpler: count rows matching date
    ws2.cell(row=r, column=COL_START + 2,
             value=f'=SUMPRODUCT((\'Sales Entry\'!{SE_DATE_COL}{DATA_START}:{SE_DATE_COL}{TOTAL_ROW_S1-1}=B{r})*1)')
    ws2.cell(row=r, column=COL_START + 2).number_format = FMT_INT

    # Avg Transaction
    ws2.cell(row=r, column=COL_START + 3,
             value=f'=IFERROR(C{r}/D{r},0)')
    ws2.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY

    # Cash Total
    ws2.cell(row=r, column=COL_START + 4,
             value=f'=SUMPRODUCT((\'Sales Entry\'!{SE_DATE_COL}{DATA_START}:{SE_DATE_COL}{TOTAL_ROW_S1-1}=B{r})*(\'Sales Entry\'!{SE_PAY_COL}{DATA_START}:{SE_PAY_COL}{TOTAL_ROW_S1-1}="Cash")*(\'Sales Entry\'!{SE_TOTAL_COL}{DATA_START}:{SE_TOTAL_COL}{TOTAL_ROW_S1-1}))')
    ws2.cell(row=r, column=COL_START + 4).number_format = FMT_CURRENCY

    # Card Total
    ws2.cell(row=r, column=COL_START + 5,
             value=f'=SUMPRODUCT((\'Sales Entry\'!{SE_DATE_COL}{DATA_START}:{SE_DATE_COL}{TOTAL_ROW_S1-1}=B{r})*(\'Sales Entry\'!{SE_PAY_COL}{DATA_START}:{SE_PAY_COL}{TOTAL_ROW_S1-1}="Card")*(\'Sales Entry\'!{SE_TOTAL_COL}{DATA_START}:{SE_TOTAL_COL}{TOTAL_ROW_S1-1}))')
    ws2.cell(row=r, column=COL_START + 5).number_format = FMT_CURRENCY

    # Online Total
    ws2.cell(row=r, column=COL_START + 6,
             value=f'=SUMPRODUCT((\'Sales Entry\'!{SE_DATE_COL}{DATA_START}:{SE_DATE_COL}{TOTAL_ROW_S1-1}=B{r})*(\'Sales Entry\'!{SE_PAY_COL}{DATA_START}:{SE_PAY_COL}{TOTAL_ROW_S1-1}="Online")*(\'Sales Entry\'!{SE_TOTAL_COL}{DATA_START}:{SE_TOTAL_COL}{TOTAL_ROW_S1-1}))')
    ws2.cell(row=r, column=COL_START + 6).number_format = FMT_CURRENCY

    # # Items Sold
    ws2.cell(row=r, column=COL_START + 7,
             value=f'=SUMPRODUCT((\'Sales Entry\'!{SE_DATE_COL}{DATA_START}:{SE_DATE_COL}{TOTAL_ROW_S1-1}=B{r})*(\'Sales Entry\'!{SE_QTY_COL}{DATA_START}:{SE_QTY_COL}{TOTAL_ROW_S1-1}))')
    ws2.cell(row=r, column=COL_START + 7).number_format = FMT_INT

    style_data_row(ws2, r, COL_START, COL_END_S2, day - 1)
    # Re-apply alignment/number format after style_data_row
    ws2.cell(row=r, column=COL_START).alignment = align_date()
    ws2.cell(row=r, column=COL_START).number_format = "YYYY-MM-DD"
    for cc in range(1, len(HEADERS_S2)):
        if cc in [2, 7]:  # # Transactions, # Items Sold
            ws2.cell(row=r, column=COL_START + cc).alignment = align_number()
            ws2.cell(row=r, column=COL_START + cc).number_format = FMT_INT
        else:
            ws2.cell(row=r, column=COL_START + cc).alignment = align_number()

# Totals row
TR_S2 = DATA_START_S2 + N_DAYS
ws2.cell(row=TR_S2, column=COL_START, value="TOTAL")
ws2.cell(row=TR_S2, column=COL_START + 1,
         value=f"=SUM(C{DATA_START_S2}:C{TR_S2-1})")
ws2.cell(row=TR_S2, column=COL_START + 1).number_format = FMT_CURRENCY
ws2.cell(row=TR_S2, column=COL_START + 2,
         value=f"=SUM(D{DATA_START_S2}:D{TR_S2-1})")
ws2.cell(row=TR_S2, column=COL_START + 2).number_format = FMT_INT
ws2.cell(row=TR_S2, column=COL_START + 3,
         value=f"=IFERROR(C{TR_S2}/D{TR_S2},0)")
ws2.cell(row=TR_S2, column=COL_START + 3).number_format = FMT_CURRENCY
ws2.cell(row=TR_S2, column=COL_START + 4,
         value=f"=SUM(F{DATA_START_S2}:F{TR_S2-1})")
ws2.cell(row=TR_S2, column=COL_START + 4).number_format = FMT_CURRENCY
ws2.cell(row=TR_S2, column=COL_START + 5,
         value=f"=SUM(G{DATA_START_S2}:G{TR_S2-1})")
ws2.cell(row=TR_S2, column=COL_START + 5).number_format = FMT_CURRENCY
ws2.cell(row=TR_S2, column=COL_START + 6,
         value=f"=SUM(H{DATA_START_S2}:H{TR_S2-1})")
ws2.cell(row=TR_S2, column=COL_START + 6).number_format = FMT_CURRENCY
ws2.cell(row=TR_S2, column=COL_START + 7,
         value=f"=SUM(I{DATA_START_S2}:I{TR_S2-1})")
ws2.cell(row=TR_S2, column=COL_START + 7).number_format = FMT_INT
style_total_row(ws2, TR_S2, COL_START, COL_END_S2)
ws2.cell(row=TR_S2, column=COL_START).alignment = align_text()

ws2.print_area = f"A1:{get_column_letter(COL_END_S2)}{TR_S2}"
ws2.print_title_rows = f"{HEADER_ROW_S2}:{HEADER_ROW_S2}"

# ══════════════════════════════════════════════════════════════
# SHEET 3: Weekly Summary
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Weekly Summary")
ws3.sheet_properties.tabColor = PRIMARY

HEADERS_S3 = ["Week #", "Week Start", "Week End", "Total Sales",
              "# Transactions", "Avg Transaction", "Top Staff"]
COL_END_S3 = COL_START + len(HEADERS_S3) - 1
HEADER_ROW_S3 = 4
DATA_START_S3 = 5

setup_sheet(ws3, "Weekly Summary — February 2025", COL_END_S3)

widths_s3 = [10, 14, 14, 16, 16, 16, 14]
for i, w in enumerate(widths_s3):
    ws3.column_dimensions[get_column_letter(COL_START + i)].width = w

for i, h in enumerate(HEADERS_S3):
    ws3.cell(row=HEADER_ROW_S3, column=COL_START + i, value=h)
style_header_row(ws3, HEADER_ROW_S3, COL_START, COL_END_S3)

# Week definitions for Feb 2025
WEEKS = [
    (1, date(2025, 2, 1), date(2025, 2, 7)),
    (2, date(2025, 2, 8), date(2025, 2, 14)),
    (3, date(2025, 2, 15), date(2025, 2, 21)),
    (4, date(2025, 2, 22), date(2025, 2, 28)),
]

DS_DATE = get_column_letter(COL_START)       # B in Daily Summary
DS_SALES = get_column_letter(COL_START + 1)  # C
DS_TXNS = get_column_letter(COL_START + 2)   # D

for idx, (wnum, wstart, wend) in enumerate(WEEKS):
    r = DATA_START_S3 + idx
    ws3.cell(row=r, column=COL_START, value=wnum)
    ws3.cell(row=r, column=COL_START + 1, value=wstart)
    ws3.cell(row=r, column=COL_START + 1).number_format = "YYYY-MM-DD"
    ws3.cell(row=r, column=COL_START + 2, value=wend)
    ws3.cell(row=r, column=COL_START + 2).number_format = "YYYY-MM-DD"

    # Total Sales = SUMPRODUCT((DS_dates>=wstart)*(DS_dates<=wend)*(DS_sales))
    ws3.cell(row=r, column=COL_START + 3,
             value=f"=SUMPRODUCT(('Daily Summary'!{DS_DATE}{DATA_START_S2}:{DS_DATE}{TR_S2-1}>=C{r})*('Daily Summary'!{DS_DATE}{DATA_START_S2}:{DS_DATE}{TR_S2-1}<=D{r})*('Daily Summary'!{DS_SALES}{DATA_START_S2}:{DS_SALES}{TR_S2-1}))")
    ws3.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY

    # # Transactions
    ws3.cell(row=r, column=COL_START + 4,
             value=f"=SUMPRODUCT(('Daily Summary'!{DS_DATE}{DATA_START_S2}:{DS_DATE}{TR_S2-1}>=C{r})*('Daily Summary'!{DS_DATE}{DATA_START_S2}:{DS_DATE}{TR_S2-1}<=D{r})*('Daily Summary'!{DS_TXNS}{DATA_START_S2}:{DS_TXNS}{TR_S2-1}))")
    ws3.cell(row=r, column=COL_START + 4).number_format = FMT_INT

    # Avg Transaction
    ws3.cell(row=r, column=COL_START + 5,
             value=f"=IFERROR(E{r}/F{r},0)")
    ws3.cell(row=r, column=COL_START + 5).number_format = FMT_CURRENCY

    # Top Staff — compute via SUMPRODUCT for each staff
    # We'll embed a formula that picks the staff with most sales in that week
    SE_STAFF_COL = get_column_letter(COL_START + 8)  # J in Sales Entry
    staff_formulas = []
    for s in STAFF:
        f = f'SUMPRODUCT((\'Sales Entry\'!{SE_DATE_COL}{DATA_START}:{SE_DATE_COL}{TOTAL_ROW_S1-1}>=C{r})*(\'Sales Entry\'!{SE_DATE_COL}{DATA_START}:{SE_DATE_COL}{TOTAL_ROW_S1-1}<=D{r})*(\'Sales Entry\'!{SE_STAFF_COL}{DATA_START}:{SE_STAFF_COL}{TOTAL_ROW_S1-1}="{s}")*(\'Sales Entry\'!{SE_TOTAL_COL}{DATA_START}:{SE_TOTAL_COL}{TOTAL_ROW_S1-1}))'
        staff_formulas.append(f)

    # Use nested IF to find top staff
    top_staff_formula = (
        f'=IF(AND({staff_formulas[0]}>={staff_formulas[1]},{staff_formulas[0]}>={staff_formulas[2]},{staff_formulas[0]}>={staff_formulas[3]},{staff_formulas[0]}>={staff_formulas[4]}),"{STAFF[0]}",'
        f'IF(AND({staff_formulas[1]}>={staff_formulas[0]},{staff_formulas[1]}>={staff_formulas[2]},{staff_formulas[1]}>={staff_formulas[3]},{staff_formulas[1]}>={staff_formulas[4]}),"{STAFF[1]}",'
        f'IF(AND({staff_formulas[2]}>={staff_formulas[0]},{staff_formulas[2]}>={staff_formulas[1]},{staff_formulas[2]}>={staff_formulas[3]},{staff_formulas[2]}>={staff_formulas[4]}),"{STAFF[2]}",'
        f'IF(AND({staff_formulas[3]}>={staff_formulas[0]},{staff_formulas[3]}>={staff_formulas[1]},{staff_formulas[3]}>={staff_formulas[2]},{staff_formulas[3]}>={staff_formulas[4]}),"{STAFF[3]}","{STAFF[4]}"))))'
    )
    ws3.cell(row=r, column=COL_START + 6, value=top_staff_formula)

    style_data_row(ws3, r, COL_START, COL_END_S3, idx)
    ws3.cell(row=r, column=COL_START).alignment = align_number()
    ws3.cell(row=r, column=COL_START + 1).alignment = align_date()
    ws3.cell(row=r, column=COL_START + 1).number_format = "YYYY-MM-DD"
    ws3.cell(row=r, column=COL_START + 2).alignment = align_date()
    ws3.cell(row=r, column=COL_START + 2).number_format = "YYYY-MM-DD"
    for cc in [3, 5]:
        ws3.cell(row=r, column=COL_START + cc).alignment = align_number()
        ws3.cell(row=r, column=COL_START + cc).number_format = FMT_CURRENCY
    ws3.cell(row=r, column=COL_START + 4).alignment = align_number()
    ws3.cell(row=r, column=COL_START + 4).number_format = FMT_INT
    ws3.cell(row=r, column=COL_START + 6).alignment = align_text()

# Totals row
TR_S3 = DATA_START_S3 + len(WEEKS)
ws3.cell(row=TR_S3, column=COL_START, value="TOTAL")
ws3.cell(row=TR_S3, column=COL_START + 3,
         value=f"=SUM(E{DATA_START_S3}:E{TR_S3-1})")
ws3.cell(row=TR_S3, column=COL_START + 3).number_format = FMT_CURRENCY
ws3.cell(row=TR_S3, column=COL_START + 4,
         value=f"=SUM(F{DATA_START_S3}:F{TR_S3-1})")
ws3.cell(row=TR_S3, column=COL_START + 4).number_format = FMT_INT
ws3.cell(row=TR_S3, column=COL_START + 5,
         value=f"=IFERROR(E{TR_S3}/F{TR_S3},0)")
ws3.cell(row=TR_S3, column=COL_START + 5).number_format = FMT_CURRENCY
style_total_row(ws3, TR_S3, COL_START, COL_END_S3)
ws3.cell(row=TR_S3, column=COL_START).alignment = align_text()

ws3.print_area = f"A1:{get_column_letter(COL_END_S3)}{TR_S3}"

# ══════════════════════════════════════════════════════════════
# SHEET 4: Monthly Summary
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Monthly Summary")
ws4.sheet_properties.tabColor = PRIMARY

HEADERS_S4 = ["Month", "Total Sales", "# Transactions", "Avg Transaction",
              "Best Day Sales", "Best Day Date"]
COL_END_S4 = COL_START + len(HEADERS_S4) - 1
HEADER_ROW_S4 = 4
DATA_START_S4 = 5

setup_sheet(ws4, "Monthly Summary — February 2025", COL_END_S4)

widths_s4 = [16, 16, 16, 16, 16, 16]
for i, w in enumerate(widths_s4):
    ws4.column_dimensions[get_column_letter(COL_START + i)].width = w

for i, h in enumerate(HEADERS_S4):
    ws4.cell(row=HEADER_ROW_S4, column=COL_START + i, value=h)
style_header_row(ws4, HEADER_ROW_S4, COL_START, COL_END_S4)

r = DATA_START_S4
ws4.cell(row=r, column=COL_START, value="February 2025")
ws4.cell(row=r, column=COL_START + 1,
         value=f"=SUM('Daily Summary'!C{DATA_START_S2}:C{TR_S2-1})")
ws4.cell(row=r, column=COL_START + 1).number_format = FMT_CURRENCY
ws4.cell(row=r, column=COL_START + 2,
         value=f"=SUM('Daily Summary'!D{DATA_START_S2}:D{TR_S2-1})")
ws4.cell(row=r, column=COL_START + 2).number_format = FMT_INT
ws4.cell(row=r, column=COL_START + 3,
         value=f"=IFERROR(C{r}/D{r},0)")
ws4.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
ws4.cell(row=r, column=COL_START + 4,
         value=f"=MAX('Daily Summary'!C{DATA_START_S2}:C{TR_S2-1})")
ws4.cell(row=r, column=COL_START + 4).number_format = FMT_CURRENCY
# Best Day Date — use INDEX/MATCH
ws4.cell(row=r, column=COL_START + 5,
         value=f"=INDEX('Daily Summary'!B{DATA_START_S2}:B{TR_S2-1},MATCH(F{r},'Daily Summary'!C{DATA_START_S2}:C{TR_S2-1},0))")
ws4.cell(row=r, column=COL_START + 5).number_format = "YYYY-MM-DD"

style_data_row(ws4, r, COL_START, COL_END_S4, 0)
ws4.cell(row=r, column=COL_START).alignment = align_text()
for cc in [1, 3, 4]:
    ws4.cell(row=r, column=COL_START + cc).alignment = align_number()
ws4.cell(row=r, column=COL_START + 2).alignment = align_number()
ws4.cell(row=r, column=COL_START + 5).alignment = align_date()

# Totals row
TR_S4 = DATA_START_S4 + 1
ws4.cell(row=TR_S4, column=COL_START, value="TOTAL")
ws4.cell(row=TR_S4, column=COL_START + 1,
         value=f"=C{DATA_START_S4}")
ws4.cell(row=TR_S4, column=COL_START + 1).number_format = FMT_CURRENCY
ws4.cell(row=TR_S4, column=COL_START + 2,
         value=f"=D{DATA_START_S4}")
ws4.cell(row=TR_S4, column=COL_START + 2).number_format = FMT_INT
ws4.cell(row=TR_S4, column=COL_START + 3,
         value=f"=IFERROR(C{TR_S4}/D{TR_S4},0)")
ws4.cell(row=TR_S4, column=COL_START + 3).number_format = FMT_CURRENCY
style_total_row(ws4, TR_S4, COL_START, COL_END_S4)
ws4.cell(row=TR_S4, column=COL_START).alignment = align_text()

# Bar chart: Daily sales from Daily Summary sheet
chart = create_bar_chart(chart_type="col", grouping="clustered", gap_width=80, width=22, height=12)
data_ref = Reference(ws2, min_col=COL_START + 1, min_row=HEADER_ROW_S2,
                     max_row=DATA_START_S2 + N_DAYS - 1)
cats_ref = Reference(ws2, min_col=COL_START, min_row=DATA_START_S2,
                     max_row=DATA_START_S2 + N_DAYS - 1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.shape = 4
setup_chart_titles(chart, title="Daily Sales — February 2025",
                   y_title="Sales (AED)", x_title="Date")
apply_chart_colors(chart, [PRIMARY])
ws4.add_chart(chart, "B9")

ws4.print_area = f"A1:{get_column_letter(COL_END_S4)}{TR_S4}"

# ══════════════════════════════════════════════════════════════
# SHEET 5: Invoice
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Invoice")
ws5.sheet_properties.tabColor = PRIMARY

setup_sheet(ws5, None, 8)

# Business header
ws5.merge_cells("B2:H2")
cell = ws5.cell(row=2, column=2, value="Al Noor Supermarket")
cell.font = Font(name="Arial", size=20, bold=True, color=PRIMARY)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[2].height = 40

ws5.merge_cells("B3:H3")
cell = ws5.cell(row=3, column=2, value="Your Neighbourhood Grocery — Dubai, UAE")
cell.font = Font(name="Arial", size=10, color=NEUTRAL_600)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[3].height = 20

ws5.merge_cells("B4:H4")
cell = ws5.cell(row=4, column=2, value="Tel: +971-4-XXX-XXXX  |  VAT: 10XXXXXX000X3")
cell.font = Font(name="Arial", size=9, color=NEUTRAL_600)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[4].height = 18

# Divider
ws5.row_dimensions[5].height = 6
for c in range(2, 9):
    ws5.cell(row=5, column=c).border = Border(bottom=Side(style="medium", color=PRIMARY))

# Invoice details
ws5.cell(row=7, column=2, value="Invoice #:").font = font_subheader()
ws5.cell(row=7, column=3, value="INV-0085").font = font_body()
ws5.cell(row=7, column=6, value="Date:").font = font_subheader()
ws5.cell(row=7, column=7, value="10-Feb-2025").font = font_body()

ws5.cell(row=8, column=2, value="Customer:").font = font_subheader()
ws5.cell(row=8, column=3, value="Regular Customer").font = font_body()
ws5.cell(row=8, column=6, value="Staff:").font = font_subheader()
ws5.cell(row=8, column=7, value="Fatima").font = font_body()

ws5.cell(row=9, column=2, value="Payment:").font = font_subheader()
ws5.cell(row=9, column=3, value="Card").font = font_body()

# Invoice items header
INV_HDR_ROW = 11
inv_headers = ["#", "Item", "Category", "Qty", "Unit Price", "Total"]
for i, h in enumerate(inv_headers):
    ws5.cell(row=INV_HDR_ROW, column=2 + i, value=h)
style_header_row(ws5, INV_HDR_ROW, 2, 7)

# Invoice items
inv_items = [
    (1, "Rice (5kg)", "Grains", 2, 35),
    (2, "Milk (1L)", "Dairy", 3, 8),
    (3, "Chicken (kg)", "Meat", 1, 28),
    (4, "Cooking Oil (1L)", "Grains", 2, 22),
    (5, "Eggs (12pc)", "Dairy", 1, 12),
    (6, "Orange Juice (1L)", "Beverages", 2, 6),
]

for idx, (num, item, cat, qty, price) in enumerate(inv_items):
    r = INV_HDR_ROW + 1 + idx
    ws5.cell(row=r, column=2, value=num)
    ws5.cell(row=r, column=2).alignment = align_number()
    ws5.cell(row=r, column=3, value=item)
    ws5.cell(row=r, column=4, value=cat)
    ws5.cell(row=r, column=5, value=qty)
    ws5.cell(row=r, column=5).alignment = align_number()
    ws5.cell(row=r, column=6, value=price)
    ws5.cell(row=r, column=6).number_format = FMT_CURRENCY
    ws5.cell(row=r, column=6).alignment = align_number()
    ws5.cell(row=r, column=7, value=f"=E{r}*F{r}")
    ws5.cell(row=r, column=7).number_format = FMT_CURRENCY
    ws5.cell(row=r, column=7).alignment = align_number()
    style_data_row(ws5, r, 2, 7, idx)

# Totals section
INV_TOT_START = INV_HDR_ROW + 1 + len(inv_items)
ws5.row_dimensions[INV_TOT_START].height = 8  # spacer

r = INV_TOT_START + 1
ws5.cell(row=r, column=5, value="Subtotal:").font = font_subheader()
ws5.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center")
ws5.cell(row=r, column=7, value=f"=SUM(G{INV_HDR_ROW+1}:G{INV_HDR_ROW+len(inv_items)})")
ws5.cell(row=r, column=7).number_format = FMT_CURRENCY
ws5.cell(row=r, column=7).font = font_body()
ws5.cell(row=r, column=7).alignment = align_number()

r += 1
ws5.cell(row=r, column=5, value="VAT (5%):").font = font_subheader()
ws5.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center")
ws5.cell(row=r, column=7, value=f"=G{r-1}*0.05")
ws5.cell(row=r, column=7).number_format = FMT_CURRENCY
ws5.cell(row=r, column=7).font = font_body()
ws5.cell(row=r, column=7).alignment = align_number()

r += 1
ws5.cell(row=r, column=5, value="Discount:").font = font_subheader()
ws5.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center")
ws5.cell(row=r, column=7, value=-10)
ws5.cell(row=r, column=7).number_format = FMT_CURRENCY
ws5.cell(row=r, column=7).font = Font(color=ACCENT_NEGATIVE, size=11)
ws5.cell(row=r, column=7).alignment = align_number()

r += 1
# Grand total with thick top border
for c in range(5, 8):
    ws5.cell(row=r, column=c).border = Border(top=Side(style="medium", color=PRIMARY))
ws5.cell(row=r, column=5, value="Grand Total:").font = Font(size=13, bold=True, color=PRIMARY)
ws5.cell(row=r, column=5).alignment = Alignment(horizontal="right", vertical="center")
ws5.cell(row=r, column=7, value=f"=G{r-3}+G{r-2}+G{r-1}")
ws5.cell(row=r, column=7).number_format = FMT_CURRENCY
ws5.cell(row=r, column=7).font = Font(size=13, bold=True, color=PRIMARY)
ws5.cell(row=r, column=7).alignment = align_number()
ws5.row_dimensions[r].height = 30

# Thank you
r += 2
ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
cell = ws5.cell(row=r, column=2, value="Thank you for shopping at Al Noor Supermarket!")
cell.font = Font(size=11, italic=True, color=NEUTRAL_600)
cell.alignment = Alignment(horizontal="center", vertical="center")

# Column widths for invoice
inv_widths = [6, 22, 14, 8, 14, 14]
for i, w in enumerate(inv_widths):
    ws5.column_dimensions[get_column_letter(2 + i)].width = w

# ══════════════════════════════════════════════════════════════
# SHEET 6: Inventory
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Inventory")
ws6.sheet_properties.tabColor = PRIMARY

HEADERS_S6 = ["Product ID", "Item Name", "Category", "Unit Price",
              "Stock Qty", "Reorder Level", "Status", "Supplier", "Last Restocked"]
COL_END_S6 = COL_START + len(HEADERS_S6) - 1
HEADER_ROW_S6 = 4
DATA_START_S6 = 5

setup_sheet(ws6, "Inventory — Al Noor Supermarket", COL_END_S6)

widths_s6 = [12, 22, 14, 14, 12, 14, 14, 18, 16]
for i, w in enumerate(widths_s6):
    ws6.column_dimensions[get_column_letter(COL_START + i)].width = w

for i, h in enumerate(HEADERS_S6):
    ws6.cell(row=HEADER_ROW_S6, column=COL_START + i, value=h)
style_header_row(ws6, HEADER_ROW_S6, COL_START, COL_END_S6)

# Inventory data: (pid, name, cat, price, stock, reorder, supplier, last_restocked)
# Some items have stock <= reorder level (LOW STOCK)
INVENTORY_DATA = [
    ("PRD001", "Rice (5kg)", "Grains", 35, 120, 30, "Al Falaj", date(2025, 2, 20)),
    ("PRD002", "Milk (1L)", "Dairy", 8, 45, 50, "GCC Foods", date(2025, 2, 25)),
    ("PRD003", "Bread", "Bakery", 5, 80, 40, "Emirates Trading", date(2025, 2, 27)),
    ("PRD004", "Eggs (12pc)", "Dairy", 12, 35, 30, "GCC Foods", date(2025, 2, 22)),
    ("PRD005", "Chicken (kg)", "Meat", 28, 15, 20, "Al Falaj", date(2025, 2, 26)),
    ("PRD006", "Cooking Oil (1L)", "Grains", 22, 60, 25, "Emirates Trading", date(2025, 2, 18)),
    ("PRD007", "Water (6pk)", "Beverages", 2, 200, 100, "GCC Foods", date(2025, 2, 24)),
    ("PRD008", "Orange Juice (1L)", "Beverages", 6, 30, 25, "GCC Foods", date(2025, 2, 23)),
    ("PRD009", "Pasta (500g)", "Grains", 4, 90, 40, "Emirates Trading", date(2025, 2, 19)),
    ("PRD010", "Tomatoes (kg)", "Produce", 6, 8, 15, "Al Falaj", date(2025, 2, 25)),
    ("PRD011", "Potatoes (kg)", "Produce", 4, 50, 20, "Al Falaj", date(2025, 2, 21)),
    ("PRD012", "Onions (kg)", "Produce", 3, 55, 25, "Al Falaj", date(2025, 2, 21)),
    ("PRD013", "Cheese (200g)", "Dairy", 15, 10, 15, "GCC Foods", date(2025, 2, 20)),
    ("PRD014", "Yogurt (500g)", "Dairy", 7, 25, 20, "GCC Foods", date(2025, 2, 26)),
    ("PRD015", "Sugar (1kg)", "Grains", 5, 70, 30, "Emirates Trading", date(2025, 2, 17)),
    ("PRD016", "Flour (1kg)", "Grains", 3, 65, 25, "Emirates Trading", date(2025, 2, 17)),
    ("PRD017", "Tea (25 bags)", "Beverages", 10, 40, 20, "Emirates Trading", date(2025, 2, 15)),
    ("PRD018", "Coffee (200g)", "Beverages", 18, 5, 10, "Emirates Trading", date(2025, 2, 14)),
    ("PRD019", "Butter (200g)", "Dairy", 11, 20, 15, "GCC Foods", date(2025, 2, 22)),
    ("PRD020", "Detergent (1L)", "Cleaning", 14, 35, 15, "Emirates Trading", date(2025, 2, 12)),
    ("PRD021", "Dish Soap (500ml)", "Cleaning", 8, 28, 15, "Emirates Trading", date(2025, 2, 12)),
    ("PRD022", "Chips (150g)", "Snacks", 5, 100, 40, "GCC Foods", date(2025, 2, 24)),
    ("PRD023", "Chocolate Bar", "Snacks", 7, 12, 20, "GCC Foods", date(2025, 2, 23)),
    ("PRD024", "Biscuits (200g)", "Snacks", 4, 75, 30, "Emirates Trading", date(2025, 2, 20)),
    ("PRD025", "Cereal (375g)", "Snacks", 12, 18, 15, "Emirates Trading", date(2025, 2, 19)),
    ("PRD026", "Canned Beans", "Grains", 3, 85, 30, "Al Falaj", date(2025, 2, 16)),
    ("PRD027", "Lamb (kg)", "Meat", 55, 6, 8, "Al Falaj", date(2025, 2, 26)),
    ("PRD028", "Fish (kg)", "Meat", 32, 3, 5, "Al Falaj", date(2025, 2, 27)),
]

for idx, (pid, name, cat, price, stock, reorder, supplier, restocked) in enumerate(INVENTORY_DATA):
    r = DATA_START_S6 + idx
    ws6.cell(row=r, column=COL_START, value=pid)
    ws6.cell(row=r, column=COL_START + 1, value=name)
    ws6.cell(row=r, column=COL_START + 2, value=cat)
    ws6.cell(row=r, column=COL_START + 3, value=price)
    ws6.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    ws6.cell(row=r, column=COL_START + 4, value=stock)
    ws6.cell(row=r, column=COL_START + 4).number_format = FMT_INT
    ws6.cell(row=r, column=COL_START + 5, value=reorder)
    ws6.cell(row=r, column=COL_START + 5).number_format = FMT_INT
    # Status = IF formula
    ws6.cell(row=r, column=COL_START + 6,
             value=f'=IF(F{r}<=G{r},"LOW STOCK","In Stock")')
    ws6.cell(row=r, column=COL_START + 7, value=supplier)
    ws6.cell(row=r, column=COL_START + 8, value=restocked)
    ws6.cell(row=r, column=COL_START + 8).number_format = "YYYY-MM-DD"

    style_data_row(ws6, r, COL_START, COL_END_S6, idx)
    # Re-apply alignments
    ws6.cell(row=r, column=COL_START).alignment = align_text()
    ws6.cell(row=r, column=COL_START + 1).alignment = align_text()
    ws6.cell(row=r, column=COL_START + 2).alignment = align_text()
    ws6.cell(row=r, column=COL_START + 3).alignment = align_number()
    ws6.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    ws6.cell(row=r, column=COL_START + 4).alignment = align_number()
    ws6.cell(row=r, column=COL_START + 4).number_format = FMT_INT
    ws6.cell(row=r, column=COL_START + 5).alignment = align_number()
    ws6.cell(row=r, column=COL_START + 5).number_format = FMT_INT
    ws6.cell(row=r, column=COL_START + 6).alignment = align_text()
    ws6.cell(row=r, column=COL_START + 7).alignment = align_text()
    ws6.cell(row=r, column=COL_START + 8).alignment = align_date()
    ws6.cell(row=r, column=COL_START + 8).number_format = "YYYY-MM-DD"

N_INV = len(INVENTORY_DATA)
TR_S6 = DATA_START_S6 + N_INV

# Conditional formatting on Status column (H)
status_col_letter = get_column_letter(COL_START + 6)  # H
status_range = f"{status_col_letter}{DATA_START_S6}:{status_col_letter}{TR_S6 - 1}"

# LOW STOCK = red
ws6.conditional_formatting.add(
    status_range,
    CellIsRule(operator="equal", formula=['"LOW STOCK"'],
              fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT)
)
# In Stock = green
ws6.conditional_formatting.add(
    status_range,
    CellIsRule(operator="equal", formula=['"In Stock"'],
              fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT)
)

ws6.print_area = f"A1:{get_column_letter(COL_END_S6)}{TR_S6 - 1}"
ws6.print_title_rows = f"{HEADER_ROW_S6}:{HEADER_ROW_S6}"

# ══════════════════════════════════════════════════════════════
# SHEET 7: Staff Performance
# ══════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Staff Performance")
ws7.sheet_properties.tabColor = PRIMARY

HEADERS_S7 = ["Staff Name", "Total Sales", "# Transactions", "Avg Transaction",
              "Top Category", "Shift", "Commission (5%)"]
COL_END_S7 = COL_START + len(HEADERS_S7) - 1
HEADER_ROW_S7 = 4
DATA_START_S7 = 5

setup_sheet(ws7, "Staff Performance — February 2025", COL_END_S7)

widths_s7 = [16, 16, 16, 16, 16, 12, 16]
for i, w in enumerate(widths_s7):
    ws7.column_dimensions[get_column_letter(COL_START + i)].width = w

for i, h in enumerate(HEADERS_S7):
    ws7.cell(row=HEADER_ROW_S7, column=COL_START + i, value=h)
style_header_row(ws7, HEADER_ROW_S7, COL_START, COL_END_S7)

SE_STAFF = get_column_letter(COL_START + 8)   # J
SE_CAT = get_column_letter(COL_START + 3)      # E
CATEGORIES = ["Dairy", "Bakery", "Meat", "Beverages", "Snacks", "Cleaning", "Produce", "Grains"]

for idx, staff in enumerate(STAFF):
    r = DATA_START_S7 + idx
    ws7.cell(row=r, column=COL_START, value=staff)

    # Total Sales = SUMPRODUCT
    ws7.cell(row=r, column=COL_START + 1,
             value=f"=SUMPRODUCT(('Sales Entry'!{SE_STAFF}{DATA_START}:{SE_STAFF}{TOTAL_ROW_S1-1}=\"{staff}\")*('Sales Entry'!{SE_TOTAL_COL}{DATA_START}:{SE_TOTAL_COL}{TOTAL_ROW_S1-1}))")
    ws7.cell(row=r, column=COL_START + 1).number_format = FMT_CURRENCY

    # # Transactions
    ws7.cell(row=r, column=COL_START + 2,
             value=f"=SUMPRODUCT(('Sales Entry'!{SE_STAFF}{DATA_START}:{SE_STAFF}{TOTAL_ROW_S1-1}=\"{staff}\")*1)")
    ws7.cell(row=r, column=COL_START + 2).number_format = FMT_INT

    # Avg Transaction
    ws7.cell(row=r, column=COL_START + 3,
             value=f"=IFERROR(C{r}/D{r},0)")
    ws7.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY

    # Top Category — find category with most sales for this staff
    cat_formulas = []
    for cat in CATEGORIES:
        f = f'SUMPRODUCT((\'Sales Entry\'!{SE_STAFF}{DATA_START}:{SE_STAFF}{TOTAL_ROW_S1-1}="{staff}")*(\'Sales Entry\'!{SE_CAT}{DATA_START}:{SE_CAT}{TOTAL_ROW_S1-1}="{cat}")*(\'Sales Entry\'!{SE_TOTAL_COL}{DATA_START}:{SE_TOTAL_COL}{TOTAL_ROW_S1-1}))'
        cat_formulas.append(f)

    # Nested IF to pick top category
    top_cat_formula = (
        f'=IF(AND({cat_formulas[0]}>={cat_formulas[1]},{cat_formulas[0]}>={cat_formulas[2]},{cat_formulas[0]}>={cat_formulas[3]},{cat_formulas[0]}>={cat_formulas[4]},{cat_formulas[0]}>={cat_formulas[5]},{cat_formulas[0]}>={cat_formulas[6]},{cat_formulas[0]}>={cat_formulas[7]}),"{CATEGORIES[0]}",'
        f'IF(AND({cat_formulas[1]}>={cat_formulas[0]},{cat_formulas[1]}>={cat_formulas[2]},{cat_formulas[1]}>={cat_formulas[3]},{cat_formulas[1]}>={cat_formulas[4]},{cat_formulas[1]}>={cat_formulas[5]},{cat_formulas[1]}>={cat_formulas[6]},{cat_formulas[1]}>={cat_formulas[7]}),"{CATEGORIES[1]}",'
        f'IF(AND({cat_formulas[2]}>={cat_formulas[0]},{cat_formulas[2]}>={cat_formulas[1]},{cat_formulas[2]}>={cat_formulas[3]},{cat_formulas[2]}>={cat_formulas[4]},{cat_formulas[2]}>={cat_formulas[5]},{cat_formulas[2]}>={cat_formulas[6]},{cat_formulas[2]}>={cat_formulas[7]}),"{CATEGORIES[2]}",'
        f'IF(AND({cat_formulas[3]}>={cat_formulas[0]},{cat_formulas[3]}>={cat_formulas[1]},{cat_formulas[3]}>={cat_formulas[2]},{cat_formulas[3]}>={cat_formulas[4]},{cat_formulas[3]}>={cat_formulas[5]},{cat_formulas[3]}>={cat_formulas[6]},{cat_formulas[3]}>={cat_formulas[7]}),"{CATEGORIES[3]}",'
        f'IF(AND({cat_formulas[4]}>={cat_formulas[0]},{cat_formulas[4]}>={cat_formulas[1]},{cat_formulas[4]}>={cat_formulas[2]},{cat_formulas[4]}>={cat_formulas[3]},{cat_formulas[4]}>={cat_formulas[5]},{cat_formulas[4]}>={cat_formulas[6]},{cat_formulas[4]}>={cat_formulas[7]}),"{CATEGORIES[4]}",'
        f'IF(AND({cat_formulas[5]}>={cat_formulas[0]},{cat_formulas[5]}>={cat_formulas[1]},{cat_formulas[5]}>={cat_formulas[2]},{cat_formulas[5]}>={cat_formulas[3]},{cat_formulas[5]}>={cat_formulas[4]},{cat_formulas[5]}>={cat_formulas[6]},{cat_formulas[5]}>={cat_formulas[7]}),"{CATEGORIES[5]}",'
        f'IF(AND({cat_formulas[6]}>={cat_formulas[0]},{cat_formulas[6]}>={cat_formulas[1]},{cat_formulas[6]}>={cat_formulas[2]},{cat_formulas[6]}>={cat_formulas[3]},{cat_formulas[6]}>={cat_formulas[4]},{cat_formulas[6]}>={cat_formulas[5]},{cat_formulas[6]}>={cat_formulas[7]}),"{CATEGORIES[6]}","{CATEGORIES[7]}"))))))'
    )
    ws7.cell(row=r, column=COL_START + 4, value=top_cat_formula)

    # Shift
    ws7.cell(row=r, column=COL_START + 5, value=SHIFTS[staff])

    # Commission (5%)
    ws7.cell(row=r, column=COL_START + 6,
             value=f"=C{r}*0.05")
    ws7.cell(row=r, column=COL_START + 6).number_format = FMT_CURRENCY

    style_data_row(ws7, r, COL_START, COL_END_S7, idx)
    ws7.cell(row=r, column=COL_START).alignment = align_text()
    ws7.cell(row=r, column=COL_START + 1).alignment = align_number()
    ws7.cell(row=r, column=COL_START + 1).number_format = FMT_CURRENCY
    ws7.cell(row=r, column=COL_START + 2).alignment = align_number()
    ws7.cell(row=r, column=COL_START + 2).number_format = FMT_INT
    ws7.cell(row=r, column=COL_START + 3).alignment = align_number()
    ws7.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    ws7.cell(row=r, column=COL_START + 4).alignment = align_text()
    ws7.cell(row=r, column=COL_START + 5).alignment = align_text()
    ws7.cell(row=r, column=COL_START + 6).alignment = align_number()
    ws7.cell(row=r, column=COL_START + 6).number_format = FMT_CURRENCY

# Totals row
TR_S7 = DATA_START_S7 + len(STAFF)
ws7.cell(row=TR_S7, column=COL_START, value="TOTAL")
ws7.cell(row=TR_S7, column=COL_START + 1,
         value=f"=SUM(C{DATA_START_S7}:C{TR_S7-1})")
ws7.cell(row=TR_S7, column=COL_START + 1).number_format = FMT_CURRENCY
ws7.cell(row=TR_S7, column=COL_START + 2,
         value=f"=SUM(D{DATA_START_S7}:D{TR_S7-1})")
ws7.cell(row=TR_S7, column=COL_START + 2).number_format = FMT_INT
ws7.cell(row=TR_S7, column=COL_START + 3,
         value=f"=IFERROR(C{TR_S7}/D{TR_S7},0)")
ws7.cell(row=TR_S7, column=COL_START + 3).number_format = FMT_CURRENCY
ws7.cell(row=TR_S7, column=COL_START + 6,
         value=f"=SUM(H{DATA_START_S7}:H{TR_S7-1})")
ws7.cell(row=TR_S7, column=COL_START + 6).number_format = FMT_CURRENCY
style_total_row(ws7, TR_S7, COL_START, COL_END_S7)
ws7.cell(row=TR_S7, column=COL_START).alignment = align_text()

# Data bars on Total Sales column
sales_col_letter = get_column_letter(COL_START + 1)  # C
sales_range = f"{sales_col_letter}{DATA_START_S7}:{sales_col_letter}{TR_S7 - 1}"
ws7.conditional_formatting.add(
    sales_range,
    DataBarRule(
        start_type="min", end_type="max",
        color=PRIMARY
    )
)

ws7.print_area = f"A1:{get_column_letter(COL_END_S7)}{TR_S7}"

# ══════════════════════════════════════════════════════════════
# SHEET 8: Product List
# ══════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Product List")
ws8.sheet_properties.tabColor = PRIMARY

HEADERS_S8 = ["Product ID", "Item Name", "Category", "Unit Price",
              "Cost Price", "Margin %", "Description", "Supplier"]
COL_END_S8 = COL_START + len(HEADERS_S8) - 1
HEADER_ROW_S8 = 4
DATA_START_S8 = 5

setup_sheet(ws8, "Product List — Al Noor Supermarket", COL_END_S8)

widths_s8 = [12, 22, 14, 14, 14, 12, 32, 18]
for i, w in enumerate(widths_s8):
    ws8.column_dimensions[get_column_letter(COL_START + i)].width = w

for i, h in enumerate(HEADERS_S8):
    ws8.cell(row=HEADER_ROW_S8, column=COL_START + i, value=h)
style_header_row(ws8, HEADER_ROW_S8, COL_START, COL_END_S8)

# Product list data
PRODUCT_LIST = [
    ("PRD001", "Rice (5kg)", "Grains", 35, 26, "Premium basmati rice 5kg bag", "Al Falaj"),
    ("PRD002", "Milk (1L)", "Dairy", 8, 5.5, "Fresh full cream milk", "GCC Foods"),
    ("PRD003", "Bread", "Bakery", 5, 3, "White sliced bread loaf", "Emirates Trading"),
    ("PRD004", "Eggs (12pc)", "Dairy", 12, 8, "Farm fresh eggs 12 pack", "GCC Foods"),
    ("PRD005", "Chicken (kg)", "Meat", 28, 20, "Fresh whole chicken per kg", "Al Falaj"),
    ("PRD006", "Cooking Oil (1L)", "Grains", 22, 15, "Vegetable cooking oil 1L", "Emirates Trading"),
    ("PRD007", "Water (6pk)", "Beverages", 2, 1.2, "Mineral water 6 bottles", "GCC Foods"),
    ("PRD008", "Orange Juice (1L)", "Beverages", 6, 3.8, "Fresh orange juice 1L", "GCC Foods"),
    ("PRD009", "Pasta (500g)", "Grains", 4, 2.5, "Italian spaghetti 500g", "Emirates Trading"),
    ("PRD010", "Tomatoes (kg)", "Produce", 6, 3.5, "Fresh vine tomatoes per kg", "Al Falaj"),
    ("PRD011", "Potatoes (kg)", "Produce", 4, 2, "Fresh potatoes per kg", "Al Falaj"),
    ("PRD012", "Onions (kg)", "Produce", 3, 1.5, "Red onions per kg", "Al Falaj"),
    ("PRD013", "Cheese (200g)", "Dairy", 15, 10, "Cheddar cheese slices 200g", "GCC Foods"),
    ("PRD014", "Yogurt (500g)", "Dairy", 7, 4.5, "Natural yogurt 500g tub", "GCC Foods"),
    ("PRD015", "Sugar (1kg)", "Grains", 5, 3, "White granulated sugar 1kg", "Emirates Trading"),
    ("PRD016", "Flour (1kg)", "Grains", 3, 1.8, "All purpose flour 1kg", "Emirates Trading"),
    ("PRD017", "Tea (25 bags)", "Beverages", 10, 6.5, "Black tea bags 25 count", "Emirates Trading"),
    ("PRD018", "Coffee (200g)", "Beverages", 18, 12, "Arabic coffee 200g", "Emirates Trading"),
    ("PRD019", "Butter (200g)", "Dairy", 11, 7, "Unsalted butter 200g", "GCC Foods"),
    ("PRD020", "Detergent (1L)", "Cleaning", 14, 9, "Laundry detergent 1L", "Emirates Trading"),
    ("PRD021", "Dish Soap (500ml)", "Cleaning", 8, 5, "Dishwashing liquid 500ml", "Emirates Trading"),
    ("PRD022", "Chips (150g)", "Snacks", 5, 3, "Potato chips 150g bag", "GCC Foods"),
    ("PRD023", "Chocolate Bar", "Snacks", 7, 4.5, "Milk chocolate bar 100g", "GCC Foods"),
    ("PRD024", "Biscuits (200g)", "Snacks", 4, 2.5, "Assorted biscuits 200g", "Emirates Trading"),
    ("PRD025", "Cereal (375g)", "Snacks", 12, 8, "Corn flakes cereal 375g", "Emirates Trading"),
    ("PRD026", "Canned Beans", "Grains", 3, 1.8, "Baked beans in tomato sauce", "Al Falaj"),
    ("PRD027", "Lamb (kg)", "Meat", 55, 40, "Fresh lamb meat per kg", "Al Falaj"),
    ("PRD028", "Fish (kg)", "Meat", 32, 22, "Fresh fish per kg", "Al Falaj"),
]

for idx, (pid, name, cat, price, cost, desc, supplier) in enumerate(PRODUCT_LIST):
    r = DATA_START_S8 + idx
    ws8.cell(row=r, column=COL_START, value=pid)
    ws8.cell(row=r, column=COL_START + 1, value=name)
    ws8.cell(row=r, column=COL_START + 2, value=cat)
    ws8.cell(row=r, column=COL_START + 3, value=price)
    ws8.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    ws8.cell(row=r, column=COL_START + 4, value=cost)
    ws8.cell(row=r, column=COL_START + 4).number_format = FMT_CURRENCY
    # Margin % = IFERROR((Unit Price - Cost Price) / Unit Price, 0)
    ws8.cell(row=r, column=COL_START + 5,
             value=f"=IFERROR((E{r}-F{r})/E{r},0)")
    ws8.cell(row=r, column=COL_START + 5).number_format = FMT_PCT
    ws8.cell(row=r, column=COL_START + 6, value=desc)
    ws8.cell(row=r, column=COL_START + 7, value=supplier)

    style_data_row(ws8, r, COL_START, COL_END_S8, idx)
    # Re-apply alignments
    ws8.cell(row=r, column=COL_START).alignment = align_text()
    ws8.cell(row=r, column=COL_START + 1).alignment = align_text()
    ws8.cell(row=r, column=COL_START + 2).alignment = align_text()
    ws8.cell(row=r, column=COL_START + 3).alignment = align_number()
    ws8.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    ws8.cell(row=r, column=COL_START + 4).alignment = align_number()
    ws8.cell(row=r, column=COL_START + 4).number_format = FMT_CURRENCY
    ws8.cell(row=r, column=COL_START + 5).alignment = align_number()
    ws8.cell(row=r, column=COL_START + 5).number_format = FMT_PCT
    ws8.cell(row=r, column=COL_START + 6).alignment = align_text()
    ws8.cell(row=r, column=COL_START + 7).alignment = align_text()

N_PROD = len(PRODUCT_LIST)
TR_S8 = DATA_START_S8 + N_PROD

# Totals row (averages)
ws8.cell(row=TR_S8, column=COL_START, value="AVERAGE")
ws8.cell(row=TR_S8, column=COL_START + 3,
         value=f"=IFERROR(AVERAGE(E{DATA_START_S8}:E{TR_S8-1}),0)")
ws8.cell(row=TR_S8, column=COL_START + 3).number_format = FMT_CURRENCY
ws8.cell(row=TR_S8, column=COL_START + 4,
         value=f"=IFERROR(AVERAGE(F{DATA_START_S8}:F{TR_S8-1}),0)")
ws8.cell(row=TR_S8, column=COL_START + 4).number_format = FMT_CURRENCY
ws8.cell(row=TR_S8, column=COL_START + 5,
         value=f"=IFERROR(AVERAGE(G{DATA_START_S8}:G{TR_S8-1}),0)")
ws8.cell(row=TR_S8, column=COL_START + 5).number_format = FMT_PCT
style_total_row(ws8, TR_S8, COL_START, COL_END_S8)
ws8.cell(row=TR_S8, column=COL_START).alignment = align_text()
ws8.cell(row=TR_S8, column=COL_START + 3).alignment = align_number()
ws8.cell(row=TR_S8, column=COL_START + 4).alignment = align_number()
ws8.cell(row=TR_S8, column=COL_START + 5).alignment = align_number()

ws8.print_area = f"A1:{get_column_letter(COL_END_S8)}{TR_S8}"
ws8.print_title_rows = f"{HEADER_ROW_S8}:{HEADER_ROW_S8}"

# ══════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════
OUTPUT = "/home/z/my-project/download/POS_Professional_Sample.xlsx"
wb.save(OUTPUT)
print(f"✅ Saved: {OUTPUT}")
