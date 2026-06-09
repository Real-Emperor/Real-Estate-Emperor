#!/usr/bin/env python3
"""
Real Estate Emperor — Real Estate Management System (SAMPLE)
AED 2000 Value Production-Ready Excel System
Bottega Palette — Dark Forest Green (2D4A3E)
"""

import sys, os
sys.path.insert(0, "/home/z/my-project/skills/xlsx")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta
import random

import templates.base as base

# Import functions (these are stable references)
from templates.base import (
    use_palette_explicit, setup_sheet, style_header_row, style_data_row,
    style_total_row, font_title, font_header, font_body, font_caption,
    font_kpi, font_kpi_label, font_subheader,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    COLUMN_WIDTHS, ROW_HEIGHTS,
    create_bar_chart, create_line_chart, create_pie_chart,
    setup_chart_titles, apply_chart_colors, apply_pie_colors,
    make_chart_title, auto_fit_columns,
)

# ============================================================
# §0  Activate Bottega Palette FIRST
# ============================================================
use_palette_explicit("bottega")

# Read color tokens FROM the module (not from local imports which are stale)
PAL_PRIMARY = base.PRIMARY
PAL_PRIMARY_LIGHT = base.PRIMARY_LIGHT
PAL_SECONDARY = base.SECONDARY
PAL_ACCENT_POSITIVE = base.ACCENT_POSITIVE
PAL_ACCENT_NEGATIVE = base.ACCENT_NEGATIVE
PAL_ACCENT_WARNING = base.ACCENT_WARNING
PAL_NEUTRAL_900 = base.NEUTRAL_900
PAL_NEUTRAL_600 = base.NEUTRAL_600
PAL_NEUTRAL_200 = base.NEUTRAL_200
PAL_NEUTRAL_100 = base.NEUTRAL_100
PAL_NEUTRAL_0 = base.NEUTRAL_0
PAL_CHART_COLORS = base.CHART_COLORS
PAL_CF_POSITIVE_FILL = base.CF_POSITIVE_FILL
PAL_CF_POSITIVE_FONT = base.CF_POSITIVE_FONT
PAL_CF_NEGATIVE_FILL = base.CF_NEGATIVE_FILL
PAL_CF_NEGATIVE_FONT = base.CF_NEGATIVE_FONT
PAL_CF_WARNING_FILL = base.CF_WARNING_FILL
PAL_CF_WARNING_FONT = base.CF_WARNING_FONT

# ============================================================
# §0.1  Workbook Init
# ============================================================
wb = Workbook()
wb.properties.creator = "Z.ai"

FMT_AED = '#,##0.00'
FMT_PCT = '0.0%'
FMT_INT = '#,##0'
FMT_DATE = 'YYYY-MM-DD'

def set_print(ws, area, title_rows):
    ws.print_area = area
    ws.print_title_rows = title_rows
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

def set_tab_color(ws):
    ws.sheet_properties.tabColor = PAL_PRIMARY

# ============================================================
# §1  DEMO DATA — Units
# ============================================================
# Building A: 10 Studios (A-101..A-110) + 5 One-Bedroom (A-201..A-205)
# Building B: 8 Studios (B-101..B-108) + 6 One-Bedroom (B-201..B-206)
# Building C: 6 Studios (C-101..C-106) + 8 One-Bedroom (C-201..C-208) + 2 Two-Bedroom (C-301..C-302)
# Building D: 4 Studios (D-101..D-104) + 4 One-Bedroom (D-201..D-204) + 2 Shops (D-S01..D-S02)

STUDIO_RENTS  = [2200, 2300, 2400, 2500, 2600, 2700, 2800, 2350, 2450, 2550,
                 2250, 2650, 2750, 2400, 2500, 2600, 2300, 2550, 2650, 2350,
                 2450, 2550, 2200, 2300, 2450, 2500, 2600, 2700]
ONEBR_RENTS   = [3500, 3600, 3700, 3800, 3900, 4000, 4100, 4200,
                 3550, 3650, 3750, 3850, 3950, 4050, 4150,
                 3600, 3800, 4000, 3700, 3900, 4100, 3500, 3800]
TWOBR_RENTS   = [5000, 5500, 5200, 5300]
SHOP_RENTS    = [6000, 8000, 7000, 7500]

STUDIO_SIZE   = [440, 450, 460, 450, 470, 480, 450, 460, 440, 450,
                 450, 460, 470, 450, 440, 460, 450, 470, 440, 450,
                 460, 450, 440, 460, 450, 470, 450, 460]
ONEBR_SIZE    = [740, 750, 760, 750, 770, 780, 750, 760,
                 740, 750, 760, 770, 750, 780, 760,
                 740, 760, 750, 780, 760, 750, 740, 770]
TWOBR_SIZE    = [1100, 1150, 1120, 1130]
SHOP_SIZE     = [480, 520, 500, 510]

# Build all 55 units
units = []
idx = 0
# Building A: 10 Studios
for i in range(10):
    units.append({
        'id': f'A-{101+i}', 'building': 'A', 'floor': 1 + (i // 4),
        'type': 'Studio', 'size': STUDIO_SIZE[idx], 'rent': STUDIO_RENTS[idx]
    })
    idx += 1
# Building A: 5 One-Bedroom
for i in range(5):
    units.append({
        'id': f'A-{201+i}', 'building': 'A', 'floor': 2 + (i // 3),
        'type': '1 Bedroom', 'size': ONEBR_SIZE[i], 'rent': ONEBR_RENTS[i]
    })
a1br = 5

# Building B: 8 Studios
for i in range(8):
    units.append({
        'id': f'B-{101+i}', 'building': 'B', 'floor': 1 + (i // 4),
        'type': 'Studio', 'size': STUDIO_SIZE[idx], 'rent': STUDIO_RENTS[idx]
    })
    idx += 1
# Building B: 6 One-Bedroom
for i in range(6):
    units.append({
        'id': f'B-{201+i}', 'building': 'B', 'floor': 2 + (i // 3),
        'type': '1 Bedroom', 'size': ONEBR_SIZE[a1br + i], 'rent': ONEBR_RENTS[a1br + i]
    })
b1br = a1br + 6

# Building C: 6 Studios
for i in range(6):
    units.append({
        'id': f'C-{101+i}', 'building': 'C', 'floor': 1 + (i // 3),
        'type': 'Studio', 'size': STUDIO_SIZE[idx], 'rent': STUDIO_RENTS[idx]
    })
    idx += 1
# Building C: 8 One-Bedroom
for i in range(8):
    units.append({
        'id': f'C-{201+i}', 'building': 'C', 'floor': 2 + (i // 4),
        'type': '1 Bedroom', 'size': ONEBR_SIZE[b1br + i], 'rent': ONEBR_RENTS[b1br + i]
    })
c1br = b1br + 8
# Building C: 2 Two-Bedroom
for i in range(2):
    units.append({
        'id': f'C-{301+i}', 'building': 'C', 'floor': 4,
        'type': '2 Bedroom', 'size': TWOBR_SIZE[i], 'rent': TWOBR_RENTS[i]
    })

# Building D: 4 Studios
for i in range(4):
    units.append({
        'id': f'D-{101+i}', 'building': 'D', 'floor': 1 + (i // 2),
        'type': 'Studio', 'size': STUDIO_SIZE[idx], 'rent': STUDIO_RENTS[idx]
    })
    idx += 1
# Building D: 4 One-Bedroom
for i in range(4):
    units.append({
        'id': f'D-{201+i}', 'building': 'D', 'floor': 2 + (i // 2),
        'type': '1 Bedroom', 'size': ONEBR_SIZE[c1br + i], 'rent': ONEBR_RENTS[c1br + i]
    })
# Building D: 2 Shops
for i in range(2):
    units.append({
        'id': f'D-S0{i+1}', 'building': 'D', 'floor': 0,
        'type': 'Shop', 'size': SHOP_SIZE[i], 'rent': SHOP_RENTS[i]
    })

# Assign vacant units (7 total)
VACANT_INDICES = [3, 8, 16, 23, 31, 40, 52]  # scattered across buildings
vacant_set = set(VACANT_INDICES)

# Tenant names (48 tenants)
TENANT_NAMES = [
    "Muhammad Ali", "Ahmed Khan", "Fatima Noor", "Rajesh Kumar", "Priya Sharma",
    "Omar Hassan", "Layla Mahmoud", "Youssef Ibrahim", "Sunil Patel", "Maria Santos",
    "Jose Reyes", "Nadia Ahmed", "Hassan Malik", "Deepak Sharma", "Amina Yusuf",
    "Khalid Al Mansouri", "Sunita Devi", "Arjun Reddy", "Zainab Khan", "Rashid Ahmad",
    "Meena Kumari", "Tariq Hussain", "Sara Mahmoud", "Vikram Singh", "Noor Fatima",
    "Imran Sheikh", "Ritu Patel", "Samir Abbas", "Pooja Sharma", "Farhan Ali",
    "Amira Hassan", "Raj Patel", "Leila Omar", "Arun Kumar", "Sanjay Gupta",
    "Huda Mahmoud", "Waqar Ahmed", "Neeta Joshi", "Bilal Siddiqui", "Priya Nair",
    "Faisal Rahman", "Anita Desai", "Karim Othman", "Ramesh Iyer", "Salma Khatoon",
    "Naveed Akhtar", "Deepa Menon", "Usman Ghani"
]

TENANT_PHONES = [
    "050-123-4567", "052-234-5678", "054-345-6789", "055-456-7890", "050-567-8901",
    "052-678-9012", "054-789-0123", "055-890-1234", "050-901-2345", "052-012-3456",
    "054-123-4560", "055-234-5670", "050-345-6780", "052-456-7890", "054-567-8900",
    "055-678-9010", "050-789-0120", "052-890-1230", "054-901-2340", "055-012-3450",
    "050-111-2233", "052-222-3344", "054-333-4455", "055-444-5566", "050-555-6677",
    "052-666-7788", "054-777-8899", "055-888-9900", "050-999-0011", "052-100-2233",
    "054-211-3344", "055-322-4455", "050-433-5566", "052-544-6677", "054-655-7788",
    "055-766-8899", "050-877-9900", "052-988-0011", "054-099-1122", "055-210-2233",
    "050-321-3344", "052-432-4455", "054-543-5566", "055-654-6677", "050-765-7788",
    "052-876-8899", "054-987-9900", "055-098-0011"
]

NATIONALITIES = [
    "Pakistani", "Pakistani", "Pakistani", "Indian", "Indian",
    "Syrian", "Egyptian", "Jordanian", "Indian", "Filipino",
    "Filipino", "Pakistani", "Pakistani", "Indian", "Egyptian",
    "Emirati", "Indian", "Indian", "Pakistani", "Pakistani",
    "Indian", "Pakistani", "Jordanian", "Indian", "Pakistani",
    "Pakistani", "Indian", "Jordanian", "Indian", "Pakistani",
    "Syrian", "Indian", "Egyptian", "Indian", "Indian",
    "Syrian", "Pakistani", "Indian", "Pakistani", "Indian",
    "Pakistani", "Indian", "Jordanian", "Indian", "Pakistani",
    "Pakistani", "Indian", "Pakistani"
]

EMPLOYERS = [
    "Al Fardan Group", "Emirates NBD", "Dubai Municipality", "Lulu Group", "Landmark Group",
    "Al Shafar United", "Arabtec Construction", "Etisalat", "DP World", "Jollibee UAE",
    "Al Naboodah Group", "Engro Corporation", "Emirates Airlines", "Tata Consultancy", "Al Futtaim Group",
    "Abu Dhabi Govt", "Infosys UAE", "Wipro Technologies", "Careem", "Meezan Bank UAE",
    "Tech Mahindra", "J. Brothers Trading", "Rotana Hotels", "HCL Technologies", "Dubai Islamic Bank",
    "Sharaf Group", "Larsen & Toubro", "Aramex", "ICICI Bank UAE", "Al Haramain Trading",
    "Damac Properties", "Capgemini UAE", "Al Habtoor Group", "Wipro UAE", "Emirates Airline",
    "Emaar Properties", "Flydubai", "HCL UAE", "Habib Bank AG Zurich", "TCS Dubai",
    "Gulftainer", "Mphasis UAE", "Al Jaber Group", "Wipro Middle East", "Al Ghurair Group",
    "Samba Bank UAE", "Cyient UAE", "Fauji Foundation UAE"
]

PAYMENT_METHODS = [
    "Bank Transfer", "Cheque", "Cash", "Bank Transfer", "Cheque",
    "Cash", "Bank Transfer", "Cheque", "Bank Transfer", "Cash",
    "Bank Transfer", "Cheque", "Cash", "Bank Transfer", "Bank Transfer",
    "Cheque", "Cash", "Bank Transfer", "Cheque", "Cash",
    "Bank Transfer", "Bank Transfer", "Cheque", "Cash", "Bank Transfer",
    "Cheque", "Cash", "Bank Transfer", "Bank Transfer", "Cheque",
    "Cash", "Bank Transfer", "Cheque", "Bank Transfer", "Cash",
    "Bank Transfer", "Cheque", "Cash", "Bank Transfer", "Cheque",
    "Cash", "Bank Transfer", "Bank Transfer", "Cheque", "Cash",
    "Bank Transfer", "Cheque", "Bank Transfer"
]

# Build tenant assignments (only for occupied units)
tenants = []
t_idx = 0
contract_starts = []
contract_ends = []

# Define contract dates for each tenant
base_date = date(2024, 1, 1)
for i in range(48):
    # Vary contract start dates
    month_offset = (i * 3) % 12
    day = (i % 28) + 1
    cs = date(2024 + (month_offset // 12), 1 + month_offset, min(day, 28))
    ce = date(cs.year + 1, cs.month, cs.day)
    contract_starts.append(cs)
    contract_ends.append(ce)

for i, u in enumerate(units):
    if i not in vacant_set:
        u['tenant_idx'] = t_idx
        u['status'] = 'Rented'
        u['tenant'] = TENANT_NAMES[t_idx]
        u['phone'] = TENANT_PHONES[t_idx]
        u['contract_start'] = contract_starts[t_idx]
        u['contract_end'] = contract_ends[t_idx]
        u['nationality'] = NATIONALITIES[t_idx]
        u['employer'] = EMPLOYERS[t_idx]
        u['payment_method'] = PAYMENT_METHODS[t_idx]
        u['security_deposit'] = u['rent']
        t_idx += 1
    else:
        u['tenant_idx'] = None
        u['status'] = 'Vacant'
        u['tenant'] = ''
        u['phone'] = ''
        u['contract_start'] = None
        u['contract_end'] = None
        u['nationality'] = ''
        u['employer'] = ''
        u['payment_method'] = ''
        u['security_deposit'] = 0

# Adjust some contract end dates to create expiring/expired scenarios
# 5 contracts expiring soon (ending in April 2025)
# 3 contracts renewal due (ending May-June 2025)
# 5 contracts expired (past end date)
# These are tenants at specific indices

# Expired contracts (5) — ended before April 2025
expired_tenant_indices = [0, 10, 22, 35, 45]
for ti in expired_tenant_indices:
    contract_ends[ti] = date(2025, 2, 28)
    contract_starts[ti] = date(2024, 3, 1)
    # Update the unit
    for u in units:
        if u.get('tenant_idx') == ti:
            u['contract_end'] = contract_ends[ti]
            u['contract_start'] = contract_starts[ti]

# Expiring soon (5) — ending in April 2025
expiring_tenant_indices = [3, 15, 27, 38, 47]
for ti in expiring_tenant_indices:
    contract_ends[ti] = date(2025, 4, 30)
    contract_starts[ti] = date(2024, 5, 1)
    for u in units:
        if u.get('tenant_idx') == ti:
            u['contract_end'] = contract_ends[ti]
            u['contract_start'] = contract_starts[ti]

# Renewal due (3) — ending May-June 2025
renewal_tenant_indices = [5, 19, 42]
for ti in renewal_tenant_indices:
    end_month = 5 + (ti % 2)
    contract_ends[ti] = date(2025, end_month, 15)
    contract_starts[ti] = date(2024, end_month, 15)
    for u in units:
        if u.get('tenant_idx') == ti:
            u['contract_end'] = contract_ends[ti]
            u['contract_start'] = contract_starts[ti]

# ============================================================
# §2  SHEET 1: Dashboard
# ============================================================
ws_dash = wb.active
ws_dash.title = "Dashboard"
set_tab_color(ws_dash)
last_col_dash = 14

setup_sheet(ws_dash, "Real Estate Emperor — Real Estate Dashboard", last_col_dash)

# KPI Cards in row 4-6
kpi_labels = ["Total Units", "Occupied", "Vacant", "Occupancy Rate", "Monthly Revenue", "Net Profit"]
kpi_values = [55, 48, 7, 0.873, 185340.00, 95200.00]
kpi_formats = [FMT_INT, FMT_INT, FMT_INT, FMT_PCT, FMT_AED, FMT_AED]

for i in range(6):
    col = 2 + i * 2
    # Label row (4)
    ws_dash.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    lbl = ws_dash.cell(row=4, column=col, value=kpi_labels[i])
    lbl.font = font_kpi_label()
    lbl.alignment = Alignment(horizontal="center", vertical="bottom")
    # Value row (5)
    ws_dash.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
    val = ws_dash.cell(row=5, column=col, value=kpi_values[i])
    val.font = font_kpi()
    val.alignment = Alignment(horizontal="center", vertical="center")
    val.number_format = kpi_formats[i]
    # Underline row (6) — thin border
    for c in range(col, col + 2):
        cell6 = ws_dash.cell(row=6, column=c)
        cell6.border = Border(top=Side(style="medium", color=PAL_PRIMARY))

ws_dash.row_dimensions[4].height = 18
ws_dash.row_dimensions[5].height = 36
ws_dash.row_dimensions[6].height = 8

# Spacer
ws_dash.row_dimensions[7].height = 16

# ---- Chart Data Tables (hidden helper data starting row 8) ----
# a) Occupancy by Building
ws_dash.cell(row=8, column=2, value="Building").font = font_caption()
ws_dash.cell(row=8, column=3, value="Occupied").font = font_caption()
ws_dash.cell(row=8, column=4, value="Vacant").font = font_caption()

# Count occupied/vacant by building
bldg_occ = {'A': 0, 'B': 0, 'C': 0, 'D': 0}
bldg_vac = {'A': 0, 'B': 0, 'C': 0, 'D': 0}
for u in units:
    b = u['building']
    if u['status'] == 'Rented':
        bldg_occ[b] += 1
    else:
        bldg_vac[b] += 1

for i, bld in enumerate(['A', 'B', 'C', 'D']):
    r = 9 + i
    ws_dash.cell(row=r, column=2, value=f"Building {bld}").font = font_body()
    ws_dash.cell(row=r, column=3, value=bldg_occ[bld]).font = font_body()
    ws_dash.cell(row=r, column=4, value=bldg_vac[bld]).font = font_body()

# b) Monthly Revenue Trend (Jan-Apr 2025)
ws_dash.cell(row=8, column=6, value="Month").font = font_caption()
ws_dash.cell(row=8, column=7, value="Revenue").font = font_caption()
rev_months = [("Jan 2025", 172500), ("Feb 2025", 175800), ("Mar 2025", 181200), ("Apr 2025", 185340)]
for i, (m, v) in enumerate(rev_months):
    ws_dash.cell(row=9 + i, column=6, value=m).font = font_body()
    ws_dash.cell(row=9 + i, column=7, value=v).font = font_body()
    ws_dash.cell(row=9 + i, column=7).number_format = FMT_AED

# c) Expense Breakdown
ws_dash.cell(row=8, column=9, value="Category").font = font_caption()
ws_dash.cell(row=8, column=10, value="Amount").font = font_caption()
expense_cats = [
    ("Manpower", 24000), ("Maintenance", 10500), ("Municipality Fees", 8500),
    ("Leasing Commissions", 5200), ("Insurance", 4200), ("Utilities (DEWA)", 6800),
    ("Other Operating", 13000)
]
for i, (cat, amt) in enumerate(expense_cats):
    ws_dash.cell(row=9 + i, column=9, value=cat).font = font_body()
    ws_dash.cell(row=9 + i, column=10, value=amt).font = font_body()
    ws_dash.cell(row=9 + i, column=10).number_format = FMT_AED

# d) Unit Type Distribution
ws_dash.cell(row=14, column=2, value="Unit Type").font = font_caption()
ws_dash.cell(row=14, column=3, value="Count").font = font_caption()
type_counts = {'Studio': 0, '1 Bedroom': 0, '2 Bedroom': 0, 'Shop': 0}
for u in units:
    type_counts[u['type']] += 1
for i, (tp, cnt) in enumerate(type_counts.items()):
    ws_dash.cell(row=15 + i, column=2, value=tp).font = font_body()
    ws_dash.cell(row=15 + i, column=3, value=cnt).font = font_body()

# ---- Charts ----
# Chart a: Occupancy by Building — Bar
chart_a = create_bar_chart(width=16, height=10)
data_a = Reference(ws_dash, min_col=3, max_col=4, min_row=8, max_row=12)
cats_a = Reference(ws_dash, min_col=2, min_row=9, max_row=12)
chart_a.add_data(data_a, titles_from_data=True)
chart_a.set_categories(cats_a)
setup_chart_titles(chart_a, title="Occupancy by Building", y_title="Units")
apply_chart_colors(chart_a, [PAL_PRIMARY, PAL_ACCENT_WARNING])
chart_a.legend.position = 'b'
ws_dash.add_chart(chart_a, "B20")

# Chart b: Monthly Revenue Trend — Line
chart_b = create_line_chart(width=16, height=10)
data_b = Reference(ws_dash, min_col=7, min_row=8, max_row=12)
cats_b = Reference(ws_dash, min_col=6, min_row=9, max_row=12)
chart_b.add_data(data_b, titles_from_data=True)
chart_b.set_categories(cats_b)
setup_chart_titles(chart_b, title="Monthly Revenue Trend", y_title="AED")
apply_chart_colors(chart_b, [PAL_PRIMARY])
chart_b.legend = None
ws_dash.add_chart(chart_b, "I20")

# Chart c: Expense Breakdown — Pie
chart_c = create_pie_chart(width=14, height=10)
data_c = Reference(ws_dash, min_col=10, min_row=8, max_row=15)
cats_c = Reference(ws_dash, min_col=9, min_row=9, max_row=15)
chart_c.add_data(data_c, titles_from_data=True)
chart_c.set_categories(cats_c)
setup_chart_titles(chart_c, title="Expense Breakdown")
apply_pie_colors(chart_c, 7)
chart_c.dataLabels = DataLabelList()
chart_c.dataLabels.showPercent = True
chart_c.dataLabels.showCatName = True
ws_dash.add_chart(chart_c, "B36")

# Chart d: Unit Type Distribution — Pie
chart_d = create_pie_chart(width=14, height=10)
data_d = Reference(ws_dash, min_col=3, min_row=14, max_row=18)
cats_d = Reference(ws_dash, min_col=2, min_row=15, max_row=18)
chart_d.add_data(data_d, titles_from_data=True)
chart_d.set_categories(cats_d)
setup_chart_titles(chart_d, title="Unit Type Distribution")
apply_pie_colors(chart_d, 4)
chart_d.dataLabels = DataLabelList()
chart_d.dataLabels.showPercent = True
chart_d.dataLabels.showCatName = True
ws_dash.add_chart(chart_d, "I36")

# Column widths for dashboard
for c in range(1, last_col_dash + 1):
    ws_dash.column_dimensions[get_column_letter(c)].width = 12
ws_dash.column_dimensions['A'].width = 3

set_print(ws_dash, f"A1:{get_column_letter(last_col_dash)}52", "2:3")


# ============================================================
# §3  SHEET 2: Property Registry
# ============================================================
ws_prop = wb.create_sheet("Property Registry")
set_tab_color(ws_prop)

prop_headers = ["Unit ID", "Building", "Floor", "Unit Type", "Size (sqft)",
                "Monthly Rent", "Municipality Fee", "Status", "Current Tenant",
                "Contract Start", "Contract End", "Notes"]
prop_last_col = 1 + len(prop_headers)  # col B to M (13 cols)

setup_sheet(ws_prop, "Property Registry — Unit Master List", prop_last_col)

# Headers at row 4
for i, h in enumerate(prop_headers):
    ws_prop.cell(row=4, column=2 + i, value=h)
style_header_row(ws_prop, 4, 2, prop_last_col)

# Data rows (5 onwards)
for idx_u, u in enumerate(units):
    r = 5 + idx_u
    muni_fee = round(u['rent'] * 0.05, 2)
    notes = ""
    if idx_u in vacant_set:
        notes = "Available for rent"
    elif u.get('contract_end') and u['contract_end'] <= date(2025, 4, 30):
        notes = "Contract expiring soon"

    vals = [
        u['id'], f"Building {u['building']}", u['floor'], u['type'],
        u['size'], u['rent'], muni_fee, u['status'],
        u['tenant'] if u['tenant'] else "",
        u['contract_start'] if u['contract_start'] else "",
        u['contract_end'] if u['contract_end'] else "",
        notes
    ]
    for i, v in enumerate(vals):
        cell = ws_prop.cell(row=r, column=2 + i, value=v)
        if i in (5, 6):  # rent, muni fee
            cell.number_format = FMT_AED
            cell.alignment = align_number()
        elif i == 4:  # size
            cell.number_format = FMT_INT
            cell.alignment = align_number()
        elif i == 2:  # floor
            cell.alignment = align_date()
        elif i in (9, 10):  # dates
            cell.number_format = FMT_DATE
            cell.alignment = align_date()
        elif i == 3:  # type
            cell.alignment = align_text()
        else:
            cell.alignment = align_text()

    style_data_row(ws_prop, r, 2, prop_last_col, idx_u)

# Totals row
total_r = 5 + len(units)
ws_prop.cell(row=total_r, column=2, value="TOTAL").font = font_subheader()
ws_prop.cell(row=total_r, column=2).alignment = align_text()
# Total units
ws_prop.cell(row=total_r, column=5, value=len(units)).number_format = FMT_INT
ws_prop.cell(row=total_r, column=5).alignment = align_number()
# Sum rents
ws_prop.cell(row=total_r, column=7, value=f"=SUM(G5:G{total_r-1})").number_format = FMT_AED
ws_prop.cell(row=total_r, column=7).alignment = align_number()
# Sum muni fees
ws_prop.cell(row=total_r, column=8, value=f"=SUM(H5:H{total_r-1})").number_format = FMT_AED
ws_prop.cell(row=total_r, column=8).alignment = align_number()
style_total_row(ws_prop, total_r, 2, prop_last_col)

# Conditional formatting on Status (col I = 9)
ws_prop.conditional_formatting.add(
    f"I5:I{total_r-1}",
    CellIsRule(operator='equal', formula=['"Rented"'], fill=PAL_CF_POSITIVE_FILL, font=PAL_CF_POSITIVE_FONT)
)
ws_prop.conditional_formatting.add(
    f"I5:I{total_r-1}",
    CellIsRule(operator='equal', formula=['"Vacant"'], fill=PAL_CF_NEGATIVE_FILL, font=PAL_CF_NEGATIVE_FONT)
)

# Data validation
dv_type = DataValidation(type="list", formula1='"Studio,1 Bedroom,2 Bedroom,Shop"', allow_blank=True)
dv_type.error = "Please select a valid unit type"
dv_type.errorTitle = "Invalid Unit Type"
ws_prop.add_data_validation(dv_type)
dv_type.add(f"E5:E{total_r-1}")

dv_bldg = DataValidation(type="list", formula1='"Building A,Building B,Building C,Building D"', allow_blank=True)
ws_prop.add_data_validation(dv_bldg)
dv_bldg.add(f"C5:C{total_r-1}")

dv_status = DataValidation(type="list", formula1='"Rented,Vacant"', allow_blank=True)
ws_prop.add_data_validation(dv_status)
dv_status.add(f"I5:I{total_r-1}")

# Column widths
col_widths_prop = [3, 10, 14, 8, 14, 12, 14, 14, 12, 20, 14, 14, 22]
for i, w in enumerate(col_widths_prop):
    ws_prop.column_dimensions[get_column_letter(i + 1)].width = w

# Freeze panes at C5
ws_prop.freeze_panes = "C5"

set_print(ws_prop, f"A1:{get_column_letter(prop_last_col)}{total_r}", "2:4")


# ============================================================
# §4  SHEET 3: Tenants
# ============================================================
ws_tenant = wb.create_sheet("Tenants")
set_tab_color(ws_tenant)

tenant_headers = ["Tenant ID", "Tenant Name", "Phone", "Email", "Unit ID", "Unit Type",
                  "Lease Start", "Lease End", "Duration (Months)", "Monthly Rent",
                  "Security Deposit", "Payment Method", "Nationality", "Employer", "Emergency Contact"]
tenant_last_col = 1 + len(tenant_headers)

setup_sheet(ws_tenant, "Tenant Directory", tenant_last_col)

for i, h in enumerate(tenant_headers):
    ws_tenant.cell(row=4, column=2 + i, value=h)
style_header_row(ws_tenant, 4, 2, tenant_last_col)

# Email generator
def make_email(name, idx):
    parts = name.lower().replace(" ", ".")
    domains = ["gmail.com", "yahoo.com", "hotmail.com", "outlook.com"]
    return f"{parts}{idx}@{domains[idx % len(domains)]}"

# Emergency contacts
emerg_contacts = [
    "050-987-6543", "052-876-5432", "054-765-4321", "055-654-3210", "050-543-2109",
    "052-432-1098", "054-321-0987", "055-210-9876", "050-109-8765", "052-098-7654",
    "054-887-6655", "055-776-5544", "050-665-4433", "052-554-3322", "054-443-2211",
    "055-332-1100", "050-221-1009", "052-110-9988", "054-099-8877", "055-988-7766",
    "050-877-6655", "052-766-5544", "054-655-4433", "055-544-3322", "050-433-2211",
    "052-322-1100", "054-211-0998", "055-100-9887", "050-099-8776", "052-988-7665",
    "054-877-6554", "055-766-5443", "050-655-4332", "052-544-3221", "054-433-2110",
    "055-322-1009", "050-211-0998", "052-100-9887", "054-099-8776", "055-988-7665",
    "050-877-6554", "052-766-5443", "054-655-4332", "055-544-3221", "050-433-2110",
    "052-322-1009", "054-211-0998", "055-100-9887"
]

t_idx = 0
for idx_u, u in enumerate(units):
    if u['status'] != 'Rented':
        continue
    r = 5 + t_idx
    ti = u['tenant_idx']
    cs = contract_starts[ti]
    ce = contract_ends[ti]
    duration = (ce.year - cs.year) * 12 + ce.month - cs.month

    vals = [
        f"T-{ti+1:03d}", u['tenant'], u['phone'], make_email(u['tenant'], ti),
        u['id'], u['type'], cs, ce, duration, u['rent'],
        u['security_deposit'], u['payment_method'], u['nationality'],
        u['employer'], emerg_contacts[ti]
    ]
    for i, v in enumerate(vals):
        cell = ws_tenant.cell(row=r, column=2 + i, value=v)
        if i in (9, 10):  # rent, deposit
            cell.number_format = FMT_AED
            cell.alignment = align_number()
        elif i == 8:  # duration
            cell.number_format = FMT_INT
            cell.alignment = align_number()
        elif i in (6, 7):  # dates
            cell.number_format = FMT_DATE
            cell.alignment = align_date()
        else:
            cell.alignment = align_text()

    style_data_row(ws_tenant, r, 2, tenant_last_col, t_idx)
    t_idx += 1

total_r_t = 5 + 48

# Conditional formatting on Lease End — highlight if within 30 days
ws_tenant.conditional_formatting.add(
    f"I5:I{total_r_t-1}",
    CellIsRule(operator='lessThanOrEqual', formula=['DATE(2025,4,30)'],
              fill=PAL_CF_WARNING_FILL, font=PAL_CF_WARNING_FONT)
)

# Column widths
col_widths_t = [3, 10, 20, 16, 26, 10, 14, 14, 14, 16, 14, 14, 14, 14, 18, 18]
for i, w in enumerate(col_widths_t):
    ws_tenant.column_dimensions[get_column_letter(i + 1)].width = w

ws_tenant.freeze_panes = "C5"
set_print(ws_tenant, f"A1:{get_column_letter(tenant_last_col)}{total_r_t}", "2:4")


# ============================================================
# §5  SHEET 4: Rent Collection — April 2025
# ============================================================
ws_rent = wb.create_sheet("Rent Collection")
set_tab_color(ws_rent)

rent_headers = ["Unit ID", "Tenant Name", "Monthly Rent", "Due Date", "Payment Date",
                "Amount Paid", "Outstanding", "Payment Status", "Days Late",
                "Late Fee", "Total Due", "Payment Method", "Receipt #", "Notes"]
rent_last_col = 1 + len(rent_headers)

setup_sheet(ws_rent, "Rent Collection — April 2025", rent_last_col)

for i, h in enumerate(rent_headers):
    ws_rent.cell(row=4, column=2 + i, value=h)
style_header_row(ws_rent, 4, 2, rent_last_col)

# Determine payment statuses for tenants
# 39 paid in full, 4 partial, 5 unpaid
paid_full_tenant_indices = list(range(48))
# Remove 9 tenants for partial/unpaid
partial_tenant_indices = [7, 18, 33, 41]   # 4 partial
unpaid_tenant_indices = [2, 12, 25, 37, 44]  # 5 unpaid
for ti in partial_tenant_indices + unpaid_tenant_indices:
    paid_full_tenant_indices.remove(ti)

due_date = date(2025, 4, 1)
receipt_counter = 1001

rent_data_rows = []
for idx_u, u in enumerate(units):
    r = 5 + idx_u
    if u['status'] == 'Vacant':
        vals = [u['id'], "", u['rent'], due_date, "", 0, u['rent'], "Unpaid",
                5, 0, u['rent'], "", "", "Unit Vacant"]
    else:
        ti = u['tenant_idx']
        if ti in unpaid_tenant_indices:
            # Unpaid — no payment by Apr 5
            late_fee = min(round(u['rent'] * 0.05, 2), 500)
            vals = [u['id'], u['tenant'], u['rent'], due_date, "", 0,
                    u['rent'], "Unpaid", 5, late_fee, u['rent'] + late_fee,
                    u['payment_method'], "", "Follow up immediately"]
        elif ti in partial_tenant_indices:
            # Partial — paid some
            pct = [0.6, 0.5, 0.7, 0.65][partial_tenant_indices.index(ti)]
            amt_paid = round(u['rent'] * pct, 2)
            outstanding = round(u['rent'] - amt_paid, 2)
            late_fee = min(round(outstanding * 0.05, 2), 500)
            pay_date = date(2025, 4, 3)
            vals = [u['id'], u['tenant'], u['rent'], due_date, pay_date,
                    amt_paid, outstanding, "Partial", 2, late_fee,
                    outstanding + late_fee, u['payment_method'],
                    f"RCP-{receipt_counter}", "Partial payment received"]
            receipt_counter += 1
        else:
            # Paid in full
            day = min((ti % 4) + 1, 4)
            pay_date = date(2025, 4, day)
            vals = [u['id'], u['tenant'], u['rent'], due_date, pay_date,
                    u['rent'], 0, "Paid", 0, 0, u['rent'],
                    u['payment_method'], f"RCP-{receipt_counter}", ""]
            receipt_counter += 1

    for i, v in enumerate(vals):
        cell = ws_rent.cell(row=r, column=2 + i, value=v)
        if i in (2, 5, 6, 9, 10):  # monetary
            cell.number_format = FMT_AED
            cell.alignment = align_number()
        elif i in (3, 4):  # dates
            cell.number_format = FMT_DATE
            cell.alignment = align_date()
        elif i == 8:  # days late
            cell.number_format = FMT_INT
            cell.alignment = align_number()
        else:
            cell.alignment = align_text()

    style_data_row(ws_rent, r, 2, rent_last_col, idx_u)

total_r_rent = 5 + 55

# Now replace Outstanding, Payment Status, Days Late, Late Fee, Total Due with FORMULAS
# Column map: B=UnitID, C=Tenant, D=Rent, E=DueDate, F=PayDate, G=AmountPaid,
#             H=Outstanding, I=PaymentStatus, J=DaysLate, K=LateFee, L=TotalDue
for idx_u in range(55):
    r = 5 + idx_u
    # Outstanding = Rent (D) - Amount Paid (G)
    ws_rent.cell(row=r, column=8, value=f"=D{r}-G{r}").number_format = FMT_AED
    ws_rent.cell(row=r, column=8).alignment = align_number()
    # Payment Status = IF logic based on Amount Paid vs Rent
    ws_rent.cell(row=r, column=9, value=f'=IF(C{r}="","N/A",IF(G{r}=0,"Unpaid",IF(G{r}<D{r},"Partial","Paid")))').alignment = align_text()
    # Days Late (col J=10) = IF paid, 0, else count
    ws_rent.cell(row=r, column=10, value=f'=IF(I{r}="Paid",0,IF(I{r}="Partial",2,5))').number_format = FMT_INT
    ws_rent.cell(row=r, column=10).alignment = align_number()
    # Late Fee (col K=11) = MIN(5% of Outstanding (H), 500)
    ws_rent.cell(row=r, column=11, value=f'=IFERROR(MIN(H{r}*0.05,500),0)').number_format = FMT_AED
    ws_rent.cell(row=r, column=11).alignment = align_number()
    # Total Due (col L=12) = Outstanding (H) + Late Fee (K)
    ws_rent.cell(row=r, column=12, value=f"=H{r}+K{r}").number_format = FMT_AED
    ws_rent.cell(row=r, column=12).alignment = align_number()

# Summary section
sum_start = total_r_rent + 1
ws_rent.cell(row=sum_start, column=2, value="COLLECTION SUMMARY").font = font_subheader()
ws_rent.merge_cells(start_row=sum_start, start_column=2, end_row=sum_start, end_column=5)

summary_items = [
    ("Total Expected", f"=SUM(D5:D{total_r_rent-1})"),
    ("Total Collected", f"=SUM(G5:G{total_r_rent-1})"),
    ("Total Outstanding", f"=SUM(H5:H{total_r_rent-1})"),
    ("Collection Rate", f"=IFERROR(SUM(G5:G{total_r_rent-1})/SUM(D5:D{total_r_rent-1}),0)"),
    ("# Paid", f'=COUNTIF(I5:I{total_r_rent-1},"Paid")'),
    ("# Partial", f'=COUNTIF(I5:I{total_r_rent-1},"Partial")'),
    ("# Unpaid", f'=COUNTIF(I5:I{total_r_rent-1},"Unpaid")'),
]
for i, (label, formula) in enumerate(summary_items):
    r = sum_start + 1 + i
    ws_rent.cell(row=r, column=2, value=label).font = font_body()
    ws_rent.cell(row=r, column=2).alignment = align_text()
    ws_rent.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    cell_v = ws_rent.cell(row=r, column=4, value=formula)
    cell_v.font = font_subheader()
    cell_v.alignment = align_number()
    if label == "Collection Rate":
        cell_v.number_format = FMT_PCT
    elif label.startswith("#"):
        cell_v.number_format = FMT_INT
    else:
        cell_v.number_format = FMT_AED

# Conditional formatting: Paid=green, Partial=amber, Unpaid=red
ws_rent.conditional_formatting.add(
    f"I5:I{total_r_rent-1}",
    CellIsRule(operator='equal', formula=['"Paid"'], fill=PAL_CF_POSITIVE_FILL, font=PAL_CF_POSITIVE_FONT)
)
ws_rent.conditional_formatting.add(
    f"I5:I{total_r_rent-1}",
    CellIsRule(operator='equal', formula=['"Partial"'], fill=PAL_CF_WARNING_FILL, font=PAL_CF_WARNING_FONT)
)
ws_rent.conditional_formatting.add(
    f"I5:I{total_r_rent-1}",
    CellIsRule(operator='equal', formula=['"Unpaid"'], fill=PAL_CF_NEGATIVE_FILL, font=PAL_CF_NEGATIVE_FONT)
)

col_widths_rent = [3, 10, 20, 14, 14, 14, 14, 14, 14, 12, 12, 14, 14, 12, 22]
for i, w in enumerate(col_widths_rent):
    ws_rent.column_dimensions[get_column_letter(i + 1)].width = w

ws_rent.freeze_panes = "C5"
set_print(ws_rent, f"A1:{get_column_letter(rent_last_col)}{sum_start+8}", "2:4")


# ============================================================
# §6  SHEET 5: Payment Alerts
# ============================================================
ws_alert = wb.create_sheet("Payment Alerts")
set_tab_color(ws_alert)

alert_headers = ["Unit ID", "Tenant Name", "Phone", "Monthly Rent", "Due Date",
                 "Days Late", "Late Fee", "Total Due", "SMS Sent?", "Follow-up Date", "Notes"]
alert_last_col = 1 + len(alert_headers)

setup_sheet(ws_alert, "Payment Alerts — Tenants Who Haven't Paid by the 5th", alert_last_col)

# Instruction row
ws_alert.merge_cells(start_row=2, start_column=2, end_row=2, end_column=alert_last_col)
instr = ws_alert.cell(row=2, column=2, value="⚠ Contact all tenants listed here immediately. Send SMS/WhatsApp reminder.")
instr.font = Font(name="Calibri", size=11, bold=True, color=PAL_ACCENT_NEGATIVE)
instr.alignment = align_text()

# Re-setup title at row 3
ws_alert.row_dimensions[3].height = ROW_HEIGHTS["title"]
ws_alert.merge_cells(start_row=3, start_column=2, end_row=3, end_column=alert_last_col)
t_cell = ws_alert.cell(row=3, column=2, value="Payment Alerts — Tenants Who Haven't Paid by the 5th")
t_cell.font = font_title()
t_cell.alignment = align_title()

# Header at row 5
ws_alert.row_dimensions[4].height = ROW_HEIGHTS["spacer"]
for i, h in enumerate(alert_headers):
    ws_alert.cell(row=5, column=2 + i, value=h)
style_header_row(ws_alert, 5, 2, alert_last_col)

# Build alert data: 5 unpaid + 4 partial
alert_tenants = []
for idx_u, u in enumerate(units):
    if u['status'] == 'Vacant':
        continue
    ti = u['tenant_idx']
    if ti in unpaid_tenant_indices or ti in partial_tenant_indices:
        if ti in unpaid_tenant_indices:
            days_late = 5
            amt_outstanding = u['rent']
            late_fee = min(round(amt_outstanding * 0.05, 2), 500)
            sms = "No" if ti in [2, 25, 44] else "Yes"
            notes = "No response to calls"
            if ti == 12:
                notes = "Promised to pay by 10th"
            elif ti == 37:
                notes = "Requesting payment plan"
        else:
            days_late = 2
            pct = [0.6, 0.5, 0.7, 0.65][partial_tenant_indices.index(ti)]
            amt_outstanding = round(u['rent'] * (1 - pct), 2)
            late_fee = min(round(amt_outstanding * 0.05, 2), 500)
            sms = "No" if ti in [7, 33] else "Yes"
            notes = "Partial payment received"

        alert_tenants.append({
            'unit_id': u['id'], 'tenant': u['tenant'], 'phone': u['phone'],
            'rent': u['rent'], 'due_date': due_date, 'days_late': days_late,
            'late_fee': late_fee, 'total_due': amt_outstanding + late_fee,
            'sms': sms, 'follow_up': date(2025, 4, 7), 'notes': notes
        })

for idx_a, at in enumerate(alert_tenants):
    r = 6 + idx_a
    vals = [
        at['unit_id'], at['tenant'], at['phone'], at['rent'], at['due_date'],
        at['days_late'], at['late_fee'], at['total_due'], at['sms'],
        at['follow_up'], at['notes']
    ]
    for i, v in enumerate(vals):
        cell = ws_alert.cell(row=r, column=2 + i, value=v)
        if i in (3, 6, 7):  # monetary
            cell.number_format = FMT_AED
            cell.alignment = align_number()
        elif i in (4, 9):  # dates
            cell.number_format = FMT_DATE
            cell.alignment = align_date()
        elif i == 5:  # days late
            cell.number_format = FMT_INT
            cell.alignment = align_number()
        else:
            cell.alignment = align_text()
    style_data_row(ws_alert, r, 2, alert_last_col, idx_a)

# Conditional formatting on Days Late
last_alert_r = 6 + len(alert_tenants) - 1
ws_alert.conditional_formatting.add(
    f"G6:G{last_alert_r}",
    CellIsRule(operator='greaterThan', formula=['30'],
              fill=PAL_CF_NEGATIVE_FILL, font=PAL_CF_NEGATIVE_FONT)
)
ws_alert.conditional_formatting.add(
    f"G6:G{last_alert_r}",
    CellIsRule(operator='greaterThan', formula=['15'],
              fill=PAL_CF_WARNING_FILL, font=PAL_CF_WARNING_FONT)
)

col_widths_alert = [3, 10, 20, 16, 14, 14, 12, 12, 14, 12, 16, 24]
for i, w in enumerate(col_widths_alert):
    ws_alert.column_dimensions[get_column_letter(i + 1)].width = w

ws_alert.freeze_panes = "C6"
set_print(ws_alert, f"A1:{get_column_letter(alert_last_col)}{last_alert_r+1}", "3:5")


# ============================================================
# §7  SHEET 6: Maintenance
# ============================================================
ws_maint = wb.create_sheet("Maintenance")
set_tab_color(ws_maint)

maint_headers = ["Request #", "Date Reported", "Unit/Building", "Issue Description",
                 "Priority", "Status", "Assigned To", "Date Completed",
                 "Estimated Cost", "Actual Cost", "Variance", "Notes"]
maint_last_col = 1 + len(maint_headers)

setup_sheet(ws_maint, "Maintenance Log & Cost Tracker", maint_last_col)

for i, h in enumerate(maint_headers):
    ws_maint.cell(row=4, column=2 + i, value=h)
style_header_row(ws_maint, 4, 2, maint_last_col)

maint_data = [
    ["MR-001", date(2025,3,28), "B-108", "AC not cooling", "Urgent", "Completed", "Al Amal AC Services", date(2025,3,30), 350, 380, None, "Compressor issue, extra refrigerant"],
    ["MR-002", date(2025,3,29), "A-103", "Water leak in bathroom", "High", "In Progress", "Al Jazira Plumbing", None, 200, None, None, "Leak from upper floor"],
    ["MR-003", date(2025,3,30), "C-112", "Door lock broken", "Medium", "Open", "Secure Locksmith", None, 150, None, None, "Tenant locked out twice"],
    ["MR-004", date(2025,4,1), "A-110", "Paint touch-up needed", "Low", "Completed", "Emperor Painters", date(2025,4,3), 100, 100, None, "Hallway wall scuffs"],
    ["MR-005", date(2025,4,1), "D-103", "Kitchen pipe blocked", "High", "Completed", "Quick Fix Plumbing", date(2025,4,2), 250, 270, None, "Grease buildup in pipe"],
    ["MR-006", date(2025,4,2), "B-105", "AC filter replacement", "Medium", "Completed", "Cool Air Maintenance", date(2025,4,3), 120, 120, None, "Routine filter change"],
    ["MR-007", date(2025,4,2), "C-109", "Electrical socket sparking", "Urgent", "Completed", "Safe Wire Electric", date(2025,4,2), 180, 200, None, "Replaced socket and wiring"],
    ["MR-008", date(2025,4,3), "A-106", "Bathroom faucet leak", "Medium", "In Progress", "Al Jazira Plumbing", None, 90, None, None, "Washer replacement needed"],
    ["MR-009", date(2025,4,3), "D-108", "Window glass crack", "High", "Open", "Crystal Glass Works", None, 300, None, None, "Ordered glass panel"],
    ["MR-010", date(2025,4,3), "Building B", "Elevator maintenance", "Medium", "Completed", "Otis Elevators", date(2025,4,4), 2500, 2500, None, "Quarterly service"],
    ["MR-011", date(2025,4,4), "Building A", "Common area lighting", "Low", "In Progress", "Safe Wire Electric", None, 400, None, None, "3rd floor corridor lights"],
    ["MR-012", date(2025,4,4), "Building C", "Pest control treatment", "Medium", "Completed", "Emperor Pest Control", date(2025,4,5), 800, 800, None, "Monthly treatment all floors"],
    ["MR-013", date(2025,4,4), "Building D", "Roof waterproofing", "High", "In Progress", "Al Shamil Waterproofing", None, 3500, None, None, "Monsoon preparation"],
    ["MR-014", date(2025,4,5), "Building A", "Intercom system repair", "Medium", "Open", "Secure Comm Tech", None, 600, None, None, "3rd-4th floor intercom down"],
    ["MR-015", date(2025,4,5), "All Buildings", "Parking area cleaning", "Low", "Completed", "Emperor Cleaning", date(2025,4,5), 500, 500, None, "Monthly deep clean"],
]

for idx_m, md in enumerate(maint_data):
    r = 5 + idx_m
    for i, v in enumerate(md):
        cell = ws_maint.cell(row=r, column=2 + i, value=v)
        if i in (8, 9):  # costs
            if v is not None:
                cell.number_format = FMT_AED
            cell.alignment = align_number()
        elif i in (1, 7):  # dates
            if v is not None:
                cell.number_format = FMT_DATE
            cell.alignment = align_date()
        else:
            cell.alignment = align_text()

    # Variance formula (col L = 13)
    ws_maint.cell(row=r, column=13, value=f"=IFERROR(K{r}-J{r},0)").number_format = FMT_AED
    ws_maint.cell(row=r, column=13).alignment = align_number()

    style_data_row(ws_maint, r, 2, maint_last_col, idx_m)

# Totals row
total_r_m = 5 + len(maint_data)
ws_maint.cell(row=total_r_m, column=2, value="TOTAL").font = font_subheader()
ws_maint.cell(row=total_r_m, column=10, value=f"=SUM(J5:J{total_r_m-1})").number_format = FMT_AED
ws_maint.cell(row=total_r_m, column=10).alignment = align_number()
ws_maint.cell(row=total_r_m, column=11, value=f"=SUM(K5:K{total_r_m-1})").number_format = FMT_AED
ws_maint.cell(row=total_r_m, column=11).alignment = align_number()
ws_maint.cell(row=total_r_m, column=12, value=f"=SUM(L5:L{total_r_m-1})").number_format = FMT_AED
ws_maint.cell(row=total_r_m, column=12).alignment = align_number()
style_total_row(ws_maint, total_r_m, 2, maint_last_col)

# Conditional formatting on Priority (col F=6) and Status (col G=7)
ws_maint.conditional_formatting.add(
    f"F5:F{total_r_m-1}",
    CellIsRule(operator='equal', formula=['"Urgent"'], fill=PAL_CF_NEGATIVE_FILL, font=PAL_CF_NEGATIVE_FONT)
)
ws_maint.conditional_formatting.add(
    f"F5:F{total_r_m-1}",
    CellIsRule(operator='equal', formula=['"High"'], fill=PAL_CF_WARNING_FILL, font=PAL_CF_WARNING_FONT)
)
ws_maint.conditional_formatting.add(
    f"G5:G{total_r_m-1}",
    CellIsRule(operator='equal', formula=['"Open"'], fill=PAL_CF_NEGATIVE_FILL, font=PAL_CF_NEGATIVE_FONT)
)
ws_maint.conditional_formatting.add(
    f"G5:G{total_r_m-1}",
    CellIsRule(operator='equal', formula=['"In Progress"'], fill=PAL_CF_WARNING_FILL, font=PAL_CF_WARNING_FONT)
)
ws_maint.conditional_formatting.add(
    f"G5:G{total_r_m-1}",
    CellIsRule(operator='equal', formula=['"Completed"'], fill=PAL_CF_POSITIVE_FILL, font=PAL_CF_POSITIVE_FONT)
)

col_widths_m = [3, 10, 14, 16, 26, 12, 14, 22, 14, 14, 14, 14, 28]
for i, w in enumerate(col_widths_m):
    ws_maint.column_dimensions[get_column_letter(i + 1)].width = w

ws_maint.freeze_panes = "C5"
set_print(ws_maint, f"A1:{get_column_letter(maint_last_col)}{total_r_m}", "2:4")


# ============================================================
# §8  SHEET 7: Expenses
# ============================================================
ws_exp = wb.create_sheet("Expenses")
set_tab_color(ws_exp)

exp_headers = ["#", "Date", "Category", "Description", "Vendor/Payee", "Amount", "Payment Method", "Reference", "Notes"]
exp_last_col = 1 + len(exp_headers)

setup_sheet(ws_exp, "Operating Expenses — April 2025", exp_last_col)

for i, h in enumerate(exp_headers):
    ws_exp.cell(row=4, column=2 + i, value=h)
style_header_row(ws_exp, 4, 2, exp_last_col)

expenses_data = [
    # Manpower/Staff
    [1, date(2025,4,1), "Manpower", "Security Guard — Building A & B", "Al Ameen Security", 3500, "Bank Transfer", "PAY-0401", "Monthly salary"],
    [2, date(2025,4,1), "Manpower", "Security Guard — Building C & D", "Al Ameen Security", 3500, "Bank Transfer", "PAY-0402", "Monthly salary"],
    [3, date(2025,4,1), "Manpower", "Security Guard — Night Shift", "Al Ameen Security", 3500, "Bank Transfer", "PAY-0403", "Monthly salary"],
    [4, date(2025,4,1), "Manpower", "Cleaner — Building A & B", "Emperor Services", 2500, "Bank Transfer", "PAY-0404", "Monthly salary"],
    [5, date(2025,4,1), "Manpower", "Cleaner — Building C & D", "Emperor Services", 2500, "Bank Transfer", "PAY-0405", "Monthly salary"],
    [6, date(2025,4,1), "Manpower", "Building Supervisor", "Ahmad Al Rashid", 4500, "Bank Transfer", "PAY-0406", "Monthly salary"],
    [7, date(2025,4,1), "Manpower", "Receptionist/Admin", "Fatima Al Suwaidi", 3000, "Bank Transfer", "PAY-0407", "Monthly salary"],
    # Municipality Fees
    [8, date(2025,4,2), "Municipality Fees", "Ajman Municipality — Building A", "Ajman Municipality", 2250, "Bank Transfer", "MUN-0401", "5% of rental income"],
    [9, date(2025,4,2), "Municipality Fees", "Ajman Municipality — Building B", "Ajman Municipality", 2100, "Bank Transfer", "MUN-0402", "5% of rental income"],
    [10, date(2025,4,2), "Municipality Fees", "Ajman Municipality — Building C", "Ajman Municipality", 2350, "Bank Transfer", "MUN-0403", "5% of rental income"],
    [11, date(2025,4,2), "Municipality Fees", "Ajman Municipality — Building D", "Ajman Municipality", 1800, "Bank Transfer", "MUN-0404", "5% of rental income"],
    # Maintenance
    [12, date(2025,4,2), "Maintenance", "AC repair — Unit B-108", "Al Amal AC Services", 380, "Cash", "INV-M001", "MR-001 completed"],
    [13, date(2025,4,3), "Maintenance", "Water leak repair — Unit A-103", "Al Jazira Plumbing", 270, "Cash", "INV-M005", "MR-005 completed"],
    [14, date(2025,4,3), "Maintenance", "Electrical socket — Unit C-109", "Safe Wire Electric", 200, "Cash", "INV-M007", "MR-007 completed"],
    [15, date(2025,4,4), "Maintenance", "AC filter replacement — Unit B-105", "Cool Air Maintenance", 120, "Cash", "INV-M006", "MR-006 completed"],
    [16, date(2025,4,4), "Maintenance", "Paint touch-up — Unit A-110", "Emperor Painters", 100, "Cash", "INV-M004", "MR-004 completed"],
    [17, date(2025,4,5), "Maintenance", "Pest control — Building C", "Emperor Pest Control", 800, "Cheque", "INV-M012", "MR-012 completed"],
    [18, date(2025,4,5), "Maintenance", "Elevator service — Building B", "Otis Elevators", 2500, "Cheque", "INV-M010", "MR-010 quarterly"],
    [19, date(2025,4,5), "Maintenance", "Parking cleaning — All Buildings", "Emperor Cleaning", 500, "Cash", "INV-M015", "MR-015 monthly"],
    [20, date(2025,4,5), "Maintenance", "Kitchen pipe — Unit D-103", "Quick Fix Plumbing", 270, "Cash", "INV-M005B", "Additional work"],
    # Leasing Commission
    [21, date(2025,4,1), "Leasing Commission", "New tenant placement — 3 units", "Emperor Leasing", 5200, "Cheque", "COM-0401", "Commission for March placements"],
    # Insurance
    [22, date(2025,4,1), "Insurance", "Building insurance — All Buildings", "Oman Insurance Co.", 4200, "Bank Transfer", "INS-0401", "Quarterly premium"],
    # Utilities
    [23, date(2025,4,5), "Utilities", "DEWA — Common areas Building A", "DEWA", 1800, "Bank Transfer", "DEWA-A04", "Electricity & water"],
    [24, date(2025,4,5), "Utilities", "DEWA — Common areas Building B", "DEWA", 1700, "Bank Transfer", "DEWA-B04", "Electricity & water"],
    [25, date(2025,4,5), "Utilities", "DEWA — Common areas Building C", "DEWA", 1900, "Bank Transfer", "DEWA-C04", "Electricity & water"],
    [26, date(2025,4,5), "Utilities", "DEWA — Common areas Building D", "DEWA", 1400, "Bank Transfer", "DEWA-D04", "Electricity & water"],
    # Marketing
    [27, date(2025,4,3), "Marketing", "Property listing — Bayut & Dubizzle", "Bayut/Dubizzle", 1000, "Credit Card", "MKT-0401", "Premium listing 30 days"],
    [28, date(2025,4,3), "Marketing", "Social media ads", "Meta Ads", 500, "Credit Card", "MKT-0402", "Instagram/Facebook campaigns"],
    # Legal
    [29, date(2025,4,2), "Legal", "Contract preparation — 4 new tenants", "Al Suwaidi Advocates", 2000, "Cheque", "LEG-0401", "Ejari registration included"],
    # Cleaning
    [30, date(2025,4,1), "Cleaning", "Deep cleaning — Vacant units", "Emperor Cleaning", 1200, "Cash", "CLN-0401", "7 vacant units prep"],
    [31, date(2025,4,1), "Cleaning", "Monthly deep cleaning — Common areas", "Emperor Cleaning", 2000, "Cheque", "CLN-0402", "All 4 buildings"],
    # Security
    [32, date(2025,4,1), "Security", "CCTV monitoring service", "Secure Vision LLC", 1800, "Bank Transfer", "SEC-0401", "Monthly monitoring fee"],
    # Pest Control
    [33, date(2025,4,4), "Pest Control", "Monthly treatment — All Buildings", "Emperor Pest Control", 1200, "Cheque", "PST-0401", "Preventive treatment"],
    # Elevator Service
    [34, date(2025,4,1), "Elevator Service", "Monthly maintenance — All elevators", "Otis Elevators", 2800, "Bank Transfer", "ELV-0401", "4 elevators across buildings"],
    # Other
    [35, date(2025,4,5), "Other", "Office supplies & stationery", "Al Gulf Stationery", 200, "Cash", "OTH-0401", "Printer paper, ink, folders"],
    [36, date(2025,4,5), "Other", "Internet & phone — Office", "Etisalat", 300, "Bank Transfer", "OTH-0402", "Monthly telecom bill"],
]

for idx_e, ed in enumerate(expenses_data):
    r = 5 + idx_e
    for i, v in enumerate(ed):
        cell = ws_exp.cell(row=r, column=2 + i, value=v)
        if i == 5:  # amount
            cell.number_format = FMT_AED
            cell.alignment = align_number()
        elif i == 1:  # date
            cell.number_format = FMT_DATE
            cell.alignment = align_date()
        else:
            cell.alignment = align_text()
    style_data_row(ws_exp, r, 2, exp_last_col, idx_e)

# Totals row
total_r_e = 5 + len(expenses_data)
ws_exp.cell(row=total_r_e, column=2, value="GRAND TOTAL").font = font_subheader()
ws_exp.cell(row=total_r_e, column=2).alignment = align_text()
ws_exp.cell(row=total_r_e, column=7, value=f"=SUM(G5:G{total_r_e-1})").number_format = FMT_AED
ws_exp.cell(row=total_r_e, column=7).alignment = align_number()
style_total_row(ws_exp, total_r_e, 2, exp_last_col)

# Summary section by category
sum_start_e = total_r_e + 2
ws_exp.cell(row=sum_start_e, column=2, value="SUMMARY BY CATEGORY").font = font_subheader()
ws_exp.merge_cells(start_row=sum_start_e, start_column=2, end_row=sum_start_e, end_column=5)

categories = ["Manpower", "Municipality Fees", "Maintenance", "Leasing Commission",
              "Insurance", "Utilities", "Marketing", "Legal", "Cleaning",
              "Security", "Pest Control", "Elevator Service", "Other"]
for i, cat in enumerate(categories):
    r = sum_start_e + 1 + i
    ws_exp.cell(row=r, column=2, value=cat).font = font_body()
    ws_exp.cell(row=r, column=2).alignment = align_text()
    ws_exp.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    formula = f'=SUMPRODUCT((D5:D{total_r_e-1}="{cat}")*G5:G{total_r_e-1})'
    cell_v = ws_exp.cell(row=r, column=4, value=formula)
    cell_v.number_format = FMT_AED
    cell_v.font = font_subheader()
    cell_v.alignment = align_number()

# Conditional formatting on amounts
ws_exp.conditional_formatting.add(
    f"G5:G{total_r_e-1}",
    CellIsRule(operator='greaterThan', formula=['3000'],
              fill=PAL_CF_WARNING_FILL, font=PAL_CF_WARNING_FONT)
)

col_widths_exp = [3, 6, 14, 18, 30, 22, 14, 16, 14, 24]
for i, w in enumerate(col_widths_exp):
    ws_exp.column_dimensions[get_column_letter(i + 1)].width = w

ws_exp.freeze_panes = "C5"
set_print(ws_exp, f"A1:{get_column_letter(exp_last_col)}{sum_start_e+len(categories)+1}", "2:4")


# ============================================================
# §9  SHEET 8: Revenue Analysis
# ============================================================
ws_rev = wb.create_sheet("Revenue Analysis")
set_tab_color(ws_rev)

rev_headers = ["Category", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "YTD"]
rev_last_col = 1 + len(rev_headers)

setup_sheet(ws_rev, "Revenue Analysis — 2025", rev_last_col)

for i, h in enumerate(rev_headers):
    ws_rev.cell(row=4, column=2 + i, value=h)
style_header_row(ws_rev, 4, 2, rev_last_col)

# Revenue data rows
rev_rows = [
    ["Rental Income", 172500, 175800, 181200, 185340, 0, 0, 0, 0, 0, 0, 0, 0],
    ["Late Fee Income", 2800, 3200, 3500, 3200, 0, 0, 0, 0, 0, 0, 0, 0],
    ["Municipality Fees Collected", 8625, 8790, 9060, 9267, 0, 0, 0, 0, 0, 0, 0, 0],
    ["Other Income", 500, 500, 500, 500, 0, 0, 0, 0, 0, 0, 0, 0],
    ["Gross Revenue", None, None, None, None, None, None, None, None, None, None, None, None],
    ["", None, None, None, None, None, None, None, None, None, None, None, None],
    ["Less: Vacancy Loss", -22000, -18500, -16800, -15960, 0, 0, 0, 0, 0, 0, 0, 0],
    ["Less: Bad Debt", -5000, -3500, -4500, -7000, 0, 0, 0, 0, 0, 0, 0, 0],
    ["Net Revenue", None, None, None, None, None, None, None, None, None, None, None, None],
]

for idx_rv, rv in enumerate(rev_rows):
    r = 5 + idx_rv
    ws_rev.cell(row=r, column=2, value=rv[0]).alignment = align_text()
    for i in range(12):
        cell = ws_rev.cell(row=r, column=3 + i, value=rv[1 + i])
        cell.number_format = FMT_AED
        cell.alignment = align_number()

    # YTD formula (col 15 = O)
    if rv[0] in ("Gross Revenue", "Net Revenue"):
        ws_rev.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").number_format = FMT_AED
    elif rv[0] and rv[0] != "":
        ws_rev.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").number_format = FMT_AED
    ws_rev.cell(row=r, column=15).alignment = align_number()

    style_data_row(ws_rev, r, 2, rev_last_col, idx_rv)

# Gross Revenue formula row (row 9)
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_rev.cell(row=9, column=col, value=f"=SUM({cl}5:{cl}8)").number_format = FMT_AED

# Net Revenue formula row (row 13)
for i in range(12):
    col = 3 + i
    cl = get_column_letter(col)
    ws_rev.cell(row=13, column=col, value=f"={cl}9+{cl}11+{cl}12").number_format = FMT_AED

# Make Gross Revenue and Net Revenue rows bold
for r_bold in [9, 13]:
    for col in range(2, rev_last_col + 1):
        ws_rev.cell(row=r_bold, column=col).font = font_subheader()

# Conditional formatting: negative red, positive green
for rr in range(5, 14):
    ws_rev.conditional_formatting.add(
        f"C{rr}:O{rr}",
        CellIsRule(operator='lessThan', formula=['0'],
                  fill=PAL_CF_NEGATIVE_FILL, font=PAL_CF_NEGATIVE_FONT)
    )
    ws_rev.conditional_formatting.add(
        f"C{rr}:O{rr}",
        CellIsRule(operator='greaterThan', formula=['0'],
                  fill=PAL_CF_POSITIVE_FILL, font=PAL_CF_POSITIVE_FONT)
    )

# Revenue vs Vacancy Loss bar chart
# Write helper data for chart with proper series labels
chart_helper_row = 30
ws_rev.cell(row=chart_helper_row, column=2, value="Month").font = font_caption()
ws_rev.cell(row=chart_helper_row + 1, column=2, value="Rental Income").font = font_caption()
ws_rev.cell(row=chart_helper_row + 2, column=2, value="Vacancy Loss").font = font_caption()

for i, m in enumerate(["Jan", "Feb", "Mar", "Apr"]):
    ws_rev.cell(row=chart_helper_row, column=3 + i, value=m).font = font_caption()
    ws_rev.cell(row=chart_helper_row + 1, column=3 + i, value=[172500, 175800, 181200, 185340][i])
    ws_rev.cell(row=chart_helper_row + 2, column=3 + i, value=[-22000, -18500, -16800, -15960][i])

chart_rev = create_bar_chart(width=20, height=11)
data_rev_rent = Reference(ws_rev, min_col=3, max_col=6, min_row=chart_helper_row + 1, max_row=chart_helper_row + 1)
data_rev_vl = Reference(ws_rev, min_col=3, max_col=6, min_row=chart_helper_row + 2, max_row=chart_helper_row + 2)
cats_rev = Reference(ws_rev, min_col=3, max_col=6, min_row=chart_helper_row)
chart_rev.add_data(data_rev_rent, from_rows=True, titles_from_data=True)
chart_rev.add_data(data_rev_vl, from_rows=True, titles_from_data=True)
chart_rev.set_categories(cats_rev)
setup_chart_titles(chart_rev, title="Revenue vs Vacancy Loss", y_title="AED")
apply_chart_colors(chart_rev, [PAL_PRIMARY, PAL_ACCENT_NEGATIVE])
chart_rev.legend.position = 'b'
ws_rev.add_chart(chart_rev, "B16")

col_widths_rev = [3, 26, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 16]
for i, w in enumerate(col_widths_rev):
    ws_rev.column_dimensions[get_column_letter(i + 1)].width = w

ws_rev.freeze_panes = "C5"
set_print(ws_rev, f"A1:{get_column_letter(rev_last_col)}30", "2:4")


# ============================================================
# §10  SHEET 9: Profit & Loss
# ============================================================
ws_pnl = wb.create_sheet("Profit & Loss")
set_tab_color(ws_pnl)

pnl_last_col = 4  # B=Category, C=Sub-category, D=Amount, E=Margin

setup_sheet(ws_pnl, "Profit & Loss Statement — Real Estate Emperor — April 2025", pnl_last_col)

# Custom layout for P&L
ws_pnl.cell(row=4, column=2, value="Category").font = font_header()
ws_pnl.cell(row=4, column=2).fill = fill_header()
ws_pnl.cell(row=4, column=2).alignment = align_header()
ws_pnl.cell(row=4, column=2).border = border_header()

ws_pnl.cell(row=4, column=3, value="Sub-Category").font = font_header()
ws_pnl.cell(row=4, column=3).fill = fill_header()
ws_pnl.cell(row=4, column=3).alignment = align_header()
ws_pnl.cell(row=4, column=3).border = border_header()

ws_pnl.cell(row=4, column=4, value="Amount (AED)").font = font_header()
ws_pnl.cell(row=4, column=4).fill = fill_header()
ws_pnl.cell(row=4, column=4).alignment = align_header()
ws_pnl.cell(row=4, column=4).border = border_header()

ws_pnl.cell(row=4, column=5, value="Margin").font = font_header()
ws_pnl.cell(row=4, column=5).fill = fill_header()
ws_pnl.cell(row=4, column=5).alignment = align_header()
ws_pnl.cell(row=4, column=5).border = border_header()

ws_pnl.row_dimensions[4].height = ROW_HEIGHTS["header"]

# P&L data
pnl_data = [
    # (category, sub, amount, margin_formula_or_none, is_header, is_total)
    ("REVENUE", None, None, None, True, False),
    (None, "Rental Income", 185340, None, False, False),
    (None, "Late Fee Income", 3200, None, False, False),
    (None, "Other Income", 500, None, False, False),
    ("Total Revenue", None, None, None, False, True),  # row 9
    ("", None, None, None, False, False),
    ("COST OF OPERATIONS", None, None, None, True, False),
    (None, "Manpower/Staff", 24000, None, False, False),
    (None, "Maintenance", 10500, None, False, False),
    (None, "Municipality Fees", 8500, None, False, False),
    (None, "Leasing Commissions", 5200, None, False, False),
    (None, "Insurance", 4200, None, False, False),
    (None, "Utilities (DEWA)", 6800, None, False, False),
    ("Total COO", None, None, None, False, True),  # row 19
    ("", None, None, None, False, False),
    ("GROSS PROFIT", None, None, None, True, False),  # row 21
    (None, "Gross Profit", None, None, False, True),
    (None, "Gross Margin", None, None, False, True),
    ("", None, None, None, False, False),
    ("OPERATING EXPENSES", None, None, None, True, False),
    (None, "Marketing", 1500, None, False, False),
    (None, "Legal", 2000, None, False, False),
    (None, "Cleaning", 3200, None, False, False),
    (None, "Security (CCTV)", 1800, None, False, False),
    (None, "Pest Control", 1200, None, False, False),
    (None, "Elevator Service", 2800, None, False, False),
    (None, "Other", 500, None, False, False),
    ("Total OpEx", None, None, None, False, True),  # row 33
    ("", None, None, None, False, False),
    ("NET PROFIT", None, None, None, True, False),  # row 35
    (None, "Net Profit", None, None, False, True),
    (None, "Net Margin", None, None, False, True),
]

pnl_row = 5
for item in pnl_data:
    cat, sub, amt, margin_f, is_header, is_total = item
    r = pnl_row

    if is_header:
        ws_pnl.cell(row=r, column=2, value=cat).font = font_subheader()
        ws_pnl.cell(row=r, column=2).alignment = align_text()
        ws_pnl.row_dimensions[r].height = 24
    elif is_total:
        label = cat if cat else sub
        ws_pnl.cell(row=r, column=2, value=label).font = font_subheader()
        ws_pnl.cell(row=r, column=2).alignment = align_text()
        if amt is not None:
            ws_pnl.cell(row=r, column=4, value=amt).number_format = FMT_AED
            ws_pnl.cell(row=r, column=4).alignment = align_number()
        ws_pnl.cell(row=r, column=2).border = border_total()
        ws_pnl.cell(row=r, column=3).border = border_total()
        ws_pnl.cell(row=r, column=4).border = border_total()
        ws_pnl.cell(row=r, column=5).border = border_total()
        ws_pnl.row_dimensions[r].height = ROW_HEIGHTS["total"]
    elif cat == "":
        ws_pnl.row_dimensions[r].height = ROW_HEIGHTS["spacer"]
    else:
        if sub:
            ws_pnl.cell(row=r, column=3, value=sub).font = font_body()
            ws_pnl.cell(row=r, column=3).alignment = align_text()
        if amt is not None:
            ws_pnl.cell(row=r, column=4, value=amt).number_format = FMT_AED
            ws_pnl.cell(row=r, column=4).alignment = align_number()
        ws_pnl.row_dimensions[r].height = ROW_HEIGHTS["data"]

    pnl_row += 1

# Now set formulas for totals
# Total Revenue (row 9) = sum of rows 6+7+8
ws_pnl.cell(row=9, column=4, value="=SUM(D6:D8)").number_format = FMT_AED
ws_pnl.cell(row=9, column=4).alignment = align_number()
ws_pnl.cell(row=9, column=4).font = font_subheader()

# Total COO (row 19) = sum of rows 13:18
ws_pnl.cell(row=19, column=4, value="=SUM(D13:D18)").number_format = FMT_AED
ws_pnl.cell(row=19, column=4).alignment = align_number()
ws_pnl.cell(row=19, column=4).font = font_subheader()

# Gross Profit (row 22) = Total Revenue - Total COO
ws_pnl.cell(row=22, column=4, value="=D9-D19").number_format = FMT_AED
ws_pnl.cell(row=22, column=4).alignment = align_number()
ws_pnl.cell(row=22, column=4).font = font_subheader()
# Gross Margin
ws_pnl.cell(row=23, column=4, value="=IFERROR(D22/D9,0)").number_format = FMT_PCT
ws_pnl.cell(row=23, column=4).alignment = align_number()
ws_pnl.cell(row=23, column=4).font = font_subheader()

# Total OpEx (row 33) = sum of rows 27:32
ws_pnl.cell(row=33, column=4, value="=SUM(D27:D32)").number_format = FMT_AED
ws_pnl.cell(row=33, column=4).alignment = align_number()
ws_pnl.cell(row=33, column=4).font = font_subheader()

# Net Profit (row 36) = Gross Profit - Total OpEx
ws_pnl.cell(row=36, column=4, value="=D22-D33").number_format = FMT_AED
ws_pnl.cell(row=36, column=4).alignment = align_number()
ws_pnl.cell(row=36, column=4).font = font_subheader()
# Net Margin
ws_pnl.cell(row=37, column=4, value="=IFERROR(D36/D9,0)").number_format = FMT_PCT
ws_pnl.cell(row=37, column=4).alignment = align_number()
ws_pnl.cell(row=37, column=4).font = font_subheader()

# Conditional formatting: positive green, negative red
for pnl_r in [6,7,8,9,13,14,15,16,17,18,19,22,23,27,28,29,30,31,32,33,36,37]:
    ws_pnl.conditional_formatting.add(
        f"D{pnl_r}",
        CellIsRule(operator='lessThan', formula=['0'],
                  fill=PAL_CF_NEGATIVE_FILL, font=PAL_CF_NEGATIVE_FONT)
    )

# Column widths
ws_pnl.column_dimensions['A'].width = 3
ws_pnl.column_dimensions['B'].width = 26
ws_pnl.column_dimensions['C'].width = 22
ws_pnl.column_dimensions['D'].width = 18
ws_pnl.column_dimensions['E'].width = 12

set_print(ws_pnl, "A1:E38", "2:4")


# ============================================================
# §11  SHEET 10: Contracts
# ============================================================
ws_con = wb.create_sheet("Contracts")
set_tab_color(ws_con)

con_headers = ["Unit ID", "Tenant Name", "Contract Start", "Contract End",
               "Duration (Months)", "Monthly Rent", "Total Contract Value",
               "Days Remaining", "Status", "Renewal Status", "New Rent", "Notes"]
con_last_col = 1 + len(con_headers)

setup_sheet(ws_con, "Lease Contract Tracker", con_last_col)

for i, h in enumerate(con_headers):
    ws_con.cell(row=4, column=2 + i, value=h)
style_header_row(ws_con, 4, 2, con_last_col)

# Determine renewal status for different contract categories
# expired (5): tenants 0,10,22,35,45
# expiring soon (5): tenants 3,15,27,38,47
# renewal due (3): tenants 5,19,42
# remaining: active (35)

renewal_map = {}
# Expired
for ti in expired_tenant_indices:
    renewal_map[ti] = ("Expired", "Not Renewed" if ti in [0,22] else ("Tenant Leaving" if ti == 45 else ("Renewed" if ti == 10 else "In Negotiation")), None)
# Fix: 2 Renewed, 3 In Negotiation, 2 Not Renewed, 1 Tenant Leaving
renewal_map[0] = ("Expired", "Not Renewed", None)
renewal_map[10] = ("Expired", "Renewed", 3800)
renewal_map[22] = ("Expired", "Not Renewed", None)
renewal_map[35] = ("Expired", "In Negotiation", None)
renewal_map[45] = ("Expired", "Tenant Leaving", None)

# Expiring soon
for ti in expiring_tenant_indices:
    renewal_map[ti] = ("Expiring Soon", "In Negotiation" if ti in [3,38] else "Pending", None)
renewal_map[3] = ("Expiring Soon", "In Negotiation", 3900)
renewal_map[38] = ("Expiring Soon", "In Negotiation", 4200)

# Renewal due
renewal_map[5] = ("Renewal Due", "In Negotiation", 3750)
renewal_map[19] = ("Renewal Due", "Pending", None)
renewal_map[42] = ("Renewal Due", "Pending", None)

c_idx = 0
for idx_u, u in enumerate(units):
    if u['status'] != 'Rented':
        continue
    r = 5 + c_idx
    ti = u['tenant_idx']
    cs = contract_starts[ti]
    ce = contract_ends[ti]
    duration = (ce.year - cs.year) * 12 + ce.month - cs.month
    total_val = u['rent'] * duration

    # Days remaining from April 5, 2025
    ref_date = date(2025, 4, 5)
    if ce <= ref_date:
        days_remaining = 0
        status = "Expired"
    elif (ce - ref_date).days <= 30:
        days_remaining = (ce - ref_date).days
        status = "Expiring Soon"
    elif (ce - ref_date).days <= 90:
        days_remaining = (ce - ref_date).days
        status = "Renewal Due"
    else:
        days_remaining = (ce - ref_date).days
        status = "Active"

    # Override status from renewal_map
    if ti in renewal_map:
        status = renewal_map[ti][0]
        renewal_status = renewal_map[ti][1]
        new_rent = renewal_map[ti][2]
    else:
        renewal_status = ""
        new_rent = None

    # Determine notes
    notes = ""
    if status == "Expired" and renewal_status == "Not Renewed":
        notes = "Tenant vacating, unit available"
    elif status == "Expired" and renewal_status == "Tenant Leaving":
        notes = "Tenant confirmed leaving"
    elif status == "Expired" and renewal_status == "Renewed":
        notes = "New contract signed"
    elif status == "Expiring Soon":
        notes = "Contact tenant for renewal"
    elif status == "Renewal Due":
        notes = "Renewal discussion needed"

    vals = [
        u['id'], u['tenant'], cs, ce, duration, u['rent'], total_val,
        days_remaining, status, renewal_status,
        new_rent if new_rent else "", notes
    ]

    for i, v in enumerate(vals):
        cell = ws_con.cell(row=r, column=2 + i, value=v)
        if i in (5, 6):  # monetary
            cell.number_format = FMT_AED
            cell.alignment = align_number()
        elif i == 10:  # new rent
            if v != "":
                cell.number_format = FMT_AED
                cell.alignment = align_number()
            else:
                cell.alignment = align_number()
        elif i in (2, 3):  # dates
            cell.number_format = FMT_DATE
            cell.alignment = align_date()
        elif i in (4, 7):  # duration, days
            cell.number_format = FMT_INT
            cell.alignment = align_number()
        else:
            cell.alignment = align_text()

    style_data_row(ws_con, r, 2, con_last_col, c_idx)
    c_idx += 1

total_r_c = 5 + 48

# Conditional formatting on Status (col J=10)
ws_con.conditional_formatting.add(
    f"J5:J{total_r_c-1}",
    CellIsRule(operator='equal', formula=['"Active"'], fill=PAL_CF_POSITIVE_FILL, font=PAL_CF_POSITIVE_FONT)
)
ws_con.conditional_formatting.add(
    f"J5:J{total_r_c-1}",
    CellIsRule(operator='equal', formula=['"Expiring Soon"'], fill=PAL_CF_WARNING_FILL, font=PAL_CF_WARNING_FONT)
)
ws_con.conditional_formatting.add(
    f"J5:J{total_r_c-1}",
    CellIsRule(operator='equal', formula=['"Renewal Due"'], fill=PAL_CF_WARNING_FILL, font=PAL_CF_WARNING_FONT)
)
ws_con.conditional_formatting.add(
    f"J5:J{total_r_c-1}",
    CellIsRule(operator='equal', formula=['"Expired"'], fill=PAL_CF_NEGATIVE_FILL, font=PAL_CF_NEGATIVE_FONT)
)

# Summary at bottom
sum_start_c = total_r_c + 1
ws_con.cell(row=sum_start_c, column=2, value="CONTRACT SUMMARY").font = font_subheader()
ws_con.merge_cells(start_row=sum_start_c, start_column=2, end_row=sum_start_c, end_column=4)

contract_summary = [
    ("Active", f'=COUNTIF(J5:J{total_r_c-1},"Active")'),
    ("Expiring Soon", f'=COUNTIF(J5:J{total_r_c-1},"Expiring Soon")'),
    ("Renewal Due", f'=COUNTIF(J5:J{total_r_c-1},"Renewal Due")'),
    ("Expired", f'=COUNTIF(J5:J{total_r_c-1},"Expired")'),
    ("Total Contracts", 48),
]
for i, (label, val) in enumerate(contract_summary):
    r = sum_start_c + 1 + i
    ws_con.cell(row=r, column=2, value=label).font = font_body()
    ws_con.cell(row=r, column=2).alignment = align_text()
    ws_con.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws_con.cell(row=r, column=4, value=val).font = font_subheader()
    ws_con.cell(row=r, column=4).alignment = align_number()
    ws_con.cell(row=r, column=4).number_format = FMT_INT

# Contract Status Distribution Pie Chart
# Helper data for chart
chart_data_start = sum_start_c + 7
ws_con.cell(row=chart_data_start, column=2, value="Status").font = font_caption()
ws_con.cell(row=chart_data_start, column=3, value="Count").font = font_caption()
for i, (label, _) in enumerate(contract_summary[:4]):
    ws_con.cell(row=chart_data_start + 1 + i, column=2, value=label).font = font_body()
    ws_con.cell(row=chart_data_start + 1 + i, column=3, value=label).font = font_body()

# Actually write counts directly for chart
active_count = 35
expiring_count = 5
renewal_due_count = 3
expired_count = 5
for i, (label, cnt) in enumerate([("Active", active_count), ("Expiring Soon", expiring_count), ("Renewal Due", renewal_due_count), ("Expired", expired_count)]):
    ws_con.cell(row=chart_data_start + 1 + i, column=2, value=label)
    ws_con.cell(row=chart_data_start + 1 + i, column=3, value=cnt)

chart_con = create_pie_chart(width=14, height=10)
data_con = Reference(ws_con, min_col=3, min_row=chart_data_start, max_row=chart_data_start + 4)
cats_con = Reference(ws_con, min_col=2, min_row=chart_data_start + 1, max_row=chart_data_start + 4)
chart_con.add_data(data_con, titles_from_data=True)
chart_con.set_categories(cats_con)
setup_chart_titles(chart_con, title="Contract Status Distribution")
apply_pie_colors(chart_con, 4)
chart_con.dataLabels = DataLabelList()
chart_con.dataLabels.showPercent = True
chart_con.dataLabels.showCatName = True
ws_con.add_chart(chart_con, f"G{sum_start_c}")

col_widths_con = [3, 10, 20, 14, 14, 16, 14, 18, 14, 14, 16, 14, 26]
for i, w in enumerate(col_widths_con):
    ws_con.column_dimensions[get_column_letter(i + 1)].width = w

ws_con.freeze_panes = "C5"
set_print(ws_con, f"A1:{get_column_letter(con_last_col)}{sum_start_c+7}", "2:4")


# ============================================================
# §12  Final: Sheet Order, Save
# ============================================================
# Sheet order is already correct (Dashboard first)
# Move sheets to proper order
desired_order = ["Dashboard", "Property Registry", "Tenants", "Rent Collection",
                 "Payment Alerts", "Maintenance", "Expenses", "Revenue Analysis",
                 "Profit & Loss", "Contracts"]
for i, name in enumerate(desired_order):
    idx = wb.sheetnames.index(name)
    wb.move_sheet(name, offset=i - idx)

# Save
OUTPUT = "/home/z/my-project/download/Real_Estate_Emperor_Sample.xlsx"
wb.save(OUTPUT)
print(f"✅ Saved: {OUTPUT}")
print(f"   Sheets: {wb.sheetnames}")
print(f"   Units: {len(units)}, Tenants: {sum(1 for u in units if u['status']=='Rented')}")
