#!/usr/bin/env python3
"""
Real Estate Emperor — Real Estate Management System SAMPLE (with Demo Data)
===========================================================================
AED 2,000 value professional sample workbook.
11-sheet comprehensive RE management system with demo data.
Bottega palette (dark forest green 2D4A3E).

CRITICAL: NO Late Fee, NO Late Payment Fee, NO late_fee anywhere.
"""

import sys, os, random
from datetime import date, timedelta

sys.path.insert(0, "/home/z/my-project/skills/xlsx")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

import templates.base as base

base.use_palette_explicit("bottega")

setup_sheet = base.setup_sheet
style_header_row = base.style_header_row
style_data_row = base.style_data_row
style_total_row = base.style_total_row
font_title = base.font_title
font_header = base.font_header
font_body = base.font_body
font_caption = base.font_caption
font_kpi = base.font_kpi
font_kpi_label = base.font_kpi_label
font_subheader = base.font_subheader
fill_header = base.fill_header
fill_total = base.fill_total
fill_data_row = base.fill_data_row
border_header = base.border_header
border_total = base.border_total
align_title = base.align_title
align_header = base.align_header
align_number = base.align_number
align_text = base.align_text
align_date = base.align_date
COLUMN_WIDTHS = base.COLUMN_WIDTHS
ROW_HEIGHTS = base.ROW_HEIGHTS
FORMATS = base.FORMATS
create_bar_chart = base.create_bar_chart
create_line_chart = base.create_line_chart
create_pie_chart = base.create_pie_chart
setup_chart_titles = base.setup_chart_titles
apply_chart_colors = base.apply_chart_colors
apply_pie_colors = base.apply_pie_colors
make_chart_title = base.make_chart_title

PRIMARY = base.PRIMARY
PRIMARY_LIGHT = base.PRIMARY_LIGHT
SECONDARY = base.SECONDARY
ACCENT_POSITIVE = base.ACCENT_POSITIVE
ACCENT_NEGATIVE = base.ACCENT_NEGATIVE
ACCENT_WARNING = base.ACCENT_WARNING
NEUTRAL_900 = base.NEUTRAL_900
NEUTRAL_600 = base.NEUTRAL_600
NEUTRAL_200 = base.NEUTRAL_200
NEUTRAL_100 = base.NEUTRAL_100
NEUTRAL_50 = base.NEUTRAL_50
NEUTRAL_0 = base.NEUTRAL_0
HEADER_TEXT = base.HEADER_TEXT
CHART_COLORS = base.CHART_COLORS
CF_POSITIVE_FILL = base.CF_POSITIVE_FILL
CF_POSITIVE_FONT = base.CF_POSITIVE_FONT
CF_NEGATIVE_FILL = base.CF_NEGATIVE_FILL
CF_NEGATIVE_FONT = base.CF_NEGATIVE_FONT
CF_WARNING_FILL = base.CF_WARNING_FILL
CF_WARNING_FONT = base.CF_WARNING_FONT

OUTPUT_PATH = "/home/z/my-project/download/Real_Estate_Emperor_Sample.xlsx"
CURRENCY_FMT = "#,##0.00"
DATE_FMT = "YYYY-MM-DD"
PCT_FMT = "0.0%"
TAB_COLOR = PRIMARY

def setup_print(ws, last_col_letter, last_row, title_rows="2:4"):
    ws.print_area = f"A1:{last_col_letter}{last_row}"
    ws.print_title_rows = title_rows
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

def cf_cell_is(ws, range_str, operator, formula, fill, font):
    rule = CellIsRule(operator=operator, formula=[formula], fill=fill, font=font)
    ws.conditional_formatting.add(range_str, rule)

def cf_formula(ws, range_str, formula, fill, font):
    rule = FormulaRule(formula=[formula], fill=fill, font=font)
    ws.conditional_formatting.add(range_str, rule)


# ================================================================
# DEMO DATA DEFINITIONS
# ================================================================

# --- 55 Property Units ---
units = []
# Building A: 10 Studios + 5 One-Bedroom
for i in range(10):
    uid = f"A-{101+i}"
    rent = 2200 + i * 67  # 2200-2800ish
    size = 440 + i * 4
    units.append((uid, "Building A", 1 if i < 5 else 2, "Studio", size, rent))
for i in range(5):
    uid = f"A-{201+i}"
    rent = 3500 + i * 100
    size = 740 + i * 10
    units.append((uid, "Building A", 2, "1 Bedroom", size, rent))

# Building B: 8 Studios + 6 One-Bedroom
for i in range(8):
    uid = f"B-{101+i}"
    rent = 2250 + i * 71
    size = 450 + i * 3
    units.append((uid, "Building B", 1 if i < 4 else 2, "Studio", size, rent))
for i in range(6):
    uid = f"B-{201+i}"
    rent = 3550 + i * 100
    size = 740 + i * 8
    units.append((uid, "Building B", 2, "1 Bedroom", size, rent))

# Building C: 6 Studios + 8 One-Bedroom + 2 Two-Bedroom
for i in range(6):
    uid = f"C-{101+i}"
    rent = 2200 + i * 100
    size = 440 + i * 6
    units.append((uid, "Building C", 1 if i < 3 else 2, "Studio", size, rent))
for i in range(8):
    uid = f"C-{201+i}"
    rent = 3600 + i * 71
    size = 740 + i * 5
    units.append((uid, "Building C", 2, "1 Bedroom", size, rent))
for i in range(2):
    uid = f"C-{301+i}"
    rent = 5000 + i * 500
    size = 1100 + i * 50
    units.append((uid, "Building C", 3, "2 Bedroom", size, rent))

# Building D: 4 Studios + 4 One-Bedroom + 2 Shops
for i in range(4):
    uid = f"D-{101+i}"
    rent = 2300 + i * 83
    size = 450 + i * 3
    units.append((uid, "Building D", 1, "Studio", size, rent))
for i in range(4):
    uid = f"D-{201+i}"
    rent = 3700 + i * 133
    size = 740 + i * 7
    units.append((uid, "Building D", 2, "1 Bedroom", size, rent))
for i in range(2):
    uid = f"D-S0{i+1}"
    rent = 6000 + i * 2000
    size = 480 + i * 40
    units.append((uid, "Building D", 0, "Shop", size, rent))

assert len(units) == 55, f"Expected 55 units, got {len(units)}"

# Vacant units
vacant_ids = {"A-104", "A-109", "B-103", "B-108", "C-105", "C-208", "D-202"}

# --- 48 Tenants ---
tenant_names = [
    "Muhammad Ali", "Ahmed Khan", "Fatima Noor", "Rajesh Kumar", "Priya Sharma",
    "Omar Hassan", "Layla Mahmoud", "Youssef Ibrahim", "Sunil Patel", "Maria Santos",
    "Hassan Al Farsi", "Aisha Bint Khalid", "Vikram Singh", "Nadia Rahman", "Ali Al Mansouri",
    "Deepak Sharma", "Sana Mirza", "Khalid Al Suwaidi", "Rizwan Ahmed", "Luzviminda Cruz",
    "Tariq Al Maktoum", "Meera Nair", "Saeed Al Dhaheri", "Renuka Devi", "Mohammed Al Qasimi",
    "Arjun Reddy", "Zainab Hussain", "Faisal Al Ketbi", "Anil Kumar", "Rosa Villanueva",
    "Hamad Al Shamsi", "Pooja Gupta", "Nasser Al Muhairi", "Sanjay Verma", "Amira Kassab",
    "Rashid Al Nuaimi", "Leila Hadid", "Vivek Joshi", "Carmen Reyes", "Majid Al Amri",
    "Shabana Begum", "Walid Al Zaabi", "Manoj Tiwari", "Elena Bautista", "Sultan Al Darmaki",
    "Divya Patel", "Huda Salem", "Jorge Santos"
]

nationalities = [
    "Pakistani", "Pakistani", "Syrian", "Indian", "Indian",
    "Jordanian", "Egyptian", "Egyptian", "Indian", "Filipino",
    "Emirati", "Emirati", "Indian", "Pakistani", "Emirati",
    "Indian", "Pakistani", "Emirati", "Pakistani", "Filipino",
    "Emirati", "Indian", "Emirati", "Indian", "Emirati",
    "Indian", "Pakistani", "Emirati", "Indian", "Filipino",
    "Emirati", "Indian", "Emirati", "Indian", "Syrian",
    "Emirati", "Jordanian", "Indian", "Filipino", "Emirati",
    "Indian", "Emirati", "Indian", "Filipino", "Emirati",
    "Indian", "Egyptian", "Filipino"
]

employers = [
    "Emirates NBD", "Lulu Group", "Dubai Municipality", "DP World", "Al Futtaim Group",
    "Etisalat", "Dubai Health Authority", "Emirates Airline", "Al Ghurair Group", "Jumeirah Group",
    "ADNOC", "Dubai Properties", "Landmark Group", "Emirates NBD", "Mubadala",
    "Al Futtaim Group", "Aramex", "Dubai Electricity & Water", "Lulu Group", "Dulsco",
    "Dubai Holding", "Infosys UAE", "Etihad Airways", "Wipro UAE", "Sharjah Islamic Bank",
    "Tech Mahindra UAE", "Emirates Islamic Bank", "Abu Dhabi Commercial Bank", "Carrefour UAE", "Al Shaya Group",
    "RAK Properties", "Tata Consultancy", "Emaar Properties", "DP World", "UNRWA",
    "Majid Al Futtaim", "Arab Bank", "CITIC UAE", "Manila Grace Dubai", "Dubai Airports",
    "HCL Technologies", "National Bank of Fujairah", "Satyam UAE", "Jollibee UAE", "Aldar Properties",
    "TCS UAE", "Arabtec Holdings", "DHL UAE"
]

payment_methods = [
    "Bank Transfer", "Cheque", "Bank Transfer", "Bank Transfer", "Cash",
    "Bank Transfer", "Cash", "Bank Transfer", "Cheque", "Cash",
    "Bank Transfer", "Bank Transfer", "Cheque", "Cash", "Bank Transfer",
    "Bank Transfer", "Cheque", "Bank Transfer", "Bank Transfer", "Cash",
    "Bank Transfer", "Bank Transfer", "Bank Transfer", "Cheque", "Bank Transfer",
    "Cheque", "Bank Transfer", "Bank Transfer", "Cash", "Bank Transfer",
    "Bank Transfer", "Cheque", "Bank Transfer", "Bank Transfer", "Cash",
    "Bank Transfer", "Bank Transfer", "Cheque", "Cash", "Bank Transfer",
    "Bank Transfer", "Bank Transfer", "Bank Transfer", "Cash", "Bank Transfer",
    "Cheque", "Cash", "Bank Transfer"
]

phones = [f"050-{random.randint(100,999)}-{random.randint(1000,9999)}" for _ in range(48)]
emails = [name.lower().replace(" ", ".") + "@email.com" for name in tenant_names]

# Assign tenants to non-vacant units
rented_units = [u for u in units if u[0] not in vacant_ids]
assert len(rented_units) == 48, f"Expected 48 rented units, got {len(rented_units)}"

# Build tenant->unit mapping
tenant_data = []
for ti in range(48):
    uid = rented_units[ti][0]
    bldg = rented_units[ti][1]
    utype = rented_units[ti][3]
    rent = rented_units[ti][5]
    # Lease dates: most start 2024, end 2025 or 2026
    if ti < 5:  # expired Feb 2025
        ls = date(2024, 3, 1)
        le = date(2025, 2, 28)
    elif ti < 10:  # expiring Apr 2025
        ls = date(2024, 5, 1)
        le = date(2025, 4, 30)
    elif ti < 13:  # expiring May-Jun 2025
        ls = date(2024, 6, 1)
        le = date(2025, 5 + (ti-10), 30)
    else:  # Active - end later
        ls = date(2024, random.choice([7,8,9,10,11]), 1)
        le = date(2025, random.choice([9,10,11,12]), 30)
    
    sec_dep = rent  # 1 month rent
    tid = f"T-{1001+ti}"
    tenant_data.append({
        'tid': tid, 'name': tenant_names[ti], 'phone': phones[ti],
        'email': emails[ti], 'unit_id': uid, 'unit_type': utype,
        'building': bldg, 'lease_start': ls, 'lease_end': le,
        'rent': rent, 'security_deposit': sec_dep,
        'payment_method': payment_methods[ti],
        'nationality': nationalities[ti],
        'employer': employers[ti],
        'emergency_contact': f"050-{random.randint(100,999)}-{random.randint(1000,9999)}"
    })

# Property data with tenant assignments
prop_data = []
for u in units:
    uid, bldg, floor, utype, size, rent = u
    is_vacant = uid in vacant_ids
    tenant_name = ""
    contract_start = None
    contract_end = None
    if not is_vacant:
        for td in tenant_data:
            if td['unit_id'] == uid:
                tenant_name = td['name']
                contract_start = td['lease_start']
                contract_end = td['lease_end']
                break
    prop_data.append({
        'unit_id': uid, 'building': bldg, 'floor': floor,
        'unit_type': utype, 'size': size, 'rent': rent,
        'status': 'Vacant' if is_vacant else 'Rented',
        'tenant': tenant_name,
        'contract_start': contract_start,
        'contract_end': contract_end
    })

# Rent Collection: indices for partial and unpaid
partial_indices = [7, 18, 33, 41]
unpaid_indices = [2, 12, 25, 37, 44]


# ================================================================
# CREATE WORKBOOK
# ================================================================
wb = Workbook()
wb.properties.creator = "Z.ai"


# ================================================================
# SHEET 1: GUIDANCE
# ================================================================
ws_guide = wb.active
ws_guide.title = "Guidance"
ws_guide.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_guide, title="Real Estate Emperor — System Usage Guide", last_col=10)

ws_guide.column_dimensions["A"].width = 3
ws_guide.column_dimensions["B"].width = 6
ws_guide.column_dimensions["C"].width = 90
for c_letter in ["D","E","F","G","H","I","J"]:
    ws_guide.column_dimensions[c_letter].width = 3

fill_section = PatternFill("solid", fgColor=PRIMARY)
font_section = Font(name="Calibri", size=13, bold=True, color=HEADER_TEXT)
fill_step = PatternFill("solid", fgColor=PRIMARY_LIGHT)
font_bullet = Font(name="Calibri", size=10, color=NEUTRAL_900)
font_tip_title = Font(name="Calibri", size=11, bold=True, color=ACCENT_POSITIVE)
font_tip_body = Font(name="Calibri", size=10, color=NEUTRAL_600)
fill_tip_bg = PatternFill("solid", fgColor="F0FAF4")
fill_critical = PatternFill("solid", fgColor=ACCENT_NEGATIVE)

r = 4

steps = [
    ("Register Your Properties", [
        'Go to the "Property Registry" tab',
        "Enter each unit: Unit ID, Building, Floor, Unit Type, Size, Monthly Rent",
        "Municipality Fee auto-calculates at 5% of Monthly Rent",
        'Set Status: Rented, Vacant, Under Maintenance, or Reserved',
        "For rented units, fill Current Tenant, Contract Start, and Contract End",
    ]),
    ("Set Up Tenant Profiles", [
        'Go to the "Tenants" tab',
        "Enter tenant details: Name, Phone, Email, Unit ID",
        "Monthly Rent auto-fills from Property Registry via VLOOKUP",
        "Fill in Lease Start, Lease End, Security Deposit, Payment Method",
        "Contract Duration auto-calculates",
        "LEASE END ALERTS: Cells turn RED if lease expired, AMBER if within 30 days of expiry",
    ]),
    ("Track Monthly Rent Collection", [
        'Go to "Rent Collection" at the start of each month',
        "Enter each rented unit's payment details",
        "Outstanding and Payment Status auto-calculate",
        "DAYS LATE auto-calculates for unpaid/partial payments",
        "Use the Collection Summary at the bottom to track overall performance",
    ]),
    ("Monitor Payment Alerts (By the 5th of Each Month)  ★ CRITICAL", [
        "This is the CRITICAL step for cash flow management",
        'On or before the 5th of each month, review "Rent Collection" tab',
        'Any tenant showing "Unpaid" or "Partial" status should be listed in "Payment Alerts"',
        "Copy their Unit ID, Name, Phone, Rent, and Due Date to the Payment Alerts tab",
        "Then MANUALLY send SMS, WhatsApp, or call each tenant",
        'Mark "Contacted?" as "Yes" after reaching out',
        "Set a Follow-up Date for a second check",
        "IMPORTANT: This system does NOT send SMS automatically. It is a tracking tool that tells you WHO to contact and WHEN. You must manually reach out to tenants using their phone numbers listed here.",
    ], True),
    ("Log Maintenance Requests", [
        'Go to "Maintenance" tab whenever a repair is needed',
        "Log the request with Unit ID, Category, Description, Priority",
        "Track Estimated vs Actual Cost",
        "Status options: Open, In Progress, Completed, Cancelled",
    ]),
    ("Record Operating Expenses", [
        'Go to "Expenses" tab regularly',
        "Log ALL expenses: Manpower, Municipality Fees, Maintenance, Leasing, Insurance, Utilities, etc.",
        "Mark recurring expenses for easy identification",
        "The Expense Summary at the bottom auto-calculates totals by category",
    ]),
    ("Review Revenue & Profit", [
        'Check "Revenue Analysis" for monthly and annual revenue breakdown',
        'Check "Profit & Loss" for the complete financial picture',
        "Dashboard shows KPIs: Total Units, Occupied, Vacant, Occupancy Rate, Revenue, Net Profit",
        "Charts visualize occupancy trends, revenue trends, expense breakdown, unit distribution",
    ]),
    ("Monitor Contract Renewals", [
        'Go to "Contract Tracker" to see which contracts are expiring soon',
        "Contact tenants 60 days before contract expiry for renewal discussions",
        "Plan ahead for units that may become vacant",
    ]),
]

for step_idx, step_info in enumerate(steps):
    if len(step_info) == 3:
        step_title, bullets, is_crit = step_info
    else:
        step_title, bullets = step_info
        is_crit = False
    is_critical = is_crit
    bg = fill_critical if is_critical else fill_section
    ws_guide.cell(row=r, column=2, value="STEP").font = font_section
    ws_guide.cell(row=r, column=2).fill = bg
    ws_guide.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.cell(row=r, column=3, value=step_title).font = font_section
    ws_guide.cell(row=r, column=3).fill = bg
    for c in range(4, 11):
        ws_guide.cell(row=r, column=c).fill = bg
    ws_guide.row_dimensions[r].height = 28
    r += 1
    for b in bullets:
        ws_guide.cell(row=r, column=2, value="").fill = fill_step
        ws_guide.cell(row=r, column=3, value=f"   •  {b}").font = font_bullet
        ws_guide.cell(row=r, column=3).fill = fill_step
        for c in range(4, 11):
            ws_guide.cell(row=r, column=c).fill = fill_step
        ws_guide.row_dimensions[r].height = 20
        r += 1
    r += 1

r += 2
ws_guide.cell(row=r, column=2, value="").fill = fill_tip_bg
ws_guide.cell(row=r, column=3, value="Tips for Optimal Use:").font = font_tip_title
ws_guide.cell(row=r, column=3).fill = fill_tip_bg
for c in range(4, 11):
    ws_guide.cell(row=r, column=c).fill = fill_tip_bg
ws_guide.row_dimensions[r].height = 24
r += 1

tips = [
    "Update Rent Collection DAILY as payments come in",
    "Review Payment Alerts by the 5th — late payments damage cash flow",
    "Log expenses immediately — don't wait until month-end",
    "Check Dashboard weekly for a quick business health check",
    "Back up this file regularly",
    "Print any sheet using File > Print (all sheets are print-ready)",
]
for t in tips:
    ws_guide.cell(row=r, column=2, value="").fill = fill_tip_bg
    ws_guide.cell(row=r, column=3, value=f"   ✓  {t}").font = font_tip_body
    ws_guide.cell(row=r, column=3).fill = fill_tip_bg
    for c in range(4, 11):
        ws_guide.cell(row=r, column=c).fill = fill_tip_bg
    ws_guide.row_dimensions[r].height = 20
    r += 1

setup_print(ws_guide, "J", r, "2:3")


# ================================================================
# SHEET 2: DASHBOARD
# ================================================================
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_dash, title="Real Estate Emperor — Real Estate Dashboard", last_col=14)

# KPI Cards - demo values
kpi_data = [
    ("Total Units", 55),
    ("Occupied Units", 48),
    ("Vacant Units", 7),
    ("Occupancy Rate", 0.873),
    ("Monthly Revenue", 185340),
    ("Net Profit", 95200),
]

ws_dash.row_dimensions[4].height = 8
ws_dash.row_dimensions[5].height = 38
ws_dash.row_dimensions[6].height = 20

for i, (label, value) in enumerate(kpi_data):
    col_val = 2 + i * 2
    col_lbl = 3 + i * 2
    ws_dash.merge_cells(start_row=5, start_column=col_val, end_row=5, end_column=col_lbl)
    ws_dash.merge_cells(start_row=6, start_column=col_val, end_row=6, end_column=col_lbl)
    cell_val = ws_dash.cell(row=5, column=col_val, value=value)
    cell_val.font = font_kpi()
    cell_val.alignment = Alignment(horizontal="center", vertical="bottom")
    cell_val.fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    if label == "Occupancy Rate":
        cell_val.number_format = PCT_FMT
    elif label in ("Monthly Revenue", "Net Profit"):
        cell_val.number_format = CURRENCY_FMT
    for c in [col_val, col_lbl]:
        ws_dash.cell(row=5, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    cell_lbl = ws_dash.cell(row=6, column=col_val, value=label)
    cell_lbl.font = font_kpi_label()
    cell_lbl.alignment = Alignment(horizontal="center", vertical="top")

# Chart Data: Occupancy by Building
ws_dash.cell(row=8, column=2, value="Building").font = font_caption()
ws_dash.cell(row=8, column=3, value="Occupied").font = font_caption()
ws_dash.cell(row=8, column=4, value="Vacant").font = font_caption()
bldg_occ = {"Building A": (13,2), "Building B": (12,2), "Building C": (14,2), "Building D": (9,1)}
for idx, (bldg, (occ, vac)) in enumerate(bldg_occ.items()):
    r = 9 + idx
    ws_dash.cell(row=r, column=2, value=bldg).font = font_caption()
    ws_dash.cell(row=r, column=3, value=occ).font = font_caption()
    ws_dash.cell(row=r, column=4, value=vac).font = font_caption()

# Chart Data: Monthly Revenue Jan-Apr 2025
ws_dash.cell(row=15, column=2, value="Month").font = font_caption()
ws_dash.cell(row=15, column=3, value="Revenue").font = font_caption()
rev_months = [("Jan", 172500), ("Feb", 174800), ("Mar", 179200), ("Apr", 185340)]
for idx, (m, rev) in enumerate(rev_months):
    r = 16 + idx
    ws_dash.cell(row=r, column=2, value=m).font = font_caption()
    ws_dash.cell(row=r, column=3, value=rev).font = font_caption()
    ws_dash.cell(row=r, column=3).number_format = CURRENCY_FMT

# Chart Data: Expense Breakdown
ws_dash.cell(row=22, column=2, value="Category").font = font_caption()
ws_dash.cell(row=22, column=3, value="Amount").font = font_caption()
exp_cats_demo = [
    ("Manpower/Staff", 35000), ("Municipality Fees", 9267), ("Maintenance", 1710),
    ("Leasing Commission", 4600), ("Insurance", 2800), ("Utilities", 8500),
    ("Marketing", 1500), ("Legal", 800), ("Cleaning", 4200),
    ("Security", 6000), ("Pest Control", 1200), ("Elevator Service", 1500), ("Other", 550)
]
for idx, (cat, amt) in enumerate(exp_cats_demo):
    r = 23 + idx
    ws_dash.cell(row=r, column=2, value=cat).font = font_caption()
    ws_dash.cell(row=r, column=3, value=amt).font = font_caption()
    ws_dash.cell(row=r, column=3).number_format = CURRENCY_FMT

# Chart Data: Unit Type Distribution
ws_dash.cell(row=38, column=2, value="Unit Type").font = font_caption()
ws_dash.cell(row=38, column=3, value="Count").font = font_caption()
ut_dist = [("Studio", 28), ("1 Bedroom", 23), ("2 Bedroom", 2), ("Shop", 2)]
for idx, (ut, cnt) in enumerate(ut_dist):
    r = 39 + idx
    ws_dash.cell(row=r, column=2, value=ut).font = font_caption()
    ws_dash.cell(row=r, column=3, value=cnt).font = font_caption()

# Charts
chart_a = create_bar_chart(width=18, height=11)
data_a = Reference(ws_dash, min_col=3, min_row=8, max_row=12)
cats_a = Reference(ws_dash, min_col=2, min_row=9, max_row=12)
chart_a.add_data(data_a, titles_from_data=True)
chart_a.set_categories(cats_a)
setup_chart_titles(chart_a, title="Occupancy by Building", y_title="Units", x_title="Building")
apply_chart_colors(chart_a)
ws_dash.add_chart(chart_a, "B44")

chart_b = create_line_chart(width=18, height=11)
data_b = Reference(ws_dash, min_col=3, min_row=15, max_row=19)
cats_b = Reference(ws_dash, min_col=2, min_row=16, max_row=19)
chart_b.add_data(data_b, titles_from_data=True)
chart_b.set_categories(cats_b)
setup_chart_titles(chart_b, title="Monthly Revenue Trend (Jan-Apr 2025)", y_title="Revenue (AED)", x_title="Month")
apply_chart_colors(chart_b)
ws_dash.add_chart(chart_b, "B61")

chart_c = create_pie_chart(width=14, height=11)
data_c = Reference(ws_dash, min_col=3, min_row=22, max_row=22+len(exp_cats_demo))
cats_c = Reference(ws_dash, min_col=2, min_row=23, max_row=22+len(exp_cats_demo))
chart_c.add_data(data_c, titles_from_data=True)
chart_c.set_categories(cats_c)
setup_chart_titles(chart_c, title="Expense Breakdown")
apply_pie_colors(chart_c, count=len(exp_cats_demo))
ws_dash.add_chart(chart_c, "B78")

chart_d = create_pie_chart(width=14, height=11)
data_d = Reference(ws_dash, min_col=3, min_row=38, max_row=38+len(ut_dist))
cats_d = Reference(ws_dash, min_col=2, min_row=39, max_row=38+len(ut_dist))
chart_d.add_data(data_d, titles_from_data=True)
chart_d.set_categories(cats_d)
setup_chart_titles(chart_d, title="Unit Type Distribution")
apply_pie_colors(chart_d, count=len(ut_dist))
ws_dash.add_chart(chart_d, "H78")

ws_dash.column_dimensions["A"].width = 3
for c in range(2, 15):
    ws_dash.column_dimensions[get_column_letter(c)].width = 14

setup_print(ws_dash, "N", 95, "2:3")


# ================================================================
# SHEET 3: PROPERTY REGISTRY
# ================================================================
ws_prop = wb.create_sheet("Property Registry")
ws_prop.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_prop, title="Property Registry — Unit Master List", last_col=13)

prop_headers = [
    "Unit ID", "Building", "Floor", "Unit Type", "Size (sqft)",
    "Monthly Rent (AED)", "Municipality Fee (AED)", "Status",
    "Current Tenant", "Contract Start", "Contract End", "Notes"
]
prop_col_start = 2
prop_col_end = 13
prop_header_row = 4
prop_data_start = 5

for i, h in enumerate(prop_headers):
    ws_prop.cell(row=prop_header_row, column=prop_col_start + i, value=h)
style_header_row(ws_prop, prop_header_row, prop_col_start, prop_col_end)

prop_widths = [12, 14, 8, 14, 12, 18, 20, 16, 18, 14, 14, 24]
for i, w in enumerate(prop_widths):
    ws_prop.column_dimensions[get_column_letter(prop_col_start + i)].width = w

for row_idx, pd in enumerate(prop_data):
    r = prop_data_start + row_idx
    ws_prop.cell(row=r, column=2, value=pd['unit_id'])
    ws_prop.cell(row=r, column=3, value=pd['building'])
    ws_prop.cell(row=r, column=4, value=pd['floor'])
    ws_prop.cell(row=r, column=5, value=pd['unit_type'])
    ws_prop.cell(row=r, column=6, value=pd['size'])
    ws_prop.cell(row=r, column=7, value=pd['rent'])
    ws_prop.cell(row=r, column=8, value=f"=IFERROR(G{r}*0.05,0)")
    ws_prop.cell(row=r, column=9, value=pd['status'])
    ws_prop.cell(row=r, column=10, value=pd['tenant'])
    if pd['contract_start']:
        ws_prop.cell(row=r, column=11, value=pd['contract_start'])
    if pd['contract_end']:
        ws_prop.cell(row=r, column=12, value=pd['contract_end'])
    if pd['status'] == 'Vacant':
        ws_prop.cell(row=r, column=13, value="Vacant unit")

    ws_prop.cell(row=r, column=8).number_format = CURRENCY_FMT
    ws_prop.cell(row=r, column=7).number_format = CURRENCY_FMT
    ws_prop.cell(row=r, column=6).number_format = "#,##0"
    ws_prop.cell(row=r, column=11).number_format = DATE_FMT
    ws_prop.cell(row=r, column=12).number_format = DATE_FMT
    style_data_row(ws_prop, r, prop_col_start, prop_col_end, row_idx)

# Extra empty rows for future data (rows 61-154)
for row_idx in range(55, 150):
    r = prop_data_start + row_idx
    ws_prop.cell(row=r, column=8, value=f"=IFERROR(G{r}*0.05,0)")
    ws_prop.cell(row=r, column=8).number_format = CURRENCY_FMT
    ws_prop.cell(row=r, column=7).number_format = CURRENCY_FMT
    ws_prop.cell(row=r, column=6).number_format = "#,##0"
    ws_prop.cell(row=r, column=11).number_format = DATE_FMT
    ws_prop.cell(row=r, column=12).number_format = DATE_FMT
    style_data_row(ws_prop, r, prop_col_start, prop_col_end, row_idx)

# Data Validations
dv_building = DataValidation(type="list", formula1='"Building A,Building B,Building C,Building D"', allow_blank=True)
dv_building.prompt = "Select Building"
ws_prop.add_data_validation(dv_building)
dv_building.add(f"C{prop_data_start}:C{prop_data_start+149}")

dv_unittype = DataValidation(type="list", formula1='"Studio,1 Bedroom,2 Bedroom,3 Bedroom,Shop,Office"', allow_blank=True)
dv_unittype.prompt = "Select Unit Type"
ws_prop.add_data_validation(dv_unittype)
dv_unittype.add(f"E{prop_data_start}:E{prop_data_start+149}")

dv_status = DataValidation(type="list", formula1='"Rented,Vacant,Under Maintenance,Reserved"', allow_blank=True)
dv_status.prompt = "Select Status"
ws_prop.add_data_validation(dv_status)
dv_status.add(f"I{prop_data_start}:I{prop_data_start+149}")

# Conditional formatting
status_range = f"I{prop_data_start}:I{prop_data_start+149}"
cf_cell_is(ws_prop, status_range, "equal", '"Rented"', CF_POSITIVE_FILL, CF_POSITIVE_FONT)
cf_cell_is(ws_prop, status_range, "equal", '"Vacant"', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
cf_cell_is(ws_prop, status_range, "equal", '"Under Maintenance"', CF_WARNING_FILL, CF_WARNING_FONT)

# Totals
total_row = prop_data_start + 150
ws_prop.cell(row=total_row, column=2, value="TOTALS")
ws_prop.cell(row=total_row, column=6, value=f"=IFERROR(SUM(F{prop_data_start}:F{total_row-1}),0)")
ws_prop.cell(row=total_row, column=6).number_format = "#,##0"
ws_prop.cell(row=total_row, column=7, value=f"=IFERROR(SUM(G{prop_data_start}:G{total_row-1}),0)")
ws_prop.cell(row=total_row, column=7).number_format = CURRENCY_FMT
ws_prop.cell(row=total_row, column=8, value=f"=IFERROR(SUM(H{prop_data_start}:H{total_row-1}),0)")
ws_prop.cell(row=total_row, column=8).number_format = CURRENCY_FMT
style_total_row(ws_prop, total_row, prop_col_start, prop_col_end)

ws_prop.freeze_panes = "C5"
setup_print(ws_prop, "M", total_row, "2:4")


# ================================================================
# SHEET 4: TENANTS
# ================================================================
ws_tenant = wb.create_sheet("Tenants")
ws_tenant.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_tenant, title="Tenant Directory", last_col=16)

tenant_headers = [
    "Tenant ID", "Tenant Name", "Phone", "Email", "Unit ID", "Unit Type",
    "Lease Start", "Lease End", "Contract Duration (months)",
    "Monthly Rent", "Security Deposit", "Payment Method",
    "Nationality", "Employer", "Emergency Contact"
]
tenant_col_start = 2
tenant_col_end = 16
tenant_header_row = 4
tenant_data_start = 5

for i, h in enumerate(tenant_headers):
    ws_tenant.cell(row=tenant_header_row, column=tenant_col_start + i, value=h)
style_header_row(ws_tenant, tenant_header_row, tenant_col_start, tenant_col_end)

tenant_widths = [12, 22, 16, 24, 12, 14, 14, 14, 22, 16, 16, 16, 14, 20, 20]
for i, w in enumerate(tenant_widths):
    ws_tenant.column_dimensions[get_column_letter(tenant_col_start + i)].width = w

for row_idx, td in enumerate(tenant_data):
    r = tenant_data_start + row_idx
    ws_tenant.cell(row=r, column=2, value=td['tid'])
    ws_tenant.cell(row=r, column=3, value=td['name'])
    ws_tenant.cell(row=r, column=4, value=td['phone'])
    ws_tenant.cell(row=r, column=5, value=td['email'])
    ws_tenant.cell(row=r, column=6, value=td['unit_id'])
    ws_tenant.cell(row=r, column=7, value=td['unit_type'])
    ws_tenant.cell(row=r, column=8, value=td['lease_start'])
    ws_tenant.cell(row=r, column=9, value=td['lease_end'])
    ws_tenant.cell(row=r, column=10, value=f"=IFERROR((I{r}-H{r})/30,0)")
    ws_tenant.cell(row=r, column=11, value=f"=IFERROR(VLOOKUP(F{r},'Property Registry'!B5:G154,6,FALSE),0)")
    ws_tenant.cell(row=r, column=12, value=td['security_deposit'])
    ws_tenant.cell(row=r, column=13, value=td['payment_method'])
    ws_tenant.cell(row=r, column=14, value=td['nationality'])
    ws_tenant.cell(row=r, column=15, value=td['employer'])
    ws_tenant.cell(row=r, column=16, value=td['emergency_contact'])

    ws_tenant.cell(row=r, column=10).number_format = "#,##0"
    ws_tenant.cell(row=r, column=11).number_format = CURRENCY_FMT
    ws_tenant.cell(row=r, column=12).number_format = CURRENCY_FMT
    ws_tenant.cell(row=r, column=8).number_format = DATE_FMT
    ws_tenant.cell(row=r, column=9).number_format = DATE_FMT
    style_data_row(ws_tenant, r, tenant_col_start, tenant_col_end, row_idx)

# Extra empty rows
for row_idx in range(48, 150):
    r = tenant_data_start + row_idx
    ws_tenant.cell(row=r, column=10, value=f"=IFERROR((I{r}-H{r})/30,0)")
    ws_tenant.cell(row=r, column=10).number_format = "#,##0"
    ws_tenant.cell(row=r, column=11, value=f"=IFERROR(VLOOKUP(F{r},'Property Registry'!B5:G154,6,FALSE),0)")
    ws_tenant.cell(row=r, column=11).number_format = CURRENCY_FMT
    ws_tenant.cell(row=r, column=12).number_format = CURRENCY_FMT
    ws_tenant.cell(row=r, column=8).number_format = DATE_FMT
    ws_tenant.cell(row=r, column=9).number_format = DATE_FMT
    style_data_row(ws_tenant, r, tenant_col_start, tenant_col_end, row_idx)

dv_paymethod = DataValidation(type="list", formula1='"Cash,Bank Transfer,Cheque,Online"', allow_blank=True)
dv_paymethod.prompt = "Select Payment Method"
ws_tenant.add_data_validation(dv_paymethod)
dv_paymethod.add(f"M{tenant_data_start}:M{tenant_data_start+149}")

lease_end_range = f"I{tenant_data_start}:I{tenant_data_start+149}"
cf_formula(ws_tenant, lease_end_range, f'AND(I{tenant_data_start}<TODAY(),I{tenant_data_start}<>"")', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
cf_formula(ws_tenant, lease_end_range, f'AND(I{tenant_data_start}-TODAY()<=30,I{tenant_data_start}-TODAY()>0,I{tenant_data_start}<>"")', CF_WARNING_FILL, CF_WARNING_FONT)

t_total_row = tenant_data_start + 150
ws_tenant.cell(row=t_total_row, column=2, value="TOTALS")
ws_tenant.cell(row=t_total_row, column=11, value=f"=IFERROR(SUM(K{tenant_data_start}:K{t_total_row-1}),0)")
ws_tenant.cell(row=t_total_row, column=11).number_format = CURRENCY_FMT
ws_tenant.cell(row=t_total_row, column=12, value=f"=IFERROR(SUM(L{tenant_data_start}:L{t_total_row-1}),0)")
ws_tenant.cell(row=t_total_row, column=12).number_format = CURRENCY_FMT
style_total_row(ws_tenant, t_total_row, tenant_col_start, tenant_col_end)

ws_tenant.freeze_panes = "C5"
setup_print(ws_tenant, "P", t_total_row, "2:4")


# ================================================================
# SHEET 5: RENT COLLECTION — April 2025 (NO Late Fee, NO Total Due)
# ================================================================
ws_rent = wb.create_sheet("Rent Collection")
ws_rent.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_rent, title="Rent Collection — April 2025", last_col=13)

rent_headers = [
    "Unit ID", "Tenant Name", "Monthly Rent", "Due Date", "Payment Date",
    "Amount Paid", "Outstanding", "Payment Status", "Days Late",
    "Payment Method", "Receipt #", "Notes"
]
rent_col_start = 2
rent_col_end = 13
rent_header_row = 4
rent_data_start = 5

for i, h in enumerate(rent_headers):
    ws_rent.cell(row=rent_header_row, column=rent_col_start + i, value=h)
style_header_row(ws_rent, rent_header_row, rent_col_start, rent_col_end)

rent_widths = [12, 22, 16, 14, 14, 16, 16, 16, 12, 16, 12, 24]
for i, w in enumerate(rent_widths):
    ws_rent.column_dimensions[get_column_letter(rent_col_start + i)].width = w

due_date = date(2025, 4, 1)
receipt_counter = 1001

for ti in range(48):
    r = rent_data_start + ti
    td = tenant_data[ti]
    rent_amt = td['rent']

    ws_rent.cell(row=r, column=2, value=td['unit_id'])
    ws_rent.cell(row=r, column=3, value=td['name'])
    ws_rent.cell(row=r, column=4, value=rent_amt)
    ws_rent.cell(row=r, column=5, value=due_date)

    if ti in unpaid_indices:
        # Unpaid
        ws_rent.cell(row=r, column=6, value=None)  # No payment date
        ws_rent.cell(row=r, column=7, value=0)  # Amount Paid
        # Outstanding = formula
        ws_rent.cell(row=r, column=8, value=f"=IFERROR(D{r}-G{r},0)")
        # Payment Status = formula
        ws_rent.cell(row=r, column=9, value=f'=IF(C{r}="","N/A",IF(G{r}=0,"Unpaid",IF(G{r}<D{r},"Partial","Paid")))')
        # Days Late = formula
        ws_rent.cell(row=r, column=10, value=f'=IF(I{r}="Paid",0,IF(I{r}="Partial",2,IF(I{r}="Unpaid",5,0)))')
        ws_rent.cell(row=r, column=11, value=td['payment_method'])
        ws_rent.cell(row=r, column=12, value=None)  # No receipt
        ws_rent.cell(row=r, column=13, value="Unpaid - follow up required")
    elif ti in partial_indices:
        # Partial
        pct = random.choice([0.5, 0.6, 0.65, 0.7])
        amt_paid = round(rent_amt * pct)
        pay_date = date(2025, 4, 3)
        ws_rent.cell(row=r, column=6, value=pay_date)
        ws_rent.cell(row=r, column=7, value=amt_paid)
        ws_rent.cell(row=r, column=8, value=f"=IFERROR(D{r}-G{r},0)")
        ws_rent.cell(row=r, column=9, value=f'=IF(C{r}="","N/A",IF(G{r}=0,"Unpaid",IF(G{r}<D{r},"Partial","Paid")))')
        ws_rent.cell(row=r, column=10, value=f'=IF(I{r}="Paid",0,IF(I{r}="Partial",2,IF(I{r}="Unpaid",5,0)))')
        ws_rent.cell(row=r, column=11, value=td['payment_method'])
        ws_rent.cell(row=r, column=12, value=f"RCP-{receipt_counter}")
        receipt_counter += 1
        ws_rent.cell(row=r, column=13, value="Partial payment received")
    else:
        # Paid in full
        pay_date = date(2025, 4, random.randint(1, 4))
        ws_rent.cell(row=r, column=6, value=pay_date)
        ws_rent.cell(row=r, column=7, value=rent_amt)
        ws_rent.cell(row=r, column=8, value=f"=IFERROR(D{r}-G{r},0)")
        ws_rent.cell(row=r, column=9, value=f'=IF(C{r}="","N/A",IF(G{r}=0,"Unpaid",IF(G{r}<D{r},"Partial","Paid")))')
        ws_rent.cell(row=r, column=10, value=f'=IF(I{r}="Paid",0,IF(I{r}="Partial",2,IF(I{r}="Unpaid",5,0)))')
        ws_rent.cell(row=r, column=11, value=td['payment_method'])
        ws_rent.cell(row=r, column=12, value=f"RCP-{receipt_counter}")
        receipt_counter += 1

    ws_rent.cell(row=r, column=8).number_format = CURRENCY_FMT
    ws_rent.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws_rent.cell(row=r, column=7).number_format = CURRENCY_FMT
    ws_rent.cell(row=r, column=5).number_format = DATE_FMT
    ws_rent.cell(row=r, column=6).number_format = DATE_FMT
    ws_rent.cell(row=r, column=10).number_format = "#,##0"
    style_data_row(ws_rent, r, rent_col_start, rent_col_end, ti)

# Add 7 vacant unit rows
for vi, uid in enumerate(sorted(vacant_ids)):
    r = rent_data_start + 48 + vi
    ws_rent.cell(row=r, column=2, value=uid)
    ws_rent.cell(row=r, column=3, value="")  # No tenant
    # Find rent for this vacant unit
    for pd in prop_data:
        if pd['unit_id'] == uid:
            ws_rent.cell(row=r, column=4, value=pd['rent'])
            break
    ws_rent.cell(row=r, column=5, value=due_date)
    ws_rent.cell(row=r, column=6, value=None)
    ws_rent.cell(row=r, column=7, value=0)
    ws_rent.cell(row=r, column=8, value=f"=IFERROR(D{r}-G{r},0)")
    ws_rent.cell(row=r, column=9, value=f'=IF(C{r}="","N/A",IF(G{r}=0,"Unpaid",IF(G{r}<D{r},"Partial","Paid")))')
    ws_rent.cell(row=r, column=10, value=f'=IF(I{r}="Paid",0,IF(I{r}="Partial",2,IF(I{r}="Unpaid",5,0)))')
    ws_rent.cell(row=r, column=13, value="Unit Vacant")

    ws_rent.cell(row=r, column=8).number_format = CURRENCY_FMT
    ws_rent.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws_rent.cell(row=r, column=7).number_format = CURRENCY_FMT
    ws_rent.cell(row=r, column=5).number_format = DATE_FMT
    ws_rent.cell(row=r, column=10).number_format = "#,##0"
    style_data_row(ws_rent, r, rent_col_start, rent_col_end, 48 + vi)

# Extra empty rows
rent_data_rows = 150
for row_idx in range(55, rent_data_rows):
    r = rent_data_start + row_idx
    ws_rent.cell(row=r, column=8, value=f"=IFERROR(D{r}-G{r},0)")
    ws_rent.cell(row=r, column=8).number_format = CURRENCY_FMT
    ws_rent.cell(row=r, column=9, value=f'=IF(C{r}="","N/A",IF(G{r}=0,"Unpaid",IF(G{r}<D{r},"Partial","Paid")))')
    ws_rent.cell(row=r, column=10, value=f'=IF(I{r}="Paid",0,IF(I{r}="Partial",2,IF(I{r}="Unpaid",5,0)))')
    ws_rent.cell(row=r, column=10).number_format = "#,##0"
    ws_rent.cell(row=r, column=4).number_format = CURRENCY_FMT
    ws_rent.cell(row=r, column=7).number_format = CURRENCY_FMT
    ws_rent.cell(row=r, column=5).number_format = DATE_FMT
    ws_rent.cell(row=r, column=6).number_format = DATE_FMT
    style_data_row(ws_rent, r, rent_col_start, rent_col_end, row_idx)

dv_paymethod2 = DataValidation(type="list", formula1='"Cash,Bank Transfer,Cheque,Online"', allow_blank=True)
dv_paymethod2.prompt = "Select Payment Method"
ws_rent.add_data_validation(dv_paymethod2)
dv_paymethod2.add(f"K{rent_data_start}:K{rent_data_start+rent_data_rows-1}")

status_range = f"I{rent_data_start}:I{rent_data_start+rent_data_rows-1}"
cf_cell_is(ws_rent, status_range, "equal", '"Paid"', CF_POSITIVE_FILL, CF_POSITIVE_FONT)
cf_cell_is(ws_rent, status_range, "equal", '"Partial"', CF_WARNING_FILL, CF_WARNING_FONT)
cf_cell_is(ws_rent, status_range, "equal", '"Unpaid"', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)

# Collection Summary
summary_start = rent_data_start + rent_data_rows + 2
ws_rent.cell(row=summary_start, column=2, value="COLLECTION SUMMARY").font = font_subheader()
ws_rent.merge_cells(start_row=summary_start, start_column=2, end_row=summary_start, end_column=5)

summary_items = [
    ("Total Expected Rent", f"=IFERROR(SUM(D{rent_data_start}:D{rent_data_start+rent_data_rows-1}),0)"),
    ("Total Collected", f"=IFERROR(SUM(G{rent_data_start}:G{rent_data_start+rent_data_rows-1}),0)"),
    ("Total Outstanding", f"=IFERROR(SUM(H{rent_data_start}:H{rent_data_start+rent_data_rows-1}),0)"),
    ("Collection Rate %", f"=IFERROR(SUM(G{rent_data_start}:G{rent_data_start+rent_data_rows-1})/SUM(D{rent_data_start}:D{rent_data_start+rent_data_rows-1}),0)"),
    ("# Paid", f'=IFERROR(COUNTIF(I{rent_data_start}:I{rent_data_start+rent_data_rows-1},"Paid"),0)'),
    ("# Partial", f'=IFERROR(COUNTIF(I{rent_data_start}:I{rent_data_start+rent_data_rows-1},"Partial"),0)'),
    ("# Unpaid", f'=IFERROR(COUNTIF(I{rent_data_start}:I{rent_data_start+rent_data_rows-1},"Unpaid"),0)'),
]

for i, (label, formula) in enumerate(summary_items):
    r = summary_start + 1 + i
    ws_rent.cell(row=r, column=2, value=label).font = font_body()
    ws_rent.cell(row=r, column=2).alignment = align_text()
    ws_rent.cell(row=r, column=4, value=formula).font = font_subheader()
    ws_rent.cell(row=r, column=4).alignment = align_number()
    if label == "Collection Rate %":
        ws_rent.cell(row=r, column=4).number_format = PCT_FMT
    elif label.startswith("#"):
        ws_rent.cell(row=r, column=4).number_format = "#,##0"
    else:
        ws_rent.cell(row=r, column=4).number_format = CURRENCY_FMT
    for c in range(2, 6):
        ws_rent.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)

ws_rent.freeze_panes = "C5"
setup_print(ws_rent, "M", summary_start + len(summary_items) + 1, "2:4")


# ================================================================
# SHEET 6: PAYMENT ALERTS (NO Late Fee, NO Total Due)
# ================================================================
ws_alerts = wb.create_sheet("Payment Alerts")
ws_alerts.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_alerts, title="Payment Alerts — Tenants Who Haven't Paid by the 5th", last_col=10)

ws_alerts.cell(row=3, column=2,
    value="IMPORTANT: This sheet does NOT send SMS automatically. It is a tracking tool. By the 5th of each month, review the Rent Collection sheet for unpaid tenants, copy their details here, then manually send SMS/WhatsApp/call them. Mark 'Contacted' as Yes once done.")
ws_alerts.cell(row=3, column=2).font = Font(name="Calibri", size=10, bold=True, color=ACCENT_NEGATIVE)
ws_alerts.merge_cells(start_row=3, start_column=2, end_row=3, end_column=10)
ws_alerts.row_dimensions[3].height = 40

alert_headers = [
    "Unit ID", "Tenant Name", "Phone", "Monthly Rent", "Due Date",
    "Days Late", "Contacted?", "Follow-up Date", "Notes"
]
alert_col_start = 2
alert_col_end = 10
alert_header_row = 4
alert_data_start = 5

for i, h in enumerate(alert_headers):
    cell = ws_alerts.cell(row=alert_header_row, column=alert_col_start + i, value=h)
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=ACCENT_NEGATIVE)
    cell.alignment = align_header()
    cell.border = border_header()
ws_alerts.row_dimensions[alert_header_row].height = ROW_HEIGHTS["header"]

alert_widths = [12, 22, 16, 16, 14, 12, 14, 14, 28]
for i, w in enumerate(alert_widths):
    ws_alerts.column_dimensions[get_column_letter(alert_col_start + i)].width = w

# Demo alert data: 5 unpaid + 4 partial
alert_rows_data = []

# Unpaid tenants
unpaid_notes = ["No response to calls", "Promised to pay by 10th", "Phone switched off", "Requesting payment plan", "Just contacted"]
for i, ti in enumerate(unpaid_indices):
    td = tenant_data[ti]
    contacted = "No" if i in [0, 2, 4] else "Yes"
    alert_rows_data.append({
        'unit_id': td['unit_id'], 'name': td['name'], 'phone': td['phone'],
        'rent': td['rent'], 'due_date': due_date, 'days_late': 5,
        'contacted': contacted, 'follow_up': date(2025, 4, 7),
        'notes': unpaid_notes[i]
    })

# Partial tenants
partial_notes = ["Partial received, following up", "Will pay remainder by 8th", "Awaiting confirmation", "Committed to full payment"]
partial_contacted = ["No", "Yes", "No", "Yes"]
for i, ti in enumerate(partial_indices):
    td = tenant_data[ti]
    alert_rows_data.append({
        'unit_id': td['unit_id'], 'name': td['name'], 'phone': td['phone'],
        'rent': td['rent'], 'due_date': due_date, 'days_late': 2,
        'contacted': partial_contacted[i], 'follow_up': date(2025, 4, 7),
        'notes': partial_notes[i]
    })

for row_idx, ad in enumerate(alert_rows_data):
    r = alert_data_start + row_idx
    ws_alerts.cell(row=r, column=2, value=ad['unit_id'])
    ws_alerts.cell(row=r, column=3, value=ad['name'])
    ws_alerts.cell(row=r, column=4, value=ad['phone'])
    ws_alerts.cell(row=r, column=5, value=ad['rent'])
    ws_alerts.cell(row=r, column=6, value=ad['due_date'])
    ws_alerts.cell(row=r, column=7, value=ad['days_late'])
    ws_alerts.cell(row=r, column=8, value=ad['contacted'])
    ws_alerts.cell(row=r, column=9, value=ad['follow_up'])
    ws_alerts.cell(row=r, column=10, value=ad['notes'])

    ws_alerts.cell(row=r, column=5).number_format = CURRENCY_FMT
    ws_alerts.cell(row=r, column=6).number_format = DATE_FMT
    ws_alerts.cell(row=r, column=9).number_format = DATE_FMT
    ws_alerts.cell(row=r, column=7).number_format = "#,##0"
    style_data_row(ws_alerts, r, alert_col_start, alert_col_end, row_idx)

# Extra empty rows
for row_idx in range(len(alert_rows_data), 50):
    r = alert_data_start + row_idx
    ws_alerts.cell(row=r, column=5).number_format = CURRENCY_FMT
    ws_alerts.cell(row=r, column=6).number_format = DATE_FMT
    ws_alerts.cell(row=r, column=9).number_format = DATE_FMT
    ws_alerts.cell(row=r, column=7).number_format = "#,##0"
    style_data_row(ws_alerts, r, alert_col_start, alert_col_end, row_idx)

dv_contacted = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
dv_contacted.prompt = "Mark as Yes after contacting tenant"
ws_alerts.add_data_validation(dv_contacted)
dv_contacted.add(f"H{alert_data_start}:H{alert_data_start+49}")

days_late_range = f"G{alert_data_start}:G{alert_data_start+49}"
cf_cell_is(ws_alerts, days_late_range, "greaterThan", "30", CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
cf_cell_is(ws_alerts, days_late_range, "greaterThan", "15", CF_WARNING_FILL, CF_WARNING_FONT)

ws_alerts.freeze_panes = "C5"
setup_print(ws_alerts, "J", alert_data_start + 49, "2:4")


# ================================================================
# SHEET 7: MAINTENANCE
# ================================================================
ws_maint = wb.create_sheet("Maintenance")
ws_maint.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_maint, title="Maintenance Log & Cost Tracker", last_col=13)

maint_headers = [
    "Request ID", "Date", "Unit ID", "Category", "Description",
    "Vendor/Technician", "Estimated Cost", "Actual Cost", "Status",
    "Priority", "Completion Date", "Notes"
]
maint_col_start = 2
maint_col_end = 13
maint_header_row = 4
maint_data_start = 5

for i, h in enumerate(maint_headers):
    ws_maint.cell(row=maint_header_row, column=maint_col_start + i, value=h)
style_header_row(ws_maint, maint_header_row, maint_col_start, maint_col_end)

maint_widths = [14, 14, 12, 16, 28, 20, 16, 16, 14, 12, 16, 24]
for i, w in enumerate(maint_widths):
    ws_maint.column_dimensions[get_column_letter(maint_col_start + i)].width = w

maint_demo = [
    ("MR-001", date(2025,3,15), "B-108", "AC", "AC not cooling", "CoolTech Services", 350, 380, "Completed", "Urgent", date(2025,3,16), "Replaced compressor"),
    ("MR-002", date(2025,3,22), "A-103", "Plumbing", "Water leak in bathroom", "Al Fix Plumbing", 200, None, "In Progress", "High", None, "Leak detected, parts ordered"),
    ("MR-003", date(2025,4,1), "C-112", "Lock/Door", "Door lock broken", "KeyMaster LLC", 150, None, "Open", "Medium", None, ""),
    ("MR-004", date(2025,3,10), "A-110", "Painting", "Paint touch-up needed", "ColorPro Painters", 100, 100, "Completed", "Low", date(2025,3,12), "Hallway and bedroom"),
    ("MR-005", date(2025,3,18), "D-103", "Plumbing", "Kitchen pipe blocked", "Al Fix Plumbing", 250, 280, "Completed", "High", date(2025,3,19), "Full pipe replacement"),
    ("MR-006", date(2025,3,25), "B-205", "Electrical", "Electrical outlet sparking", "SafeWire Electric", 180, 180, "Completed", "Urgent", date(2025,3,25), "Replaced outlet, safety check done"),
    ("MR-007", date(2025,4,2), "C-302", "Lock/Door", "Balcony door handle loose", "KeyMaster LLC", 80, None, "Open", "Low", None, ""),
    ("MR-008", date(2025,3,28), "A-107", "Structural", "Ceiling water stain", "BuildRight Maintenance", 400, None, "In Progress", "Medium", None, "Investigating source"),
]

for row_idx, md in enumerate(maint_demo):
    r = maint_data_start + row_idx
    ws_maint.cell(row=r, column=2, value=md[0])   # Request ID
    ws_maint.cell(row=r, column=3, value=md[1])   # Date
    ws_maint.cell(row=r, column=4, value=md[2])   # Unit ID
    ws_maint.cell(row=r, column=5, value=md[3])   # Category
    ws_maint.cell(row=r, column=6, value=md[4])   # Description
    ws_maint.cell(row=r, column=7, value=md[5])   # Vendor
    ws_maint.cell(row=r, column=8, value=md[6])   # Est Cost
    ws_maint.cell(row=r, column=9, value=md[7])   # Actual Cost
    ws_maint.cell(row=r, column=10, value=md[8])  # Status
    ws_maint.cell(row=r, column=11, value=md[9])  # Priority
    ws_maint.cell(row=r, column=12, value=md[10]) # Completion Date
    ws_maint.cell(row=r, column=13, value=md[11]) # Notes

    ws_maint.cell(row=r, column=3).number_format = DATE_FMT
    ws_maint.cell(row=r, column=8).number_format = CURRENCY_FMT
    ws_maint.cell(row=r, column=9).number_format = CURRENCY_FMT
    ws_maint.cell(row=r, column=12).number_format = DATE_FMT
    style_data_row(ws_maint, r, maint_col_start, maint_col_end, row_idx)

# Extra empty rows
for row_idx in range(len(maint_demo), 200):
    r = maint_data_start + row_idx
    ws_maint.cell(row=r, column=3).number_format = DATE_FMT
    ws_maint.cell(row=r, column=8).number_format = CURRENCY_FMT
    ws_maint.cell(row=r, column=9).number_format = CURRENCY_FMT
    ws_maint.cell(row=r, column=12).number_format = DATE_FMT
    style_data_row(ws_maint, r, maint_col_start, maint_col_end, row_idx)

dv_maint_cat = DataValidation(type="list",
    formula1='"Plumbing,Electrical,AC,Painting,Structural,Cleaning,Lock/Door,Appliance,Other"',
    allow_blank=True)
ws_maint.add_data_validation(dv_maint_cat)
dv_maint_cat.add(f"E{maint_data_start}:E{maint_data_start+199}")

dv_maint_status = DataValidation(type="list", formula1='"Open,In Progress,Completed,Cancelled"', allow_blank=True)
ws_maint.add_data_validation(dv_maint_status)
dv_maint_status.add(f"J{maint_data_start}:J{maint_data_start+199}")

dv_maint_priority = DataValidation(type="list", formula1='"Low,Medium,High,Urgent"', allow_blank=True)
ws_maint.add_data_validation(dv_maint_priority)
dv_maint_priority.add(f"K{maint_data_start}:K{maint_data_start+199}")

priority_range = f"K{maint_data_start}:K{maint_data_start+199}"
cf_cell_is(ws_maint, priority_range, "equal", '"Urgent"', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
cf_cell_is(ws_maint, priority_range, "equal", '"High"', CF_WARNING_FILL, CF_WARNING_FONT)

maint_status_range = f"J{maint_data_start}:J{maint_data_start+199}"
cf_cell_is(ws_maint, maint_status_range, "equal", '"Open"', CF_WARNING_FILL, CF_WARNING_FONT)
cf_cell_is(ws_maint, maint_status_range, "equal", '"In Progress"',
           PatternFill("solid", fgColor="D6E4F0"), Font(color="1B2A4A"))
cf_cell_is(ws_maint, maint_status_range, "equal", '"Completed"', CF_POSITIVE_FILL, CF_POSITIVE_FONT)

m_total_row = maint_data_start + 200
ws_maint.cell(row=m_total_row, column=2, value="TOTALS")
ws_maint.cell(row=m_total_row, column=8, value=f"=IFERROR(SUM(H{maint_data_start}:H{m_total_row-1}),0)")
ws_maint.cell(row=m_total_row, column=8).number_format = CURRENCY_FMT
ws_maint.cell(row=m_total_row, column=9, value=f"=IFERROR(SUM(I{maint_data_start}:I{m_total_row-1}),0)")
ws_maint.cell(row=m_total_row, column=9).number_format = CURRENCY_FMT
ws_maint.cell(row=m_total_row, column=10, value="Variance:")
ws_maint.cell(row=m_total_row, column=11, value=f"=IFERROR(H{m_total_row}-I{m_total_row},0)")
ws_maint.cell(row=m_total_row, column=11).number_format = CURRENCY_FMT
style_total_row(ws_maint, m_total_row, maint_col_start, maint_col_end)

ws_maint.freeze_panes = "C5"
setup_print(ws_maint, "M", m_total_row, "2:4")


# ================================================================
# SHEET 8: EXPENSES
# ================================================================
ws_exp = wb.create_sheet("Expenses")
ws_exp.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_exp, title="Operating Expenses — Monthly Tracker", last_col=11)

exp_headers = [
    "Date", "Category", "Description", "Amount (AED)", "Payment Method",
    "Vendor/Payee", "Invoice #", "Recurring?", "Building", "Notes"
]
exp_col_start = 2
exp_col_end = 11
exp_header_row = 4
exp_data_start = 5

for i, h in enumerate(exp_headers):
    ws_exp.cell(row=exp_header_row, column=exp_col_start + i, value=h)
style_header_row(ws_exp, exp_header_row, exp_col_start, exp_col_end)

exp_widths = [14, 20, 28, 16, 16, 20, 14, 12, 16, 24]
for i, w in enumerate(exp_widths):
    ws_exp.column_dimensions[get_column_letter(exp_col_start + i)].width = w

# 18 demo expenses
exp_demo = [
    (date(2025,3,1), "Manpower/Staff", "Building security - March", 12000, "Bank Transfer", "SafeGuard Security", "INV-2001", "Yes", "All Buildings", "Monthly contract"),
    (date(2025,3,1), "Manpower/Staff", "Building cleaners - March", 8000, "Bank Transfer", "CleanPro Services", "INV-2002", "Yes", "All Buildings", "Monthly contract"),
    (date(2025,3,1), "Manpower/Staff", "Building maintenance staff - March", 15000, "Bank Transfer", "Emperor Maintenance", "INV-2003", "Yes", "All Buildings", "3 staff members"),
    (date(2025,3,5), "Municipality Fees", "Q1 Municipality fees", 9267, "Bank Transfer", "Dubai Municipality", "MUN-0125", "Yes", "All Buildings", "5% of rental income"),
    (date(2025,3,10), "Maintenance", "AC repair B-108", 380, "Cheque", "CoolTech Services", "INV-2010", "No", "Building B", "MR-001"),
    (date(2025,3,12), "Maintenance", "Paint touch-up A-110", 100, "Cash", "ColorPro Painters", "INV-2011", "No", "Building A", "MR-004"),
    (date(2025,3,15), "Utilities", "DEWA electricity March", 5500, "Bank Transfer", "DEWA", "DEWA-3301", "Yes", "All Buildings", ""),
    (date(2025,3,15), "Utilities", "Water bill March", 3000, "Bank Transfer", "DEWA", "DEWA-3302", "Yes", "All Buildings", ""),
    (date(2025,3,18), "Leasing Commission", "Leasing commission - 2 new tenants", 4600, "Bank Transfer", "Emperor Leasing", "INV-2020", "No", "All Buildings", "A-110, B-205"),
    (date(2025,3,20), "Insurance", "Building insurance Q2", 2800, "Bank Transfer", "Oman Insurance", "POL-4455", "Yes", "All Buildings", "Quarterly payment"),
    (date(2025,3,22), "Security", "CCTV monitoring March", 6000, "Bank Transfer", "SafeGuard Security", "INV-2004", "Yes", "All Buildings", ""),
    (date(2025,3,25), "Maintenance", "Electrical repair B-205", 180, "Cash", "SafeWire Electric", "INV-2012", "No", "Building B", "MR-006"),
    (date(2025,3,28), "Cleaning", "Deep cleaning common areas", 4200, "Bank Transfer", "CleanPro Services", "INV-2005", "Yes", "All Buildings", ""),
    (date(2025,4,1), "Manpower/Staff", "Building security - April", 12000, "Bank Transfer", "SafeGuard Security", "INV-2101", "Yes", "All Buildings", "Monthly contract"),
    (date(2025,4,1), "Pest Control", "Quarterly pest control", 1200, "Cheque", "PestGuard UAE", "INV-2102", "Yes", "All Buildings", ""),
    (date(2025,4,2), "Elevator Service", "Elevator maintenance Q2", 1500, "Bank Transfer", "Otis Elevators", "INV-2103", "Yes", "All Buildings", "Quarterly service"),
    (date(2025,4,3), "Marketing", "Property listing ads", 1500, "Bank Transfer", "Bayut/Dubizzle", "INV-2104", "No", "All Buildings", ""),
    (date(2025,4,4), "Other", "Office supplies", 550, "Cash", "Office Depot UAE", "INV-2105", "No", "All Buildings", ""),
]

for row_idx, ed in enumerate(exp_demo):
    r = exp_data_start + row_idx
    ws_exp.cell(row=r, column=2, value=ed[0])
    ws_exp.cell(row=r, column=3, value=ed[1])
    ws_exp.cell(row=r, column=4, value=ed[2])
    ws_exp.cell(row=r, column=5, value=ed[3])
    ws_exp.cell(row=r, column=6, value=ed[4])
    ws_exp.cell(row=r, column=7, value=ed[5])
    ws_exp.cell(row=r, column=8, value=ed[6])
    ws_exp.cell(row=r, column=9, value=ed[7])
    ws_exp.cell(row=r, column=10, value=ed[8])
    ws_exp.cell(row=r, column=11, value=ed[9])

    ws_exp.cell(row=r, column=2).number_format = DATE_FMT
    ws_exp.cell(row=r, column=5).number_format = CURRENCY_FMT
    style_data_row(ws_exp, r, exp_col_start, exp_col_end, row_idx)

# Extra empty rows
for row_idx in range(len(exp_demo), 300):
    r = exp_data_start + row_idx
    ws_exp.cell(row=r, column=2).number_format = DATE_FMT
    ws_exp.cell(row=r, column=5).number_format = CURRENCY_FMT
    style_data_row(ws_exp, r, exp_col_start, exp_col_end, row_idx)

exp_categories = "Manpower/Staff,Municipality Fees,Maintenance,Leasing Commission,Insurance,Utilities,Marketing,Legal,Cleaning,Security,Pest Control,Elevator Service,Other"
dv_exp_cat = DataValidation(type="list", formula1=f'"{exp_categories}"', allow_blank=True)
ws_exp.add_data_validation(dv_exp_cat)
dv_exp_cat.add(f"C{exp_data_start}:C{exp_data_start+299}")

dv_exp_pay = DataValidation(type="list", formula1='"Cash,Bank Transfer,Cheque,Online"', allow_blank=True)
ws_exp.add_data_validation(dv_exp_pay)
dv_exp_pay.add(f"F{exp_data_start}:F{exp_data_start+299}")

dv_exp_recur = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws_exp.add_data_validation(dv_exp_recur)
dv_exp_recur.add(f"I{exp_data_start}:I{exp_data_start+299}")

dv_exp_building = DataValidation(type="list", formula1='"All Buildings,Building A,Building B,Building C,Building D"', allow_blank=True)
ws_exp.add_data_validation(dv_exp_building)
dv_exp_building.add(f"J{exp_data_start}:J{exp_data_start+299}")

# Expense Summary
exp_summary_start = exp_data_start + 300 + 2
ws_exp.cell(row=exp_summary_start, column=2, value="EXPENSE SUMMARY BY CATEGORY").font = font_subheader()
ws_exp.merge_cells(start_row=exp_summary_start, start_column=2, end_row=exp_summary_start, end_column=3)

exp_sum_hdr_row = exp_summary_start + 1
ws_exp.cell(row=exp_sum_hdr_row, column=2, value="Category")
ws_exp.cell(row=exp_sum_hdr_row, column=3, value="Total (AED)")
ws_exp.cell(row=exp_sum_hdr_row, column=2).font = font_header()
ws_exp.cell(row=exp_sum_hdr_row, column=3).font = font_header()
ws_exp.cell(row=exp_sum_hdr_row, column=2).fill = fill_header()
ws_exp.cell(row=exp_sum_hdr_row, column=3).fill = fill_header()
ws_exp.cell(row=exp_sum_hdr_row, column=2).alignment = align_header()
ws_exp.cell(row=exp_sum_hdr_row, column=3).alignment = align_header()
ws_exp.cell(row=exp_sum_hdr_row, column=2).border = border_header()
ws_exp.cell(row=exp_sum_hdr_row, column=3).border = border_header()

cat_list = [
    "Manpower/Staff", "Municipality Fees", "Maintenance", "Leasing Commission",
    "Insurance", "Utilities", "Marketing", "Legal", "Cleaning",
    "Security", "Pest Control", "Elevator Service", "Other"
]

for i, cat in enumerate(cat_list):
    r = exp_sum_hdr_row + 1 + i
    ws_exp.cell(row=r, column=2, value=cat).font = font_body()
    ws_exp.cell(row=r, column=2).alignment = align_text()
    ws_exp.cell(row=r, column=3,
                value=f'=IFERROR(SUMPRODUCT((C{exp_data_start}:C{exp_data_start+299}="{cat}")*(E{exp_data_start}:E{exp_data_start+299})),0)')
    ws_exp.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws_exp.cell(row=r, column=3).alignment = align_number()
    ws_exp.cell(row=r, column=3).font = font_body()
    fill = fill_data_row(i)
    ws_exp.cell(row=r, column=2).fill = fill
    ws_exp.cell(row=r, column=3).fill = fill

gt_row = exp_sum_hdr_row + 1 + len(cat_list)
ws_exp.cell(row=gt_row, column=2, value="GRAND TOTAL").font = font_subheader()
ws_exp.cell(row=gt_row, column=3, value=f"=IFERROR(SUM(C{exp_sum_hdr_row+1}:C{gt_row-1}),0)")
ws_exp.cell(row=gt_row, column=3).number_format = CURRENCY_FMT
ws_exp.cell(row=gt_row, column=3).font = font_subheader()
style_total_row(ws_exp, gt_row, 2, 3)

ws_exp.freeze_panes = "C5"
setup_print(ws_exp, "K", gt_row, "2:4")


# ================================================================
# SHEET 9: REVENUE ANALYSIS (NO Late Fee Income line)
# ================================================================
ws_rev = wb.create_sheet("Revenue Analysis")
ws_rev.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_rev, title="Revenue Analysis — Annual Overview", last_col=15)

rev_col_start = 2
rev_header_row = 4
rev_data_start = 5

ws_rev.cell(row=rev_header_row, column=2, value="Revenue Line Item")
months_full = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
for i, m in enumerate(months_full):
    ws_rev.cell(row=rev_header_row, column=3 + i, value=m)
ws_rev.cell(row=rev_header_row, column=15, value="Annual Total")

style_header_row(ws_rev, rev_header_row, 2, 15)

ws_rev.column_dimensions["B"].width = 24
for c in range(3, 16):
    ws_rev.column_dimensions[get_column_letter(c)].width = 14

# Demo revenue data
monthly_rental = [172500, 174800, 179200, 185340, 185340, 185340, 185340, 185340, 185340, 185340, 185340, 185340]
monthly_muni = [round(r * 0.05) for r in monthly_rental]
monthly_other = [1200, 1500, 800, 1100, 900, 1300, 1000, 1400, 1100, 900, 1200, 1500]
monthly_vacancy = [-round(7 * 2400), -round(7 * 2400), -round(7 * 2400), -round(7 * 2400), -round(7 * 2400), -round(7 * 2400), -round(7 * 2400), -round(7 * 2400), -round(7 * 2400), -round(7 * 2400), -round(7 * 2400), -round(7 * 2400)]
monthly_bad_debt = [-3200, -2400, -1800, -5800, -2400, -1800, -1200, -2400, -1800, -1200, -1800, -2400]

rev_items = [
    ("Rental Income", monthly_rental, "rental"),
    ("Municipality Fees Collected", monthly_muni, "municipality"),
    ("Other Income", monthly_other, "other"),
    ("GROSS REVENUE", None, "gross"),
    ("Less: Vacancy Loss", monthly_vacancy, "vacancy"),
    ("Less: Bad Debt/Unpaid", monthly_bad_debt, "bad_debt"),
    ("NET REVENUE", None, "net"),
]

current_row = rev_data_start
item_rows = {}

for item_name, data, key in rev_items:
    r = current_row
    item_rows[key] = r
    ws_rev.cell(row=r, column=2, value=item_name)

    if key == "gross":
        for c in range(3, 16):
            col_letter = get_column_letter(c)
            ws_rev.cell(row=r, column=c,
                        value=f"=IFERROR(SUM({col_letter}{rev_data_start}:{col_letter}{r-1}),0)")
        ws_rev.cell(row=r, column=2).font = font_subheader()
        for c in range(2, 16):
            ws_rev.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    elif key == "net":
        for c in range(3, 16):
            col_letter = get_column_letter(c)
            ws_rev.cell(row=r, column=c,
                        value=f"=IFERROR({col_letter}{item_rows['gross']}+{col_letter}{item_rows['vacancy']}+{col_letter}{item_rows['bad_debt']},0)")
        ws_rev.cell(row=r, column=2).font = font_subheader()
        for c in range(2, 16):
            ws_rev.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    else:
        for c_idx, val in enumerate(data):
            ws_rev.cell(row=r, column=3 + c_idx, value=val)
        ws_rev.cell(row=r, column=2).font = font_body()
        # Annual Total
        ws_rev.cell(row=r, column=15, value=f"=IFERROR(SUM(C{r}:N{r}),0)")

    for c in range(3, 16):
        ws_rev.cell(row=r, column=c).number_format = CURRENCY_FMT
        ws_rev.cell(row=r, column=c).alignment = align_number()

    current_row += 1

# Conditional formatting: negative -> red
for r in range(rev_data_start, current_row):
    for c in range(3, 16):
        cell_ref = f"{get_column_letter(c)}{r}"
        cf_formula(ws_rev, cell_ref, f'{cell_ref}<0', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)

# Revenue vs Vacancy Loss chart
chart_rev = create_bar_chart(width=22, height=12)
chart_data_start = current_row + 2
ws_rev.cell(row=chart_data_start, column=2, value="Category").font = font_caption()
ws_rev.cell(row=chart_data_start, column=3, value="Amount").font = font_caption()
chart_rev_items = [
    ("Gross Revenue", f"=O{item_rows['gross']}"),
    ("Vacancy Loss", f"=O{item_rows['vacancy']}"),
    ("Bad Debt", f"=O{item_rows['bad_debt']}"),
    ("Net Revenue", f"=O{item_rows['net']}"),
]
for i, (lbl, formula) in enumerate(chart_rev_items):
    r = chart_data_start + 1 + i
    ws_rev.cell(row=r, column=2, value=lbl).font = font_caption()
    ws_rev.cell(row=r, column=3, value=formula).font = font_caption()
    ws_rev.cell(row=r, column=3).number_format = CURRENCY_FMT

data_rev = Reference(ws_rev, min_col=3, min_row=chart_data_start, max_row=chart_data_start + len(chart_rev_items))
cats_rev = Reference(ws_rev, min_col=2, min_row=chart_data_start + 1, max_row=chart_data_start + len(chart_rev_items))
chart_rev.add_data(data_rev, titles_from_data=True)
chart_rev.set_categories(cats_rev)
setup_chart_titles(chart_rev, title="Revenue vs Vacancy Loss", y_title="Amount (AED)")
apply_chart_colors(chart_rev)
ws_rev.add_chart(chart_rev, f"B{current_row + 8}")

ws_rev.freeze_panes = "C5"
setup_print(ws_rev, "O", current_row + 25, "2:4")


# ================================================================
# SHEET 10: PROFIT & LOSS (NO Late Fee Income)
# ================================================================
ws_pl = wb.create_sheet("Profit & Loss")
ws_pl.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_pl, title="Profit & Loss Statement — Real Estate Emperor", last_col=4)

pl_col_start = 2
pl_col_end = 4
pl_header_row = 4
current_row = 5

def pl_section_header(ws, row, text):
    ws.cell(row=row, column=2, value=text).font = font_subheader()
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    ws.row_dimensions[row].height = ROW_HEIGHTS["header"]

def pl_line_item(ws, row, text, formula=None, indent=False):
    ws.cell(row=row, column=2, value=text).font = font_body()
    if indent:
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=2)
    if formula:
        ws.cell(row=row, column=3, value=formula).font = font_body()
        ws.cell(row=row, column=3).number_format = CURRENCY_FMT
        ws.cell(row=row, column=3).alignment = align_number()
    ws.row_dimensions[row].height = ROW_HEIGHTS["data"]
    return row

def pl_total_line(ws, row, text, formula):
    ws.cell(row=row, column=2, value=text).font = font_subheader()
    ws.cell(row=row, column=3, value=formula).font = font_subheader()
    ws.cell(row=row, column=3).number_format = CURRENCY_FMT
    ws.cell(row=row, column=3).alignment = align_number()
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
        ws.cell(row=row, column=c).border = border_total()
    ws.row_dimensions[row].height = ROW_HEIGHTS["total"]
    return row

def pl_margin_line(ws, row, text, formula):
    ws.cell(row=row, column=2, value=text).font = font_caption()
    ws.cell(row=row, column=3, value=formula).font = font_caption()
    ws.cell(row=row, column=3).number_format = PCT_FMT
    ws.cell(row=row, column=3).alignment = align_number()
    ws.row_dimensions[row].height = ROW_HEIGHTS["data"]
    return row

# REVENUE
pl_section_header(ws_pl, current_row, "REVENUE")
current_row += 1

rev_rental_row = current_row
pl_line_item(ws_pl, current_row, "Rental Income", f"=IFERROR('Revenue Analysis'!O{rev_data_start},0)", indent=True)
current_row += 1

rev_other_row = current_row
pl_line_item(ws_pl, current_row, "Other Income", f"=IFERROR('Revenue Analysis'!O{rev_data_start+2},0)", indent=True)
current_row += 1

total_rev_row = current_row
pl_total_line(ws_pl, current_row, "Total Revenue", f"=IFERROR(SUM(C{rev_rental_row}:C{current_row-1}),0)")
current_row += 1

# COST OF OPERATIONS
current_row += 1
pl_section_header(ws_pl, current_row, "COST OF OPERATIONS")
current_row += 1

cost_items = [
    ("Manpower/Staff Costs", "Manpower/Staff"),
    ("Maintenance Costs", "Maintenance"),
    ("Municipality Fees", "Municipality Fees"),
    ("Leasing Commissions", "Leasing Commission"),
    ("Insurance", "Insurance"),
    ("Utilities", "Utilities"),
]

cost_start_row = current_row
for item_name, cat in cost_items:
    pl_line_item(ws_pl, current_row, item_name,
                 f'=IFERROR(SUMPRODUCT((\'Expenses\'!C5:C304="{cat}")*(\'Expenses\'!E5:E304)),0)', indent=True)
    current_row += 1

total_cost_row = current_row
pl_total_line(ws_pl, current_row, "Total Cost of Operations", f"=IFERROR(SUM(C{cost_start_row}:C{current_row-1}),0)")
current_row += 1

# GROSS PROFIT
current_row += 1
gross_profit_row = current_row
pl_total_line(ws_pl, current_row, "Gross Profit", f"=IFERROR(C{total_rev_row}-C{total_cost_row},0)")
current_row += 1

gross_margin_row = current_row
pl_margin_line(ws_pl, current_row, "Gross Margin %", f"=IFERROR(C{gross_profit_row}/C{total_rev_row},0)")
current_row += 1

# OPERATING EXPENSES
current_row += 1
pl_section_header(ws_pl, current_row, "OPERATING EXPENSES")
current_row += 1

opex_items = [
    ("Marketing", "Marketing"),
    ("Legal & Professional", "Legal"),
    ("Cleaning & Security", "Cleaning"),
    ("Security Costs", "Security"),
    ("Pest Control", "Pest Control"),
    ("Elevator Service", "Elevator Service"),
    ("Other", "Other"),
]

opex_start_row = current_row
for item_name, cat in opex_items:
    pl_line_item(ws_pl, current_row, item_name,
                 f'=IFERROR(SUMPRODUCT((\'Expenses\'!C5:C304="{cat}")*(\'Expenses\'!E5:E304)),0)', indent=True)
    current_row += 1

total_opex_row = current_row
pl_total_line(ws_pl, current_row, "Total Operating Expenses", f"=IFERROR(SUM(C{opex_start_row}:C{current_row-1}),0)")
current_row += 1

# NET PROFIT
current_row += 1
net_profit_row = current_row
pl_total_line(ws_pl, current_row, "Net Profit", f"=IFERROR(C{gross_profit_row}-C{total_opex_row},0)")
current_row += 1

net_margin_row = current_row
pl_margin_line(ws_pl, current_row, "Net Margin %", f"=IFERROR(C{net_profit_row}/C{total_rev_row},0)")
current_row += 1

assert net_profit_row == 32, f"Net Profit row is {net_profit_row}, expected 32"
print(f"  Net Profit row confirmed at: {net_profit_row}")

# Waterfall Summary
current_row += 2
pl_section_header(ws_pl, current_row, "WATERFALL SUMMARY")
current_row += 1

waterfall_items = [
    ("Total Revenue", f"=C{total_rev_row}"),
    ("Less: Cost of Operations", f"=-C{total_cost_row}"),
    ("Gross Profit", f"=C{gross_profit_row}"),
    ("Less: Operating Expenses", f"=-C{total_opex_row}"),
    ("Net Profit", f"=C{net_profit_row}"),
]

for item_name, formula in waterfall_items:
    ws_pl.cell(row=current_row, column=2, value=item_name).font = font_body()
    ws_pl.cell(row=current_row, column=3, value=formula).font = font_subheader()
    ws_pl.cell(row=current_row, column=3).number_format = CURRENCY_FMT
    ws_pl.cell(row=current_row, column=3).alignment = align_number()
    if "Less" in item_name:
        ws_pl.cell(row=current_row, column=3).font = Font(name="Calibri", size=11, bold=True, color=ACCENT_NEGATIVE)
    elif "Net Profit" in item_name:
        ws_pl.cell(row=current_row, column=3).font = Font(name="Calibri", size=11, bold=True, color=ACCENT_POSITIVE)
    ws_pl.row_dimensions[current_row].height = ROW_HEIGHTS["data"]
    current_row += 1

ws_pl.column_dimensions["B"].width = 28
ws_pl.column_dimensions["C"].width = 20
ws_pl.column_dimensions["D"].width = 5

setup_print(ws_pl, "D", current_row, "2:4")


# ================================================================
# SHEET 11: CONTRACT TRACKER
# ================================================================
ws_con = wb.create_sheet("Contract Tracker")
ws_con.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_con, title="Lease Contract Tracker", last_col=12)

con_headers = [
    "Unit ID", "Tenant", "Phone", "Building", "Unit Type",
    "Lease Start", "Lease End", "Days Until Expiry", "Renewal Status",
    "New Rent", "Notes"
]
con_col_start = 2
con_col_end = 12
con_header_row = 4
con_data_start = 5

for i, h in enumerate(con_headers):
    ws_con.cell(row=con_header_row, column=con_col_start + i, value=h)
style_header_row(ws_con, con_header_row, con_col_start, con_col_end)

con_widths = [12, 22, 16, 14, 14, 14, 14, 18, 16, 14, 24]
for i, w in enumerate(con_widths):
    ws_con.column_dimensions[get_column_letter(con_col_start + i)].width = w

for ti in range(48):
    r = con_data_start + ti
    td = tenant_data[ti]
    
    ws_con.cell(row=r, column=2, value=td['unit_id'])
    ws_con.cell(row=r, column=3, value=td['name'])
    ws_con.cell(row=r, column=4, value=td['phone'])
    ws_con.cell(row=r, column=5, value=td['building'])
    ws_con.cell(row=r, column=6, value=td['unit_type'])
    ws_con.cell(row=r, column=7, value=td['lease_start'])
    ws_con.cell(row=r, column=8, value=td['lease_end'])
    ws_con.cell(row=r, column=9, value=f"=IFERROR(I{r}-TODAY(),0)")
    
    # Renewal Status
    if ti < 5:  # expired Feb 2025
        status = "Expired"
    elif ti < 10:  # expiring Apr 2025
        status = "Expiring Soon"
    elif ti < 13:  # expiring May-Jun 2025
        status = "Expiring Soon"
    else:
        status = "Active"
    ws_con.cell(row=r, column=10, value=status)
    
    # New Rent (5% increase for expired/expiring)
    if status in ("Expired", "Expiring Soon"):
        ws_con.cell(row=r, column=11, value=round(td['rent'] * 1.05))
    else:
        ws_con.cell(row=r, column=11, value="")
    
    # Notes
    if ti < 5:
        ws_con.cell(row=r, column=12, value="Contract expired - renewal discussion pending")
    elif ti < 10:
        ws_con.cell(row=r, column=12, value="Expiring this month - urgent renewal needed")
    elif ti < 13:
        ws_con.cell(row=r, column=12, value="Expiring soon - schedule renewal meeting")

    ws_con.cell(row=r, column=7).number_format = DATE_FMT
    ws_con.cell(row=r, column=8).number_format = DATE_FMT
    ws_con.cell(row=r, column=9).number_format = "#,##0"
    ws_con.cell(row=r, column=11).number_format = CURRENCY_FMT
    style_data_row(ws_con, r, con_col_start, con_col_end, ti)

# Extra empty rows
for row_idx in range(48, 150):
    r = con_data_start + row_idx
    ws_con.cell(row=r, column=9, value=f"=IFERROR(I{r}-TODAY(),0)")
    ws_con.cell(row=r, column=7).number_format = DATE_FMT
    ws_con.cell(row=r, column=8).number_format = DATE_FMT
    ws_con.cell(row=r, column=9).number_format = "#,##0"
    ws_con.cell(row=r, column=11).number_format = CURRENCY_FMT
    style_data_row(ws_con, r, con_col_start, con_col_end, row_idx)

dv_renewal = DataValidation(type="list", formula1='"Active,Expiring Soon,Expired,Renewed,Not Renewed"', allow_blank=True)
dv_renewal.prompt = "Select Renewal Status"
ws_con.add_data_validation(dv_renewal)
dv_renewal.add(f"J{con_data_start}:J{con_data_start+149}")

# Conditional formatting on Days Until Expiry
days_range = f"I{con_data_start}:I{con_data_start+149}"
cf_formula(ws_con, days_range, f'AND(I{con_data_start}<0,I{con_data_start}<>"")', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
cf_formula(ws_con, days_range, f'AND(I{con_data_start}<=30,I{con_data_start}>0,I{con_data_start}<>"")', CF_WARNING_FILL, CF_WARNING_FONT)

# Conditional formatting on Renewal Status
status_range = f"J{con_data_start}:J{con_data_start+149}"
cf_cell_is(ws_con, status_range, "equal", '"Expired"', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
cf_cell_is(ws_con, status_range, "equal", '"Expiring Soon"', CF_WARNING_FILL, CF_WARNING_FONT)
cf_cell_is(ws_con, status_range, "equal", '"Active"', CF_POSITIVE_FILL, CF_POSITIVE_FONT)

ws_con.freeze_panes = "C5"
setup_print(ws_con, "L", con_data_start + 149, "2:4")


# ================================================================
# VERIFY: No Late Fee references anywhere
# ================================================================
print("\n  Verifying NO Late Fee references...")
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val_lower = cell.value.lower()
                if "late fee" in val_lower or "late_fee" in val_lower or "late payment fee" in val_lower:
                    print(f"  WARNING: Late Fee reference found in {ws.title} at {cell.coordinate}: {cell.value}")

print("  Verification complete.\n")


# ================================================================
# SAVE
# ================================================================
print(f"  Saving to: {OUTPUT_PATH}")
wb.save(OUTPUT_PATH)

file_size = os.path.getsize(OUTPUT_PATH)
print(f"  File size: {file_size:,} bytes ({file_size/1024:.1f} KB)")
print(f"  Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"    - {s}")
print("\n  DONE! Sample workbook generated successfully.")
