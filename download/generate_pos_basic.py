#!/usr/bin/env python3
"""
Generate POS_Basic_Empty.xlsx — Professional POS Empty Template
Uses warm palette and base.py styling system.
"""

import sys
sys.path.insert(0, "/home/z/my-project/skills/xlsx")

from openpyxl import Workbook
from openpyxl.styles import Border as OpenpyxlBorder, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule

from templates.base import (
    use_palette_explicit,
    setup_sheet, style_header_row, style_data_row, style_total_row,
    font_title, font_header, font_body, font_caption, font_subheader,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    COLUMN_WIDTHS, ROW_HEIGHTS, FORMATS,
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    create_bar_chart, setup_chart_titles, apply_chart_colors,
    CHART_COLORS,
)

# ─── Activate warm palette BEFORE creating any styles ───
use_palette_explicit("warm")

wb = Workbook()
wb.properties.creator = "Z.ai"

CURRENCY_FMT = "#,##0.00"
DATE_FMT = "YYYY-MM-DD"
INTEGER_FMT = "#,##0"

# ─── Helper: set column widths from a list of (width_key_or_value, ...) ───
def set_col_widths(ws, widths):
    """widths: list of int or COLUMN_WIDTHS key strings, starting from col B."""
    for i, w in enumerate(widths, start=2):
        if isinstance(w, str):
            ws.column_dimensions[get_column_letter(i)].width = COLUMN_WIDTHS.get(w, w)
        else:
            ws.column_dimensions[get_column_letter(i)].width = w

# ─── Helper: apply print setup ───
def apply_print_setup(ws, orientation="landscape", fit_to_width=1, fit_to_height=0,
                      print_area=None, print_title_rows=None):
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = fit_to_width
    ws.page_setup.fitToHeight = fit_to_height
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    if print_area:
        ws.print_area = print_area
    if print_title_rows:
        ws.print_title_rows = print_title_rows

# ─── Helper: set warm tab color ───
def set_tab_color(ws):
    ws.sheet_properties.tabColor = "B85C1E"


# ═══════════════════════════════════════════════════════════
# SHEET 1: Sales Entry
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Sales Entry"

setup_sheet(ws1, title="Sales Entry", last_col=11)
set_tab_color(ws1)

# Column widths (B–K = cols 2–11)
# Date, Invoice#, Item/Service, Category, Qty, Unit Price, Total, Payment Method, Customer Name, Notes
col_widths_1 = ["date", "id_short", "name_en", "status", "id_short", "number", "number", "status", "name_en", "description"]
set_col_widths(ws1, col_widths_1)

# Header row (row 4)
headers_1 = ["Date", "Invoice #", "Item / Service", "Category", "Qty", "Unit Price", "Total", "Payment Method", "Customer Name", "Notes"]
for i, h in enumerate(headers_1, start=2):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, 2, 11)

# Data rows (rows 5–204, 200 rows)
for r in range(5, 205):
    row_idx = r - 5
    style_data_row(ws1, r, 2, 11, row_idx)
    # Total formula: =IFERROR(Qty*UnitPrice,"")
    # Qty = col F(6), UnitPrice = col G(7), Total = col H(8)
    ws1.cell(row=r, column=8).value = f'=IFERROR(F{r}*G{r},"")'
    ws1.cell(row=r, column=8).number_format = CURRENCY_FMT
    ws1.cell(row=r, column=8).alignment = align_number()
    # Date column formatting
    ws1.cell(row=r, column=2).number_format = DATE_FMT
    ws1.cell(row=r, column=2).alignment = align_date()
    # Invoice # alignment
    ws1.cell(row=r, column=3).alignment = align_text()
    # Item/Service alignment
    ws1.cell(row=r, column=4).alignment = align_text()
    # Category alignment
    ws1.cell(row=r, column=5).alignment = align_text()
    # Qty alignment
    ws1.cell(row=r, column=6).alignment = align_number()
    # Unit Price formatting
    ws1.cell(row=r, column=7).number_format = CURRENCY_FMT
    ws1.cell(row=r, column=7).alignment = align_number()
    # Payment Method alignment
    ws1.cell(row=r, column=9).alignment = align_text()
    # Customer Name alignment
    ws1.cell(row=r, column=10).alignment = align_text()
    # Notes alignment
    ws1.cell(row=r, column=11).alignment = align_text()

# Data validation: Payment Method dropdown (col I = col 9)
dv_payment = DataValidation(type="list", formula1='"Cash,Card,Online,Cheque"', allow_blank=True)
dv_payment.error = "Please select a valid payment method"
dv_payment.errorTitle = "Invalid Payment Method"
dv_payment.prompt = "Select payment method"
dv_payment.promptTitle = "Payment Method"
ws1.add_data_validation(dv_payment)
dv_payment.add(f"I5:I204")

# Totals row (row 205)
total_row_1 = 205
ws1.cell(row=total_row_1, column=2, value="TOTALS")
ws1.cell(row=total_row_1, column=2).alignment = align_text()
ws1.cell(row=total_row_1, column=6).value = f'=SUMPRODUCT((F5:F204<>"")*1)'
ws1.cell(row=total_row_1, column=6).alignment = align_number()
ws1.cell(row=total_row_1, column=7).value = ""
ws1.cell(row=total_row_1, column=8).value = f'=IFERROR(SUM(H5:H204),"")'
ws1.cell(row=total_row_1, column=8).number_format = CURRENCY_FMT
ws1.cell(row=total_row_1, column=8).alignment = align_number()
style_total_row(ws1, total_row_1, 2, 11)

# Freeze panes at C5
ws1.freeze_panes = "C5"

# Print setup
apply_print_setup(ws1, orientation="landscape", fit_to_width=1,
                  print_area="B1:K204",
                  print_title_rows="1:4")


# ═══════════════════════════════════════════════════════════
# SHEET 2: Daily Summary
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Daily Summary")

setup_sheet(ws2, title="Daily Summary", last_col=8)
set_tab_color(ws2)

# Columns: Date, Total Sales, # Transactions, Avg Transaction, Cash Total, Card Total, Online Total
# B=Date, C=Total Sales, D=# Transactions, E=Avg Transaction, F=Cash Total, G=Card Total, H=Online Total
col_widths_2 = ["date", "number", "id_short", "number", "number", "number", "number"]
set_col_widths(ws2, col_widths_2)

headers_2 = ["Date", "Total Sales", "# Transactions", "Avg Transaction", "Cash Total", "Card Total", "Online Total"]
for i, h in enumerate(headers_2, start=2):
    ws2.cell(row=4, column=i, value=h)
style_header_row(ws2, 4, 2, 8)

# 31 rows for daily tracking (rows 5–35)
for r in range(5, 36):
    row_idx = r - 5
    style_data_row(ws2, r, 2, 8, row_idx)
    # Date col
    ws2.cell(row=r, column=2).number_format = DATE_FMT
    ws2.cell(row=r, column=2).alignment = align_date()
    # Total Sales = SUMPRODUCT on 'Sales Entry' matching date
    # Date in Sales Entry is col B(2), Total is col H(8)
    ws2.cell(row=r, column=3).value = (
        f'=IFERROR(SUMPRODUCT(("\'Sales Entry\'!B$5:B$204"=B{r})*("\'Sales Entry\'!H$5:H$204")),"")'
    )
    ws2.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws2.cell(row=r, column=3).alignment = align_number()
    # # Transactions = SUMPRODUCT count matching date
    ws2.cell(row=r, column=4).value = (
        f'=IFERROR(SUMPRODUCT(("\'Sales Entry\'!B$5:B$204"=B{r})*1),"")'
    )
    ws2.cell(row=r, column=4).alignment = align_number()
    # Avg Transaction = Total Sales / # Transactions
    ws2.cell(row=r, column=5).value = f'=IFERROR(C{r}/D{r},"")'
    ws2.cell(row=r, column=5).number_format = CURRENCY_FMT
    ws2.cell(row=r, column=5).alignment = align_number()
    # Cash Total
    ws2.cell(row=r, column=6).value = (
        f'=IFERROR(SUMPRODUCT(("\'Sales Entry\'!B$5:B$204"=B{r})*("\'Sales Entry\'!I$5:I$204"="Cash")*("\'Sales Entry\'!H$5:H$204")),"")'
    )
    ws2.cell(row=r, column=6).number_format = CURRENCY_FMT
    ws2.cell(row=r, column=6).alignment = align_number()
    # Card Total
    ws2.cell(row=r, column=7).value = (
        f'=IFERROR(SUMPRODUCT(("\'Sales Entry\'!B$5:B$204"=B{r})*("\'Sales Entry\'!I$5:I$204"="Card")*("\'Sales Entry\'!H$5:H$204")),"")'
    )
    ws2.cell(row=r, column=7).number_format = CURRENCY_FMT
    ws2.cell(row=r, column=7).alignment = align_number()
    # Online Total
    ws2.cell(row=r, column=8).value = (
        f'=IFERROR(SUMPRODUCT(("\'Sales Entry\'!B$5:B$204"=B{r})*("\'Sales Entry\'!I$5:I$204"="Online")*("\'Sales Entry\'!H$5:H$204")),"")'
    )
    ws2.cell(row=r, column=8).number_format = CURRENCY_FMT
    ws2.cell(row=r, column=8).alignment = align_number()

# Totals row (row 36)
total_row_2 = 36
ws2.cell(row=total_row_2, column=2, value="TOTALS")
ws2.cell(row=total_row_2, column=2).alignment = align_text()
ws2.cell(row=total_row_2, column=3).value = f'=IFERROR(SUM(C5:C35),"")'
ws2.cell(row=total_row_2, column=3).number_format = CURRENCY_FMT
ws2.cell(row=total_row_2, column=3).alignment = align_number()
ws2.cell(row=total_row_2, column=4).value = f'=IFERROR(SUM(D5:D35),"")'
ws2.cell(row=total_row_2, column=4).alignment = align_number()
ws2.cell(row=total_row_2, column=5).value = f'=IFERROR(C{total_row_2}/D{total_row_2},"")'
ws2.cell(row=total_row_2, column=5).number_format = CURRENCY_FMT
ws2.cell(row=total_row_2, column=5).alignment = align_number()
ws2.cell(row=total_row_2, column=6).value = f'=IFERROR(SUM(F5:F35),"")'
ws2.cell(row=total_row_2, column=6).number_format = CURRENCY_FMT
ws2.cell(row=total_row_2, column=6).alignment = align_number()
ws2.cell(row=total_row_2, column=7).value = f'=IFERROR(SUM(G5:G35),"")'
ws2.cell(row=total_row_2, column=7).number_format = CURRENCY_FMT
ws2.cell(row=total_row_2, column=7).alignment = align_number()
ws2.cell(row=total_row_2, column=8).value = f'=IFERROR(SUM(H5:H35),"")'
ws2.cell(row=total_row_2, column=8).number_format = CURRENCY_FMT
ws2.cell(row=total_row_2, column=8).alignment = align_number()
style_total_row(ws2, total_row_2, 2, 8)

apply_print_setup(ws2, orientation="landscape", fit_to_width=1,
                  print_area="B1:H35",
                  print_title_rows="1:4")


# ═══════════════════════════════════════════════════════════
# SHEET 3: Weekly Summary
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Weekly Summary")

setup_sheet(ws3, title="Weekly Summary", last_col=7)
set_tab_color(ws3)

# Columns: Week #, Week Start Date, Week End Date, Total Sales, # Transactions, Avg Transaction
# B=Week#, C=Week Start, D=Week End, E=Total Sales, F=# Transactions, G=Avg Transaction
col_widths_3 = ["id_short", "date", "date", "number", "id_short", "number"]
set_col_widths(ws3, col_widths_3)

headers_3 = ["Week #", "Week Start Date", "Week End Date", "Total Sales", "# Transactions", "Avg Transaction"]
for i, h in enumerate(headers_3, start=2):
    ws3.cell(row=4, column=i, value=h)
style_header_row(ws3, 4, 2, 7)

# 52 rows for full year (rows 5–56)
for r in range(5, 57):
    row_idx = r - 5
    style_data_row(ws3, r, 2, 7, row_idx)
    # Week #
    ws3.cell(row=r, column=2).alignment = align_number()
    # Dates
    ws3.cell(row=r, column=3).number_format = DATE_FMT
    ws3.cell(row=r, column=3).alignment = align_date()
    ws3.cell(row=r, column=4).number_format = DATE_FMT
    ws3.cell(row=r, column=4).alignment = align_date()
    # Total Sales = SUMPRODUCT matching dates in Daily Summary between week start and end
    # Daily Summary: Date=B5:B35, Total Sales=C5:C35
    ws3.cell(row=r, column=5).value = (
        f'=IFERROR(SUMPRODUCT(("\'Daily Summary\'!B$5:B$35">=C{r})*("\'Daily Summary\'!B$5:B$35"<=D{r})*("\'Daily Summary\'!C$5:C$35")),"")'
    )
    ws3.cell(row=r, column=5).number_format = CURRENCY_FMT
    ws3.cell(row=r, column=5).alignment = align_number()
    # # Transactions
    ws3.cell(row=r, column=6).value = (
        f'=IFERROR(SUMPRODUCT(("\'Daily Summary\'!B$5:B$35">=C{r})*("\'Daily Summary\'!B$5:B$35"<=D{r})*("\'Daily Summary\'!D$5:D$35")),"")'
    )
    ws3.cell(row=r, column=6).alignment = align_number()
    # Avg Transaction
    ws3.cell(row=r, column=7).value = f'=IFERROR(E{r}/F{r},"")'
    ws3.cell(row=r, column=7).number_format = CURRENCY_FMT
    ws3.cell(row=r, column=7).alignment = align_number()

# Totals row (row 57)
total_row_3 = 57
ws3.cell(row=total_row_3, column=2, value="TOTALS")
ws3.cell(row=total_row_3, column=2).alignment = align_text()
ws3.cell(row=total_row_3, column=5).value = f'=IFERROR(SUM(E5:E56),"")'
ws3.cell(row=total_row_3, column=5).number_format = CURRENCY_FMT
ws3.cell(row=total_row_3, column=5).alignment = align_number()
ws3.cell(row=total_row_3, column=6).value = f'=IFERROR(SUM(F5:F56),"")'
ws3.cell(row=total_row_3, column=6).alignment = align_number()
ws3.cell(row=total_row_3, column=7).value = f'=IFERROR(E{total_row_3}/F{total_row_3},"")'
ws3.cell(row=total_row_3, column=7).number_format = CURRENCY_FMT
ws3.cell(row=total_row_3, column=7).alignment = align_number()
style_total_row(ws3, total_row_3, 2, 7)

apply_print_setup(ws3, orientation="landscape", fit_to_width=1,
                  print_area="B1:G56",
                  print_title_rows="1:4")


# ═══════════════════════════════════════════════════════════
# SHEET 4: Monthly Summary
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Monthly Summary")

setup_sheet(ws4, title="Monthly Summary", last_col=7)
set_tab_color(ws4)

# Columns: Month, Total Sales, # Transactions, Avg Transaction, Best Day Sales, Best Day Date
# B=Month, C=Total Sales, D=# Transactions, E=Avg Transaction, F=Best Day Sales, G=Best Day Date
col_widths_4 = ["name_en", "number", "id_short", "number", "number", "date"]
set_col_widths(ws4, col_widths_4)

headers_4 = ["Month", "Total Sales", "# Transactions", "Avg Transaction", "Best Day Sales", "Best Day Date"]
for i, h in enumerate(headers_4, start=2):
    ws4.cell(row=4, column=i, value=h)
style_header_row(ws4, 4, 2, 7)

# 12 rows for months (rows 5–16)
month_names = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]

for r in range(5, 17):
    row_idx = r - 5
    style_data_row(ws4, r, 2, 7, row_idx)
    m = row_idx + 1
    # Month name
    ws4.cell(row=r, column=2).value = month_names[row_idx]
    ws4.cell(row=r, column=2).alignment = align_text()
    # Total Sales = SUMPRODUCT matching MONTH on Daily Summary
    # Daily Summary: Date=B5:B35, Total Sales=C5:C35
    ws4.cell(row=r, column=3).value = (
        f'=IFERROR(SUMPRODUCT((MONTH("\'Daily Summary\'!B$5:B$35")={m})*("\'Daily Summary\'!C$5:C$35")),"")'
    )
    ws4.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws4.cell(row=r, column=3).alignment = align_number()
    # # Transactions
    ws4.cell(row=r, column=4).value = (
        f'=IFERROR(SUMPRODUCT((MONTH("\'Daily Summary\'!B$5:B$35")={m})*("\'Daily Summary\'!D$5:D$35")),"")'
    )
    ws4.cell(row=r, column=4).alignment = align_number()
    # Avg Transaction
    ws4.cell(row=r, column=5).value = f'=IFERROR(C{r}/D{r},"")'
    ws4.cell(row=r, column=5).number_format = CURRENCY_FMT
    ws4.cell(row=r, column=5).alignment = align_number()
    # Best Day Sales = MAX of matching month
    ws4.cell(row=r, column=6).value = (
        f'=IFERROR(MAX((MONTH("\'Daily Summary\'!B$5:B$35")={m})*("\'Daily Summary\'!C$5:C$35")),"")'
    )
    ws4.cell(row=r, column=6).number_format = CURRENCY_FMT
    ws4.cell(row=r, column=6).alignment = align_number()
    # Best Day Date — we use SUMPRODUCT to find the date where sales = best day sales for that month
    ws4.cell(row=r, column=7).value = (
        f'=IFERROR(SUMPRODUCT((MONTH("\'Daily Summary\'!B$5:B$35")={m})*("\'Daily Summary\'!C$5:C$35"=F{r})*("\'Daily Summary\'!B$5:B$35")),"")'
    )
    ws4.cell(row=r, column=7).number_format = DATE_FMT
    ws4.cell(row=r, column=7).alignment = align_date()

# Totals row (row 17)
total_row_4 = 17
ws4.cell(row=total_row_4, column=2, value="TOTALS")
ws4.cell(row=total_row_4, column=2).alignment = align_text()
ws4.cell(row=total_row_4, column=3).value = f'=IFERROR(SUM(C5:C16),"")'
ws4.cell(row=total_row_4, column=3).number_format = CURRENCY_FMT
ws4.cell(row=total_row_4, column=3).alignment = align_number()
ws4.cell(row=total_row_4, column=4).value = f'=IFERROR(SUM(D5:D16),"")'
ws4.cell(row=total_row_4, column=4).alignment = align_number()
ws4.cell(row=total_row_4, column=5).value = f'=IFERROR(C{total_row_4}/D{total_row_4},"")'
ws4.cell(row=total_row_4, column=5).number_format = CURRENCY_FMT
ws4.cell(row=total_row_4, column=5).alignment = align_number()
ws4.cell(row=total_row_4, column=6).value = ""
ws4.cell(row=total_row_4, column=7).value = ""
style_total_row(ws4, total_row_4, 2, 7)

# Bar chart of monthly sales
chart = create_bar_chart(chart_type="col", grouping="clustered", gap_width=80, width=18, height=10)
data_ref = Reference(ws4, min_col=3, min_row=4, max_row=16)
cats_ref = Reference(ws4, min_col=2, min_row=5, max_row=16)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
setup_chart_titles(chart, title="Monthly Sales", y_title="Sales", x_title="Month")
apply_chart_colors(chart)
ws4.add_chart(chart, "B19")

apply_print_setup(ws4, orientation="landscape", fit_to_width=1,
                  print_area="B1:G17",
                  print_title_rows="1:4")


# ═══════════════════════════════════════════════════════════
# SHEET 5: Invoice
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Invoice")

setup_sheet(ws5, title="", last_col=6)
set_tab_color(ws5)

# Wider columns for invoice layout
# B=Field labels, C=Values, D=spacer, E=Item, F=Qty, G=Price, H=Total
ws5.column_dimensions["B"].width = 18
ws5.column_dimensions["C"].width = 22
ws5.column_dimensions["D"].width = 3
ws5.column_dimensions["E"].width = 28
ws5.column_dimensions["F"].width = 10
ws5.column_dimensions["G"].width = 14
ws5.column_dimensions["H"].width = 14

# Company header placeholder (row 2 is title row from setup_sheet)
ws5.merge_cells("B2:C2")
ws5.cell(row=2, column=2).value = "[Your Company Name]"
ws5.cell(row=2, column=2).font = font_title()
ws5.cell(row=2, column=2).alignment = align_title()

ws5.merge_cells("B3:C3")
ws5.cell(row=3, column=2).value = "[Company Address | Phone | Email]"
ws5.cell(row=3, column=2).font = font_caption()
ws5.cell(row=3, column=2).alignment = align_text()

# INVOICE title
ws5.merge_cells("E2:H2")
ws5.cell(row=5, column=2).value = "INVOICE"
ws5.cell(row=5, column=2).font = font_title()
ws5.cell(row=5, column=2).alignment = align_title()

# Invoice fields (row 7–9)
ws5.cell(row=7, column=2).value = "Invoice #"
ws5.cell(row=7, column=2).font = font_subheader()
ws5.cell(row=7, column=3).alignment = align_text()
ws5.cell(row=7, column=3).font = font_body()

ws5.cell(row=8, column=2).value = "Date"
ws5.cell(row=8, column=2).font = font_subheader()
ws5.cell(row=8, column=3).number_format = DATE_FMT
ws5.cell(row=8, column=3).alignment = align_date()
ws5.cell(row=8, column=3).font = font_body()

ws5.cell(row=9, column=2).value = "Customer"
ws5.cell(row=9, column=2).font = font_subheader()
ws5.cell(row=9, column=3).alignment = align_text()
ws5.cell(row=9, column=3).font = font_body()

# Spacer row 10
ws5.row_dimensions[10].height = ROW_HEIGHTS["spacer"]

# Items table header (row 11)
item_headers = ["Item", "Qty", "Price", "Total"]
item_header_cols = [5, 6, 7, 8]  # E, F, G, H
for i, h in enumerate(item_headers):
    col = item_header_cols[i]
    ws5.cell(row=11, column=col, value=h)
style_header_row(ws5, 11, 5, 8)

# Item rows (rows 12–21, 10 line items)
for r in range(12, 22):
    row_idx = r - 12
    style_data_row(ws5, r, 5, 8, row_idx)
    # Total = Qty * Price
    ws5.cell(row=r, column=8).value = f'=IFERROR(F{r}*G{r},"")'
    ws5.cell(row=r, column=8).number_format = CURRENCY_FMT
    ws5.cell(row=r, column=8).alignment = align_number()
    ws5.cell(row=r, column=5).alignment = align_text()
    ws5.cell(row=r, column=6).alignment = align_number()
    ws5.cell(row=r, column=7).number_format = CURRENCY_FMT
    ws5.cell(row=r, column=7).alignment = align_number()

# Spacer row 22
ws5.row_dimensions[22].height = ROW_HEIGHTS["spacer"]

# Summary rows
ws5.cell(row=23, column=7, value="Subtotal")
ws5.cell(row=23, column=7).font = font_subheader()
ws5.cell(row=23, column=7).alignment = align_text()
ws5.cell(row=23, column=8).value = '=IFERROR(SUM(H12:H21),"")'
ws5.cell(row=23, column=8).number_format = CURRENCY_FMT
ws5.cell(row=23, column=8).font = font_body()
ws5.cell(row=23, column=8).alignment = align_number()

ws5.cell(row=24, column=7, value="VAT 5%")
ws5.cell(row=24, column=7).font = font_subheader()
ws5.cell(row=24, column=7).alignment = align_text()
ws5.cell(row=24, column=8).value = '=IFERROR(H23*0.05,"")'
ws5.cell(row=24, column=8).number_format = CURRENCY_FMT
ws5.cell(row=24, column=8).font = font_body()
ws5.cell(row=24, column=8).alignment = align_number()

# Grand total with special styling
ws5.cell(row=25, column=7, value="Grand Total")
ws5.cell(row=25, column=7).font = font_subheader()
ws5.cell(row=25, column=7).alignment = align_text()
ws5.cell(row=25, column=8).value = '=IFERROR(H23+H24,"")'
ws5.cell(row=25, column=8).number_format = CURRENCY_FMT
ws5.cell(row=25, column=8).font = font_subheader()
ws5.cell(row=25, column=8).alignment = align_number()
style_total_row(ws5, 25, 7, 8)

# Apply total fill and border to the summary area for consistency
for r in [23, 24]:
    for c in [7, 8]:
        ws5.cell(row=r, column=c).border = OpenpyxlBorder()

# Footer note
ws5.merge_cells("B27:H27")
ws5.cell(row=27, column=2).value = "Thank you for your business!"
ws5.cell(row=27, column=2).font = font_caption()
ws5.cell(row=27, column=2).alignment = Alignment(horizontal="center", vertical="center")

apply_print_setup(ws5, orientation="portrait", fit_to_width=1, fit_to_height=1,
                  print_area="B1:H27")


# ═══════════════════════════════════════════════════════════
# SHEET 6: Product List
# ═══════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Product List")

setup_sheet(ws6, title="Product List", last_col=6)
set_tab_color(ws6)

# Columns: Product ID, Item Name, Category, Unit Price, Description
# B=Product ID, C=Item Name, D=Category, E=Unit Price, F=Description
col_widths_6 = ["id_short", "name_en", "status", "number", "description"]
set_col_widths(ws6, col_widths_6)

headers_6 = ["Product ID", "Item Name", "Category", "Unit Price", "Description"]
for i, h in enumerate(headers_6, start=2):
    ws6.cell(row=4, column=i, value=h)
style_header_row(ws6, 4, 2, 6)

# 100 empty rows for products (rows 5–104)
for r in range(5, 105):
    row_idx = r - 5
    style_data_row(ws6, r, 2, 6, row_idx)
    ws6.cell(row=r, column=2).alignment = align_text()
    ws6.cell(row=r, column=3).alignment = align_text()
    ws6.cell(row=r, column=4).alignment = align_text()
    ws6.cell(row=r, column=5).number_format = CURRENCY_FMT
    ws6.cell(row=r, column=5).alignment = align_number()
    ws6.cell(row=r, column=6).alignment = align_text()

# Data validation: Category dropdown (col D = col 4)
# Common POS categories
dv_category = DataValidation(
    type="list",
    formula1='"Food,Beverage,Service,Merchandise,Electronics,Clothing,Other"',
    allow_blank=True
)
dv_category.error = "Please select a valid category"
dv_category.errorTitle = "Invalid Category"
dv_category.prompt = "Select product category"
dv_category.promptTitle = "Category"
ws6.add_data_validation(dv_category)
dv_category.add("D5:D104")

apply_print_setup(ws6, orientation="landscape", fit_to_width=1,
                  print_area="B1:F104",
                  print_title_rows="1:4")


# ═══════════════════════════════════════════════════════════
# Save workbook
# ═══════════════════════════════════════════════════════════
output_path = "/home/z/my-project/download/POS_Basic_Empty.xlsx"
wb.save(output_path)
print(f"✅ Saved: {output_path}")
