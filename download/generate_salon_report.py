#!/usr/bin/env python3
"""Generate Luxe Beauty Salon sample Excel report."""

import sys, os
sys.path.insert(0, '/home/z/my-project/skills/xlsx/templates')
sys.path.insert(0, '/home/z/my-project/skills/xlsx')
from base import *

from openpyxl import Workbook
from openpyxl.chart import Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# ── Activate elegant palette ──
use_palette_explicit("elegant")

wb = Workbook()
wb.properties.creator = "Z.ai"

# ════════════════════════════════════════════════════════════
# Helper: write footer row
# ════════════════════════════════════════════════════════════
def write_footer(ws, row, col_start, col_end):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start,
                   value="Prepared by Ahmed Ali | Data Analysis Services | ahmed-ali-ops.vercel.app")
    cell.font = font_caption()
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ════════════════════════════════════════════════════════════
# SHEET 1: Dashboard
# ════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Dashboard"

LAST_COL_DASH = 6  # columns B-F (A is margin)
setup_sheet(ws1, title="Luxe Beauty Salon — Monthly Business Report", last_col=LAST_COL_DASH)

# Subtitle row 3
ws1.merge_cells(start_row=3, start_column=2, end_row=3, end_column=LAST_COL_DASH)
sub_cell = ws1.cell(row=3, column=2, value="January 2026 | Dubai, UAE")
sub_cell.font = font_subheader()
sub_cell.alignment = align_title()
ws1.row_dimensions[3].height = ROW_HEIGHTS["title"]

# ── KPI Row (row 5) ──
kpi_data = [
    ("Total Revenue", 98600, "AED #,##0"),
    ("Total Expenses", 72400, "AED #,##0"),
    ("Net Profit", 26200, "AED #,##0"),
    ("Profit Margin", 0.266, "0.0%"),
]

kpi_row = 5
ws1.row_dimensions[kpi_row - 1].height = ROW_HEIGHTS["spacer"]  # row 4 spacer
ws1.row_dimensions[kpi_row].height = 40
ws1.row_dimensions[kpi_row + 1].height = 18  # label row

for i, (label, value, fmt) in enumerate(kpi_data):
    col = 2 + i  # B, C, D, E
    # Big number
    num_cell = ws1.cell(row=kpi_row, column=col, value=value)
    num_cell.font = font_kpi()
    num_cell.alignment = Alignment(horizontal="center", vertical="bottom")
    num_cell.number_format = fmt
    # Label
    lbl_cell = ws1.cell(row=kpi_row + 1, column=col, value=label)
    lbl_cell.font = font_kpi_label()
    lbl_cell.alignment = Alignment(horizontal="center", vertical="top")

# ── Monthly data table (starting row 8) ──
table_start = 8
headers_dash = ["Month", "Revenue (AED)", "Expenses (AED)", "Net Profit (AED)", "Profit Margin"]
for i, h in enumerate(headers_dash):
    ws1.cell(row=table_start, column=2 + i, value=h)
style_header_row(ws1, table_start, 2, 2 + len(headers_dash) - 1)

data_dash = [
    ("Oct 2025", 82300, 68500, 13800, 0.168),
    ("Nov 2025", 88900, 70200, 18700, 0.210),
    ("Dec 2025", 112400, 76800, 35600, 0.317),
    ("Jan 2026", 98600, 72400, 26200, 0.266),
]

for ri, row_data in enumerate(data_dash):
    r = table_start + 1 + ri
    for ci, val in enumerate(row_data):
        cell = ws1.cell(row=r, column=2 + ci, value=val)
        if ci == 0:
            cell.alignment = align_text()
        elif ci == 4:
            cell.number_format = "0.0%"
            cell.alignment = align_number()
        else:
            cell.number_format = "#,##0"
            cell.alignment = align_number()
    style_data_row(ws1, r, 2, 2 + len(headers_dash) - 1, ri)
    # Re-apply alignment since style_data_row overrides
    ws1.cell(row=r, column=2).alignment = align_text()
    for ci in range(1, 4):
        ws1.cell(row=r, column=2 + ci).alignment = align_number()
    ws1.cell(row=r, column=2 + 4).alignment = align_number()

# ── Charts ──
# Bar Chart: Revenue vs Expenses
chart_bar = create_bar_chart(chart_type="col", grouping="clustered", width=18, height=10)
cats = Reference(ws1, min_col=2, min_row=table_start + 1, max_row=table_start + 4)
rev_ref = Reference(ws1, min_col=3, min_row=table_start, max_row=table_start + 4)
exp_ref = Reference(ws1, min_col=4, min_row=table_start, max_row=table_start + 4)
chart_bar.add_data(rev_ref, titles_from_data=True)
chart_bar.add_data(exp_ref, titles_from_data=True)
chart_bar.set_categories(cats)
setup_chart_titles(chart_bar, title="Revenue vs Expenses", y_title="AED", x_title="Month")
apply_chart_colors(chart_bar, [CHART_COLORS[0], CHART_COLORS[2]])

# Line Chart: Profit Margin trend
chart_line = create_line_chart(width=18, height=10)
margin_ref = Reference(ws1, min_col=6, min_row=table_start, max_row=table_start + 4)
chart_line.add_data(margin_ref, titles_from_data=True)
chart_line.set_categories(cats)
setup_chart_titles(chart_line, title="Profit Margin Trend", y_title="Margin", x_title="Month")
apply_chart_colors(chart_line, [CHART_COLORS[1]])
chart_line.y_axis.numFmt = "0%"

# Place charts
ws1.add_chart(chart_bar, "B14")
ws1.add_chart(chart_line, "B26")

# Footer
write_footer(ws1, 38, 2, LAST_COL_DASH)

auto_fit_columns(ws1, header_row=table_start, data_start_row=table_start + 1)


# ════════════════════════════════════════════════════════════
# SHEET 2: Service Analysis
# ════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Service Analysis")

LAST_COL_SVC = 7  # B-G
setup_sheet(ws2, title="Revenue by Service — January 2026", last_col=LAST_COL_SVC)

headers_svc = ["Service", "Bookings", "Revenue (AED)", "Avg Price (AED)", "% of Revenue", "Utilization"]
svc_start = 4
for i, h in enumerate(headers_svc):
    ws2.cell(row=svc_start, column=2 + i, value=h)
style_header_row(ws2, svc_start, 2, 2 + len(headers_svc) - 1)

data_svc = [
    ("Hair Coloring",       180, 27000, 150.0, 0.274, 0.85),
    ("Haircut & Styling",   320, 19200, 60.0,  0.195, 0.90),
    ("Bridal Package",       12, 16800, 1400.0, 0.170, 1.00),
    ("Facial Treatment",    210, 12600, 60.0,  0.128, 0.72),
    ("Manicure & Pedicure", 280, 11200, 40.0,  0.114, 0.80),
    ("Keratin Treatment",    45,  8550, 190.0, 0.087, 0.65),
    ("Henna Design",         60,  3250,  54.2, 0.032, 0.50),
]

for ri, row_data in enumerate(data_svc):
    r = svc_start + 1 + ri
    for ci, val in enumerate(row_data):
        cell = ws2.cell(row=r, column=2 + ci, value=val)
        if ci == 0:
            cell.alignment = align_text()
        elif ci in (4, 5):
            cell.number_format = "0.0%"
            cell.alignment = align_number()
        else:
            cell.number_format = "#,##0"
            cell.alignment = align_number()
    style_data_row(ws2, r, 2, 2 + len(headers_svc) - 1, ri)
    ws2.cell(row=r, column=2).alignment = align_text()
    for ci in range(1, 4):
        ws2.cell(row=r, column=2 + ci).alignment = align_number()
    ws2.cell(row=r, column=2 + 4).alignment = align_number()
    ws2.cell(row=r, column=2 + 5).alignment = align_number()

# Total row
total_row_svc = svc_start + 1 + len(data_svc)
ws2.cell(row=total_row_svc, column=2, value="TOTAL")
ws2.cell(row=total_row_svc, column=3, value=1107)
ws2.cell(row=total_row_svc, column=4, value=98600)
ws2.cell(row=total_row_svc, column=5, value=89.0)
ws2.cell(row=total_row_svc, column=6, value=1.0)
ws2.cell(row=total_row_svc, column=7, value="—")
style_total_row(ws2, total_row_svc, 2, 2 + len(headers_svc) - 1)
ws2.cell(row=total_row_svc, column=2).alignment = align_text()
for ci in range(1, 4):
    ws2.cell(row=total_row_svc, column=2 + ci).alignment = align_number()
    ws2.cell(row=total_row_svc, column=2 + ci).number_format = "#,##0"
ws2.cell(row=total_row_svc, column=2 + 4).alignment = align_number()
ws2.cell(row=total_row_svc, column=2 + 4).number_format = "0.0%"
ws2.cell(row=total_row_svc, column=2 + 5).alignment = align_text()

# ── Charts ──
# Pie Chart: Revenue share
chart_pie = create_pie_chart(width=16, height=12)
pie_cats = Reference(ws2, min_col=2, min_row=svc_start + 1, max_row=svc_start + len(data_svc))
pie_vals = Reference(ws2, min_col=4, min_row=svc_start + 1, max_row=svc_start + len(data_svc))
chart_pie.add_data(pie_vals, titles_from_data=False)
chart_pie.set_categories(pie_cats)
setup_chart_titles(chart_pie, title="Revenue Share by Service")
apply_pie_colors(chart_pie, len(data_svc))
chart_pie.dataLabels = chart_pie.dataLabels or None
from openpyxl.chart.label import DataLabelList
chart_pie.dataLabels = DataLabelList(showPercent=True, showCatName=False, showVal=False)

# Bar Chart: Bookings by service
chart_book = create_bar_chart(chart_type="col", width=18, height=10)
book_cats = Reference(ws2, min_col=2, min_row=svc_start + 1, max_row=svc_start + len(data_svc))
book_vals = Reference(ws2, min_col=3, min_row=svc_start, max_row=svc_start + len(data_svc))
chart_book.add_data(book_vals, titles_from_data=True)
chart_book.set_categories(book_cats)
setup_chart_titles(chart_book, title="Bookings by Service", y_title="Bookings")
apply_chart_colors(chart_book, [CHART_COLORS[0]])

ws2.add_chart(chart_pie, "B14")
ws2.add_chart(chart_book, "B27")

# Footer
write_footer(ws2, 39, 2, LAST_COL_SVC)

auto_fit_columns(ws2, header_row=svc_start, data_start_row=svc_start + 1)


# ════════════════════════════════════════════════════════════
# SHEET 3: Staff Performance
# ════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Staff Performance")

LAST_COL_STAFF = 7  # B-G
setup_sheet(ws3, title="Staff Performance — January 2026", last_col=LAST_COL_STAFF)

headers_staff = ["Staff Name", "Role", "Clients Served", "Revenue (AED)", "Avg Ticket (AED)", "Rating"]
staff_start = 4
for i, h in enumerate(headers_staff):
    ws3.cell(row=staff_start, column=2 + i, value=h)
style_header_row(ws3, staff_start, 2, 2 + len(headers_staff) - 1)

data_staff = [
    ("Fatima K.",  "Senior Stylist",    95, 18500, 194.7, 4.9),
    ("Noor A.",    "Colorist",           78, 15200, 194.9, 4.8),
    ("Sara M.",    "Bridal Specialist",  22, 16800, 763.6, 5.0),
    ("Maryam H.",  "Nail Technician",   120,  8400,  70.0, 4.7),
    ("Aisha R.",   "Facial Specialist",  88,  7920,  90.0, 4.6),
    ("Layla S.",   "Junior Stylist",     65,  5850,  90.0, 4.5),
    ("Huda B.",    "Henna Artist",       50,  2700,  54.0, 4.8),
]

for ri, row_data in enumerate(data_staff):
    r = staff_start + 1 + ri
    for ci, val in enumerate(row_data):
        cell = ws3.cell(row=r, column=2 + ci, value=val)
        if ci <= 1:
            cell.alignment = align_text()
        elif ci in (2, 3):
            cell.number_format = "#,##0"
            cell.alignment = align_number()
        elif ci == 4:
            cell.number_format = "#,##0.0"
            cell.alignment = align_number()
        elif ci == 5:
            cell.number_format = "0.0"
            cell.alignment = align_number()
    style_data_row(ws3, r, 2, 2 + len(headers_staff) - 1, ri)
    ws3.cell(row=r, column=2).alignment = align_text()
    ws3.cell(row=r, column=3).alignment = align_text()
    for ci in range(2, 6):
        ws3.cell(row=r, column=2 + ci).alignment = align_number()

# Total row
total_row_staff = staff_start + 1 + len(data_staff)
ws3.cell(row=total_row_staff, column=2, value="TOTAL")
ws3.cell(row=total_row_staff, column=3, value="—")
ws3.cell(row=total_row_staff, column=4, value=518)
ws3.cell(row=total_row_staff, column=5, value=75370)
ws3.cell(row=total_row_staff, column=6, value=145.5)
ws3.cell(row=total_row_staff, column=7, value="—")
style_total_row(ws3, total_row_staff, 2, 2 + len(headers_staff) - 1)
ws3.cell(row=total_row_staff, column=2).alignment = align_text()
ws3.cell(row=total_row_staff, column=3).alignment = align_text()
for ci in range(2, 5):
    ws3.cell(row=total_row_staff, column=2 + ci).alignment = align_number()
    if ci == 2:
        ws3.cell(row=total_row_staff, column=2 + ci).number_format = "#,##0"
    elif ci == 3:
        ws3.cell(row=total_row_staff, column=2 + ci).number_format = "#,##0"
    elif ci == 4:
        ws3.cell(row=total_row_staff, column=2 + ci).number_format = "#,##0.0"
ws3.cell(row=total_row_staff, column=7).alignment = align_text()

# Note row
note_row = total_row_staff + 1
ws3.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=LAST_COL_STAFF)
note_cell = ws3.cell(row=note_row, column=2,
                     value="Note: Staff revenue does not equal total salon revenue (some bookings have multiple staff)")
note_cell.font = font_caption()
note_cell.alignment = align_text()

# ── Conditional formatting on Rating column (G) ──
rating_col_letter = "G"
rating_data_start = staff_start + 1
rating_data_end = staff_start + len(data_staff)
rating_range = f"{rating_col_letter}{rating_data_start}:{rating_col_letter}{rating_data_end}"

# Green for >= 4.8
ws3.conditional_formatting.add(
    rating_range,
    CellIsRule(operator="greaterThanOrEqual", formula=["4.8"],
              fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT)
)
# Amber for >= 4.6 and < 4.8
ws3.conditional_formatting.add(
    rating_range,
    CellIsRule(operator="between", formula=["4.6", "4.79"],
              fill=CF_WARNING_FILL, font=CF_WARNING_FONT)
)
# Red for < 4.6
ws3.conditional_formatting.add(
    rating_range,
    CellIsRule(operator="lessThan", formula=["4.6"],
              fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT)
)

# ── Horizontal Bar Chart: Revenue by Staff ──
chart_hbar = create_bar_chart(chart_type="bar", grouping="clustered", width=18, height=10)
hbar_cats = Reference(ws3, min_col=2, min_row=staff_start + 1, max_row=staff_start + len(data_staff))
hbar_vals = Reference(ws3, min_col=5, min_row=staff_start, max_row=staff_start + len(data_staff))
chart_hbar.add_data(hbar_vals, titles_from_data=True)
chart_hbar.set_categories(hbar_cats)
setup_chart_titles(chart_hbar, title="Revenue by Staff Member", x_title="AED")
apply_chart_colors(chart_hbar, [CHART_COLORS[0]])

ws3.add_chart(chart_hbar, "B15")

# Footer
write_footer(ws3, 27, 2, LAST_COL_STAFF)

auto_fit_columns(ws3, header_row=staff_start, data_start_row=staff_start + 1)


# ════════════════════════════════════════════════════════════
# SHEET 4: Recommendations
# ════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Recommendations")

LAST_COL_REC = 4  # B-D
setup_sheet(ws4, title="Key Insights & Recommendations", last_col=LAST_COL_REC)

headers_rec = ["#", "Insight", "Priority"]
rec_start = 4
for i, h in enumerate(headers_rec):
    ws4.cell(row=rec_start, column=2 + i, value=h)
style_header_row(ws4, rec_start, 2, 2 + len(headers_rec) - 1)

data_rec = [
    (1, "Bridal packages generate AED 1,400 avg ticket — highest ROI. Increase bridal marketing budget by 20%.", "High"),
    (2, "Keratin treatment utilization is only 65%. Run a promotion to boost bookings.", "Medium"),
    (3, "Henna service utilization at 50%. Consider offering henna as add-on to bridal packages.", "Medium"),
    (4, "December showed 31.7% margin (wedding season). Plan staffing for Dec 2026 peak.", "Low"),
    (5, "Facial treatments have room for growth (72% utilization). Partner with skincare brands for events.", "Medium"),
]

for ri, row_data in enumerate(data_rec):
    r = rec_start + 1 + ri
    for ci, val in enumerate(row_data):
        cell = ws4.cell(row=r, column=2 + ci, value=val)
        if ci == 0:
            cell.alignment = align_date()
        elif ci == 1:
            cell.alignment = align_text()
        else:
            cell.alignment = align_date()
    style_data_row(ws4, r, 2, 2 + len(headers_rec) - 1, ri)
    ws4.cell(row=r, column=2).alignment = align_date()
    ws4.cell(row=r, column=3).alignment = align_text()
    ws4.cell(row=r, column=4).alignment = align_date()

# ── Conditional formatting on Priority column (D) ──
priority_col_letter = "D"
priority_data_start = rec_start + 1
priority_data_end = rec_start + len(data_rec)
priority_range = f"{priority_col_letter}{priority_data_start}:{priority_col_letter}{priority_data_end}"

# High = red/negative
ws4.conditional_formatting.add(
    priority_range,
    CellIsRule(operator="equal", formula=['"High"'],
              fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT)
)
# Medium = amber/warning
ws4.conditional_formatting.add(
    priority_range,
    CellIsRule(operator="equal", formula=['"Medium"'],
              fill=CF_WARNING_FILL, font=CF_WARNING_FONT)
)
# Low = green/positive
ws4.conditional_formatting.add(
    priority_range,
    CellIsRule(operator="equal", formula=['"Low"'],
              fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT)
)

# Footer
write_footer(ws4, rec_start + len(data_rec) + 2, 2, LAST_COL_REC)

auto_fit_columns(ws4, header_row=rec_start, data_start_row=rec_start + 1)

# ── Set wider column for Insight text ──
ws4.column_dimensions["C"].width = 70

# ════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════
output_path = "/home/z/my-project/download/sample_report_salon.xlsx"
wb.save(output_path)
print(f"✅ Report saved to {output_path}")
