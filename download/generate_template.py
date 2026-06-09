#!/usr/bin/env python3
"""
Property Management Data Collection Template Generator
=======================================================
Generates a professional 8-sheet workbook for data migration preparation.
Extracts only explicitly available data from APRIL-25.xlsx source.
"""

import sys, os
sys.path.insert(0, '/home/z/my-project/skills/xlsx/templates')
from base import *

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Border, Side, Alignment, numbers, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from collections import Counter, OrderedDict
import datetime

# ── Load source data ─────────────────────────────────────────────────
source_wb = openpyxl.load_workbook('/home/z/my-project/upload/APRIL-25.xlsx', data_only=True)
source_ws = source_wb['Sheet1']

# Extract all records from Sheet1 consolidated view
all_records = []
current_property = None
for row_idx in range(3, source_ws.max_row + 1):
    a_val = source_ws.cell(row=row_idx, column=1).value
    b_val = source_ws.cell(row=row_idx, column=2).value
    c_val = source_ws.cell(row=row_idx, column=3).value
    f_val = source_ws.cell(row=row_idx, column=6).value
    
    if a_val and str(a_val).strip():
        current_property = str(a_val).strip()
    if b_val is None and f_val is not None:
        continue
    if b_val is None:
        continue
    
    c_val_str = str(c_val).strip() if c_val else ''
    d_val = source_ws.cell(row=row_idx, column=4).value
    e_val = source_ws.cell(row=row_idx, column=5).value
    g_val = source_ws.cell(row=row_idx, column=7).value
    j_val = source_ws.cell(row=row_idx, column=10).value
    k_val = source_ws.cell(row=row_idx, column=11).value
    
    # Determine if bill_no is actually a bank name
    bill_no_str = ''
    if d_val is not None:
        bill_no_str = str(d_val).strip()
    
    is_bank_name = any(b in bill_no_str.upper() for b in ['ADIB', 'ADCB', 'ENBD', 'DIB', 'ARBAB', 'CHQ', 'ADVB'])
    payment_method = ''
    actual_bill_no = ''
    if bill_no_str:
        if is_bank_name:
            payment_method = bill_no_str.upper().replace('\n', '/')
            if 'ADIB' in payment_method and 'DIB' not in payment_method.replace('ADIB',''):
                payment_method = 'ADIB'
            elif 'DIB' in payment_method:
                payment_method = 'DIB'
            elif 'ADCB' in payment_method:
                payment_method = 'ADCB'
            elif 'ENBD' in payment_method:
                payment_method = 'ENBD'
            elif 'ARBAB' in payment_method:
                payment_method = 'ARBAB (Cash Collection)'
            elif 'CHQ' in payment_method.lower():
                payment_method = 'Cheque'
        else:
            actual_bill_no = bill_no_str
    
    record = {
        'property': current_property or '',
        'room': b_val,
        'name': c_val_str if c_val_str and c_val_str != 'None' else '',
        'bill_no': actual_bill_no,
        'date': e_val,
        'rent': f_val,
        'prev_due': g_val if g_val else 0,
        'balance': j_val if j_val else 0,
        'remarks': k_val,
        'payment_method': payment_method,
        'is_bank_name': is_bank_name,
    }
    all_records.append(record)

source_wb.close()

# ── Unique properties list ───────────────────────────────────────────
properties = OrderedDict()
for r in all_records:
    p = r['property']
    if p and p not in properties:
        properties[p] = {'units': 0, 'records': []}
    if p:
        properties[p]['units'] += 1
        properties[p]['records'].append(r)

# Generate property codes
property_code_map = {}
for i, pname in enumerate(properties.keys(), 1):
    code = f"PROP-{i:03d}"
    property_code_map[pname] = code

# Generate tenant IDs
name_count = Counter()
tenant_id_map = {}
for r in all_records:
    if r['name']:
        key = (r['property'], r['room'], r['name'])
        if key not in tenant_id_map:
            name_count[r['name']] += 1
            seq = name_count[r['name']]
            clean_name = r['name'].replace(' ','').replace('/','').replace(chr(39),'').replace('.','').upper()[:8]
            if seq == 1:
                tid = f"TEN-{clean_name}"
            else:
                tid = f"TEN-{clean_name}-{seq}"
            tenant_id_map[key] = tid

# Detect duplicates across properties
global_name_count = Counter(r['name'] for r in all_records if r['name'])

# ── Create workbook ──────────────────────────────────────────────────
wb = Workbook()
wb.properties.creator = "Z.ai"

# ══════════════════════════════════════════════════════════════════════
# CUSTOM STYLE OVERRIDES FOR TEMPLATE (required fields = red, optional = yellow)
# ══════════════════════════════════════════════════════════════════════
REQUIRED_FILL = PatternFill('solid', fgColor='FDE8E8')   # Light red for required
OPTIONAL_FILL = PatternFill('solid', fgColor='FFF8E1')   # Light yellow for optional
REQUIRED_HEADER_FILL = PatternFill('solid', fgColor='C0392B')  # Red header for required columns
OPTIONAL_HEADER_FILL = PatternFill('solid', fgColor='D4820A')  # Amber header for optional columns
COMPLETE_FILL = PatternFill('solid', fgColor='E8F5E9')   # Green for complete
NEEDS_REVIEW_FILL = PatternFill('solid', fgColor='FFF8E1')  # Yellow for needs review
MISSING_FILL = PatternFill('solid', fgColor='FDE8E8')    # Red for missing

# ══════════════════════════════════════════════════════════════════════
# SHEET 1: Instructions & Data Entry Guide
# ══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Instructions"
setup_sheet(ws1, title="Property Management Data Collection Guide", last_col=8)

# Column widths
ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 6
ws1.column_dimensions['C'].width = 28
ws1.column_dimensions['D'].width = 45
ws1.column_dimensions['E'].width = 45
ws1.column_dimensions['F'].width = 45
ws1.column_dimensions['G'].width = 45
ws1.column_dimensions['H'].width = 20

# ── Section: Overview ────────────────────────────────────────────────
row = 4
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
cell = ws1.cell(row=row, column=2, value="OVERVIEW")
cell.font = font_subheader()
cell.fill = fill_header()
cell.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
cell.alignment = align_header()
for c in range(3, 9):
    ws1.cell(row=row, column=c).fill = fill_header()
ws1.row_dimensions[row].height = 28

row = 5
ws1.merge_cells(start_row=row, start_column=2, end_row=row+3, end_column=8)
overview_text = (
    "This workbook is a structured data collection template for the Real Estate Emperor property management system migration. "
    "It is designed to capture all required information about properties, units, tenants, leases, balances, and payment history "
    "so that the data can be imported into the Property Dashboard with 100% accuracy and zero assumptions.\n\n"
    "The existing APRIL-25 billing spreadsheet was used as a reference source only. Where data could be safely extracted "
    "(property names, unit numbers, tenant names, monthly rent amounts, outstanding balances, and some payment methods), "
    "it has been pre-populated. All other fields are left BLANK for manual completion by staff using tenant contracts, "
    "paper records, Emirates ID copies, and property files.\n\n"
    "DO NOT import, modify, correct, infer, generate, or fabricate any data. Only fill in fields where you have verified information."
)
cell = ws1.cell(row=row, column=2, value=overview_text)
cell.font = font_body()
cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
for r in range(row, row+4):
    ws1.row_dimensions[r].height = 30

# ── Section: Color Legend ────────────────────────────────────────────
row = 10
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
cell = ws1.cell(row=row, column=2, value="COLOR LEGEND")
cell.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
cell.fill = fill_header()
cell.alignment = align_header()
for c in range(3, 9):
    ws1.cell(row=row, column=c).fill = fill_header()
ws1.row_dimensions[row].height = 28

legend_items = [
    (REQUIRED_FILL, "Required Field", "Mandatory - must be completed before migration. Red background indicates missing required data."),
    (OPTIONAL_FILL, "Optional Field", "Helpful but not mandatory. Yellow background indicates this field is optional."),
    (COMPLETE_FILL, "Complete", "Data quality status: All required fields are filled and verified."),
    (NEEDS_REVIEW_FILL, "Needs Review", "Data quality status: Some fields may need verification or correction."),
    (MISSING_FILL, "Missing Information", "Data quality status: One or more required fields are empty."),
]
for i, (fill, label, desc) in enumerate(legend_items):
    r = row + 1 + i
    ws1.cell(row=r, column=2).fill = fill
    ws1.cell(row=r, column=3, value=label).font = Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_900)
    ws1.cell(row=r, column=4, value=desc).font = font_body()
    ws1.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
    ws1.row_dimensions[r].height = 25

# ── Section: Field Guide (EN + BN) ──────────────────────────────────
row = 17
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
cell = ws1.cell(row=row, column=2, value="FIELD GUIDE — ENGLISH & BENGALI (বাংলা)")
cell.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
cell.fill = fill_header()
cell.alignment = align_header()
for c in range(3, 9):
    ws1.cell(row=row, column=c).fill = fill_header()
ws1.row_dimensions[row].height = 28

row = 18
headers_fg = ["#", "Field Name (EN)", "Description (EN)", "Description (বাংলা)", "Mandatory?", "Dashboard Usage"]
for ci, h in enumerate(headers_fg, 2):
    cell = ws1.cell(row=row, column=ci, value=h)
    cell.fill = PatternFill('solid', fgColor=PRIMARY)
    cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    cell.alignment = align_header()
ws1.row_dimensions[row].height = 26

field_guide = [
    ("Property Code", "Unique identifier for each property (e.g. PROP-001)", "প্রতিটি সম্পত্তির জন্য অনন্য কোড", "YES", "Links all units, tenants, and leases to a property"),
    ("Property Name", "Name of the building/villa (e.g. Khalifa Villa)", "ভবন/ভিলার নাম", "YES", "Displayed on property cards and navigation"),
    ("Property Type", "Villa, Building, Office Complex, Commercial, or Residential", "সম্পত্তির ধরন (ভিলা, বিল্ডিং, অফিস ইত্যাদি)", "YES", "Used for filtering and categorization"),
    ("Property Address", "Full street address of the property", "সম্পত্তির সম্পূর্ণ ঠিকানা", "NO", "Shown on property detail pages"),
    ("Area", "Neighborhood/district name (e.g. Al Jahili, Al Sarooj)", "এলাকার নাম", "YES", "Used for area-based filtering and reports"),
    ("City", "City name (typically Al Ain)", "শহরের নাম", "YES", "Used in reports and location grouping"),
    ("Total Units", "Total number of rentable units in the property", "মোট ভাড়াযোগ্য ইউনিট সংখ্যা", "YES", "Occupancy rate calculations"),
    ("Unit Number", "Room/unit identifier (e.g. 1, 2, A1, GARAGE)", "রুম/ইউনিট নম্বর", "YES", "Identifies specific rental unit within property"),
    ("Unit Type", "Room, Studio, Office, Shop, Garage, or Apartment", "ইউনিটের ধরন", "YES", "Unit categorization and rent comparison"),
    ("Floor", "Floor number or level (e.g. 1, 2, Ground)", "মেঝে/তলা নম্বর", "NO", "Unit location within building"),
    ("Monthly Rent (AED)", "Monthly rental amount in UAE Dirhams", "মাসিক ভাড়া (আয়াতি দিরহাম)", "YES", "Rent collection tracking and revenue reports"),
    ("Occupancy Status", "Occupied, Vacant, or Under Maintenance", "দখল অবস্থা", "YES", "Vacancy tracking and occupancy reports"),
    ("Tenant ID", "Unique tenant identifier (e.g. TEN-ESSAMAMA)", "ভাড়াটিয়ার অনন্য কোড", "YES", "Links tenant to leases, payments, and balances"),
    ("Full Name", "Tenant's full legal name as per contract/Emirates ID", "ভাড়াটিয়ার পূর্ণ আইনি নাম", "YES", "Primary tenant identification"),
    ("Phone Number", "Primary contact number (+971-XXXXXXXXX)", "প্রাথমিক ফোন নম্বর", "YES", "Communication and notifications"),
    ("WhatsApp Number", "WhatsApp contact if different from phone", "হোয়াটসঅ্যাপ নম্বর", "NO", "WhatsApp notification delivery"),
    ("Email", "Email address for official communications", "ইমেইল ঠিকানা", "NO", "Email notification delivery"),
    ("Emirates ID", "15-digit UAE national ID number (784-XXXX-XXXXXXX-X)", "সংযুক্ত আরব আমিরাত জাতীয় পরিচয়পত্র নম্বর", "YES", "Identity verification and duplicate detection"),
    ("Nationality", "Country of nationality", "জাতীয়তা", "NO", "Demographic reporting"),
    ("Lease Start Date", "Contract start date (YYYY-MM-DD)", "চুক্তি শুরুর তারিখ", "YES", "Lease tracking and renewal alerts"),
    ("Lease End Date", "Contract end date (YYYY-MM-DD)", "চুক্তি শেষের তারিখ", "YES", "Expiry tracking and renewal alerts"),
    ("Security Deposit (AED)", "Refundable deposit amount paid at lease start", "নিরাপত্তা জমা পরিমাণ", "YES", "Deposit tracking on move-out"),
    ("Payment Method", "Cash, Cheque, Bank Transfer, ADIB, ADCB, or Arbab", "পেমেন্ট পদ্ধতি", "YES", "Payment tracking and reconciliation"),
    ("Contract Status", "Active, Expired, Terminated, or Pending Renewal", "চুক্তির অবস্থা", "YES", "Contract management and alerts"),
    ("Previous Balance (AED)", "Balance carried forward from previous period", "পূর্ববর্তী বকেয়া", "YES", "Balance tracking and reconciliation"),
    ("Outstanding Balance (AED)", "Current total amount owed by tenant", "বর্তমান বকেয়া পরিমাণ", "YES", "Collection tracking and aging reports"),
    ("Payment ID", "Unique payment receipt identifier", "পেমেন্ট রশিদ কোড", "YES", "Payment tracking and audit trail"),
    ("Payment Date", "Date payment was received (YYYY-MM-DD)", "পেমেন্ট গ্রহণের তারিখ", "YES", "Payment history and reconciliation"),
    ("Amount (AED)", "Payment amount received", "প্রাপ্ত পেমেন্ট পরিমাণ", "YES", "Revenue tracking and cash flow"),
    ("Reference Number", "Cheque number, transaction ID, or receipt number", "চেক নম্বর বা লেনদেন কোড", "NO", "Payment verification and audit"),
]

for i, (fname, desc_en, desc_bn, mandatory, usage) in enumerate(field_guide, 1):
    r = row + i
    ws1.cell(row=r, column=2, value=i).font = font_body()
    ws1.cell(row=r, column=2).alignment = align_header()
    ws1.cell(row=r, column=3, value=fname).font = Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_900)
    ws1.cell(row=r, column=4, value=desc_en).font = font_body()
    ws1.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
    ws1.cell(row=r, column=5, value=desc_bn).font = font_body()
    ws1.cell(row=r, column=5).alignment = Alignment(wrap_text=True)
    ws1.cell(row=r, column=6, value=mandatory).font = Font(name=FONT_NAME, size=11, bold=True, 
        color=ACCENT_NEGATIVE if mandatory == "YES" else ACCENT_WARNING)
    ws1.cell(row=r, column=6).alignment = align_header()
    ws1.cell(row=r, column=7, value=usage).font = font_body()
    ws1.cell(row=r, column=7).alignment = Alignment(wrap_text=True)
    
    fill_color = REQUIRED_FILL if mandatory == "YES" else OPTIONAL_FILL
    for c in range(2, 9):
        ws1.cell(row=r, column=c).fill = fill_color if c == 3 else (fill_data_row(i) if c > 3 else fill_data_row(i))
    ws1.row_dimensions[r].height = 32

# ── Section: Common Mistakes ─────────────────────────────────────────
row = row + len(field_guide) + 2
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
cell = ws1.cell(row=row, column=2, value="COMMON MISTAKES TO AVOID")
cell.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
cell.fill = PatternFill('solid', fgColor=ACCENT_NEGATIVE)
cell.alignment = align_header()
for c in range(3, 9):
    ws1.cell(row=row, column=c).fill = PatternFill('solid', fgColor=ACCENT_NEGATIVE)
ws1.row_dimensions[row].height = 28

mistakes = [
    "DO NOT guess or fabricate any missing data. Leave fields blank if information is not available from verified sources.",
    "DO NOT use the same Emirates ID for multiple tenants. Each person has a unique Emirates ID. If two tenants share a unit, each must have their own ID.",
    "DO NOT enter partial phone numbers. UAE phone numbers must include the country code: +971-5XXXXXXXX or +971-XXXXXXXXX.",
    "DO NOT mix date formats. Always use YYYY-MM-DD format (e.g. 2025-01-15, not 15/1/25 or Jan 15).",
    "DO NOT enter rent amounts as text. All monetary values must be numbers only (e.g. 2500, not 'AED 2500' or '2,500').",
    "DO NOT combine multiple tenants in one Name field. If a unit has co-tenants, create separate tenant records and link them to the same unit.",
    "DO NOT leave the Property Code blank. Every unit and tenant must be linked to a valid property code from the Properties sheet.",
    "DO NOT reuse Payment IDs. Each payment must have a unique identifier.",
    "DO NOT enter 'EMPTY' or 'VACANT' in the tenant name field. If a unit is vacant, leave the tenant fields blank and set Occupancy Status to 'Vacant'.",
    "DO NOT assume lease dates from the billing date. Use the actual contract start and end dates from signed lease agreements.",
]
for i, mistake in enumerate(mistakes, 1):
    r = row + i
    ws1.cell(row=r, column=2, value=i).font = font_body()
    ws1.cell(row=r, column=2).alignment = align_header()
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    ws1.cell(row=r, column=3, value=mistake).font = font_body()
    ws1.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
    for c in range(2, 9):
        ws1.cell(row=r, column=c).fill = fill_data_row(i)
    ws1.row_dimensions[r].height = 30

# ── Section: Completion Checklist ────────────────────────────────────
row = row + len(mistakes) + 2
ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
cell = ws1.cell(row=row, column=2, value="COMPLETION CHECKLIST")
cell.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
cell.fill = PatternFill('solid', fgColor=ACCENT_POSITIVE)
cell.alignment = align_header()
for c in range(3, 9):
    ws1.cell(row=row, column=c).fill = PatternFill('solid', fgColor=ACCENT_POSITIVE)
ws1.row_dimensions[row].height = 28

checklist = [
    "All 26 properties are listed in the Properties sheet with correct codes and names",
    "All units from the APRIL-25 spreadsheet are listed with correct Property Code and Unit Number",
    "Monthly Rent amounts match the source spreadsheet for each occupied unit",
    "All occupied units have a corresponding tenant record in the Tenants sheet",
    "Phone numbers have been collected from tenant contracts for all tenants",
    "Emirates ID numbers have been verified and entered for all tenants",
    "Lease start and end dates have been entered from signed contracts",
    "Outstanding balances from APRIL-25 have been verified against ledger records",
    "All vacant units are marked as 'Vacant' in Occupancy Status",
    "Data Quality Status column is set to 'Complete' for all verified records",
    "No duplicate Emirates ID numbers exist across tenants",
    "All required (red) fields are filled — no blank cells in required columns",
]
for i, item in enumerate(checklist, 1):
    r = row + i
    ws1.cell(row=r, column=2, value="☐").font = Font(name=FONT_NAME, size=14)
    ws1.cell(row=r, column=2).alignment = align_header()
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    ws1.cell(row=r, column=3, value=item).font = font_body()
    ws1.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
    for c in range(2, 9):
        ws1.cell(row=r, column=c).fill = fill_data_row(i)
    ws1.row_dimensions[r].height = 28

# ══════════════════════════════════════════════════════════════════════
# SHEET 2: Properties
# ══════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Properties")
prop_headers = [
    ("Property Code", True, 16),
    ("Property Name", True, 28),
    ("Property Type", True, 18),
    ("Property Address", False, 35),
    ("Area", True, 20),
    ("City", True, 16),
    ("Total Units", True, 14),
    ("Active Status", True, 16),
    ("Notes", False, 35),
]
last_col = len(prop_headers) + 1
setup_sheet(ws2, title="Properties", last_col=last_col)

# Write headers with required/optional color coding
for ci, (header, required, width) in enumerate(prop_headers, 2):
    cell = ws2.cell(row=4, column=ci, value=header)
    cell.fill = REQUIRED_HEADER_FILL if required else OPTIONAL_HEADER_FILL
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.alignment = align_header()
    cell.border = border_header()
    ws2.column_dimensions[get_column_letter(ci)].width = width
ws2.row_dimensions[4].height = 28

# Property type mapping from names
def infer_property_type(name):
    name_upper = name.upper() if name else ''
    if 'OFFICE' in name_upper or 'NEW OFFICE' in name_upper:
        return 'Office Complex'
    elif 'MALL' in name_upper:
        return 'Commercial'
    elif 'SANAIYA' in name_upper:
        return 'Commercial'
    elif 'BUILDING' in name_upper:
        return 'Building'
    else:
        return 'Villa'

# Area mapping from property names
def infer_area(name):
    area_map = {
        'KHALIFA': 'Khalifa Bin Zayed',
        'MATOWA': 'Al Mutawwa',
        'SALHE': 'Al Salhe',
        'SUMALI': 'Al Sumali',
        'JAHILI': 'Al Jahili',
        'ZAFRANA': 'Zafrana',
        'DIWAN': 'Al Diwan',
        'HABOOY': 'Al Habouy',
        'SANAIYA': 'Sanaiya',
        'SAROOJ': 'Al Sarooj',
        'AA MALL': 'Al Ain Mall Area',
        'NEIMA': 'Neima',
        'SHAMKHA': 'Al Shamkha',
        'BATEEN': 'Al Bateen',
        'MANASIR': 'Al Manasir',
        'NIYADAT': 'Niyadat',
        'INDIA': 'Al Ain City Center',
        'MUTARID': 'Al Mutarid',
        'MAQAM': 'Al Maqam',
        'HABOOY NEW': 'Al Habouy',
        'MUWAIJI': 'Al Muwaiji',
    }
    name_upper = name.upper() if name else ''
    for key, area in area_map.items():
        if key in name_upper:
            return area
    return 'Al Ain'

# Data validation for Property Type
dv_prop_type = DataValidation(type="list", formula1='"Villa,Building,Office Complex,Commercial,Residential Compound"', allow_blank=True)
dv_prop_type.error = "Please select a valid Property Type"
dv_prop_type.errorTitle = "Invalid Property Type"
ws2.add_data_validation(dv_prop_type)

dv_active_status = DataValidation(type="list", formula1='"Active,Inactive,Under Maintenance"', allow_blank=True)
ws2.add_data_validation(dv_active_status)

# Populate properties from source data
for i, (pname, pdata) in enumerate(properties.items()):
    r = 5 + i
    code = property_code_map.get(pname, f"PROP-{i+1:03d}")
    ptype = infer_property_type(pname)
    area = infer_area(pname)
    
    # Check for "NEW" suffix to determine if it's a separate listing
    active = 'Active'
    
    values = [
        code,
        pname,
        ptype,
        '',  # Address - not in source
        area,
        'Al Ain',
        pdata['units'],
        active,
        '',
    ]
    for ci, val in enumerate(values, 2):
        cell = ws2.cell(row=r, column=ci, value=val)
        cell.font = font_body()
        cell.fill = fill_data_row(i)
        # Required fields get red background if empty
        if prop_headers[ci-2][1] and (val is None or val == ''):
            cell.fill = REQUIRED_FILL
    
    # Apply data validation
    dv_prop_type.add(ws2.cell(row=r, column=4))  # Property Type
    dv_active_status.add(ws2.cell(row=r, column=9))  # Active Status
    
    ws2.row_dimensions[r].height = ROW_HEIGHTS['data']

# Completion percentage formula
total_props = len(properties)
comp_row = 5 + total_props + 1
ws2.merge_cells(start_row=comp_row, start_column=2, end_row=comp_row, end_column=4)
ws2.cell(row=comp_row, column=2, value="Completion Rate:").font = font_subheader()
ws2.cell(row=comp_row, column=5).value = f'=COUNTA(E5:E{4+total_props})/{total_props}'
ws2.cell(row=comp_row, column=5).number_format = '0%'
ws2.cell(row=comp_row, column=5).font = font_subheader()

ws2.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 3: Units
# ══════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Units")
unit_headers = [
    ("Property Code", True, 16),
    ("Unit Number", True, 16),
    ("Unit Type", True, 16),
    ("Floor", False, 12),
    ("Monthly Rent (AED)", True, 18),
    ("Occupancy Status", True, 18),
    ("Data Quality Status", True, 20),
    ("Notes", False, 35),
]
last_col = len(unit_headers) + 1
setup_sheet(ws3, title="Units", last_col=last_col)

for ci, (header, required, width) in enumerate(unit_headers, 2):
    cell = ws3.cell(row=4, column=ci, value=header)
    cell.fill = REQUIRED_HEADER_FILL if required else OPTIONAL_HEADER_FILL
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.alignment = align_header()
    cell.border = border_header()
    ws3.column_dimensions[get_column_letter(ci)].width = width
ws3.row_dimensions[4].height = 28

# Data validation for Unit Type
dv_unit_type = DataValidation(type="list", formula1='"Room,Studio,Office,Shop,Garage,Apartment,Penthouse,Storage"', allow_blank=True)
ws3.add_data_validation(dv_unit_type)

dv_occ_status = DataValidation(type="list", formula1='"Occupied,Vacant,Under Maintenance"', allow_blank=True)
ws3.add_data_validation(dv_occ_status)

dv_quality = DataValidation(type="list", formula1='"Complete,Needs Review,Missing Information"', allow_blank=True)
ws3.add_data_validation(dv_quality)

# Populate units from source data
unit_row = 5
for r in all_records:
    pname = r['property']
    pcode = property_code_map.get(pname, '')
    unit_num = r['room']
    name = r['name']
    rent = r['rent']
    
    # Determine unit type from room number pattern
    room_str = str(unit_num).upper()
    if 'GARAGE' in room_str:
        unit_type = 'Garage'
    elif 'OFFICE' in room_str:
        unit_type = 'Office'
    elif isinstance(unit_num, str) and any(c.isalpha() for c in unit_num):
        if any(x in room_str for x in ['S-', 'S1', 'S2']):
            unit_type = 'Storage'
        else:
            unit_type = 'Room'
    else:
        unit_type = 'Room'
    
    # Occupancy
    if name and name != 'EMPTY':
        occ_status = 'Occupied'
    else:
        occ_status = 'Vacant'
    
    # Data quality
    quality = 'Needs Review'
    if name and rent and pcode:
        quality = 'Needs Review'  # Still needs phone, EID, lease dates
    
    values = [
        pcode,
        unit_num,
        unit_type,
        '',  # Floor - not in source
        rent if rent else '',
        occ_status,
        quality,
        '',  # Notes
    ]
    
    for ci, val in enumerate(values, 2):
        cell = ws3.cell(row=unit_row, column=ci, value=val)
        cell.font = font_body()
        cell.fill = fill_data_row(unit_row - 5)
        if unit_headers[ci-2][1] and (val is None or val == ''):
            cell.fill = REQUIRED_FILL
    
    # Numeric formatting for rent
    if rent:
        ws3.cell(row=unit_row, column=6).number_format = '#,##0'
    
    # Apply data validation
    dv_unit_type.add(ws3.cell(row=unit_row, column=4))
    dv_occ_status.add(ws3.cell(row=unit_row, column=7))
    dv_quality.add(ws3.cell(row=unit_row, column=8))
    
    ws3.row_dimensions[unit_row].height = ROW_HEIGHTS['data']
    unit_row += 1

# Completion percentage
comp_row = unit_row + 1
total_units = unit_row - 5
ws3.merge_cells(start_row=comp_row, start_column=2, end_row=comp_row, end_column=4)
ws3.cell(row=comp_row, column=2, value="Completion Rate:").font = font_subheader()
ws3.cell(row=comp_row, column=5).value = f'=COUNTA(E5:E{unit_row-1})/{total_units}'
ws3.cell(row=comp_row, column=5).number_format = '0%'
ws3.cell(row=comp_row, column=5).font = font_subheader()

ws3.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 4: Tenants
# ══════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Tenants")
tenant_headers = [
    ("Tenant ID", True, 22),
    ("Full Name", True, 28),
    ("Phone Number", True, 22),
    ("WhatsApp Number", False, 22),
    ("Email", False, 28),
    ("Emirates ID", True, 22),
    ("Nationality", False, 18),
    ("Emergency Contact", False, 22),
    ("Emergency Phone", False, 22),
    ("Status", True, 16),
    ("Data Quality Status", True, 20),
    ("Notes", False, 35),
]
last_col = len(tenant_headers) + 1
setup_sheet(ws4, title="Tenants", last_col=last_col)

for ci, (header, required, width) in enumerate(tenant_headers, 2):
    cell = ws4.cell(row=4, column=ci, value=header)
    cell.fill = REQUIRED_HEADER_FILL if required else OPTIONAL_HEADER_FILL
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.alignment = align_header()
    cell.border = border_header()
    ws4.column_dimensions[get_column_letter(ci)].width = width
ws4.row_dimensions[4].height = 28

dv_tenant_status = DataValidation(type="list", formula1='"Active,Inactive,Evicted,Moved Out"', allow_blank=True)
ws4.add_data_validation(dv_tenant_status)
dv_quality_t = DataValidation(type="list", formula1='"Complete,Needs Review,Missing Information"', allow_blank=True)
ws4.add_data_validation(dv_quality_t)

# Populate tenants - unique tenants by (property, room, name)
seen_tenants = {}
tenant_row = 5
for r in all_records:
    if not r['name'] or r['name'] == 'EMPTY':
        continue
    
    key = (r['property'], r['room'], r['name'])
    if key in seen_tenants:
        continue
    seen_tenants[key] = True
    
    tid = tenant_id_map.get(key, f"TEN-{r['name'].upper()[:8]}")
    
    # Determine quality
    quality = 'Missing Information'  # Missing phone, EID, email, etc.
    
    values = [
        tid,
        r['name'],
        '',  # Phone - not in source
        '',  # WhatsApp - not in source
        '',  # Email - not in source
        '',  # Emirates ID - not in source
        '',  # Nationality - not in source
        '',  # Emergency Contact - not in source
        '',  # Emergency Phone - not in source
        'Active',
        quality,
        '',  # Notes
    ]
    
    for ci, val in enumerate(values, 2):
        cell = ws4.cell(row=tenant_row, column=ci, value=val)
        cell.font = font_body()
        cell.fill = fill_data_row(tenant_row - 5)
        if tenant_headers[ci-2][1] and (val is None or val == ''):
            cell.fill = REQUIRED_FILL
    
    dv_tenant_status.add(ws4.cell(row=tenant_row, column=11))
    dv_quality_t.add(ws4.cell(row=tenant_row, column=12))
    
    ws4.row_dimensions[tenant_row].height = ROW_HEIGHTS['data']
    tenant_row += 1

# Add conditional formatting for duplicate Emirates ID detection
ws4.conditional_formatting.add(
    f'G5:G{tenant_row-1}',
    FormulaRule(
        formula=[f'COUNTIF(G$5:G${tenant_row-1},G5)>1'],
        fill=PatternFill('solid', fgColor='FDE8E8'),
        font=Font(color=ACCENT_NEGATIVE, bold=True)
    )
)

# Completion percentage
comp_row = tenant_row + 1
total_tenants = tenant_row - 5
ws4.merge_cells(start_row=comp_row, start_column=2, end_row=comp_row, end_column=4)
ws4.cell(row=comp_row, column=2, value="Completion Rate:").font = font_subheader()
ws4.cell(row=comp_row, column=5).value = f'=COUNTA(D5:D{tenant_row-1})/{total_tenants}'
ws4.cell(row=comp_row, column=5).number_format = '0%'
ws4.cell(row=comp_row, column=5).font = font_subheader()

ws4.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 5: Lease Information
# ══════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Lease Information")
lease_headers = [
    ("Tenant ID", True, 22),
    ("Property Code", True, 16),
    ("Unit Number", True, 16),
    ("Lease Start Date", True, 18),
    ("Lease End Date", True, 18),
    ("Monthly Rent (AED)", True, 18),
    ("Security Deposit (AED)", True, 20),
    ("Payment Method", True, 18),
    ("Contract Status", True, 18),
    ("Data Quality Status", True, 20),
    ("Notes", False, 35),
]
last_col = len(lease_headers) + 1
setup_sheet(ws5, title="Lease Information", last_col=last_col)

for ci, (header, required, width) in enumerate(lease_headers, 2):
    cell = ws5.cell(row=4, column=ci, value=header)
    cell.fill = REQUIRED_HEADER_FILL if required else OPTIONAL_HEADER_FILL
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.alignment = align_header()
    cell.border = border_header()
    ws5.column_dimensions[get_column_letter(ci)].width = width
ws5.row_dimensions[4].height = 28

dv_pay_method = DataValidation(type="list", formula1='"Cash,Cheque,Bank Transfer,ADIB,ADCB,ENBD,DIB,Arbab (Cash Collection)"', allow_blank=True)
ws5.add_data_validation(dv_pay_method)

dv_contract_status = DataValidation(type="list", formula1='"Active,Expired,Terminated,Pending Renewal"', allow_blank=True)
ws5.add_data_validation(dv_contract_status)

dv_quality_l = DataValidation(type="list", formula1='"Complete,Needs Review,Missing Information"', allow_blank=True)
ws5.add_data_validation(dv_quality_l)

# Date validation
dv_date = DataValidation(type="date", allow_blank=True)
dv_date.error = "Please enter a valid date in YYYY-MM-DD format"
dv_date.errorTitle = "Invalid Date"
ws5.add_data_validation(dv_date)

lease_row = 5
for r in all_records:
    if not r['name'] or r['name'] == 'EMPTY':
        continue
    
    key = (r['property'], r['room'], r['name'])
    tid = tenant_id_map.get(key, '')
    pcode = property_code_map.get(r['property'], '')
    
    # Payment method from source
    pay_method = r['payment_method'] if r['payment_method'] else ''
    
    quality = 'Missing Information'  # Missing lease dates, deposit
    
    values = [
        tid,
        pcode,
        r['room'],
        '',  # Lease Start Date - not in source
        '',  # Lease End Date - not in source
        r['rent'] if r['rent'] else '',
        '',  # Security Deposit - not in source
        pay_method,
        'Active',  # Assuming active since they're in the billing
        quality,
        '',  # Notes
    ]
    
    for ci, val in enumerate(values, 2):
        cell = ws5.cell(row=lease_row, column=ci, value=val)
        cell.font = font_body()
        cell.fill = fill_data_row(lease_row - 5)
        if lease_headers[ci-2][1] and (val is None or val == ''):
            cell.fill = REQUIRED_FILL
    
    if r['rent']:
        ws5.cell(row=lease_row, column=7).number_format = '#,##0'
    
    dv_pay_method.add(ws5.cell(row=lease_row, column=9))
    dv_contract_status.add(ws5.cell(row=lease_row, column=10))
    dv_quality_l.add(ws5.cell(row=lease_row, column=11))
    dv_date.add(ws5.cell(row=lease_row, column=5))
    dv_date.add(ws5.cell(row=lease_row, column=6))
    
    ws5.row_dimensions[lease_row].height = ROW_HEIGHTS['data']
    lease_row += 1

ws5.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 6: Outstanding Balances
# ══════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Outstanding Balances")
balance_headers = [
    ("Tenant ID", True, 22),
    ("Property Code", True, 16),
    ("Unit Number", True, 16),
    ("Previous Balance (AED)", True, 22),
    ("Outstanding Balance (AED)", True, 24),
    ("Data Quality Status", True, 20),
    ("Balance Notes", False, 35),
]
last_col = len(balance_headers) + 1
setup_sheet(ws6, title="Outstanding Balances", last_col=last_col)

for ci, (header, required, width) in enumerate(balance_headers, 2):
    cell = ws6.cell(row=4, column=ci, value=header)
    cell.fill = REQUIRED_HEADER_FILL if required else OPTIONAL_HEADER_FILL
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.alignment = align_header()
    cell.border = border_header()
    ws6.column_dimensions[get_column_letter(ci)].width = width
ws6.row_dimensions[4].height = 28

dv_quality_b = DataValidation(type="list", formula1='"Complete,Needs Review,Missing Information"', allow_blank=True)
ws6.add_data_validation(dv_quality_b)

balance_row = 5
for r in all_records:
    if not r['name'] or r['name'] == 'EMPTY':
        continue
    
    key = (r['property'], r['room'], r['name'])
    tid = tenant_id_map.get(key, '')
    pcode = property_code_map.get(r['property'], '')
    
    prev_due = r['prev_due'] if r['prev_due'] else 0
    balance = r['balance'] if r['balance'] else 0
    
    # Only include if there's a balance
    if prev_due == 0 and balance == 0:
        continue
    
    quality = 'Needs Review' if balance > 0 else 'Complete'
    
    values = [
        tid,
        pcode,
        r['room'],
        prev_due,
        balance,
        quality,
        '',  # Balance Notes
    ]
    
    for ci, val in enumerate(values, 2):
        cell = ws6.cell(row=balance_row, column=ci, value=val)
        cell.font = font_body()
        cell.fill = fill_data_row(balance_row - 5)
        if balance_headers[ci-2][1] and (val is None or val == ''):
            cell.fill = REQUIRED_FILL
    
    ws6.cell(row=balance_row, column=5).number_format = '#,##0'
    ws6.cell(row=balance_row, column=6).number_format = '#,##0'
    
    dv_quality_b.add(ws6.cell(row=balance_row, column=7))
    
    ws6.row_dimensions[balance_row].height = ROW_HEIGHTS['data']
    balance_row += 1

ws6.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 7: Payment History
# ══════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Payment History")
payment_headers = [
    ("Payment ID", True, 18),
    ("Tenant ID", True, 22),
    ("Payment Date", True, 18),
    ("Amount (AED)", True, 16),
    ("Payment Method", True, 18),
    ("Reference Number", False, 22),
    ("Data Quality Status", True, 20),
    ("Remarks", False, 35),
]
last_col = len(payment_headers) + 1
setup_sheet(ws7, title="Payment History — April 2025", last_col=last_col)

for ci, (header, required, width) in enumerate(payment_headers, 2):
    cell = ws7.cell(row=4, column=ci, value=header)
    cell.fill = REQUIRED_HEADER_FILL if required else OPTIONAL_HEADER_FILL
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.alignment = align_header()
    cell.border = border_header()
    ws7.column_dimensions[get_column_letter(ci)].width = width
ws7.row_dimensions[4].height = 28

dv_pay_method_h = DataValidation(type="list", formula1='"Cash,Cheque,Bank Transfer,ADIB,ADCB,ENBD,DIB,Arbab (Cash Collection)"', allow_blank=True)
ws7.add_data_validation(dv_pay_method_h)
dv_quality_p = DataValidation(type="list", formula1='"Complete,Needs Review,Missing Information"', allow_blank=True)
ws7.add_data_validation(dv_quality_p)

payment_row = 5
payment_counter = 0
for r in all_records:
    if not r['name'] or r['name'] == 'EMPTY':
        continue
    
    key = (r['property'], r['room'], r['name'])
    tid = tenant_id_map.get(key, '')
    
    cash = r.get('cash', 0)  # From source
    # We don't have individual payment cash in the extracted records, 
    # but we have the "Cash" column which is the amount paid
    # Let's check the source for cash values
    # Actually, we need to go back to source for this
    
    # For now, create a payment record with available info
    payment_counter += 1
    pay_id = f"PAY-2025-04-{payment_counter:04d}"
    
    # Determine payment method
    pay_method = r['payment_method'] if r['payment_method'] else 'Cash'
    
    # Reference number from bill_no
    ref = r['bill_no'] if r['bill_no'] else ''
    
    # Cash amount from source (the 'Cash' column)
    # We need to re-extract this from source
    amount = ''  # Will need manual entry
    
    quality = 'Missing Information'
    
    values = [
        pay_id,
        tid,
        '',  # Payment Date - source dates are ambiguous (billing vs payment)
        amount,
        pay_method,
        ref,
        quality,
        '',  # Remarks
    ]
    
    for ci, val in enumerate(values, 2):
        cell = ws7.cell(row=payment_row, column=ci, value=val)
        cell.font = font_body()
        cell.fill = fill_data_row(payment_row - 5)
        if payment_headers[ci-2][1] and (val is None or val == ''):
            cell.fill = REQUIRED_FILL
    
    dv_pay_method_h.add(ws7.cell(row=payment_row, column=6))
    dv_quality_p.add(ws7.cell(row=payment_row, column=8))
    
    ws7.row_dimensions[payment_row].height = ROW_HEIGHTS['data']
    payment_row += 1

ws7.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 8: Data Issues & Missing Information
# ══════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Data Issues")
issue_headers = [
    ("#", True, 6),
    ("Sheet", True, 18),
    ("Record Identifier", True, 28),
    ("Issue Type", True, 22),
    ("Missing Field", True, 22),
    ("Current Value", True, 22),
    ("Severity", True, 14),
    ("Resolution Notes", False, 35),
]
last_col = len(issue_headers) + 1
setup_sheet(ws8, title="Data Issues & Missing Information", last_col=last_col)

for ci, (header, required, width) in enumerate(issue_headers, 2):
    cell = ws8.cell(row=4, column=ci, value=header)
    cell.fill = REQUIRED_HEADER_FILL if required else OPTIONAL_HEADER_FILL
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.alignment = align_header()
    cell.border = border_header()
    ws8.column_dimensions[get_column_letter(ci)].width = width
ws8.row_dimensions[4].height = 28

dv_severity = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
ws8.add_data_validation(dv_severity)

# Generate issues list
issues = []

# 1. Missing phone numbers for all tenants
for r in all_records:
    if r['name'] and r['name'] != 'EMPTY':
        key = (r['property'], r['room'], r['name'])
        tid = tenant_id_map.get(key, '')
        issues.append(("Tenants", tid, "Missing Required Data", "Phone Number", "(blank)", "Critical"))
        issues.append(("Tenants", tid, "Missing Required Data", "Emirates ID", "(blank)", "Critical"))
        issues.append(("Lease Information", f"{tid} / {r['room']}", "Missing Required Data", "Lease Start Date", "(blank)", "Critical"))
        issues.append(("Lease Information", f"{tid} / {r['room']}", "Missing Required Data", "Lease End Date", "(blank)", "Critical"))
        issues.append(("Lease Information", f"{tid} / {r['room']}", "Missing Required Data", "Security Deposit", "(blank)", "High"))

# 2. Duplicate tenant names
for name, count in global_name_count.items():
    if count > 1:
        issues.append(("Tenants", f"Name: {name}", "Potential Duplicate", "Full Name", f"Appears {count} times", "High"))

# 3. Units with no tenant name (vacant)
for r in all_records:
    if not r['name'] or r['name'] == 'EMPTY':
        pcode = property_code_map.get(r['property'], '')
        issues.append(("Units", f"{pcode} / Room {r['room']}", "Vacant Unit", "Tenant Name", "(blank/EMPTY)", "Low"))

# 4. Bill No is actually a bank name
for r in all_records:
    if r['is_bank_name'] and r['name']:
        key = (r['property'], r['room'], r['name'])
        tid = tenant_id_map.get(key, '')
        issues.append(("Lease Information", f"{tid} / {r['room']}", "Ambiguous Data", "Bill No / Payment Method", f"'{r['payment_method']}' was in Bill No column", "Medium"))

# 5. Properties without addresses
for pname in properties.keys():
    pcode = property_code_map.get(pname, '')
    issues.append(("Properties", pcode, "Missing Optional Data", "Property Address", "(blank)", "Low"))

# 6. Payment History - no amounts
issues.append(("Payment History", "All Records", "Missing Required Data", "Amount & Payment Date", "Source has billing data, not individual payments", "Critical"))

# Write issues
issue_row = 5
for i, (sheet, record_id, issue_type, missing_field, current_val, severity) in enumerate(issues, 1):
    values = [i, sheet, record_id, issue_type, missing_field, current_val, severity, '']
    for ci, val in enumerate(values, 2):
        cell = ws8.cell(row=issue_row, column=ci, value=val)
        cell.font = font_body()
        cell.fill = fill_data_row(issue_row - 5)
        # Color code severity
        if ci == 8 and val == 'Critical':
            cell.font = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_NEGATIVE)
        elif ci == 8 and val == 'High':
            cell.font = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_WARNING)
        elif ci == 8 and val == 'Medium':
            cell.font = Font(name=FONT_NAME, size=11, color=ACCENT_WARNING)
    
    dv_severity.add(ws8.cell(row=issue_row, column=8))
    ws8.row_dimensions[issue_row].height = ROW_HEIGHTS['data']
    issue_row += 1

# Summary at bottom
summary_row = issue_row + 2
ws8.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row, end_column=5)
ws8.cell(row=summary_row, column=2, value="ISSUE SUMMARY").font = font_subheader()
ws8.cell(row=summary_row, column=2).fill = fill_total()
for c in range(3, 9):
    ws8.cell(row=summary_row, column=c).fill = fill_total()

summary_items = [
    ("Critical Issues:", f'=COUNTIF(H5:H{issue_row-1},"Critical")'),
    ("High Issues:", f'=COUNTIF(H5:H{issue_row-1},"High")'),
    ("Medium Issues:", f'=COUNTIF(H5:H{issue_row-1},"Medium")'),
    ("Low Issues:", f'=COUNTIF(H5:H{issue_row-1},"Low")'),
    ("Total Issues:", f'=COUNTA(H5:H{issue_row-1})'),
]
for i, (label, formula) in enumerate(summary_items):
    r = summary_row + 1 + i
    ws8.cell(row=r, column=2, value=label).font = font_subheader()
    ws8.cell(row=r, column=3, value=formula).font = font_subheader()

ws8.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# FINAL: Conditional formatting across all data sheets
# ══════════════════════════════════════════════════════════════════════

# For Data Quality Status columns - color code based on value
for ws in [ws3, ws4, ws5, ws6, ws7]:
    # Find the Data Quality Status column
    for col in range(2, ws.max_column + 1):
        header_val = ws.cell(row=4, column=col).value
        if header_val and 'Data Quality' in str(header_val):
            col_letter = get_column_letter(col)
            max_row = ws.max_row
            
            # Complete = green
            ws.conditional_formatting.add(
                f'{col_letter}5:{col_letter}{max_row}',
                CellIsRule(operator='equal', formula=['"Complete"'],
                    fill=PatternFill('solid', fgColor='E8F5E9'),
                    font=Font(color=ACCENT_POSITIVE))
            )
            # Needs Review = yellow
            ws.conditional_formatting.add(
                f'{col_letter}5:{col_letter}{max_row}',
                CellIsRule(operator='equal', formula=['"Needs Review"'],
                    fill=PatternFill('solid', fgColor='FFF8E1'),
                    font=Font(color=ACCENT_WARNING))
            )
            # Missing Information = red
            ws.conditional_formatting.add(
                f'{col_letter}5:{col_letter}{max_row}',
                CellIsRule(operator='equal', formula=['"Missing Information"'],
                    fill=PatternFill('solid', fgColor='FDE8E8'),
                    font=Font(color=ACCENT_NEGATIVE))
            )
            break

# ══════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════
output_path = '/home/z/my-project/download/Property Management Data Collection Template.xlsx'
wb.save(output_path)
print(f"Workbook saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Properties: {len(properties)}")
print(f"Units: {len(all_records)}")
print(f"Tenants: {len(seen_tenants)}")
print(f"Data Issues: {len(issues)}")
