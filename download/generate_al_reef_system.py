#!/usr/bin/env python3
"""
Real Estate Emperor — Real Estate Management System
====================================================
AED 2,000 value professional empty template.
10-sheet comprehensive RE management system.
Bottega palette (dark forest green 2D4A3E).
"""

import sys
import os

# Add skills path so we can import base
sys.path.insert(0, "/home/z/my-project/skills/xlsx")

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# Import base module (import the module, not individual values, so palette switch works)
import templates.base as base

# ── Activate Bottega palette BEFORE anything else ──
base.use_palette_explicit("bottega")

# ── Now pull values AFTER palette is active ──
# Functions (these read from module globals at call time, so they're fine)
use_palette_explicit = base.use_palette_explicit
setup_sheet = base.setup_sheet
style_header_row = base.style_header_row
style_data_row = base.style_data_row
style_total_row = base.style_total_row
font_title = base.font_title
font_header = base.font_header
font_body = base.font_body
font_caption = base.font_caption
font_kpi = base.font_kpi
font_kpi_label = base.font_kpi_label
font_subheader = base.font_subheader
fill_header = base.fill_header
fill_total = base.fill_total
fill_data_row = base.fill_data_row
border_header = base.border_header
border_total = base.border_total
align_title = base.align_title
align_header = base.align_header
align_number = base.align_number
align_text = base.align_text
align_date = base.align_date
COLUMN_WIDTHS = base.COLUMN_WIDTHS
ROW_HEIGHTS = base.ROW_HEIGHTS
FORMATS = base.FORMATS
create_bar_chart = base.create_bar_chart
create_line_chart = base.create_line_chart
create_pie_chart = base.create_pie_chart
setup_chart_titles = base.setup_chart_titles
apply_chart_colors = base.apply_chart_colors
apply_pie_colors = base.apply_pie_colors
make_chart_title = base.make_chart_title

# Color tokens — read AFTER palette switch so they're bottega values
PRIMARY = base.PRIMARY
PRIMARY_LIGHT = base.PRIMARY_LIGHT
SECONDARY = base.SECONDARY
ACCENT_POSITIVE = base.ACCENT_POSITIVE
ACCENT_NEGATIVE = base.ACCENT_NEGATIVE
ACCENT_WARNING = base.ACCENT_WARNING
NEUTRAL_900 = base.NEUTRAL_900
NEUTRAL_600 = base.NEUTRAL_600
NEUTRAL_200 = base.NEUTRAL_200
NEUTRAL_100 = base.NEUTRAL_100
NEUTRAL_50 = base.NEUTRAL_50
NEUTRAL_0 = base.NEUTRAL_0
HEADER_TEXT = base.HEADER_TEXT
CHART_COLORS = base.CHART_COLORS
CF_POSITIVE_FILL = base.CF_POSITIVE_FILL
CF_POSITIVE_FONT = base.CF_POSITIVE_FONT
CF_NEGATIVE_FILL = base.CF_NEGATIVE_FILL
CF_NEGATIVE_FONT = base.CF_NEGATIVE_FONT
CF_WARNING_FILL = base.CF_WARNING_FILL
CF_WARNING_FONT = base.CF_WARNING_FONT

# ── Constants ──
OUTPUT_PATH = "/home/z/my-project/download/Real_Estate_Emperor_System.xlsx"
COMPANY = "Real Estate Emperor"
CURRENCY_FMT = "#,##0.00"
DATE_FMT = "YYYY-MM-DD"
PCT_FMT = "0.0%"

# ── Tab color from bottega PRIMARY ──
TAB_COLOR = PRIMARY  # Should be 2D4A3E

# ── Helper: set print area, landscape, fit to width ──
def setup_print(ws, last_col_letter, last_row, title_rows="2:4"):
    ws.print_area = f"A1:{last_col_letter}{last_row}"
    ws.print_title_rows = title_rows
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

# ── Helper: add conditional formatting fills ──
def cf_cell_is(ws, range_str, operator, formula, fill, font):
    """Add CellIsRule conditional formatting."""
    rule = CellIsRule(operator=operator, formula=[formula], fill=fill, font=font)
    ws.conditional_formatting.add(range_str, rule)

def cf_formula(ws, range_str, formula, fill, font):
    """Add FormulaRule conditional formatting."""
    rule = FormulaRule(formula=[formula], fill=fill, font=font)
    ws.conditional_formatting.add(range_str, rule)


# ================================================================
# CREATE WORKBOOK
# ================================================================
wb = Workbook()
wb.properties.creator = "Z.ai"

# ================================================================
# SHEET 1: DASHBOARD
# ================================================================
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_dash.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_dash, title="Real Estate Emperor — Real Estate Dashboard", last_col=14)

# ── KPI Cards Row ──
# Row 4-6: 6 KPIs side by side starting at B4
# Each KPI: value (row 5) on top, label (row 6) below
# KPI columns: B-C, D-E, F-G, H-I, J-K, L-M

kpi_data = [
    ("Total Units", "=IFERROR(COUNTA('Property Registry'!B5:B154)-COUNTBLANK('Property Registry'!B5:B154),0)"),
    ("Occupied Units", '=IFERROR(COUNTIF(\'Property Registry\'!I5:I154,"Rented"),0)'),
    ("Vacant Units", '=IFERROR(COUNTIF(\'Property Registry\'!I5:I154,"Vacant"),0)'),
    ("Occupancy Rate", '=IFERROR(COUNTIF(\'Property Registry\'!I5:I154,"Rented")/COUNTA(\'Property Registry\'!B5:B154),0)'),
    ("Monthly Revenue", "=IFERROR(SUM('Property Registry'!G5:G154),0)"),
    ("Net Profit", "=IFERROR('Profit & Loss'!C33,0)"),
]

# Set row heights for KPI area
ws_dash.row_dimensions[4].height = 8  # spacer
ws_dash.row_dimensions[5].height = 38  # KPI value
ws_dash.row_dimensions[6].height = 20  # KPI label

for i, (label, formula) in enumerate(kpi_data):
    col_val = 2 + i * 2    # B, D, F, H, J, L
    col_lbl = 3 + i * 2    # C, E, G, I, K, M

    # Merge two columns for each KPI
    ws_dash.merge_cells(start_row=5, start_column=col_val, end_row=5, end_column=col_lbl)
    ws_dash.merge_cells(start_row=6, start_column=col_val, end_row=6, end_column=col_lbl)

    # KPI Value cell
    cell_val = ws_dash.cell(row=5, column=col_val, value=formula)
    cell_val.font = font_kpi()
    cell_val.alignment = Alignment(horizontal="center", vertical="bottom")
    cell_val.fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    if label == "Occupancy Rate":
        cell_val.number_format = PCT_FMT
    elif label in ("Monthly Revenue", "Net Profit"):
        cell_val.number_format = CURRENCY_FMT

    # Also style the merged partner
    for c in [col_val, col_lbl]:
        ws_dash.cell(row=5, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)

    # KPI Label cell
    cell_lbl = ws_dash.cell(row=6, column=col_val, value=label)
    cell_lbl.font = font_kpi_label()
    cell_lbl.alignment = Alignment(horizontal="center", vertical="top")

# ── Chart Data Areas (hidden rows for chart references) ──
# We'll place chart source data starting at row 8 for convenience
# Chart a: Occupancy Rate by Building (rows 8-13)
ws_dash.cell(row=8, column=2, value="Building").font = font_caption()
ws_dash.cell(row=8, column=3, value="Occupancy Rate").font = font_caption()
buildings = ["Building A", "Building B", "Building C", "Building D"]
for idx, b in enumerate(buildings):
    r = 9 + idx
    ws_dash.cell(row=r, column=2, value=b).font = font_caption()
    ws_dash.cell(row=r, column=3,
                 value=f'=IFERROR(COUNTIF(\'Property Registry\'!C5:C154,"{b}")/COUNTIF(\'Property Registry\'!C5:C154,"{b}"),0)').font = font_caption()
    ws_dash.cell(row=r, column=3).number_format = PCT_FMT

# Chart b: Monthly Revenue Trend (rows 15-27)
ws_dash.cell(row=15, column=2, value="Month").font = font_caption()
ws_dash.cell(row=15, column=3, value="Revenue").font = font_caption()
months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
for idx, m in enumerate(months):
    r = 16 + idx
    ws_dash.cell(row=r, column=2, value=m).font = font_caption()
    ws_dash.cell(row=r, column=3,
                 value=f"=IFERROR('Revenue Analysis'!C{5+idx},0)").font = font_caption()
    ws_dash.cell(row=r, column=3).number_format = CURRENCY_FMT

# Chart c: Expense Breakdown (rows 29-42)
ws_dash.cell(row=29, column=2, value="Category").font = font_caption()
ws_dash.cell(row=29, column=3, value="Amount").font = font_caption()
expense_cats = [
    "Manpower/Staff", "Municipality Fees", "Maintenance", "Leasing Commission",
    "Insurance", "Utilities", "Marketing", "Legal", "Cleaning",
    "Security", "Pest Control", "Elevator Service", "Other"
]
for idx, cat in enumerate(expense_cats):
    r = 30 + idx
    ws_dash.cell(row=r, column=2, value=cat).font = font_caption()
    ws_dash.cell(row=r, column=3,
                 value=f'=IFERROR(SUMPRODUCT((\'Expenses\'!C5:C304="{cat}")*(\'Expenses\'!E5:E304)),0)').font = font_caption()
    ws_dash.cell(row=r, column=3).number_format = CURRENCY_FMT

# Chart d: Unit Type Distribution (rows 44-50)
ws_dash.cell(row=44, column=2, value="Unit Type").font = font_caption()
ws_dash.cell(row=44, column=3, value="Count").font = font_caption()
unit_types = ["Studio", "1 Bedroom", "2 Bedroom", "3 Bedroom", "Shop", "Office"]
for idx, ut in enumerate(unit_types):
    r = 45 + idx
    ws_dash.cell(row=r, column=2, value=ut).font = font_caption()
    ws_dash.cell(row=r, column=3,
                 value=f'=IFERROR(COUNTIF(\'Property Registry\'!E5:E154,"{ut}"),0)').font = font_caption()

# ── Create Charts ──
# Chart a: Occupancy Rate by Building — Bar chart
chart_a = create_bar_chart(width=18, height=11)
data_a = Reference(ws_dash, min_col=3, min_row=8, max_row=12)
cats_a = Reference(ws_dash, min_col=2, min_row=9, max_row=12)
chart_a.add_data(data_a, titles_from_data=True)
chart_a.set_categories(cats_a)
setup_chart_titles(chart_a, title="Occupancy Rate by Building", y_title="Rate", x_title="Building")
apply_chart_colors(chart_a)
ws_dash.add_chart(chart_a, "B53")

# Chart b: Monthly Revenue Trend — Line chart
chart_b = create_line_chart(width=18, height=11)
data_b = Reference(ws_dash, min_col=3, min_row=15, max_row=27)
cats_b = Reference(ws_dash, min_col=2, min_row=16, max_row=27)
chart_b.add_data(data_b, titles_from_data=True)
chart_b.set_categories(cats_b)
setup_chart_titles(chart_b, title="Monthly Revenue Trend", y_title="Revenue (AED)", x_title="Month")
apply_chart_colors(chart_b)
ws_dash.add_chart(chart_b, "B70")

# Chart c: Expense Breakdown — Pie chart
chart_c = create_pie_chart(width=14, height=11)
data_c = Reference(ws_dash, min_col=3, min_row=29, max_row=42)
cats_c = Reference(ws_dash, min_col=2, min_row=30, max_row=42)
chart_c.add_data(data_c, titles_from_data=True)
chart_c.set_categories(cats_c)
setup_chart_titles(chart_c, title="Expense Breakdown")
apply_pie_colors(chart_c, count=len(expense_cats))
ws_dash.add_chart(chart_c, "B87")

# Chart d: Unit Type Distribution — Pie chart
chart_d = create_pie_chart(width=14, height=11)
data_d = Reference(ws_dash, min_col=3, min_row=44, max_row=50)
cats_d = Reference(ws_dash, min_col=2, min_row=45, max_row=50)
chart_d.add_data(data_d, titles_from_data=True)
chart_d.set_categories(cats_d)
setup_chart_titles(chart_d, title="Unit Type Distribution")
apply_pie_colors(chart_d, count=len(unit_types))
ws_dash.add_chart(chart_d, "H87")

# Column widths for dashboard
ws_dash.column_dimensions["A"].width = 3
for c in range(2, 15):
    ws_dash.column_dimensions[get_column_letter(c)].width = 14

# Print area
setup_print(ws_dash, "N", 104, "2:3")


# ================================================================
# SHEET 2: PROPERTY REGISTRY
# ================================================================
ws_prop = wb.create_sheet("Property Registry")
ws_prop.sheet_properties.tabColor = TAB_COLOR

prop_title = "Property Registry — Unit Master List"
setup_sheet(ws_prop, title=prop_title, last_col=13)

prop_headers = [
    "Unit ID", "Building", "Floor", "Unit Type", "Size (sqft)",
    "Monthly Rent (AED)", "Municipality Fee (AED)", "Status",
    "Current Tenant", "Contract Start", "Contract End", "Notes"
]
prop_col_start = 2  # B
prop_col_end = 13   # M
prop_header_row = 4
prop_data_start = 5
prop_data_rows = 150

# Write headers
for i, h in enumerate(prop_headers):
    ws_prop.cell(row=prop_header_row, column=prop_col_start + i, value=h)
style_header_row(ws_prop, prop_header_row, prop_col_start, prop_col_end)

# Column widths
prop_widths = [12, 14, 8, 14, 12, 18, 20, 16, 18, 14, 14, 24]
for i, w in enumerate(prop_widths):
    ws_prop.column_dimensions[get_column_letter(prop_col_start + i)].width = w

# Data rows with formulas and styling
for row_idx in range(prop_data_rows):
    r = prop_data_start + row_idx
    # Municipality Fee formula (col H = col 8 → column H in sheet, but col index 8 from B=2 → G is col 7, H is col 8)
    # Columns: B=UnitID, C=Building, D=Floor, E=UnitType, F=Size, G=MonthlyRent, H=MunicipalityFee
    # Municipality Fee = IFERROR(G5*0.05,0)
    ws_prop.cell(row=r, column=8, value=f"=IFERROR(G{r}*0.05,0)")
    ws_prop.cell(row=r, column=8).number_format = CURRENCY_FMT
    # Monthly Rent format
    ws_prop.cell(row=r, column=7).number_format = CURRENCY_FMT
    # Size format
    ws_prop.cell(row=r, column=6).number_format = "#,##0"
    # Date formats
    ws_prop.cell(row=r, column=10).number_format = DATE_FMT  # Contract Start
    ws_prop.cell(row=r, column=11).number_format = DATE_FMT  # Contract End

    style_data_row(ws_prop, r, prop_col_start, prop_col_end, row_idx)

# Data Validations
dv_building = DataValidation(type="list", formula1='"Building A,Building B,Building C,Building D"', allow_blank=True)
dv_building.prompt = "Select Building"
dv_building.promptTitle = "Building"
ws_prop.add_data_validation(dv_building)
dv_building.add(f"C{prop_data_start}:C{prop_data_start + prop_data_rows - 1}")

dv_unittype = DataValidation(type="list", formula1='"Studio,1 Bedroom,2 Bedroom,3 Bedroom,Shop,Office"', allow_blank=True)
dv_unittype.prompt = "Select Unit Type"
dv_unittype.promptTitle = "Unit Type"
ws_prop.add_data_validation(dv_unittype)
dv_unittype.add(f"E{prop_data_start}:E{prop_data_start + prop_data_rows - 1}")

dv_status = DataValidation(type="list", formula1='"Rented,Vacant,Under Maintenance,Reserved"', allow_blank=True)
dv_status.prompt = "Select Status"
dv_status.promptTitle = "Status"
ws_prop.add_data_validation(dv_status)
dv_status.add(f"I{prop_data_start}:I{prop_data_start + prop_data_rows - 1}")

# Conditional formatting for Status column (I)
status_range = f"I{prop_data_start}:I{prop_data_start + prop_data_rows - 1}"
cf_cell_is(ws_prop, status_range, "equal", '"Rented"', CF_POSITIVE_FILL, CF_POSITIVE_FONT)
cf_cell_is(ws_prop, status_range, "equal", '"Vacant"', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
cf_cell_is(ws_prop, status_range, "equal", '"Under Maintenance"', CF_WARNING_FILL, CF_WARNING_FONT)

# Totals row
total_row = prop_data_start + prop_data_rows
ws_prop.cell(row=total_row, column=2, value="TOTALS")
ws_prop.cell(row=total_row, column=6, value=f"=IFERROR(SUM(F{prop_data_start}:F{total_row-1}),0)")
ws_prop.cell(row=total_row, column=6).number_format = "#,##0"
ws_prop.cell(row=total_row, column=7, value=f"=IFERROR(SUM(G{prop_data_start}:G{total_row-1}),0)")
ws_prop.cell(row=total_row, column=7).number_format = CURRENCY_FMT
ws_prop.cell(row=total_row, column=8, value=f"=IFERROR(SUM(H{prop_data_start}:H{total_row-1}),0)")
ws_prop.cell(row=total_row, column=8).number_format = CURRENCY_FMT
ws_prop.cell(row=total_row, column=9, value=f"=IFERROR(COUNTA(B{prop_data_start}:B{total_row-1})-COUNTBLANK(B{prop_data_start}:B{total_row-1}),0)&\" units\"")
style_total_row(ws_prop, total_row, prop_col_start, prop_col_end)

# Freeze panes
ws_prop.freeze_panes = "C5"

# Print
setup_print(ws_prop, "M", total_row, "2:4")


# ================================================================
# SHEET 3: TENANTS
# ================================================================
ws_tenant = wb.create_sheet("Tenants")
ws_tenant.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_tenant, title="Tenant Directory", last_col=16)

tenant_headers = [
    "Tenant ID", "Tenant Name", "Phone", "Email", "Unit ID", "Unit Type",
    "Lease Start", "Lease End", "Contract Duration (months)",
    "Monthly Rent", "Security Deposit", "Payment Method",
    "Nationality", "Employer", "Emergency Contact"
]
tenant_col_start = 2
tenant_col_end = 16
tenant_header_row = 4
tenant_data_start = 5
tenant_data_rows = 150

for i, h in enumerate(tenant_headers):
    ws_tenant.cell(row=tenant_header_row, column=tenant_col_start + i, value=h)
style_header_row(ws_tenant, tenant_header_row, tenant_col_start, tenant_col_end)

tenant_widths = [12, 22, 16, 24, 12, 14, 14, 14, 22, 16, 16, 16, 14, 20, 20]
for i, w in enumerate(tenant_widths):
    ws_tenant.column_dimensions[get_column_letter(tenant_col_start + i)].width = w

for row_idx in range(tenant_data_rows):
    r = tenant_data_start + row_idx
    # Contract Duration: =IFERROR((I5-H5)/30,0)
    ws_tenant.cell(row=r, column=10, value=f"=IFERROR((I{r}-H{r})/30,0)")
    ws_tenant.cell(row=r, column=10).number_format = "#,##0"
    # Monthly Rent: VLOOKUP to Property Registry
    ws_tenant.cell(row=r, column=11, value=f"=IFERROR(VLOOKUP(F{r},'Property Registry'!B5:G154,6,FALSE),0)")
    ws_tenant.cell(row=r, column=11).number_format = CURRENCY_FMT
    # Security Deposit format
    ws_tenant.cell(row=r, column=12).number_format = CURRENCY_FMT
    # Date formats
    ws_tenant.cell(row=r, column=8).number_format = DATE_FMT  # Lease Start
    ws_tenant.cell(row=r, column=9).number_format = DATE_FMT  # Lease End

    style_data_row(ws_tenant, r, tenant_col_start, tenant_col_end, row_idx)

# Data Validation: Payment Method
dv_paymethod = DataValidation(type="list", formula1='"Cash,Bank Transfer,Cheque,Online"', allow_blank=True)
dv_paymethod.prompt = "Select Payment Method"
dv_paymethod.promptTitle = "Payment Method"
ws_tenant.add_data_validation(dv_paymethod)
dv_paymethod.add(f"L{tenant_data_start}:L{tenant_data_start + tenant_data_rows - 1}")

# Conditional formatting: Lease End within 30 days → amber, past today → red
lease_end_range = f"I{tenant_data_start}:I{tenant_data_start + tenant_data_rows - 1}"
# Lease End past today
cf_formula(ws_tenant, lease_end_range, f'AND(I{tenant_data_start}<TODAY(),I{tenant_data_start}<>"")', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
# Lease End within 30 days
cf_formula(ws_tenant, lease_end_range, f'AND(I{tenant_data_start}-TODAY()<=30,I{tenant_data_start}-TODAY()>0,I{tenant_data_start}<>"")', CF_WARNING_FILL, CF_WARNING_FONT)

# Totals row
t_total_row = tenant_data_start + tenant_data_rows
ws_tenant.cell(row=t_total_row, column=2, value="TOTALS")
ws_tenant.cell(row=t_total_row, column=11, value=f"=IFERROR(SUM(K{tenant_data_start}:K{t_total_row-1}),0)")
ws_tenant.cell(row=t_total_row, column=11).number_format = CURRENCY_FMT
ws_tenant.cell(row=t_total_row, column=12, value=f"=IFERROR(SUM(L{tenant_data_start}:L{t_total_row-1}),0)")
ws_tenant.cell(row=t_total_row, column=12).number_format = CURRENCY_FMT
style_total_row(ws_tenant, t_total_row, tenant_col_start, tenant_col_end)

ws_tenant.freeze_panes = "C5"
setup_print(ws_tenant, "P", t_total_row, "2:4")


# ================================================================
# SHEET 4: RENT COLLECTION
# ================================================================
ws_rent = wb.create_sheet("Rent Collection")
ws_rent.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_rent, title="Rent Collection — [Month/Year]", last_col=14)

rent_headers = [
    "Unit ID", "Tenant Name", "Monthly Rent", "Due Date", "Payment Date",
    "Amount Paid", "Outstanding", "Payment Status", "Days Late",
    "Late Fee", "Total Due", "Payment Method", "Receipt #", "Notes"
]
rent_col_start = 2
rent_col_end = 15
rent_header_row = 4
rent_data_start = 5
rent_data_rows = 150

for i, h in enumerate(rent_headers):
    ws_rent.cell(row=rent_header_row, column=rent_col_start + i, value=h)
style_header_row(ws_rent, rent_header_row, rent_col_start, rent_col_end)

rent_widths = [12, 22, 16, 14, 14, 16, 16, 16, 12, 14, 16, 16, 12, 24]
for i, w in enumerate(rent_widths):
    ws_rent.column_dimensions[get_column_letter(rent_col_start + i)].width = w

for row_idx in range(rent_data_rows):
    r = rent_data_start + row_idx
    # Outstanding = IFERROR(C5-F5,0) — Rent minus Amount Paid
    ws_rent.cell(row=r, column=8, value=f"=IFERROR(D{r}-G{r},0)")
    ws_rent.cell(row=r, column=8).number_format = CURRENCY_FMT
    # Payment Status = IF(G5>=D5,"Paid",IF(G5>0,"Partial","Unpaid"))
    ws_rent.cell(row=r, column=9, value=f'=IF(G{r}>=D{r},"Paid",IF(G{r}>0,"Partial","Unpaid"))')
    # Days Late = IFERROR(IF(I5="Unpaid",MAX(0,TODAY()-E5),IF(I5="Partial",MAX(0,TODAY()-E5),0)),0)
    ws_rent.cell(row=r, column=10, value=f'=IFERROR(IF(I{r}="Unpaid",MAX(0,TODAY()-E{r}),IF(I{r}="Partial",MAX(0,TODAY()-E{r}),0)),0)')
    ws_rent.cell(row=r, column=10).number_format = "#,##0"
    # Late Fee = IFERROR(IF(J5>0,MIN(D5*0.05,500),0),0) — 5% late fee, max 500 AED
    ws_rent.cell(row=r, column=11, value=f"=IFERROR(IF(J{r}>0,MIN(D{r}*0.05,500),0),0)")
    ws_rent.cell(row=r, column=11).number_format = CURRENCY_FMT
    # Total Due = IFERROR(D5+K5,0) — Rent + Late Fee
    ws_rent.cell(row=r, column=12, value=f"=IFERROR(D{r}+K{r},0)")
    ws_rent.cell(row=r, column=12).number_format = CURRENCY_FMT
    # Currency formats
    ws_rent.cell(row=r, column=4).number_format = CURRENCY_FMT   # Monthly Rent
    ws_rent.cell(row=r, column=7).number_format = CURRENCY_FMT   # Amount Paid
    # Date formats
    ws_rent.cell(row=r, column=5).number_format = DATE_FMT  # Due Date
    ws_rent.cell(row=r, column=6).number_format = DATE_FMT  # Payment Date

    style_data_row(ws_rent, r, rent_col_start, rent_col_end, row_idx)

# Data Validation: Payment Method
dv_paymethod2 = DataValidation(type="list", formula1='"Cash,Bank Transfer,Cheque,Online"', allow_blank=True)
dv_paymethod2.prompt = "Select Payment Method"
dv_paymethod2.promptTitle = "Payment Method"
ws_rent.add_data_validation(dv_paymethod2)
dv_paymethod2.add(f"L{rent_data_start}:L{rent_data_start + rent_data_rows - 1}")

# Conditional formatting: Payment Status
status_range = f"I{rent_data_start}:I{rent_data_start + rent_data_rows - 1}"
cf_cell_is(ws_rent, status_range, "equal", '"Paid"', CF_POSITIVE_FILL, CF_POSITIVE_FONT)
cf_cell_is(ws_rent, status_range, "equal", '"Partial"', CF_WARNING_FILL, CF_WARNING_FONT)
cf_cell_is(ws_rent, status_range, "equal", '"Unpaid"', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)

# Summary section
summary_start = rent_data_start + rent_data_rows + 2
ws_rent.cell(row=summary_start, column=2, value="COLLECTION SUMMARY").font = font_subheader()
ws_rent.merge_cells(start_row=summary_start, start_column=2, end_row=summary_start, end_column=5)

summary_items = [
    ("Total Expected Rent", f"=IFERROR(SUM(D{rent_data_start}:D{rent_data_start+rent_data_rows-1}),0)"),
    ("Total Collected", f"=IFERROR(SUM(G{rent_data_start}:G{rent_data_start+rent_data_rows-1}),0)"),
    ("Total Outstanding", f"=IFERROR(SUM(H{rent_data_start}:H{rent_data_start+rent_data_rows-1}),0)"),
    ("Collection Rate %", f"=IFERROR(SUM(G{rent_data_start}:G{rent_data_start+rent_data_rows-1})/SUM(D{rent_data_start}:D{rent_data_start+rent_data_rows-1}),0)"),
    ("# Paid", f'=IFERROR(COUNTIF(I{rent_data_start}:I{rent_data_start+rent_data_rows-1},"Paid"),0)'),
    ("# Partial", f'=IFERROR(COUNTIF(I{rent_data_start}:I{rent_data_start+rent_data_rows-1},"Partial"),0)'),
    ("# Unpaid", f'=IFERROR(COUNTIF(I{rent_data_start}:I{rent_data_start+rent_data_rows-1},"Unpaid"),0)'),
]

for i, (label, formula) in enumerate(summary_items):
    r = summary_start + 1 + i
    ws_rent.cell(row=r, column=2, value=label).font = font_body()
    ws_rent.cell(row=r, column=2).alignment = align_text()
    ws_rent.cell(row=r, column=4, value=formula).font = font_subheader()
    ws_rent.cell(row=r, column=4).alignment = align_number()
    if label == "Collection Rate %":
        ws_rent.cell(row=r, column=4).number_format = PCT_FMT
    elif label.startswith("#"):
        ws_rent.cell(row=r, column=4).number_format = "#,##0"
    else:
        ws_rent.cell(row=r, column=4).number_format = CURRENCY_FMT
    # Style the summary rows
    for c in range(2, 6):
        ws_rent.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)

ws_rent.freeze_panes = "C5"
setup_print(ws_rent, "O", summary_start + len(summary_items) + 1, "2:4")


# ================================================================
# SHEET 5: PAYMENT ALERTS
# ================================================================
ws_alerts = wb.create_sheet("Payment Alerts")
ws_alerts.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_alerts, title="Payment Alerts — Tenants Who Haven't Paid by the 5th", last_col=12)

# Instruction row
ws_alerts.cell(row=3, column=2, value="⚠ Fill this sheet by the 5th of each month. Contact all tenants listed here immediately.")
ws_alerts.cell(row=3, column=2).font = Font(name="Calibri", size=10, bold=True, color=ACCENT_NEGATIVE)
ws_alerts.merge_cells(start_row=3, start_column=2, end_row=3, end_column=12)

alert_headers = [
    "Unit ID", "Tenant Name", "Phone", "Monthly Rent", "Due Date",
    "Days Late", "Late Fee", "Total Due", "SMS Sent?",
    "Follow-up Date", "Notes"
]
alert_col_start = 2
alert_col_end = 12
alert_header_row = 4
alert_data_start = 5
alert_data_rows = 50

# Style headers with urgent red/amber colors
for i, h in enumerate(alert_headers):
    cell = ws_alerts.cell(row=alert_header_row, column=alert_col_start + i, value=h)
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=ACCENT_NEGATIVE)  # Red header for urgency
    cell.alignment = align_header()
    cell.border = border_header()
ws_alerts.row_dimensions[alert_header_row].height = ROW_HEIGHTS["header"]

alert_widths = [12, 22, 16, 16, 14, 12, 14, 16, 12, 14, 24]
for i, w in enumerate(alert_widths):
    ws_alerts.column_dimensions[get_column_letter(alert_col_start + i)].width = w

for row_idx in range(alert_data_rows):
    r = alert_data_start + row_idx
    # Late Fee formula: IFERROR(IF(G5>0,MIN(E5*0.05,500),0),0) — same as Rent Collection
    # Days Late is manual, Late Fee based on it
    ws_alerts.cell(row=r, column=8, value=f"=IFERROR(IF(G{r}>0,MIN(E{r}*0.05,500),0),0)")
    ws_rent.cell(row=r, column=8).number_format = CURRENCY_FMT
    # Total Due = Rent + Late Fee
    ws_alerts.cell(row=r, column=9, value=f"=IFERROR(E{r}+H{r},0)")
    ws_alerts.cell(row=r, column=9).number_format = CURRENCY_FMT
    # Currency formats
    ws_alerts.cell(row=r, column=5).number_format = CURRENCY_FMT   # Monthly Rent
    ws_alerts.cell(row=r, column=6).number_format = DATE_FMT       # Due Date
    ws_alerts.cell(row=r, column=11).number_format = DATE_FMT      # Follow-up Date
    ws_alerts.cell(row=r, column=7).number_format = "#,##0"        # Days Late

    style_data_row(ws_alerts, r, alert_col_start, alert_col_end, row_idx)

# Data Validation: SMS Sent
dv_sms = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws_alerts.add_data_validation(dv_sms)
dv_sms.add(f"J{alert_data_start}:J{alert_data_start + alert_data_rows - 1}")

# Conditional formatting: Days Late > 30 → deep red, > 15 → amber
days_late_range = f"G{alert_data_start}:G{alert_data_start + alert_data_rows - 1}"
cf_cell_is(ws_alerts, days_late_range, "greaterThan", "30", CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
cf_cell_is(ws_alerts, days_late_range, "greaterThan", "15", CF_WARNING_FILL, CF_WARNING_FONT)

ws_alerts.freeze_panes = "C5"
setup_print(ws_alerts, "L", alert_data_start + alert_data_rows, "2:4")


# ================================================================
# SHEET 6: MAINTENANCE
# ================================================================
ws_maint = wb.create_sheet("Maintenance")
ws_maint.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_maint, title="Maintenance Log & Cost Tracker", last_col=13)

maint_headers = [
    "Request ID", "Date", "Unit ID", "Category", "Description",
    "Vendor/Technician", "Estimated Cost", "Actual Cost", "Status",
    "Priority", "Completion Date", "Notes"
]
maint_col_start = 2
maint_col_end = 13
maint_header_row = 4
maint_data_start = 5
maint_data_rows = 200

for i, h in enumerate(maint_headers):
    ws_maint.cell(row=maint_header_row, column=maint_col_start + i, value=h)
style_header_row(ws_maint, maint_header_row, maint_col_start, maint_col_end)

maint_widths = [14, 14, 12, 16, 28, 20, 16, 16, 14, 12, 16, 24]
for i, w in enumerate(maint_widths):
    ws_maint.column_dimensions[get_column_letter(maint_col_start + i)].width = w

for row_idx in range(maint_data_rows):
    r = maint_data_start + row_idx
    ws_maint.cell(row=r, column=3).number_format = DATE_FMT   # Date
    ws_maint.cell(row=r, column=8).number_format = CURRENCY_FMT  # Estimated Cost
    ws_maint.cell(row=r, column=9).number_format = CURRENCY_FMT  # Actual Cost
    ws_maint.cell(row=r, column=13).number_format = DATE_FMT  # Completion Date
    style_data_row(ws_maint, r, maint_col_start, maint_col_end, row_idx)

# Data Validations
dv_maint_cat = DataValidation(type="list",
    formula1='"Plumbing,Electrical,AC,Painting,Structural,Cleaning,Lock/Door,Appliance,Other"',
    allow_blank=True)
dv_maint_cat.prompt = "Select Category"
ws_maint.add_data_validation(dv_maint_cat)
dv_maint_cat.add(f"E{maint_data_start}:E{maint_data_start + maint_data_rows - 1}")

dv_maint_status = DataValidation(type="list", formula1='"Open,In Progress,Completed,Cancelled"', allow_blank=True)
dv_maint_status.prompt = "Select Status"
ws_maint.add_data_validation(dv_maint_status)
dv_maint_status.add(f"J{maint_data_start}:J{maint_data_start + maint_data_rows - 1}")

dv_maint_priority = DataValidation(type="list", formula1='"Low,Medium,High,Urgent"', allow_blank=True)
dv_maint_priority.prompt = "Select Priority"
ws_maint.add_data_validation(dv_maint_priority)
dv_maint_priority.add(f"K{maint_data_start}:K{maint_data_start + maint_data_rows - 1}")

# Conditional formatting: Priority
priority_range = f"K{maint_data_start}:K{maint_data_start + maint_data_rows - 1}"
cf_cell_is(ws_maint, priority_range, "equal", '"Urgent"', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
cf_cell_is(ws_maint, priority_range, "equal", '"High"', CF_WARNING_FILL, CF_WARNING_FONT)

# Conditional formatting: Status
maint_status_range = f"J{maint_data_start}:J{maint_data_start + maint_data_rows - 1}"
cf_cell_is(ws_maint, maint_status_range, "equal", '"Open"', CF_WARNING_FILL, CF_WARNING_FONT)
cf_cell_is(ws_maint, maint_status_range, "equal", '"In Progress"',
           PatternFill("solid", fgColor="D6E4F0"), Font(color="1B2A4A"))
cf_cell_is(ws_maint, maint_status_range, "equal", '"Completed"', CF_POSITIVE_FILL, CF_POSITIVE_FONT)

# Totals row
m_total_row = maint_data_start + maint_data_rows
ws_maint.cell(row=m_total_row, column=2, value="TOTALS")
ws_maint.cell(row=m_total_row, column=8, value=f"=IFERROR(SUM(H{maint_data_start}:H{m_total_row-1}),0)")
ws_maint.cell(row=m_total_row, column=8).number_format = CURRENCY_FMT
ws_maint.cell(row=m_total_row, column=9, value=f"=IFERROR(SUM(I{maint_data_start}:I{m_total_row-1}),0)")
ws_maint.cell(row=m_total_row, column=9).number_format = CURRENCY_FMT
# Variance
ws_maint.cell(row=m_total_row, column=10, value="Variance:")
ws_maint.cell(row=m_total_row, column=11, value=f"=IFERROR(H{m_total_row}-I{m_total_row},0)")
ws_maint.cell(row=m_total_row, column=11).number_format = CURRENCY_FMT
style_total_row(ws_maint, m_total_row, maint_col_start, maint_col_end)

ws_maint.freeze_panes = "C5"
setup_print(ws_maint, "M", m_total_row, "2:4")


# ================================================================
# SHEET 7: EXPENSES
# ================================================================
ws_exp = wb.create_sheet("Expenses")
ws_exp.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_exp, title="Operating Expenses — Monthly Tracker", last_col=11)

exp_headers = [
    "Date", "Category", "Description", "Amount (AED)", "Payment Method",
    "Vendor/Payee", "Invoice #", "Recurring?", "Building", "Notes"
]
exp_col_start = 2
exp_col_end = 11
exp_header_row = 4
exp_data_start = 5
exp_data_rows = 300

for i, h in enumerate(exp_headers):
    ws_exp.cell(row=exp_header_row, column=exp_col_start + i, value=h)
style_header_row(ws_exp, exp_header_row, exp_col_start, exp_col_end)

exp_widths = [14, 20, 28, 16, 16, 20, 14, 12, 16, 24]
for i, w in enumerate(exp_widths):
    ws_exp.column_dimensions[get_column_letter(exp_col_start + i)].width = w

for row_idx in range(exp_data_rows):
    r = exp_data_start + row_idx
    ws_exp.cell(row=r, column=2).number_format = DATE_FMT   # Date
    ws_exp.cell(row=r, column=5).number_format = CURRENCY_FMT  # Amount
    style_data_row(ws_exp, r, exp_col_start, exp_col_end, row_idx)

# Data Validations
exp_categories = "Manpower/Staff,Municipality Fees,Maintenance,Leasing Commission,Insurance,Utilities,Marketing,Legal,Cleaning,Security,Pest Control,Elevator Service,Other"
dv_exp_cat = DataValidation(type="list", formula1=f'"{exp_categories}"', allow_blank=True)
dv_exp_cat.prompt = "Select Category"
ws_exp.add_data_validation(dv_exp_cat)
dv_exp_cat.add(f"C{exp_data_start}:C{exp_data_start + exp_data_rows - 1}")

dv_exp_pay = DataValidation(type="list", formula1='"Cash,Bank Transfer,Cheque,Online"', allow_blank=True)
ws_exp.add_data_validation(dv_exp_pay)
dv_exp_pay.add(f"F{exp_data_start}:F{exp_data_start + exp_data_rows - 1}")

dv_exp_recur = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws_exp.add_data_validation(dv_exp_recur)
dv_exp_recur.add(f"I{exp_data_start}:I{exp_data_start + exp_data_rows - 1}")

dv_exp_building = DataValidation(type="list", formula1='"All Buildings,Building A,Building B,Building C,Building D"', allow_blank=True)
ws_exp.add_data_validation(dv_exp_building)
dv_exp_building.add(f"J{exp_data_start}:J{exp_data_start + exp_data_rows - 1}")

# Summary section at bottom
exp_summary_start = exp_data_start + exp_data_rows + 2
ws_exp.cell(row=exp_summary_start, column=2, value="EXPENSE SUMMARY BY CATEGORY").font = font_subheader()
ws_exp.merge_cells(start_row=exp_summary_start, start_column=2, end_row=exp_summary_start, end_column=3)

# Summary header row (next row)
exp_sum_hdr_row = exp_summary_start + 1
ws_exp.cell(row=exp_sum_hdr_row, column=2, value="Category")
ws_exp.cell(row=exp_sum_hdr_row, column=3, value="Total (AED)")
ws_exp.cell(row=exp_sum_hdr_row, column=2).font = font_header()
ws_exp.cell(row=exp_sum_hdr_row, column=3).font = font_header()
ws_exp.cell(row=exp_sum_hdr_row, column=2).fill = fill_header()
ws_exp.cell(row=exp_sum_hdr_row, column=3).fill = fill_header()
ws_exp.cell(row=exp_sum_hdr_row, column=2).alignment = align_header()
ws_exp.cell(row=exp_sum_hdr_row, column=3).alignment = align_header()
ws_exp.cell(row=exp_sum_hdr_row, column=2).border = border_header()
ws_exp.cell(row=exp_sum_hdr_row, column=3).border = border_header()

cat_list = [
    "Manpower/Staff", "Municipality Fees", "Maintenance", "Leasing Commission",
    "Insurance", "Utilities", "Marketing", "Legal", "Cleaning",
    "Security", "Pest Control", "Elevator Service", "Other"
]

for i, cat in enumerate(cat_list):
    r = exp_sum_hdr_row + 1 + i
    ws_exp.cell(row=r, column=2, value=cat).font = font_body()
    ws_exp.cell(row=r, column=2).alignment = align_text()
    ws_exp.cell(row=r, column=3,
                value=f'=IFERROR(SUMPRODUCT((C{exp_data_start}:C{exp_data_start+exp_data_rows-1}="{cat}")*(E{exp_data_start}:E{exp_data_start+exp_data_rows-1})),0)')
    ws_exp.cell(row=r, column=3).number_format = CURRENCY_FMT
    ws_exp.cell(row=r, column=3).alignment = align_number()
    ws_exp.cell(row=r, column=3).font = font_body()
    fill = fill_data_row(i)
    ws_exp.cell(row=r, column=2).fill = fill
    ws_exp.cell(row=r, column=3).fill = fill

# Grand Total
gt_row = exp_sum_hdr_row + 1 + len(cat_list)
ws_exp.cell(row=gt_row, column=2, value="GRAND TOTAL").font = font_subheader()
ws_exp.cell(row=gt_row, column=3, value=f"=IFERROR(SUM(C{exp_sum_hdr_row+1}:C{gt_row-1}),0)")
ws_exp.cell(row=gt_row, column=3).number_format = CURRENCY_FMT
ws_exp.cell(row=gt_row, column=3).font = font_subheader()
style_total_row(ws_exp, gt_row, 2, 3)

ws_exp.freeze_panes = "C5"
setup_print(ws_exp, "K", gt_row, "2:4")


# ================================================================
# SHEET 8: REVENUE ANALYSIS
# ================================================================
ws_rev = wb.create_sheet("Revenue Analysis")
ws_rev.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_rev, title="Revenue Analysis — Annual Overview", last_col=15)

# Headers: B4 = Line Item, C4-N4 = Jan-Dec, O4 = Annual Total
rev_col_start = 2
rev_header_row = 4
rev_data_start = 5

# Column headers
ws_rev.cell(row=rev_header_row, column=2, value="Revenue Line Item")
months_full = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
for i, m in enumerate(months_full):
    ws_rev.cell(row=rev_header_row, column=3 + i, value=m)
ws_rev.cell(row=rev_header_row, column=15, value="Annual Total")

style_header_row(ws_rev, rev_header_row, 2, 15)

# Column widths
ws_rev.column_dimensions["B"].width = 24
for c in range(3, 16):
    ws_rev.column_dimensions[get_column_letter(c)].width = 14

# Revenue line items with formulas
rev_items = [
    ("Rental Income", "rental"),
    ("Late Fee Income", "late_fee"),
    ("Municipality Fees Collected", "municipality"),
    ("Other Income", "other"),
    ("GROSS REVENUE", "gross"),          # separator / bold
    ("Less: Vacancy Loss", "vacancy"),
    ("Less: Bad Debt/Unpaid", "bad_debt"),
    ("NET REVENUE", "net"),              # separator / bold
]

current_row = rev_data_start
item_rows = {}  # track row numbers for formulas

for item_name, key in rev_items:
    r = current_row
    item_rows[key] = r
    ws_rev.cell(row=r, column=2, value=item_name)

    if key == "gross":
        # Gross Revenue = SUM of above 4 items
        for c in range(3, 16):
            col_letter = get_column_letter(c)
            ws_rev.cell(row=r, column=c,
                        value=f"=IFERROR(SUM({col_letter}{rev_data_start}:{col_letter}{r-1}),0)")
        ws_rev.cell(row=r, column=2).font = font_subheader()
        for c in range(2, 16):
            ws_rev.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    elif key == "net":
        # Net Revenue = Gross - Vacancy - Bad Debt
        for c in range(3, 16):
            col_letter = get_column_letter(c)
            ws_rev.cell(row=r, column=c,
                        value=f"=IFERROR({col_letter}{item_rows['gross']}+{col_letter}{item_rows['vacancy']}+{col_letter}{item_rows['bad_debt']},0)")
        ws_rev.cell(row=r, column=2).font = font_subheader()
        for c in range(2, 16):
            ws_rev.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    else:
        # Regular items - formulas reference other sheets where possible
        for c in range(3, 15):  # Jan-Dec (columns C-N)
            # Leave structured but with SUMPRODUCT references
            if key == "rental":
                ws_rev.cell(row=r, column=c,
                    value=f"=IFERROR(SUMPRODUCT(('Property Registry'!I5:I154=\"Rented\")*('Property Registry'!G5:G154)),0)")
            elif key == "late_fee":
                ws_rev.cell(row=r, column=c,
                    value=f"=IFERROR(SUM('Rent Collection'!K5:K154),0)")
            elif key == "municipality":
                ws_rev.cell(row=r, column=c,
                    value=f"=IFERROR(SUM('Property Registry'!H5:H154),0)")
            elif key == "other":
                ws_rev.cell(row=r, column=c, value=0)
            elif key == "vacancy":
                ws_rev.cell(row=r, column=c,
                    value=f"=IFERROR(-SUMPRODUCT(('Property Registry'!I5:I154=\"Vacant\")*('Property Registry'!G5:G154)),0)")
            elif key == "bad_debt":
                ws_rev.cell(row=r, column=c,
                    value=f"=IFERROR(-SUM('Rent Collection'!H5:H154),0)")

        ws_rev.cell(row=r, column=2).font = font_body()

    # Annual Total (column O = 15)
    if key not in ("gross", "net"):
        ws_rev.cell(row=r, column=15, value=f"=IFERROR(SUM(C{r}:N{r}),0)")

    # Number formats
    for c in range(3, 16):
        ws_rev.cell(row=r, column=c).number_format = CURRENCY_FMT
        ws_rev.cell(row=r, column=c).alignment = align_number()

    current_row += 1

# Conditional formatting: negative → red, positive → green
for r in range(rev_data_start, current_row):
    for c in range(3, 16):
        cell_ref = f"{get_column_letter(c)}{r}"
        cf_formula(ws_rev, cell_ref, f'{cell_ref}<0', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)

# Revenue vs Vacancy Loss bar chart
chart_rev = create_bar_chart(width=22, height=12)
# Create small data area for chart (rows below main data)
chart_data_start = current_row + 2
ws_rev.cell(row=chart_data_start, column=2, value="Category").font = font_caption()
ws_rev.cell(row=chart_data_start, column=3, value="Amount").font = font_caption()
chart_rev_items = [
    ("Gross Revenue", f"=O{item_rows['gross']}"),
    ("Vacancy Loss", f"=O{item_rows['vacancy']}"),
    ("Bad Debt", f"=O{item_rows['bad_debt']}"),
    ("Net Revenue", f"=O{item_rows['net']}"),
]
for i, (lbl, formula) in enumerate(chart_rev_items):
    r = chart_data_start + 1 + i
    ws_rev.cell(row=r, column=2, value=lbl).font = font_caption()
    ws_rev.cell(row=r, column=3, value=formula).font = font_caption()
    ws_rev.cell(row=r, column=3).number_format = CURRENCY_FMT

data_rev = Reference(ws_rev, min_col=3, min_row=chart_data_start, max_row=chart_data_start + len(chart_rev_items))
cats_rev = Reference(ws_rev, min_col=2, min_row=chart_data_start + 1, max_row=chart_data_start + len(chart_rev_items))
chart_rev.add_data(data_rev, titles_from_data=True)
chart_rev.set_categories(cats_rev)
setup_chart_titles(chart_rev, title="Revenue vs Vacancy Loss", y_title="Amount (AED)")
apply_chart_colors(chart_rev)
ws_rev.add_chart(chart_rev, f"B{current_row + 8}")

ws_rev.freeze_panes = "C5"
setup_print(ws_rev, "O", current_row + 25, "2:4")


# ================================================================
# SHEET 9: PROFIT & LOSS
# ================================================================
ws_pl = wb.create_sheet("Profit & Loss")
ws_pl.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_pl, title="Profit & Loss Statement — Real Estate Emperor", last_col=4)

pl_col_start = 2
pl_col_end = 4
pl_header_row = 4
current_row = 5

# Style helper for section headers
def pl_section_header(ws, row, text):
    ws.cell(row=row, column=2, value=text).font = font_subheader()
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
    ws.row_dimensions[row].height = ROW_HEIGHTS["header"]

def pl_line_item(ws, row, text, formula=None, indent=False):
    ws.cell(row=row, column=2, value=text).font = font_body()
    if indent:
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=2)
    if formula:
        ws.cell(row=row, column=3, value=formula).font = font_body()
        ws.cell(row=row, column=3).number_format = CURRENCY_FMT
        ws.cell(row=row, column=3).alignment = align_number()
    ws.row_dimensions[row].height = ROW_HEIGHTS["data"]
    return row

def pl_total_line(ws, row, text, formula):
    ws.cell(row=row, column=2, value=text).font = font_subheader()
    ws.cell(row=row, column=3, value=formula).font = font_subheader()
    ws.cell(row=row, column=3).number_format = CURRENCY_FMT
    ws.cell(row=row, column=3).alignment = align_number()
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)
        ws.cell(row=row, column=c).border = border_total()
    ws.row_dimensions[row].height = ROW_HEIGHTS["total"]
    return row

def pl_margin_line(ws, row, text, formula):
    ws.cell(row=row, column=2, value=text).font = font_caption()
    ws.cell(row=row, column=3, value=formula).font = font_caption()
    ws.cell(row=row, column=3).number_format = PCT_FMT
    ws.cell(row=row, column=3).alignment = align_number()
    ws.row_dimensions[row].height = ROW_HEIGHTS["data"]
    return row

# ── REVENUE ──
pl_section_header(ws_pl, current_row, "REVENUE")
current_row += 1

# Revenue items referencing Revenue Analysis sheet
rev_rental_row = current_row
pl_line_item(ws_pl, current_row, "Rental Income", f"=IFERROR('Revenue Analysis'!O{rev_data_start},0)", indent=True)
current_row += 1

rev_late_row = current_row
pl_line_item(ws_pl, current_row, "Late Fee Income", f"=IFERROR('Revenue Analysis'!O{rev_data_start+1},0)", indent=True)
current_row += 1

rev_other_row = current_row
pl_line_item(ws_pl, current_row, "Other Income", f"=IFERROR('Revenue Analysis'!O{rev_data_start+3},0)", indent=True)
current_row += 1

total_rev_row = current_row
pl_total_line(ws_pl, current_row, "Total Revenue", f"=IFERROR(SUM(C{rev_rental_row}:C{current_row-1}),0)")
current_row += 1

# ── COST OF OPERATIONS ──
current_row += 1
pl_section_header(ws_pl, current_row, "COST OF OPERATIONS")
current_row += 1

# Use SUMPRODUCT referencing Expenses sheet categories
cost_items = [
    ("Manpower/Staff Costs", "Manpower/Staff"),
    ("Maintenance Costs", "Maintenance"),
    ("Municipality Fees", "Municipality Fees"),
    ("Leasing Commissions", "Leasing Commission"),
    ("Insurance", "Insurance"),
    ("Utilities", "Utilities"),
]

cost_start_row = current_row
for item_name, cat in cost_items:
    pl_line_item(ws_pl, current_row, item_name,
                 f'=IFERROR(SUMPRODUCT((\'Expenses\'!C5:C304="{cat}")*(\'Expenses\'!E5:E304)),0)', indent=True)
    current_row += 1

total_cost_row = current_row
pl_total_line(ws_pl, current_row, "Total Cost of Operations", f"=IFERROR(SUM(C{cost_start_row}:C{current_row-1}),0)")
current_row += 1

# ── GROSS PROFIT ──
current_row += 1
gross_profit_row = current_row
pl_total_line(ws_pl, current_row, "Gross Profit", f"=IFERROR(C{total_rev_row}-C{total_cost_row},0)")
current_row += 1

gross_margin_row = current_row
pl_margin_line(ws_pl, current_row, "Gross Margin %", f"=IFERROR(C{gross_profit_row}/C{total_rev_row},0)")
current_row += 1

# ── OPERATING EXPENSES ──
current_row += 1
pl_section_header(ws_pl, current_row, "OPERATING EXPENSES")
current_row += 1

opex_items = [
    ("Marketing", "Marketing"),
    ("Legal & Professional", "Legal"),
    ("Cleaning & Security", "Cleaning"),
    ("Security Costs", "Security"),
    ("Pest Control", "Pest Control"),
    ("Elevator Service", "Elevator Service"),
    ("Other", "Other"),
]

opex_start_row = current_row
for item_name, cat in opex_items:
    pl_line_item(ws_pl, current_row, item_name,
                 f'=IFERROR(SUMPRODUCT((\'Expenses\'!C5:C304="{cat}")*(\'Expenses\'!E5:E304)),0)', indent=True)
    current_row += 1

total_opex_row = current_row
pl_total_line(ws_pl, current_row, "Total Operating Expenses", f"=IFERROR(SUM(C{opex_start_row}:C{current_row-1}),0)")
current_row += 1

# ── NET PROFIT ──
current_row += 1
net_profit_row = current_row
pl_total_line(ws_pl, current_row, "Net Profit", f"=IFERROR(C{gross_profit_row}-C{total_opex_row},0)")
current_row += 1

net_margin_row = current_row
pl_margin_line(ws_pl, current_row, "Net Margin %", f"=IFERROR(C{net_profit_row}/C{total_rev_row},0)")
current_row += 1

# ── Waterfall Summary ──
current_row += 2
pl_section_header(ws_pl, current_row, "WATERFALL SUMMARY")
current_row += 1

waterfall_items = [
    ("Total Revenue", f"=C{total_rev_row}"),
    ("Less: Cost of Operations", f"=-C{total_cost_row}"),
    ("Gross Profit", f"=C{gross_profit_row}"),
    ("Less: Operating Expenses", f"=-C{total_opex_row}"),
    ("Net Profit", f"=C{net_profit_row}"),
]

for item_name, formula in waterfall_items:
    ws_pl.cell(row=current_row, column=2, value=item_name).font = font_body()
    ws_pl.cell(row=current_row, column=3, value=formula).font = font_subheader()
    ws_pl.cell(row=current_row, column=3).number_format = CURRENCY_FMT
    ws_pl.cell(row=current_row, column=3).alignment = align_number()
    # Conditional colors inline
    if "Less" in item_name:
        ws_pl.cell(row=current_row, column=3).font = Font(name="Calibri", size=11, bold=True, color=ACCENT_NEGATIVE)
    elif "Net Profit" in item_name:
        ws_pl.cell(row=current_row, column=3).font = Font(name="Calibri", size=11, bold=True, color=ACCENT_POSITIVE)
    ws_pl.row_dimensions[current_row].height = ROW_HEIGHTS["data"]
    current_row += 1

# Column widths
ws_pl.column_dimensions["B"].width = 28
ws_pl.column_dimensions["C"].width = 20
ws_pl.column_dimensions["D"].width = 5

# Conditional formatting: positive → green, negative → red on column C
pl_data_range = f"C5:C{current_row - 1}"
cf_formula(ws_pl, pl_data_range, "C5<0", CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
cf_formula(ws_pl, pl_data_range, "AND(C5>0,ROW()>5)", CF_POSITIVE_FILL, CF_POSITIVE_FONT)

setup_print(ws_pl, "D", current_row, "2:4")


# ================================================================
# SHEET 10: CONTRACTS
# ================================================================
ws_con = wb.create_sheet("Contracts")
ws_con.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws_con, title="Lease Contract Tracker", last_col=13)

con_headers = [
    "Unit ID", "Tenant Name", "Contract Start", "Contract End",
    "Duration (months)", "Monthly Rent", "Total Contract Value",
    "Days Remaining", "Status", "Renewal Status",
    "New Rent (if renewed)", "Notes"
]
con_col_start = 2
con_col_end = 13
con_header_row = 4
con_data_start = 5
con_data_rows = 150

for i, h in enumerate(con_headers):
    ws_con.cell(row=con_header_row, column=con_col_start + i, value=h)
style_header_row(ws_con, con_header_row, con_col_start, con_col_end)

con_widths = [12, 22, 14, 14, 18, 16, 20, 16, 16, 18, 20, 24]
for i, w in enumerate(con_widths):
    ws_con.column_dimensions[get_column_letter(con_col_start + i)].width = w

for row_idx in range(con_data_rows):
    r = con_data_start + row_idx
    # Duration: =IFERROR((E5-D5)/30,0)
    ws_con.cell(row=r, column=6, value=f"=IFERROR((E{r}-D{r})/30,0)")
    ws_con.cell(row=r, column=6).number_format = "#,##0"
    # Total Contract Value: =IFERROR(G5*F5,0)
    ws_con.cell(row=r, column=8, value=f"=IFERROR(G{r}*F{r},0)")
    ws_con.cell(row=r, column=8).number_format = CURRENCY_FMT
    # Days Remaining: =IFERROR(MAX(0,E5-TODAY()),0)
    ws_con.cell(row=r, column=9, value=f"=IFERROR(MAX(0,E{r}-TODAY()),0)")
    ws_con.cell(row=r, column=9).number_format = "#,##0"
    # Status: =IF(I5<=0,"Expired",IF(I5<=30,"Expiring Soon",IF(I5<=90,"Renewal Due","Active")))
    ws_con.cell(row=r, column=10, value=f'=IF(I{r}<=0,"Expired",IF(I{r}<=30,"Expiring Soon",IF(I{r}<=90,"Renewal Due","Active")))')
    # Currency formats
    ws_con.cell(row=r, column=7).number_format = CURRENCY_FMT  # Monthly Rent
    ws_con.cell(row=r, column=12).number_format = CURRENCY_FMT  # New Rent
    # Date formats
    ws_con.cell(row=r, column=4).number_format = DATE_FMT  # Contract Start
    ws_con.cell(row=r, column=5).number_format = DATE_FMT  # Contract End

    style_data_row(ws_con, r, con_col_start, con_col_end, row_idx)

# Data Validation: Renewal Status
dv_renewal = DataValidation(type="list", formula1='"Renewed,In Negotiation,Not Renewed,Tenant Leaving"', allow_blank=True)
dv_renewal.prompt = "Select Renewal Status"
ws_con.add_data_validation(dv_renewal)
dv_renewal.add(f"J{con_data_start}:J{con_data_start + con_data_rows - 1}")

# Conditional formatting: Status
con_status_range = f"J{con_data_start}:J{con_data_start + con_data_rows - 1}"
cf_cell_is(ws_con, con_status_range, "equal", '"Expired"', CF_NEGATIVE_FILL, CF_NEGATIVE_FONT)
cf_cell_is(ws_con, con_status_range, "equal", '"Expiring Soon"', CF_WARNING_FILL, CF_WARNING_FONT)
cf_cell_is(ws_con, con_status_range, "equal", '"Renewal Due"',
           PatternFill("solid", fgColor="FEF9E7"), Font(color="B89B4A"))  # Yellow-ish
cf_cell_is(ws_con, con_status_range, "equal", '"Active"', CF_POSITIVE_FILL, CF_POSITIVE_FONT)

# Summary at bottom
con_summary_start = con_data_start + con_data_rows + 2
ws_con.cell(row=con_summary_start, column=2, value="CONTRACT SUMMARY").font = font_subheader()
ws_con.merge_cells(start_row=con_summary_start, start_column=2, end_row=con_summary_start, end_column=5)

con_summary_items = [
    ("# Active", f'=IFERROR(COUNTIF(J{con_data_start}:J{con_data_start+con_data_rows-1},"Active"),0)'),
    ("# Expiring Soon", f'=IFERROR(COUNTIF(J{con_data_start}:J{con_data_start+con_data_rows-1},"Expiring Soon"),0)'),
    ("# Renewal Due", f'=IFERROR(COUNTIF(J{con_data_start}:J{con_data_start+con_data_rows-1},"Renewal Due"),0)'),
    ("# Expired", f'=IFERROR(COUNTIF(J{con_data_start}:J{con_data_start+con_data_rows-1},"Expired"),0)'),
    ("# Total", f'=IFERROR(COUNTA(B{con_data_start}:B{con_data_start+con_data_rows-1})-COUNTBLANK(B{con_data_start}:B{con_data_start+con_data_rows-1}),0)'),
]

for i, (label, formula) in enumerate(con_summary_items):
    r = con_summary_start + 1 + i
    ws_con.cell(row=r, column=2, value=label).font = font_body()
    ws_con.cell(row=r, column=2).alignment = align_text()
    ws_con.cell(row=r, column=4, value=formula).font = font_subheader()
    ws_con.cell(row=r, column=4).alignment = align_number()
    ws_con.cell(row=r, column=4).number_format = "#,##0"
    for c in range(2, 6):
        ws_con.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PRIMARY_LIGHT)

ws_con.freeze_panes = "C5"
setup_print(ws_con, "M", con_summary_start + len(con_summary_items) + 1, "2:4")


# ================================================================
# FINAL: Set sheet order (Dashboard first)
# ================================================================
# Sheets are already in order: Dashboard, Property Registry, Tenants,
# Rent Collection, Payment Alerts, Maintenance, Expenses,
# Revenue Analysis, Profit & Loss, Contracts

# ================================================================
# SAVE
# ================================================================
wb.save(OUTPUT_PATH)
print(f"✅ Real Estate Emperor RE System saved to: {OUTPUT_PATH}")
print(f"   Sheets: {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"   - {s}")
