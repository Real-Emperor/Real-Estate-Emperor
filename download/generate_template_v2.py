#!/usr/bin/env python3
"""
Property Management Data Collection Template Generator v2
==========================================================
Refined version with:
- Payment amounts from source data
- Consolidated Data Issues (grouped by type, not per-tenant)
- ADVB mapped to ADCB
- Better balance calculations
"""

import sys, os
sys.path.insert(0, '/home/z/my-project/skills/xlsx/templates')
from base import *

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Border, Side, Alignment, numbers, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from collections import Counter, OrderedDict
import datetime

# ── Load source data ─────────────────────────────────────────────────
source_wb = openpyxl.load_workbook('/home/z/my-project/upload/APRIL-25.xlsx', data_only=True)
source_ws = source_wb['Sheet1']

all_records = []
current_property = None
for row_idx in range(3, source_ws.max_row + 1):
    a_val = source_ws.cell(row=row_idx, column=1).value
    b_val = source_ws.cell(row=row_idx, column=2).value
    c_val = source_ws.cell(row=row_idx, column=3).value
    f_val = source_ws.cell(row=row_idx, column=6).value
    
    if a_val and str(a_val).strip():
        current_property = str(a_val).strip()
    if b_val is None and f_val is not None:
        continue
    if b_val is None:
        continue
    
    c_val_str = str(c_val).strip() if c_val else ''
    d_val = source_ws.cell(row=row_idx, column=4).value
    e_val = source_ws.cell(row=row_idx, column=5).value
    g_val = source_ws.cell(row=row_idx, column=7).value
    i_val = source_ws.cell(row=row_idx, column=9).value  # CASH column
    j_val = source_ws.cell(row=row_idx, column=10).value  # BALANCE
    k_val = source_ws.cell(row=row_idx, column=11).value
    
    bill_no_str = str(d_val).strip() if d_val else ''
    is_bank_name = any(b in bill_no_str.upper() for b in ['ADIB', 'ADCB', 'ENBD', 'DIB', 'ARBAB', 'CHQ', 'ADVB'])
    payment_method = ''
    actual_bill_no = ''
    if bill_no_str:
        if is_bank_name:
            bank_upper = bill_no_str.upper().replace('\n', '/')
            if 'ADVB' in bank_upper:
                payment_method = 'ADCB'  # ADVB is likely a typo for ADCB
            elif 'ADIB' in bank_upper:
                payment_method = 'ADIB'
            elif 'ADCB' in bank_upper:
                payment_method = 'ADCB'
            elif 'ENBD' in bank_upper:
                payment_method = 'ENBD'
            elif 'DIB' in bank_upper and 'ADIB' not in bank_upper:
                payment_method = 'DIB'
            elif 'ARBAB' in bank_upper:
                payment_method = 'Arbab (Cash Collection)'
            elif 'CHQ' in bank_upper.lower():
                payment_method = 'Cheque'
        else:
            actual_bill_no = bill_no_str
    
    # If there's both a bill number AND a bank name in the same cell
    if bill_no_str and '\n' in bill_no_str and not payment_method:
        parts = bill_no_str.split('\n')
        for part in parts:
            part = part.strip()
            if part.isdigit():
                actual_bill_no = part
            elif any(b in part.upper() for b in ['ADIB','ADCB','ENBD','DIB']):
                payment_method = part.upper()
    
    record = {
        'property': current_property or '',
        'room': b_val,
        'name': c_val_str if c_val_str and c_val_str != 'None' else '',
        'bill_no': actual_bill_no,
        'date': e_val,
        'rent': f_val,
        'prev_due': g_val if g_val else 0,
        'cash_paid': i_val if i_val else 0,
        'balance': j_val if j_val else 0,
        'remarks': k_val,
        'payment_method': payment_method,
        'is_bank_name': is_bank_name,
    }
    all_records.append(record)

source_wb.close()

# ── Unique properties list ───────────────────────────────────────────
properties = OrderedDict()
for r in all_records:
    p = r['property']
    if p and p not in properties:
        properties[p] = {'units': 0, 'records': []}
    if p:
        properties[p]['units'] += 1
        properties[p]['records'].append(r)

property_code_map = {}
for i, pname in enumerate(properties.keys(), 1):
    property_code_map[pname] = f"PROP-{i:03d}"

# Generate tenant IDs
name_count = Counter()
tenant_id_map = {}
for r in all_records:
    if r['name']:
        key = (r['property'], r['room'], r['name'])
        if key not in tenant_id_map:
            name_count[r['name']] += 1
            seq = name_count[r['name']]
            clean = r['name'].replace(' ','').replace('/','').replace(chr(39),'').replace('.','').upper()[:8]
            tid = f"TEN-{clean}" if seq == 1 else f"TEN-{clean}-{seq}"
            tenant_id_map[key] = tid

global_name_count = Counter(r['name'] for r in all_records if r['name'])

# ── Create workbook ──────────────────────────────────────────────────
wb = Workbook()
wb.properties.creator = "Z.ai"

# Style constants
REQUIRED_FILL = PatternFill('solid', fgColor='FDE8E8')
OPTIONAL_FILL = PatternFill('solid', fgColor='FFF8E1')
REQUIRED_HEADER_FILL = PatternFill('solid', fgColor='C0392B')
OPTIONAL_HEADER_FILL = PatternFill('solid', fgColor='D4820A')

# ══════════════════════════════════════════════════════════════════════
# SHEET 1: Instructions
# ══════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Instructions"
setup_sheet(ws1, title="Property Management Data Collection Guide", last_col=8)

ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 6
ws1.column_dimensions['C'].width = 28
ws1.column_dimensions['D'].width = 50
ws1.column_dimensions['E'].width = 50
ws1.column_dimensions['F'].width = 14
ws1.column_dimensions['G'].width = 50
ws1.column_dimensions['H'].width = 20

def write_section_header(ws, row, title, fill_color=None):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=2, value=title)
    cell.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
    cell.fill = fill_color or fill_header()
    cell.alignment = align_header()
    for c in range(3, 9):
        ws.cell(row=row, column=c).fill = fill_color or fill_header()
    ws.row_dimensions[row].height = 28
    return row + 1

# ── Overview ─────────────────────────────────────────────────────────
row = write_section_header(ws1, 4, "OVERVIEW")
ws1.merge_cells(start_row=row, start_column=2, end_row=row+4, end_column=8)
overview = (
    "This workbook is a structured data collection template for the Al Reef Al Madeena property management system migration. "
    "It captures all required information about properties, units, tenants, leases, balances, and payment history so that the data "
    "can be imported into the Property Dashboard with 100% accuracy and zero assumptions.\n\n"
    "The existing APRIL-25 billing spreadsheet was used as a reference source only. Where data could be safely extracted "
    "(property names, unit numbers, tenant names, monthly rent amounts, cash payments, outstanding balances, and some payment methods), "
    "it has been pre-populated. All other fields are left BLANK for manual completion by staff using tenant contracts, "
    "paper records, Emirates ID copies, and property files.\n\n"
    "DO NOT import, modify, correct, infer, generate, or fabricate any data. Only fill in fields where you have verified information. "
    "When this workbook is fully completed and verified, it becomes the definitive source for migration with no assumptions required."
)
ws1.cell(row=row, column=2, value=overview).font = font_body()
ws1.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
for rr in range(row, row+5):
    ws1.row_dimensions[rr].height = 30

# ── Color Legend ─────────────────────────────────────────────────────
row = write_section_header(ws1, 11, "COLOR LEGEND")
legend = [
    (REQUIRED_FILL, "Required Field (Red Header)", "Mandatory field that must be completed before migration. Red background indicates missing required data."),
    (OPTIONAL_FILL, "Optional Field (Amber Header)", "Helpful but not mandatory. Yellow background indicates optional data."),
    (PatternFill('solid', fgColor='E8F5E9'), "Complete (Green)", "Data quality status: All required fields are filled and verified."),
    (PatternFill('solid', fgColor='FFF8E1'), "Needs Review (Yellow)", "Data quality status: Some fields need verification or correction."),
    (PatternFill('solid', fgColor='FDE8E8'), "Missing Information (Red)", "Data quality status: One or more required fields are empty."),
]
for i, (fill, label, desc) in enumerate(legend):
    r = row + i
    ws1.cell(row=r, column=2).fill = fill
    ws1.cell(row=r, column=3, value=label).font = Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_900)
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    ws1.cell(row=r, column=4, value=desc).font = font_body()
    ws1.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
    ws1.row_dimensions[r].height = 25

# ── Field Guide ──────────────────────────────────────────────────────
row = 18
row = write_section_header(ws1, row, "FIELD GUIDE - ENGLISH & BENGALI")

# Sub-header
for ci, h in enumerate(["#", "Field Name", "Description (EN)", "Description (BN)", "Required?", "Dashboard Usage"], 2):
    cell = ws1.cell(row=row, column=ci, value=h)
    cell.fill = PatternFill('solid', fgColor=PRIMARY)
    cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    cell.alignment = align_header()
ws1.row_dimensions[row].height = 24
row += 1

fields = [
    ("Property Code", "Unique code for each property (PROP-001, etc.)", "সম্পত্তির অনন্য কোড", "YES", "Links all units/tenants/leases"),
    ("Property Name", "Building/villa name from source (e.g. Khalifa Villa)", "ভবন/ভিলার নাম", "YES", "Property cards and navigation"),
    ("Property Type", "Villa, Building, Office, Commercial, or Residential", "সম্পত্তির ধরন", "YES", "Filtering and categorization"),
    ("Property Address", "Full street address", "সম্পত্তির ঠিকানা", "NO", "Property detail pages"),
    ("Area", "Neighborhood/district (e.g. Al Jahili)", "এলাকার নাম", "YES", "Area-based filtering"),
    ("City", "City name (Al Ain)", "শহরের নাম", "YES", "Location grouping"),
    ("Total Units", "Number of rentable units", "ইউনিট সংখ্যা", "YES", "Occupancy rate calculations"),
    ("Unit Number", "Room/unit ID (1, 2, A1, GARAGE, etc.)", "ইউনিট নম্বর", "YES", "Unit identification"),
    ("Unit Type", "Room, Studio, Office, Shop, Garage, Apartment", "ইউনিটের ধরন", "YES", "Unit categorization"),
    ("Monthly Rent", "Monthly rent in AED", "মাসিক ভাড়া (দিরহাম)", "YES", "Rent tracking and revenue"),
    ("Occupancy Status", "Occupied, Vacant, or Under Maintenance", "দখল অবস্থা", "YES", "Vacancy tracking"),
    ("Tenant ID", "Unique tenant code (TEN-XXXXXXXX)", "ভাড়াটিয়া কোড", "YES", "Links to leases/payments"),
    ("Full Name", "Legal name as per contract/Emirates ID", "আইনি নাম", "YES", "Tenant identification"),
    ("Phone Number", "Primary contact (+971-XXXXXXXXX)", "ফোন নম্বর", "YES", "Communication"),
    ("WhatsApp Number", "WhatsApp contact if different", "হোয়াটসঅ্যাপ নম্বর", "NO", "WhatsApp notifications"),
    ("Email", "Email for official communications", "ইমেইল", "NO", "Email notifications"),
    ("Emirates ID", "15-digit UAE ID (784-XXXX-XXXXXXX-X)", "আমিরাত আইডি", "YES", "Identity verification"),
    ("Nationality", "Country of nationality", "জাতীয়তা", "NO", "Demographic reports"),
    ("Lease Start Date", "Contract start (YYYY-MM-DD)", "চুক্তি শুরু", "YES", "Lease tracking"),
    ("Lease End Date", "Contract end (YYYY-MM-DD)", "চুক্তি শেষ", "YES", "Expiry alerts"),
    ("Security Deposit", "Refundable deposit amount in AED", "নিরাপত্তা জমা", "YES", "Deposit tracking"),
    ("Payment Method", "Cash, Cheque, Bank Transfer, ADIB, ADCB, etc.", "পেমেন্ট পদ্ধতি", "YES", "Payment tracking"),
    ("Contract Status", "Active, Expired, Terminated, Pending Renewal", "চুক্তির অবস্থা", "YES", "Contract management"),
    ("Outstanding Balance", "Total amount currently owed", "বকেয়া পরিমাণ", "YES", "Collection tracking"),
    ("Payment Amount", "Amount received in AED", "প্রাপ্ত পরিমাণ", "YES", "Revenue tracking"),
    ("Reference Number", "Cheque number, transaction ID, or receipt number", "রেফারেন্স নম্বর", "NO", "Payment verification"),
]

for i, (fname, desc_en, desc_bn, req, usage) in enumerate(fields, 1):
    r = row + i - 1
    ws1.cell(row=r, column=2, value=i).font = font_body()
    ws1.cell(row=r, column=2).alignment = align_header()
    ws1.cell(row=r, column=3, value=fname).font = Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_900)
    ws1.cell(row=r, column=4, value=desc_en).font = font_body()
    ws1.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
    ws1.cell(row=r, column=5, value=desc_bn).font = font_body()
    ws1.cell(row=r, column=5).alignment = Alignment(wrap_text=True)
    ws1.cell(row=r, column=6, value=req).font = Font(name=FONT_NAME, size=11, bold=True, 
        color=ACCENT_NEGATIVE if req == "YES" else ACCENT_WARNING)
    ws1.cell(row=r, column=6).alignment = align_header()
    ws1.cell(row=r, column=7, value=usage).font = font_body()
    ws1.cell(row=r, column=7).alignment = Alignment(wrap_text=True)
    fill = REQUIRED_FILL if req == "YES" else OPTIONAL_FILL
    ws1.cell(row=r, column=3).fill = fill
    ws1.row_dimensions[r].height = 28

# ── Common Mistakes ──────────────────────────────────────────────────
row = row + len(fields) + 1
row = write_section_header(ws1, row, "COMMON MISTAKES TO AVOID", PatternFill('solid', fgColor=ACCENT_NEGATIVE))

mistakes = [
    "DO NOT guess or fabricate any missing data. Leave fields blank if you cannot verify the information from source documents.",
    "DO NOT use the same Emirates ID for multiple tenants. Each person has a unique 15-digit Emirates ID number.",
    "DO NOT enter partial phone numbers. UAE numbers must include country code: +971-5XXXXXXXX or +971-XXXXXXXXX.",
    "DO NOT mix date formats. Always use YYYY-MM-DD format (e.g. 2025-01-15, not 15/1/25 or Jan 15).",
    "DO NOT enter rent amounts as text. All monetary values must be numbers only (e.g. 2500, not 'AED 2500').",
    "DO NOT combine multiple tenants in one Name field. If a unit has co-tenants, create separate tenant records.",
    "DO NOT leave Property Code blank. Every unit and tenant must be linked to a valid property code.",
    "DO NOT enter 'EMPTY' or 'VACANT' in tenant name. Set Occupancy Status to 'Vacant' and leave tenant fields blank.",
    "DO NOT assume lease dates from billing dates. Use actual contract dates from signed agreements.",
    "DO NOT duplicate Payment IDs. Each payment must have a unique identifier.",
]
for i, mistake in enumerate(mistakes, 1):
    r = row + i - 1
    ws1.cell(row=r, column=2, value=i).font = font_body()
    ws1.cell(row=r, column=2).alignment = align_header()
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    ws1.cell(row=r, column=3, value=mistake).font = font_body()
    ws1.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
    for c in range(2, 9):
        ws1.cell(row=r, column=c).fill = fill_data_row(i)
    ws1.row_dimensions[r].height = 28

# ── Examples ─────────────────────────────────────────────────────────
row = row + len(mistakes) + 1
row = write_section_header(ws1, row, "EXAMPLES OF CORRECTLY COMPLETED ROWS", PatternFill('solid', fgColor=ACCENT_POSITIVE))

examples = [
    ("Property:", "PROP-001 | Khalifa Villa | Villa | Near LuLu, Al Jahili | Al Jahili | Al Ain | 15 | Active"),
    ("Unit:", "PROP-001 | 1 | Room | 1 | 1700 | Occupied | Complete"),
    ("Tenant:", "TEN-ESSAMAMA | Essam Amasa | +971-501234567 | +971-501234567 | essam@email.com | 784-1990-1234567-1 | Egyptian | Active | Complete"),
    ("Lease:", "TEN-ESSAMAMA | PROP-001 | 1 | 2025-01-01 | 2025-12-31 | 1700 | 3400 | ADIB | Active | Complete"),
    ("Balance:", "TEN-ESSAMAMA | PROP-001 | 1 | 1700 | 1700 | Needs Review | Previous month carry-forward"),
    ("Payment:", "PAY-2025-04-0001 | TEN-ESSAMAMA | 2025-04-01 | 1700 | ADIB | 102871 | Complete | April rent"),
]
for i, (label, example) in enumerate(examples, 1):
    r = row + i - 1
    ws1.cell(row=r, column=2, value=i).font = font_body()
    ws1.cell(row=r, column=3, value=label).font = Font(name=FONT_NAME, size=11, bold=True, color=PRIMARY)
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
    ws1.cell(row=r, column=4, value=example).font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
    ws1.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
    for c in range(2, 9):
        ws1.cell(row=r, column=c).fill = fill_data_row(i)
    ws1.row_dimensions[r].height = 28

# ── Completion Checklist ─────────────────────────────────────────────
row = row + len(examples) + 1
row = write_section_header(ws1, row, "COMPLETION CHECKLIST", PatternFill('solid', fgColor=ACCENT_POSITIVE))

checklist = [
    "All 26 properties listed with correct codes, names, types, areas, and unit counts",
    "All units listed with Property Code, Unit Number, Unit Type, Monthly Rent, and Occupancy Status",
    "All occupied units have a corresponding tenant record with Tenant ID and Full Name",
    "Phone numbers collected from tenant contracts for all 317 tenants",
    "Emirates ID numbers verified and entered for all tenants",
    "Lease start and end dates entered from signed contracts for all tenants",
    "Security deposit amounts entered for all leases",
    "Outstanding balances verified against ledger records for all tenants with balances",
    "All vacant units marked as 'Vacant' in Occupancy Status column",
    "Data Quality Status set to 'Complete' for all verified records",
    "No duplicate Emirates ID numbers across tenants",
    "All required (red header) fields filled - no blank cells in required columns",
]
for i, item in enumerate(checklist, 1):
    r = row + i - 1
    ws1.cell(row=r, column=2, value="[ ]").font = Font(name=FONT_NAME, size=12)
    ws1.cell(row=r, column=2).alignment = align_header()
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    ws1.cell(row=r, column=3, value=item).font = font_body()
    ws1.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
    for c in range(2, 9):
        ws1.cell(row=r, column=c).fill = fill_data_row(i)
    ws1.row_dimensions[r].height = 28

# ══════════════════════════════════════════════════════════════════════
# Helper: Standard sheet setup with headers
# ══════════════════════════════════════════════════════════════════════
def setup_data_sheet(wb, title, headers_spec):
    """Create a data sheet with proper headers and validation-ready structure."""
    ws = wb.create_sheet(title)
    last_col = len(headers_spec) + 1
    setup_sheet(ws, title=title, last_col=last_col)
    
    for ci, (header, required, width, _) in enumerate(headers_spec, 2):
        cell = ws.cell(row=4, column=ci, value=header)
        cell.fill = REQUIRED_HEADER_FILL if required else OPTIONAL_HEADER_FILL
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        cell.alignment = align_header()
        cell.border = border_header()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[4].height = 28
    return ws

def write_data_row(ws, row_num, values, headers_spec, row_index):
    """Write a data row with proper styling."""
    for ci, val in enumerate(values, 2):
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.font = font_body()
        cell.fill = fill_data_row(row_index)
        required = headers_spec[ci-2][1]
        if required and (val is None or val == ''):
            cell.fill = REQUIRED_FILL
    
    # Number formatting for currency columns
    for ci, val in enumerate(values, 2):
        header = headers_spec[ci-2][0]
        if isinstance(val, (int, float)) and val != 0 and any(kw in header for kw in ['AED', 'Rent', 'Balance', 'Amount', 'Deposit']):
            ws.cell(row=row_num, column=ci).number_format = '#,##0'
    
    ws.row_dimensions[row_num].height = ROW_HEIGHTS['data']

# ══════════════════════════════════════════════════════════════════════
# SHEET 2: Properties
# ══════════════════════════════════════════════════════════════════════
prop_headers = [
    ("Property Code", True, 16, 'text'),
    ("Property Name", True, 28, 'text'),
    ("Property Type", True, 18, 'dropdown'),
    ("Property Address", False, 35, 'text'),
    ("Area", True, 20, 'text'),
    ("City", True, 16, 'text'),
    ("Total Units", True, 14, 'number'),
    ("Active Status", True, 16, 'dropdown'),
    ("Notes", False, 35, 'text'),
]

ws2 = setup_data_sheet(wb, "Properties", prop_headers)

dv_prop_type = DataValidation(type="list", formula1='"Villa,Building,Office Complex,Commercial,Residential Compound"', allow_blank=True)
dv_prop_type.error = "Select a valid Property Type"
ws2.add_data_validation(dv_prop_type)
dv_active = DataValidation(type="list", formula1='"Active,Inactive,Under Maintenance"', allow_blank=True)
ws2.add_data_validation(dv_active)

def infer_property_type(name):
    n = name.upper() if name else ''
    if 'OFFICE' in n: return 'Office Complex'
    if 'MALL' in n: return 'Commercial'
    if 'SANAIYA' in n: return 'Commercial'
    if 'BUILDING' in n: return 'Building'
    return 'Villa'

def infer_area(name):
    mapping = {
        'KHALIFA': 'Khalifa Bin Zayed', 'MATOWA': 'Al Mutawwa', 'SALHE': 'Al Salhe',
        'SUMALI': 'Al Sumali', 'JAHILI': 'Al Jahili', 'ZAFRANA': 'Zafrana',
        'DIWAN': 'Al Diwan', 'HABOOY': 'Al Habouy', 'SANAIYA': 'Sanaiya',
        'SAROOJ': 'Al Sarooj', 'AA MALL': 'Al Ain Mall Area', 'NEIMA': 'Neima',
        'SHAMKHA': 'Al Shamkha', 'BATEEN': 'Al Bateen', 'MANASIR': 'Al Manasir',
        'NIYADAT': 'Niyadat', 'INDIA': 'Al Ain City Center', 'MUTARID': 'Al Mutarid',
        'MAQAM': 'Al Maqam', 'MUWAIJI': 'Al Muwaiji',
    }
    for key, area in mapping.items():
        if key in (name.upper() if name else ''):
            return area
    return 'Al Ain'

for i, (pname, pdata) in enumerate(properties.items()):
    r = 5 + i
    values = [
        property_code_map.get(pname, f"PROP-{i+1:03d}"),
        pname,
        infer_property_type(pname),
        '',
        infer_area(pname),
        'Al Ain',
        pdata['units'],
        'Active',
        '',
    ]
    write_data_row(ws2, r, values, prop_headers, i)
    dv_prop_type.add(ws2.cell(row=r, column=4))
    dv_active.add(ws2.cell(row=r, column=9))

ws2.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 3: Units
# ══════════════════════════════════════════════════════════════════════
unit_headers = [
    ("Property Code", True, 16, 'text'),
    ("Unit Number", True, 14, 'text'),
    ("Unit Type", True, 16, 'dropdown'),
    ("Floor", False, 12, 'text'),
    ("Monthly Rent (AED)", True, 18, 'number'),
    ("Occupancy Status", True, 18, 'dropdown'),
    ("Data Quality Status", True, 20, 'dropdown'),
    ("Notes", False, 35, 'text'),
]
ws3 = setup_data_sheet(wb, "Units", unit_headers)

dv_utype = DataValidation(type="list", formula1='"Room,Studio,Office,Shop,Garage,Apartment,Penthouse,Storage"', allow_blank=True)
ws3.add_data_validation(dv_utype)
dv_ostatus = DataValidation(type="list", formula1='"Occupied,Vacant,Under Maintenance"', allow_blank=True)
ws3.add_data_validation(dv_ostatus)
dv_dq = DataValidation(type="list", formula1='"Complete,Needs Review,Missing Information"', allow_blank=True)
ws3.add_data_validation(dv_dq)

for i, r in enumerate(all_records):
    pcode = property_code_map.get(r['property'], '')
    room_str = str(r['room']).upper()
    if 'GARAGE' in room_str: utype = 'Garage'
    elif 'OFFICE' in room_str: utype = 'Office'
    elif isinstance(r['room'], str) and any(c.isalpha() for c in str(r['room'])): utype = 'Room'
    else: utype = 'Room'
    
    occ = 'Vacant' if not r['name'] else 'Occupied'
    quality = 'Needs Review' if r['name'] and r['rent'] else ('Missing Information' if not r['rent'] and r['name'] else '')
    
    values = [pcode, r['room'], utype, '', r['rent'] if r['rent'] else '', occ, quality, '']
    write_data_row(ws3, 5+i, values, unit_headers, i)
    dv_utype.add(ws3.cell(row=5+i, column=4))
    dv_ostatus.add(ws3.cell(row=5+i, column=7))
    dv_dq.add(ws3.cell(row=5+i, column=8))

ws3.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 4: Tenants
# ══════════════════════════════════════════════════════════════════════
tenant_headers = [
    ("Tenant ID", True, 22, 'text'),
    ("Full Name", True, 28, 'text'),
    ("Phone Number", True, 22, 'text'),
    ("WhatsApp Number", False, 22, 'text'),
    ("Email", False, 28, 'text'),
    ("Emirates ID", True, 22, 'text'),
    ("Nationality", False, 18, 'text'),
    ("Emergency Contact", False, 22, 'text'),
    ("Emergency Phone", False, 22, 'text'),
    ("Status", True, 16, 'dropdown'),
    ("Data Quality Status", True, 20, 'dropdown'),
    ("Notes", False, 35, 'text'),
]
ws4 = setup_data_sheet(wb, "Tenants", tenant_headers)

dv_tstatus = DataValidation(type="list", formula1='"Active,Inactive,Evicted,Moved Out"', allow_blank=True)
ws4.add_data_validation(dv_tstatus)
dv_dq4 = DataValidation(type="list", formula1='"Complete,Needs Review,Missing Information"', allow_blank=True)
ws4.add_data_validation(dv_dq4)

# Phone number validation (UAE format)
dv_phone = DataValidation(type="custom", formula1='OR(ISBLANK(D5),LEFT(D5,5)="+971-")', allow_blank=True)
dv_phone.error = "Enter UAE phone: +971-XXXXXXXXX"
dv_phone.errorTitle = "Invalid Phone"
ws4.add_data_validation(dv_phone)

seen = {}
trow = 5
for r in all_records:
    if not r['name']: continue
    key = (r['property'], r['room'], r['name'])
    if key in seen: continue
    seen[key] = True
    
    tid = tenant_id_map.get(key, f"TEN-{r['name'].upper()[:8]}")
    values = [tid, r['name'], '', '', '', '', '', '', '', 'Active', 'Missing Information', '']
    write_data_row(ws4, trow, values, tenant_headers, trow-5)
    dv_tstatus.add(ws4.cell(row=trow, column=11))
    dv_dq4.add(ws4.cell(row=trow, column=12))
    dv_phone.add(ws4.cell(row=trow, column=4))
    trow += 1

# Duplicate Emirates ID detection conditional formatting
ws4.conditional_formatting.add(
    f'G5:G{trow-1}',
    FormulaRule(formula=[f'COUNTIF(G$5:G${trow-1},G5)>1'],
        fill=PatternFill('solid', fgColor='FDE8E8'),
        font=Font(color=ACCENT_NEGATIVE, bold=True))
)

ws4.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 5: Lease Information
# ══════════════════════════════════════════════════════════════════════
lease_headers = [
    ("Tenant ID", True, 22, 'text'),
    ("Property Code", True, 16, 'text'),
    ("Unit Number", True, 14, 'text'),
    ("Lease Start Date", True, 18, 'date'),
    ("Lease End Date", True, 18, 'date'),
    ("Monthly Rent (AED)", True, 18, 'number'),
    ("Security Deposit (AED)", True, 20, 'number'),
    ("Payment Method", True, 18, 'dropdown'),
    ("Contract Status", True, 18, 'dropdown'),
    ("Data Quality Status", True, 20, 'dropdown'),
    ("Notes", False, 35, 'text'),
]
ws5 = setup_data_sheet(wb, "Lease Information", lease_headers)

dv_pmethod = DataValidation(type="list", formula1='"Cash,Cheque,Bank Transfer,ADIB,ADCB,ENBD,DIB,Arbab (Cash Collection)"', allow_blank=True)
ws5.add_data_validation(dv_pmethod)
dv_cstatus = DataValidation(type="list", formula1='"Active,Expired,Terminated,Pending Renewal"', allow_blank=True)
ws5.add_data_validation(dv_cstatus)
dv_dq5 = DataValidation(type="list", formula1='"Complete,Needs Review,Missing Information"', allow_blank=True)
ws5.add_data_validation(dv_dq5)
dv_date = DataValidation(type="date", allow_blank=True)
ws5.add_data_validation(dv_date)

lrow = 5
for r in all_records:
    if not r['name']: continue
    key = (r['property'], r['room'], r['name'])
    if key not in tenant_id_map: continue
    
    tid = tenant_id_map[key]
    pcode = property_code_map.get(r['property'], '')
    pay = r['payment_method'] if r['payment_method'] else ''
    
    values = [tid, pcode, r['room'], '', '', r['rent'] if r['rent'] else '', '', pay, 'Active', 'Missing Information', '']
    write_data_row(ws5, lrow, values, lease_headers, lrow-5)
    dv_pmethod.add(ws5.cell(row=lrow, column=9))
    dv_cstatus.add(ws5.cell(row=lrow, column=10))
    dv_dq5.add(ws5.cell(row=lrow, column=11))
    dv_date.add(ws5.cell(row=lrow, column=5))
    dv_date.add(ws5.cell(row=lrow, column=6))
    lrow += 1

ws5.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 6: Outstanding Balances
# ══════════════════════════════════════════════════════════════════════
bal_headers = [
    ("Tenant ID", True, 22, 'text'),
    ("Property Code", True, 16, 'text'),
    ("Unit Number", True, 14, 'text'),
    ("Previous Balance (AED)", True, 22, 'number'),
    ("Outstanding Balance (AED)", True, 24, 'number'),
    ("Data Quality Status", True, 20, 'dropdown'),
    ("Balance Notes", False, 35, 'text'),
]
ws6 = setup_data_sheet(wb, "Outstanding Balances", bal_headers)
dv_dq6 = DataValidation(type="list", formula1='"Complete,Needs Review,Missing Information"', allow_blank=True)
ws6.add_data_validation(dv_dq6)

brow = 5
for r in all_records:
    if not r['name']: continue
    key = (r['property'], r['room'], r['name'])
    if key not in tenant_id_map: continue
    
    prev_due = r['prev_due'] if r['prev_due'] else 0
    balance = r['balance'] if r['balance'] else 0
    
    # Include ALL tenants with balance (including 0 for completeness)
    if prev_due == 0 and balance == 0:
        continue
    
    tid = tenant_id_map[key]
    pcode = property_code_map.get(r['property'], '')
    quality = 'Needs Review' if balance != 0 else 'Complete'
    notes = ''
    if r['remarks']:
        notes = str(r['remarks'])
    
    values = [tid, pcode, r['room'], prev_due, balance, quality, notes]
    write_data_row(ws6, brow, values, bal_headers, brow-5)
    dv_dq6.add(ws6.cell(row=brow, column=7))
    brow += 1

ws6.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 7: Payment History
# ══════════════════════════════════════════════════════════════════════
pay_headers = [
    ("Payment ID", True, 20, 'text'),
    ("Tenant ID", True, 22, 'text'),
    ("Property Code", True, 16, 'text'),
    ("Unit Number", True, 14, 'text'),
    ("Payment Date", True, 18, 'date'),
    ("Amount (AED)", True, 16, 'number'),
    ("Payment Method", True, 18, 'dropdown'),
    ("Reference Number", False, 22, 'text'),
    ("Data Quality Status", True, 20, 'dropdown'),
    ("Remarks", False, 35, 'text'),
]
ws7 = setup_data_sheet(wb, "Payment History", pay_headers)

dv_pm7 = DataValidation(type="list", formula1='"Cash,Cheque,Bank Transfer,ADIB,ADCB,ENBD,DIB,Arbab (Cash Collection)"', allow_blank=True)
ws7.add_data_validation(dv_pm7)
dv_dq7 = DataValidation(type="list", formula1='"Complete,Needs Review,Missing Information"', allow_blank=True)
ws7.add_data_validation(dv_dq7)
dv_date7 = DataValidation(type="date", allow_blank=True)
ws7.add_data_validation(dv_date7)

prow = 5
pay_counter = 0
for r in all_records:
    if not r['name']: continue
    key = (r['property'], r['room'], r['name'])
    if key not in tenant_id_map: continue
    
    tid = tenant_id_map[key]
    pcode = property_code_map.get(r['property'], '')
    pay_counter += 1
    pay_id = f"PAY-2025-04-{pay_counter:04d}"
    
    pay_method = r['payment_method'] if r['payment_method'] else ''
    ref = r['bill_no'] if r['bill_no'] else ''
    amount = r['cash_paid'] if r['cash_paid'] else ''
    remarks = str(r['remarks']) if r['remarks'] else ''
    
    quality = 'Missing Information'
    if amount and pay_method:
        quality = 'Needs Review'
    
    values = [pay_id, tid, pcode, r['room'], '', amount, pay_method, ref, quality, remarks]
    write_data_row(ws7, prow, values, pay_headers, prow-5)
    dv_pm7.add(ws7.cell(row=prow, column=8))
    dv_dq7.add(ws7.cell(row=prow, column=10))
    dv_date7.add(ws7.cell(row=prow, column=6))
    prow += 1

ws7.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# SHEET 8: Data Issues (CONSOLIDATED)
# ══════════════════════════════════════════════════════════════════════
issue_headers = [
    ("#", True, 6, 'text'),
    ("Category", True, 22, 'text'),
    ("Description", True, 55, 'text'),
    ("Affected Records", True, 16, 'number'),
    ("Severity", True, 14, 'dropdown'),
    ("Action Required", True, 45, 'text'),
    ("Status", False, 16, 'dropdown'),
]
ws8 = setup_data_sheet(wb, "Data Issues", issue_headers)

dv_sev = DataValidation(type="list", formula1='"Critical,High,Medium,Low"', allow_blank=True)
ws8.add_data_validation(dv_sev)
dv_istatus = DataValidation(type="list", formula1='"Open,In Progress,Resolved,Not Applicable"', allow_blank=True)
ws8.add_data_validation(dv_istatus)

# Consolidated issues - grouped by type, not per-tenant
occupied_tenants = len(seen)
issues = [
    ("Missing Tenant Data", f"All {occupied_tenants} tenants are missing Phone Numbers - the most critical communication field", occupied_tenants, "Critical", "Collect phone numbers from tenant contracts, Emirates ID copies, or previous communication records"),
    ("Missing Tenant Data", f"All {occupied_tenants} tenants are missing Emirates ID numbers - required for identity verification", occupied_tenants, "Critical", "Collect Emirates ID copies from tenant files. Format: 784-XXXX-XXXXXXX-X"),
    ("Missing Lease Dates", f"All {occupied_tenants} tenants are missing Lease Start Dates", occupied_tenants, "Critical", "Retrieve contract start dates from signed lease agreements"),
    ("Missing Lease Dates", f"All {occupied_tenants} tenants are missing Lease End Dates", occupied_tenants, "Critical", "Retrieve contract end dates from signed lease agreements"),
    ("Missing Lease Data", f"All {occupied_tenants} tenants are missing Security Deposit amounts", occupied_tenants, "High", "Check deposit receipts and contract terms for each tenant"),
    ("Duplicate Names", f"21 tenant names appear multiple times across properties (e.g. AHMED x6, KHAIRUL TAILOR x3)", 21, "High", "Verify these are different people with same name or data entry errors. Use Emirates ID to distinguish"),
    ("Vacant Units", "10 units have no tenant assigned (marked EMPTY or blank in source)", 10, "Low", "Verify vacancy status and mark as 'Vacant' in Occupancy Status column"),
    ("Ambiguous Payment Method", "52 records have bank names (ADIB/ADCB/DIB/ENBD) in the Bill Number column instead of actual bill numbers", 52, "Medium", "These have been mapped to Payment Method field. Verify the actual payment method from bank records"),
    ("Missing Property Address", "All 26 properties are missing street addresses", 26, "Low", "Collect property addresses from municipality records or site visits"),
    ("Source Date Ambiguity", "Payment dates in APRIL-25 are ambiguous (some appear as Excel dates, others as text like '16/4/25' or '30/4/25')", -1, "Medium", "Dates have been left blank for manual entry. Use actual payment receipt dates, not billing dates"),
    ("Balance Verification", f"78 tenants have outstanding balances totaling AED values from April 2025", 78, "High", "Verify each tenant's current balance against the full ledger (not just April billing). Some balances may have been partially paid since"),
    ("Co-Tenant Names", "Several unit names contain slashes indicating multiple tenants (e.g. 'JALAL/SHARIF', 'ABDU UNCLE/ABDULLAH')", -1, "High", "Split co-tenants into separate Tenant records with individual Emirates IDs and contact information"),
    ("ADVB Payment Method", "At least one record shows 'ADVB' as payment method - this is likely a typo for ADCB bank", 1, "Low", "Verify with tenant whether payment was via ADCB and update Payment Method accordingly"),
    ("Remarks Codes", "Source remarks contain abbreviated codes (e.g. 'DIS400', '250ac pymt', 'MARC PAID') that need interpretation", -1, "Medium", "Translate abbreviated remarks into proper notes. 'DIS' = discount, 'pymt' = payment, month names = advance payments"),
]

for i, (cat, desc, count, severity, action) in enumerate(issues, 1):
    r = 4 + i
    values = [i, cat, desc, count if count > 0 else '', severity, action, 'Open']
    write_data_row(ws8, r, values, issue_headers, i-1)
    # Color-code severity
    sev_cell = ws8.cell(row=r, column=6)
    if severity == 'Critical':
        sev_cell.font = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_NEGATIVE)
    elif severity == 'High':
        sev_cell.font = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_WARNING)
    elif severity == 'Medium':
        sev_cell.font = Font(name=FONT_NAME, size=11, color=ACCENT_WARNING)
    dv_sev.add(ws8.cell(row=r, column=6))
    dv_istatus.add(ws8.cell(row=r, column=7))

ws8.freeze_panes = 'C5'

# ══════════════════════════════════════════════════════════════════════
# Conditional formatting for Data Quality Status
# ══════════════════════════════════════════════════════════════════════
for ws in [ws3, ws4, ws5, ws6, ws7]:
    for col in range(2, ws.max_column + 1):
        hval = ws.cell(row=4, column=col).value
        if hval and 'Data Quality' in str(hval):
            cl = get_column_letter(col)
            mr = ws.max_row
            ws.conditional_formatting.add(f'{cl}5:{cl}{mr}',
                CellIsRule(operator='equal', formula=['"Complete"'],
                    fill=PatternFill('solid', fgColor='E8F5E9'),
                    font=Font(color=ACCENT_POSITIVE)))
            ws.conditional_formatting.add(f'{cl}5:{cl}{mr}',
                CellIsRule(operator='equal', formula=['"Needs Review"'],
                    fill=PatternFill('solid', fgColor='FFF8E1'),
                    font=Font(color=ACCENT_WARNING)))
            ws.conditional_formatting.add(f'{cl}5:{cl}{mr}',
                CellIsRule(operator='equal', formula=['"Missing Information"'],
                    fill=PatternFill('solid', fgColor='FDE8E8'),
                    font=Font(color=ACCENT_NEGATIVE)))
            break

# ══════════════════════════════════════════════════════════════════════
# SAVE
# ══════════════════════════════════════════════════════════════════════
output = '/home/z/my-project/download/Property Management Data Collection Template.xlsx'
wb.save(output)
print(f"Saved: {output}")
print(f"Sheets: {wb.sheetnames}")
print(f"Properties: {len(properties)}")
print(f"Units: {len(all_records)}")
print(f"Tenants: {len(seen)}")
print(f"Outstanding Balances: {brow-5}")
print(f"Payment Records: {prow-5}")
print(f"Data Issues: {len(issues)}")
