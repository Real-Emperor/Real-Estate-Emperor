#!/usr/bin/env python3
"""
Generate POS_Premium_Sample.xlsx — Premium POS demo file for client presentation.
Bloomberg palette, 10 sheets, realistic electronics store data.
"""

import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'skills', 'xlsx'))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import Reference
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta
import random

from templates.base import (
    use_palette_explicit, setup_sheet, style_header_row, style_data_row,
    style_total_row, font_title, font_header, font_body, font_caption,
    font_kpi, font_kpi_label, font_subheader,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    COLUMN_WIDTHS, ROW_HEIGHTS, FORMATS,
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_50, NEUTRAL_0,
    HEADER_TEXT, CHART_COLORS,
    CF_POSITIVE_FILL, CF_POSITIVE_FONT, CF_NEGATIVE_FILL, CF_NEGATIVE_FONT, CF_WARNING_FILL, CF_WARNING_FONT,
    create_bar_chart, create_line_chart, create_pie_chart,
    setup_chart_titles, apply_chart_colors, apply_pie_colors,
    auto_fit_columns, make_chart_title,
)

# ══════════════════════════════════════════════════════════════
# Activate Bloomberg palette FIRST
# ══════════════════════════════════════════════════════════════
use_palette_explicit("bloomberg")

# Re-import tokens after palette switch (they are now updated)
from templates.base import (
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_50, NEUTRAL_0,
    HEADER_TEXT, CHART_COLORS,
)

wb = Workbook()
wb.properties.creator = "Z.ai"

CURRENCY_FMT = "#,##0.00"
INT_FMT = "#,##0"
PCT_FMT = "0.0%"
DATE_FMT = "DD-MMM-YYYY"

# Tab colour from PRIMARY
TAB_COLOR = PRIMARY  # 0D1B2A

# ══════════════════════════════════════════════════════════════
# DATA CONSTANTS
# ══════════════════════════════════════════════════════════════
STAFF = ["Ahmed", "Fatima", "Omar", "Sara", "Khalid", "Maryam"]
SHIFTS = ["Morning", "Afternoon", "Evening", "Morning", "Afternoon", "Morning"]
SALES_TARGETS = [25000, 28000, 22000, 26000, 20000, 24000]

ITEMS = [
    ("Phone Case", "Accessories", 45, 20),
    ("USB Cable", "Accessories", 25, 10),
    ("Screen Protector", "Accessories", 30, 12),
    ("Power Bank", "Accessories", 120, 55),
    ("Earbuds", "Audio", 85, 40),
    ("Bluetooth Speaker", "Audio", 150, 70),
    ("Headphones", "Audio", 200, 90),
    ("Tablet", "Devices", 899, 540),
    ("Smart Watch", "Wearables", 450, 270),
    ("Fitness Band", "Wearables", 180, 90),
    ("Laptop Stand", "Accessories", 150, 65),
    ("Keyboard", "Peripherals", 200, 100),
    ("Mouse", "Peripherals", 80, 35),
    ("Charger", "Accessories", 55, 22),
    ("HDMI Cable", "Cables", 35, 14),
    ("Type-C Hub", "Accessories", 95, 42),
    ("Wireless Charger", "Accessories", 110, 50),
    ("Memory Card 64GB", "Storage", 45, 18),
    ("USB Flash Drive", "Storage", 30, 12),
    ("Portable SSD", "Storage", 250, 140),
    ("Webcam", "Peripherals", 180, 85),
    ("Monitor Cable", "Cables", 40, 16),
    ("Phone Holder", "Accessories", 35, 14),
    ("Stylus Pen", "Accessories", 65, 28),
    ("Car Mount", "Accessories", 50, 20),
    ("Surge Protector", "Accessories", 75, 32),
    ("Cleaning Kit", "Accessories", 25, 8),
    ("Mouse Pad XL", "Peripherals", 40, 15),
    ("Adapter Plug", "Accessories", 20, 7),
    ("AV Cable", "Cables", 28, 10),
    ("Ethernet Cable", "Cables", 22, 8),
    ("Audio Splitter", "Audio", 18, 6),
    ("Ring Light", "Accessories", 120, 52),
    ("Tripod Stand", "Accessories", 90, 38),
]

PAYMENT_METHODS = ["Cash", "Card", "Online", "Bank Transfer"]
PAYMENT_WEIGHTS = [0.25, 0.45, 0.20, 0.10]

SUPPLIERS = ["Samsung Gulf", "Apple ME", "Dell Middle East", "Accessories Plus", "TechParts UAE"]

CUSTOMERS = [
    "Mohamed Al Rashid", "Aisha Al Maktoum", "Hassan Al Nuaimi", "Fatma Al Suwaidi",
    "Omar Khalaf", "Layla Al Qasimi", "Youssef Al Mansoori", "Noor Al Hashmi",
    "Saeed Al Dhaheri", "Mariam Al Shamsi", "Ali Al Falasi", "Huda Al Ameri",
    "Khalid Al Ketbi", "Salma Al Tamimi", "Rashid Al Zaabi", "Amna Al Otaiba",
    "Walk-in", "Walk-in", "Walk-in", "Walk-in", "Walk-in",
]

random.seed(42)

# ══════════════════════════════════════════════════════════════
# HELPER: set column widths
# ══════════════════════════════════════════════════════════════
def set_col_widths(ws, widths_dict):
    """Set column widths. widths_dict: {col_letter: width}"""
    for col, w in widths_dict.items():
        ws.column_dimensions[col].width = w

def apply_tab_color(ws):
    ws.sheet_properties.tabColor = TAB_COLOR

def add_print_area(ws, last_col, last_row):
    ws.print_area = f"B1:{get_column_letter(last_col)}{last_row}"
    ws.print_title_rows = '1:4'

# ══════════════════════════════════════════════════════════════
# SHEET 1: Sales Entry
# ══════════════════════════════════════════════════════════════
def create_sales_entry():
    ws = wb.active
    ws.title = "Sales Entry"
    apply_tab_color(ws)

    headers = ["Date", "Invoice #", "Item/Service", "Category", "Qty",
               "Unit Price", "Total", "Payment Method", "Staff Name",
               "Customer Name", "Discount", "Net Total", "Notes"]

    col_start = 2  # B
    col_end = col_start + len(headers) - 1  # N
    last_col = col_end

    setup_sheet(ws, title="Sales Entry — TechZone Electronics", last_col=last_col)

    # Header row = 4
    hr = 4
    for i, h in enumerate(headers):
        ws.cell(row=hr, column=col_start + i, value=h)
    style_header_row(ws, hr, col_start, col_end)

    # Generate 110 transactions
    data_start = 5
    invoice_num = 100
    rows_data = []

    for day in range(1, 32):
        d = date(2025, 3, day)
        # 3-5 transactions per day
        num_txns = random.randint(3, 5)
        for t in range(num_txns):
            invoice_num += 1
            item = random.choice(ITEMS)
            item_name, category, unit_price, cost = item
            qty = random.randint(1, 4)
            total = qty * unit_price
            payment = random.choices(PAYMENT_METHODS, weights=PAYMENT_WEIGHTS, k=1)[0]
            staff = random.choice(STAFF)
            customer = random.choice(CUSTOMERS)
            discount = 0
            if random.random() < 0.3:
                discount = random.choice([10, 15, 20, 25, 30, 40, 50])
            net_total = total - discount
            notes = ""
            if discount > 0:
                notes = f"Promo discount AED {discount}"
            rows_data.append([
                d, f"INV-{invoice_num:04d}", item_name, category, qty,
                unit_price, total, payment, staff, customer, discount, net_total, notes
            ])

    # Write data
    for idx, row in enumerate(rows_data):
        r = data_start + idx
        for c, val in enumerate(row):
            cell = ws.cell(row=r, column=col_start + c, value=val)
        style_data_row(ws, r, col_start, col_end, idx)

        # Format specific columns
        ws.cell(row=r, column=col_start + 0).number_format = DATE_FMT
        ws.cell(row=r, column=col_start + 0).alignment = align_date()
        ws.cell(row=r, column=col_start + 5).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start + 5).alignment = align_number()
        ws.cell(row=r, column=col_start + 6).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start + 6).alignment = align_number()
        ws.cell(row=r, column=col_start + 10).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start + 10).alignment = align_number()
        ws.cell(row=r, column=col_start + 11).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start + 11).alignment = align_number()

    data_end = data_start + len(rows_data) - 1

    # Totals row
    tr = data_end + 1
    ws.cell(row=tr, column=col_start, value="TOTALS")
    # SUM formulas
    ws.cell(row=tr, column=col_start + 4, value=f"=SUM({get_column_letter(col_start+4)}{data_start}:{get_column_letter(col_start+4)}{data_end})")
    ws.cell(row=tr, column=col_start + 4).number_format = INT_FMT
    ws.cell(row=tr, column=col_start + 6, value=f"=SUM({get_column_letter(col_start+6)}{data_start}:{get_column_letter(col_start+6)}{data_end})")
    ws.cell(row=tr, column=col_start + 6).number_format = CURRENCY_FMT
    ws.cell(row=tr, column=col_start + 10, value=f"=SUM({get_column_letter(col_start+10)}{data_start}:{get_column_letter(col_start+10)}{data_end})")
    ws.cell(row=tr, column=col_start + 10).number_format = CURRENCY_FMT
    ws.cell(row=tr, column=col_start + 11, value=f"=SUM({get_column_letter(col_start+11)}{data_start}:{get_column_letter(col_start+11)}{data_end})")
    ws.cell(row=tr, column=col_start + 11).number_format = CURRENCY_FMT
    style_total_row(ws, tr, col_start, col_end)

    # Column widths
    widths = {"B": 14, "C": 14, "D": 20, "E": 14, "F": 8, "G": 12,
              "H": 12, "I": 16, "J": 14, "K": 22, "L": 12, "M": 12, "N": 28}
    set_col_widths(ws, widths)

    add_print_area(ws, last_col, tr)

    # Caption
    cr = tr + 2
    ws.cell(row=cr, column=col_start, value=f"Total Transactions: {len(rows_data)} | March 2025 | TechZone Electronics")
    ws.cell(row=cr, column=col_start).font = font_caption()

    return ws, rows_data, data_start, data_end


# ══════════════════════════════════════════════════════════════
# SHEET 2: Daily Summary
# ══════════════════════════════════════════════════════════════
def create_daily_summary(se_data_end_ref=114):
    ws = wb.create_sheet("Daily Summary")
    apply_tab_color(ws)

    headers = ["Date", "Total Sales", "# Transactions", "Avg Transaction",
               "Cash", "Card", "Online", "Bank Transfer",
               "# Items Sold", "Total Discount Given"]

    col_start = 2
    col_end = col_start + len(headers) - 1
    last_col = col_end

    setup_sheet(ws, title="Daily Summary — March 2025", last_col=last_col)

    hr = 4
    for i, h in enumerate(headers):
        ws.cell(row=hr, column=col_start + i, value=h)
    style_header_row(ws, hr, col_start, col_end)

    # Reference Sales Entry sheet for SUMPRODUCT formulas
    se_sheet = "'Sales Entry'"
    se_date_col = "B"   # Date column in Sales Entry
    se_total_col = "H"  # Total column
    se_qty_col = "F"    # Qty column
    se_disc_col = "L"   # Discount column
    se_pay_col = "I"    # Payment Method column
    se_net_col = "M"    # Net Total column

    # We need to know the data range in Sales Entry
    se_data_start = 5
    se_data_end = se_data_end_ref  # passed from create_sales_entry()

    data_start = 5
    for day in range(1, 32):
        r = data_start + day - 1
        d = date(2025, 3, day)
        ws.cell(row=r, column=col_start, value=d)
        ws.cell(row=r, column=col_start).number_format = DATE_FMT
        ws.cell(row=r, column=col_start).alignment = align_date()

        # SUMPRODUCT formulas
        date_str = d.strftime("%Y-%m-%d")
        # Total Sales (Net Total)
        ws.cell(row=r, column=col_start+1).value = (
            f"=SUMPRODUCT(({se_sheet}!{se_date_col}{se_data_start}:{se_date_col}{se_data_end}=B{r})*"
            f"({se_sheet}!{se_net_col}{se_data_start}:{se_net_col}{se_data_end}))"
        )
        ws.cell(row=r, column=col_start+1).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+1).alignment = align_number()

        # # Transactions
        ws.cell(row=r, column=col_start+2).value = (
            f"=SUMPRODUCT(({se_sheet}!{se_date_col}{se_data_start}:{se_date_col}{se_data_end}=B{r})*1)"
        )
        ws.cell(row=r, column=col_start+2).number_format = INT_FMT
        ws.cell(row=r, column=col_start+2).alignment = align_number()

        # Avg Transaction
        ws.cell(row=r, column=col_start+3).value = (
            f"=IFERROR(C{r}/D{r},0)"
        )
        ws.cell(row=r, column=col_start+3).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+3).alignment = align_number()

        # Cash
        ws.cell(row=r, column=col_start+4).value = (
            f"=SUMPRODUCT(({se_sheet}!{se_date_col}{se_data_start}:{se_date_col}{se_data_end}=B{r})*"
            f"({se_sheet}!{se_pay_col}{se_data_start}:{se_pay_col}{se_data_end}=\"Cash\")*"
            f"({se_sheet}!{se_net_col}{se_data_start}:{se_net_col}{se_data_end}))"
        )
        ws.cell(row=r, column=col_start+4).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+4).alignment = align_number()

        # Card
        ws.cell(row=r, column=col_start+5).value = (
            f"=SUMPRODUCT(({se_sheet}!{se_date_col}{se_data_start}:{se_date_col}{se_data_end}=B{r})*"
            f"({se_sheet}!{se_pay_col}{se_data_start}:{se_pay_col}{se_data_end}=\"Card\")*"
            f"({se_sheet}!{se_net_col}{se_data_start}:{se_net_col}{se_data_end}))"
        )
        ws.cell(row=r, column=col_start+5).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+5).alignment = align_number()

        # Online
        ws.cell(row=r, column=col_start+6).value = (
            f"=SUMPRODUCT(({se_sheet}!{se_date_col}{se_data_start}:{se_date_col}{se_data_end}=B{r})*"
            f"({se_sheet}!{se_pay_col}{se_data_start}:{se_pay_col}{se_data_end}=\"Online\")*"
            f"({se_sheet}!{se_net_col}{se_data_start}:{se_net_col}{se_data_end}))"
        )
        ws.cell(row=r, column=col_start+6).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+6).alignment = align_number()

        # Bank Transfer
        ws.cell(row=r, column=col_start+7).value = (
            f"=SUMPRODUCT(({se_sheet}!{se_date_col}{se_data_start}:{se_date_col}{se_data_end}=B{r})*"
            f"({se_sheet}!{se_pay_col}{se_data_start}:{se_pay_col}{se_data_end}=\"Bank Transfer\")*"
            f"({se_sheet}!{se_net_col}{se_data_start}:{se_net_col}{se_data_end}))"
        )
        ws.cell(row=r, column=col_start+7).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+7).alignment = align_number()

        # # Items Sold
        ws.cell(row=r, column=col_start+8).value = (
            f"=SUMPRODUCT(({se_sheet}!{se_date_col}{se_data_start}:{se_date_col}{se_data_end}=B{r})*"
            f"({se_sheet}!{se_qty_col}{se_data_start}:{se_qty_col}{se_data_end}))"
        )
        ws.cell(row=r, column=col_start+8).number_format = INT_FMT
        ws.cell(row=r, column=col_start+8).alignment = align_number()

        # Total Discount Given
        ws.cell(row=r, column=col_start+9).value = (
            f"=SUMPRODUCT(({se_sheet}!{se_date_col}{se_data_start}:{se_date_col}{se_data_end}=B{r})*"
            f"({se_sheet}!{se_disc_col}{se_data_start}:{se_disc_col}{se_data_end}))"
        )
        ws.cell(row=r, column=col_start+9).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+9).alignment = align_number()

        style_data_row(ws, r, col_start, col_end, day - 1)

    data_end = data_start + 30  # 31 days (1-31), rows 5-35

    # Totals row
    tr = data_end + 1
    ws.cell(row=tr, column=col_start, value="TOTALS")
    for c_idx in [1, 2, 4, 5, 6, 7, 8, 9]:
        col_letter = get_column_letter(col_start + c_idx)
        ws.cell(row=tr, column=col_start + c_idx).value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        if c_idx in [2, 8]:
            ws.cell(row=tr, column=col_start + c_idx).number_format = INT_FMT
        else:
            ws.cell(row=tr, column=col_start + c_idx).number_format = CURRENCY_FMT
    # Avg Transaction = Total Sales / # Transactions
    ws.cell(row=tr, column=col_start+3).value = f"=IFERROR(C{tr}/D{tr},0)"
    ws.cell(row=tr, column=col_start+3).number_format = CURRENCY_FMT
    style_total_row(ws, tr, col_start, col_end)

    widths = {"B": 14, "C": 14, "D": 14, "E": 16, "F": 12, "G": 12, "H": 12, "I": 16, "J": 14, "K": 20}
    set_col_widths(ws, widths)
    add_print_area(ws, last_col, tr)

    return ws, data_start, data_end, tr


# ══════════════════════════════════════════════════════════════
# SHEET 3: Weekly Summary
# ══════════════════════════════════════════════════════════════
def create_weekly_summary():
    ws = wb.create_sheet("Weekly Summary")
    apply_tab_color(ws)

    headers = ["Week #", "Week Start", "Week End", "Total Sales",
               "# Transactions", "Avg Transaction", "Top Staff", "Profit Estimate"]

    col_start = 2
    col_end = col_start + len(headers) - 1
    last_col = col_end

    setup_sheet(ws, title="Weekly Summary — March 2025", last_col=last_col)

    hr = 4
    for i, h in enumerate(headers):
        ws.cell(row=hr, column=col_start + i, value=h)
    style_header_row(ws, hr, col_start, col_end)

    ds_sheet = "'Daily Summary'"
    weeks = [
        (1, date(2025, 3, 1), date(2025, 3, 7)),
        (2, date(2025, 3, 8), date(2025, 3, 14)),
        (3, date(2025, 3, 15), date(2025, 3, 21)),
        (4, date(2025, 3, 22), date(2025, 3, 28)),
        (5, date(2025, 3, 29), date(2025, 3, 31)),
    ]

    data_start = 5
    for idx, (wk, start, end) in enumerate(weeks):
        r = data_start + idx
        ws.cell(row=r, column=col_start, value=wk)
        ws.cell(row=r, column=col_start+1, value=start)
        ws.cell(row=r, column=col_start+1).number_format = DATE_FMT
        ws.cell(row=r, column=col_start+1).alignment = align_date()
        ws.cell(row=r, column=col_start+2, value=end)
        ws.cell(row=r, column=col_start+2).number_format = DATE_FMT
        ws.cell(row=r, column=col_start+2).alignment = align_date()

        # Find row range in Daily Summary for this week
        ds_start_row = 5 + (start.day - 1)
        ds_end_row = 5 + (end.day - 1)

        # Total Sales
        ws.cell(row=r, column=col_start+3).value = f"=SUM({ds_sheet}!C{ds_start_row}:C{ds_end_row})"
        ws.cell(row=r, column=col_start+3).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+3).alignment = align_number()

        # # Transactions
        ws.cell(row=r, column=col_start+4).value = f"=SUM({ds_sheet}!D{ds_start_row}:D{ds_end_row})"
        ws.cell(row=r, column=col_start+4).number_format = INT_FMT
        ws.cell(row=r, column=col_start+4).alignment = align_number()

        # Avg Transaction
        ws.cell(row=r, column=col_start+5).value = f"=IFERROR(E{r}/F{r},0)"
        ws.cell(row=r, column=col_start+5).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+5).alignment = align_number()

        # Top Staff (hardcoded for demo)
        top_staff = ["Ahmed", "Fatima", "Sara", "Khalid", "Omar"]
        ws.cell(row=r, column=col_start+6, value=top_staff[idx])

        # Profit Estimate (40% margin)
        ws.cell(row=r, column=col_start+7).value = f"=E{r}*0.4"
        ws.cell(row=r, column=col_start+7).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+7).alignment = align_number()

        style_data_row(ws, r, col_start, col_end, idx)

    data_end = data_start + len(weeks) - 1

    # Totals
    tr = data_end + 1
    ws.cell(row=tr, column=col_start, value="TOTALS")
    for c_idx in [3, 4]:
        col_letter = get_column_letter(col_start + c_idx)
        ws.cell(row=tr, column=col_start + c_idx).value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        if c_idx == 4:
            ws.cell(row=tr, column=col_start + c_idx).number_format = INT_FMT
        else:
            ws.cell(row=tr, column=col_start + c_idx).number_format = CURRENCY_FMT
    ws.cell(row=tr, column=col_start+5).value = f"=IFERROR(E{tr}/F{tr},0)"
    ws.cell(row=tr, column=col_start+5).number_format = CURRENCY_FMT
    ws.cell(row=tr, column=col_start+7).value = f"=SUM({get_column_letter(col_start+7)}{data_start}:{get_column_letter(col_start+7)}{data_end})"
    ws.cell(row=tr, column=col_start+7).number_format = CURRENCY_FMT
    style_total_row(ws, tr, col_start, col_end)

    widths = {"B": 10, "C": 14, "D": 14, "E": 14, "F": 14, "G": 16, "H": 14, "I": 16}
    set_col_widths(ws, widths)
    add_print_area(ws, last_col, tr)

    return ws, data_start, data_end


# ══════════════════════════════════════════════════════════════
# SHEET 4: Monthly Summary
# ══════════════════════════════════════════════════════════════
def create_monthly_summary():
    ws = wb.create_sheet("Monthly Summary")
    apply_tab_color(ws)

    headers = ["Month", "Total Sales", "# Transactions", "Avg Transaction",
               "Best Day Sales", "Best Day Date", "Total Expenses", "Net Profit"]

    col_start = 2
    col_end = col_start + len(headers) - 1
    last_col = col_end

    setup_sheet(ws, title="Monthly Summary — March 2025", last_col=last_col)

    hr = 4
    for i, h in enumerate(headers):
        ws.cell(row=hr, column=col_start + i, value=h)
    style_header_row(ws, hr, col_start, col_end)

    ds_sheet = "'Daily Summary'"
    r = 5
    ws.cell(row=r, column=col_start, value="March 2025")
    # Total Sales = sum of daily totals
    ws.cell(row=r, column=col_start+1).value = f"=SUM({ds_sheet}!C5:C35)"
    ws.cell(row=r, column=col_start+1).number_format = CURRENCY_FMT
    ws.cell(row=r, column=col_start+1).alignment = align_number()
    # # Transactions
    ws.cell(row=r, column=col_start+2).value = f"=SUM({ds_sheet}!D5:D35)"
    ws.cell(row=r, column=col_start+2).number_format = INT_FMT
    ws.cell(row=r, column=col_start+2).alignment = align_number()
    # Avg Transaction
    ws.cell(row=r, column=col_start+3).value = f"=IFERROR(C{r}/D{r},0)"
    ws.cell(row=r, column=col_start+3).number_format = CURRENCY_FMT
    ws.cell(row=r, column=col_start+3).alignment = align_number()
    # Best Day Sales
    ws.cell(row=r, column=col_start+4).value = f"=MAX({ds_sheet}!C5:C35)"
    ws.cell(row=r, column=col_start+4).number_format = CURRENCY_FMT
    ws.cell(row=r, column=col_start+4).alignment = align_number()
    # Best Day Date - use MATCH
    ws.cell(row=r, column=col_start+5).value = (
        f"=INDEX({ds_sheet}!B5:B35,MATCH(E{r},{ds_sheet}!C5:C35,0))"
    )
    ws.cell(row=r, column=col_start+5).number_format = DATE_FMT
    ws.cell(row=r, column=col_start+5).alignment = align_date()
    # Total Expenses - reference Expenses sheet
    ws.cell(row=r, column=col_start+6).value = f"=SUM('Expenses'!E5:E44)"
    ws.cell(row=r, column=col_start+6).number_format = CURRENCY_FMT
    ws.cell(row=r, column=col_start+6).alignment = align_number()
    # Net Profit
    ws.cell(row=r, column=col_start+7).value = f"=C{r}-G{r}"
    ws.cell(row=r, column=col_start+7).number_format = CURRENCY_FMT
    ws.cell(row=r, column=col_start+7).alignment = align_number()

    style_data_row(ws, r, col_start, col_end, 0)

    # Totals row (same as the single data row for monthly)
    tr = 6
    ws.cell(row=tr, column=col_start, value="TOTALS")
    for c_idx in [1, 2, 4, 6, 7]:
        ws.cell(row=tr, column=col_start + c_idx).value = ws.cell(row=5, column=col_start + c_idx).value
        if c_idx == 2:
            ws.cell(row=tr, column=col_start + c_idx).number_format = INT_FMT
        else:
            ws.cell(row=tr, column=col_start + c_idx).number_format = CURRENCY_FMT
    ws.cell(row=tr, column=col_start+3).value = f"=IFERROR(C{tr}/D{tr},0)"
    ws.cell(row=tr, column=col_start+3).number_format = CURRENCY_FMT
    style_total_row(ws, tr, col_start, col_end)

    # Chart: Monthly Sales vs Expenses bar chart
    # Put chart data in a small helper area
    chart_data_row = 9
    ws.cell(row=chart_data_row, column=2, value="Category")
    ws.cell(row=chart_data_row, column=3, value="Amount")
    ws.cell(row=chart_data_row + 1, column=2, value="Total Sales")
    ws.cell(row=chart_data_row + 1, column=3).value = f"=C5"
    ws.cell(row=chart_data_row + 2, column=2, value="Total Expenses")
    ws.cell(row=chart_data_row + 2, column=3).value = f"=G5"
    ws.cell(row=chart_data_row + 3, column=2, value="Net Profit")
    ws.cell(row=chart_data_row + 3, column=3).value = f"=H5"

    # Style the helper area lightly
    for rr in range(chart_data_row, chart_data_row + 4):
        for cc in [2, 3]:
            ws.cell(row=rr, column=cc).font = font_caption()

    chart = create_bar_chart(width=16, height=10)
    data_ref = Reference(ws, min_col=3, min_row=chart_data_row, max_row=chart_data_row + 3)
    cats_ref = Reference(ws, min_col=2, min_row=chart_data_row + 1, max_row=chart_data_row + 3)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    setup_chart_titles(chart, title="March 2025: Sales vs Expenses vs Profit", y_title="Amount (AED)")
    apply_chart_colors(chart)
    ws.add_chart(chart, "B14")

    widths = {"B": 16, "C": 14, "D": 14, "E": 16, "F": 16, "G": 16, "H": 16, "I": 14}
    set_col_widths(ws, widths)
    add_print_area(ws, last_col, tr)

    return ws


# ══════════════════════════════════════════════════════════════
# SHEET 5: Invoice
# ══════════════════════════════════════════════════════════════
def create_invoice():
    ws = wb.create_sheet("Invoice")
    apply_tab_color(ws)
    setup_sheet(ws, last_col=8)

    # Business header
    ws.merge_cells("B2:H2")
    cell = ws.cell(row=2, column=2, value="TechZone Electronics")
    cell.font = Font(name="Calibri", size=20, bold=True, color=PRIMARY)
    cell.alignment = align_title()
    ws.row_dimensions[2].height = 36

    ws.merge_cells("B3:H3")
    ws.cell(row=3, column=2, value="Premium Electronics & Accessories | Dubai, UAE").font = font_caption()

    # Invoice details
    ws.cell(row=5, column=2, value="INVOICE").font = Font(name="Calibri", size=14, bold=True, color=PRIMARY)
    ws.cell(row=5, column=5, value="Invoice #:").font = font_body()
    ws.cell(row=5, column=6, value="INV-0142").font = Font(name="Calibri", size=11, bold=True, color=NEUTRAL_900)
    ws.cell(row=6, column=5, value="Date:").font = font_body()
    ws.cell(row=6, column=6, value=date(2025, 3, 15)).font = font_body()
    ws.cell(row=6, column=6).number_format = DATE_FMT
    ws.cell(row=7, column=5, value="Customer:").font = font_body()
    ws.cell(row=7, column=6, value="Mohamed Al Rashid").font = Font(name="Calibri", size=11, bold=True, color=NEUTRAL_900)
    ws.cell(row=8, column=5, value="Staff:").font = font_body()
    ws.cell(row=8, column=6, value="Fatima").font = font_body()

    # Item table
    inv_headers = ["#", "Item", "Qty", "Unit Price", "Total"]
    hr = 10
    for i, h in enumerate(inv_headers):
        ws.cell(row=hr, column=2 + i, value=h)
    style_header_row(ws, hr, 2, 6)

    items = [
        (1, "Tablet", 1, 899, 899),
        (2, "Screen Protector", 2, 30, 60),
        (3, "Earbuds", 1, 85, 85),
        (4, "Phone Case", 1, 45, 45),
    ]

    for idx, (num, name, qty, price, total) in enumerate(items):
        r = 11 + idx
        ws.cell(row=r, column=2, value=num)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=qty)
        ws.cell(row=r, column=5, value=price)
        ws.cell(row=r, column=5).number_format = CURRENCY_FMT
        ws.cell(row=r, column=6, value=total)
        ws.cell(row=r, column=6).number_format = CURRENCY_FMT
        style_data_row(ws, r, 2, 6, idx)

    # Subtotal, Discount, VAT, Grand Total
    r = 15
    ws.cell(row=r, column=3, value="Subtotal")
    ws.cell(row=r, column=6, value=1089)
    ws.cell(row=r, column=6).number_format = CURRENCY_FMT
    ws.cell(row=r, column=3).font = font_body()
    ws.cell(row=r, column=6).font = font_body()

    r = 16
    ws.cell(row=r, column=3, value="Discount")
    ws.cell(row=r, column=6, value=-50)
    ws.cell(row=r, column=6).number_format = CURRENCY_FMT
    ws.cell(row=r, column=3).font = font_body()
    ws.cell(row=r, column=6).font = Font(name="Calibri", size=11, color=ACCENT_NEGATIVE)

    r = 17
    ws.cell(row=r, column=3, value="VAT (5%)")
    ws.cell(row=r, column=6, value=51.95)  # (1089-50)*0.05
    ws.cell(row=r, column=6).number_format = CURRENCY_FMT
    ws.cell(row=r, column=3).font = font_body()
    ws.cell(row=r, column=6).font = font_body()

    # Divider line
    r = 18
    for c in range(2, 7):
        ws.cell(row=r, column=c).border = Border(top=Side(style="medium", color=PRIMARY))

    r = 19
    ws.cell(row=r, column=3, value="GRAND TOTAL")
    ws.cell(row=r, column=3).font = Font(name="Calibri", size=12, bold=True, color=PRIMARY)
    ws.cell(row=r, column=6, value=1090.95)  # 1089 - 50 + 51.95
    ws.cell(row=r, column=6).number_format = CURRENCY_FMT
    ws.cell(row=r, column=6).font = Font(name="Calibri", size=14, bold=True, color=PRIMARY)

    # Footer
    r = 21
    ws.cell(row=r, column=2, value="Thank you for shopping at TechZone Electronics!").font = font_caption()
    r = 22
    ws.cell(row=r, column=2, value="Returns within 14 days with receipt. VAT Registration No: 100XXXXXX000003").font = font_caption()

    widths = {"B": 8, "C": 22, "D": 10, "E": 14, "F": 14, "G": 10, "H": 10}
    set_col_widths(ws, widths)

    return ws


# ══════════════════════════════════════════════════════════════
# SHEET 6: Inventory
# ══════════════════════════════════════════════════════════════
def create_inventory():
    ws = wb.create_sheet("Inventory")
    apply_tab_color(ws)

    headers = ["Product ID", "Item Name", "Category", "Unit Price", "Cost Price",
               "Margin", "Stock Qty", "Reorder Level", "Status",
               "Supplier", "Last Restocked", "Days Since Restock"]

    col_start = 2
    col_end = col_start + len(headers) - 1
    last_col = col_end

    setup_sheet(ws, title="Inventory — TechZone Electronics", last_col=last_col)

    hr = 4
    for i, h in enumerate(headers):
        ws.cell(row=hr, column=col_start + i, value=h)
    style_header_row(ws, hr, col_start, col_end)

    products = [
        ("P001", "Phone Case", "Accessories", 45, 20, 22, 15, "Samsung Gulf", date(2025, 3, 10)),
        ("P002", "USB Cable", "Accessories", 25, 10, 28, 20, "Accessories Plus", date(2025, 3, 8)),
        ("P003", "Screen Protector", "Accessories", 30, 12, 3, 15, "Apple ME", date(2025, 3, 12)),
        ("P004", "Power Bank", "Accessories", 120, 55, 12, 10, "Samsung Gulf", date(2025, 3, 5)),
        ("P005", "Earbuds", "Audio", 85, 40, 18, 12, "Apple ME", date(2025, 3, 15)),
        ("P006", "Bluetooth Speaker", "Audio", 150, 70, 14, 8, "Samsung Gulf", date(2025, 3, 1)),
        ("P007", "Headphones", "Audio", 200, 90, 2, 6, "Apple ME", date(2025, 3, 20)),
        ("P008", "Tablet", "Devices", 899, 540, 6, 3, "Apple ME", date(2025, 3, 18)),
        ("P009", "Smart Watch", "Wearables", 450, 270, 3, 5, "Apple ME", date(2025, 3, 14)),
        ("P010", "Fitness Band", "Wearables", 180, 90, 15, 8, "Samsung Gulf", date(2025, 3, 9)),
        ("P011", "Laptop Stand", "Accessories", 150, 65, 14, 10, "Dell Middle East", date(2025, 3, 7)),
        ("P012", "Keyboard", "Peripherals", 200, 100, 15, 8, "Dell Middle East", date(2025, 3, 11)),
        ("P013", "Mouse", "Peripherals", 80, 35, 22, 10, "Dell Middle East", date(2025, 3, 6)),
        ("P014", "Charger", "Accessories", 55, 22, 3, 12, "Accessories Plus", date(2025, 3, 13)),
        ("P015", "HDMI Cable", "Cables", 35, 14, 25, 20, "Accessories Plus", date(2025, 3, 4)),
        ("P016", "Type-C Hub", "Accessories", 95, 42, 12, 8, "TechParts UAE", date(2025, 3, 16)),
        ("P017", "Wireless Charger", "Accessories", 110, 50, 18, 8, "Samsung Gulf", date(2025, 3, 2)),
        ("P018", "Memory Card 64GB", "Storage", 45, 18, 30, 15, "TechParts UAE", date(2025, 3, 19)),
        ("P019", "USB Flash Drive", "Storage", 30, 12, 22, 18, "TechParts UAE", date(2025, 3, 3)),
        ("P020", "Portable SSD", "Storage", 250, 140, 2, 4, "Samsung Gulf", date(2025, 3, 17)),
        ("P021", "Webcam", "Peripherals", 180, 85, 10, 6, "Dell Middle East", date(2025, 3, 8)),
        ("P022", "Monitor Cable", "Cables", 40, 16, 28, 20, "Accessories Plus", date(2025, 3, 5)),
        ("P023", "Phone Holder", "Accessories", 35, 14, 15, 10, "Accessories Plus", date(2025, 3, 20)),
        ("P024", "Stylus Pen", "Accessories", 65, 28, 12, 8, "Apple ME", date(2025, 3, 10)),
        ("P025", "Car Mount", "Accessories", 50, 20, 14, 12, "Accessories Plus", date(2025, 3, 12)),
        ("P026", "Surge Protector", "Accessories", 75, 32, 16, 10, "TechParts UAE", date(2025, 3, 6)),
        ("P027", "Cleaning Kit", "Accessories", 25, 8, 35, 20, "TechParts UAE", date(2025, 3, 14)),
        ("P028", "Mouse Pad XL", "Peripherals", 40, 15, 18, 12, "Dell Middle East", date(2025, 3, 9)),
        ("P029", "Adapter Plug", "Accessories", 20, 7, 20, 15, "Accessories Plus", date(2025, 3, 15)),
        ("P030", "AV Cable", "Cables", 28, 10, 4, 10, "Accessories Plus", date(2025, 3, 11)),
        ("P031", "Ethernet Cable", "Cables", 22, 8, 40, 25, "Accessories Plus", date(2025, 3, 3)),
        ("P032", "Audio Splitter", "Audio", 18, 6, 18, 12, "TechParts UAE", date(2025, 3, 18)),
        ("P033", "Ring Light", "Accessories", 120, 52, 3, 6, "TechParts UAE", date(2025, 3, 7)),
        ("P034", "Tripod Stand", "Accessories", 90, 38, 12, 8, "TechParts UAE", date(2025, 3, 13)),
    ]

    data_start = 5
    for idx, (pid, name, cat, price, cost, stock, reorder, supplier, restock_date) in enumerate(products):
        r = data_start + idx
        ws.cell(row=r, column=col_start, value=pid)
        ws.cell(row=r, column=col_start+1, value=name)
        ws.cell(row=r, column=col_start+2, value=cat)
        ws.cell(row=r, column=col_start+3, value=price)
        ws.cell(row=r, column=col_start+3).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+3).alignment = align_number()
        ws.cell(row=r, column=col_start+4, value=cost)
        ws.cell(row=r, column=col_start+4).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+4).alignment = align_number()
        # Margin formula
        ws.cell(row=r, column=col_start+5).value = f"=IFERROR((E{r}-F{r})/E{r},0)"
        ws.cell(row=r, column=col_start+5).number_format = PCT_FMT
        ws.cell(row=r, column=col_start+5).alignment = align_number()
        ws.cell(row=r, column=col_start+6, value=stock)
        ws.cell(row=r, column=col_start+6).alignment = align_number()
        ws.cell(row=r, column=col_start+7, value=reorder)
        ws.cell(row=r, column=col_start+7).alignment = align_number()
        # Status formula
        ws.cell(row=r, column=col_start+8).value = f'=IF(H{r}<=I{r},"LOW STOCK","In Stock")'
        ws.cell(row=r, column=col_start+9, value=supplier)
        ws.cell(row=r, column=col_start+10, value=restock_date)
        ws.cell(row=r, column=col_start+10).number_format = DATE_FMT
        ws.cell(row=r, column=col_start+10).alignment = align_date()
        # Days Since Restock
        ws.cell(row=r, column=col_start+11).value = f"=MAX(0,TODAY()-L{r})"
        ws.cell(row=r, column=col_start+11).number_format = INT_FMT
        ws.cell(row=r, column=col_start+11).alignment = align_number()

        style_data_row(ws, r, col_start, col_end, idx)

    data_end = data_start + len(products) - 1

    # Conditional formatting for LOW STOCK
    from openpyxl.formatting.rule import CellIsRule
    ws.conditional_formatting.add(
        f"J{data_start}:J{data_end}",
        CellIsRule(operator="equal", formula=['"LOW STOCK"'],
                   fill=PatternFill(bgColor="FDEDEC"),
                   font=Font(color=ACCENT_NEGATIVE, bold=True))
    )
    ws.conditional_formatting.add(
        f"J{data_start}:J{data_end}",
        CellIsRule(operator="equal", formula=['"In Stock"'],
                   fill=PatternFill(bgColor="E8F5E9"),
                   font=Font(color=ACCENT_POSITIVE))
    )

    # Totals row
    tr = data_end + 1
    ws.cell(row=tr, column=col_start, value="TOTALS")
    ws.cell(row=tr, column=col_start+6).value = f"=SUM(H{data_start}:H{data_end})"
    ws.cell(row=tr, column=col_start+6).number_format = INT_FMT
    # Count low stock
    ws.cell(row=tr, column=col_start+8).value = f'=COUNTIF(J{data_start}:J{data_end},"LOW STOCK")'
    style_total_row(ws, tr, col_start, col_end)

    widths = {"B": 12, "C": 20, "D": 14, "E": 12, "F": 12, "G": 10, "H": 10,
              "I": 14, "J": 14, "K": 18, "L": 14, "M": 18}
    set_col_widths(ws, widths)
    add_print_area(ws, last_col, tr)

    return ws


# ══════════════════════════════════════════════════════════════
# SHEET 7: Staff Performance
# ══════════════════════════════════════════════════════════════
def create_staff_performance(se_data_end_ref=114):
    ws = wb.create_sheet("Staff Performance")
    apply_tab_color(ws)

    headers = ["Staff Name", "Total Sales", "# Transactions", "Avg Transaction",
               "Top Category", "Shift", "Commission (5%)", "Sales Target",
               "Target Achievement %"]

    col_start = 2
    col_end = col_start + len(headers) - 1
    last_col = col_end

    setup_sheet(ws, title="Staff Performance — March 2025", last_col=last_col)

    hr = 4
    for i, h in enumerate(headers):
        ws.cell(row=hr, column=col_start + i, value=h)
    style_header_row(ws, hr, col_start, col_end)

    se_sheet = "'Sales Entry'"
    se_data_start = 5
    se_data_end = se_data_end_ref  # passed from create_sales_entry()

    staff_data = list(zip(STAFF, SHIFTS, SALES_TARGETS))

    # Top categories per staff (for demo)
    top_cats = ["Accessories", "Devices", "Audio", "Accessories", "Peripherals", "Wearables"]

    data_start = 5
    for idx, (name, shift, target) in enumerate(staff_data):
        r = data_start + idx
        ws.cell(row=r, column=col_start, value=name)
        # Total Sales = SUMPRODUCT
        ws.cell(row=r, column=col_start+1).value = (
            f"=SUMPRODUCT(({se_sheet}!J{se_data_start}:J{se_data_end}=B{r})*"
            f"({se_sheet}!M{se_data_start}:M{se_data_end}))"
        )
        ws.cell(row=r, column=col_start+1).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+1).alignment = align_number()

        # # Transactions
        ws.cell(row=r, column=col_start+2).value = (
            f"=SUMPRODUCT(({se_sheet}!J{se_data_start}:J{se_data_end}=B{r})*1)"
        )
        ws.cell(row=r, column=col_start+2).number_format = INT_FMT
        ws.cell(row=r, column=col_start+2).alignment = align_number()

        # Avg Transaction
        ws.cell(row=r, column=col_start+3).value = f"=IFERROR(C{r}/D{r},0)"
        ws.cell(row=r, column=col_start+3).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+3).alignment = align_number()

        # Top Category
        ws.cell(row=r, column=col_start+4, value=top_cats[idx])

        # Shift
        ws.cell(row=r, column=col_start+5, value=shift)

        # Commission (5%)
        ws.cell(row=r, column=col_start+6).value = f"=C{r}*0.05"
        ws.cell(row=r, column=col_start+6).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+6).alignment = align_number()

        # Sales Target
        ws.cell(row=r, column=col_start+7, value=target)
        ws.cell(row=r, column=col_start+7).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+7).alignment = align_number()

        # Target Achievement %
        ws.cell(row=r, column=col_start+8).value = f"=IFERROR(C{r}/I{r},0)"
        ws.cell(row=r, column=col_start+8).number_format = PCT_FMT
        ws.cell(row=r, column=col_start+8).alignment = align_number()

        style_data_row(ws, r, col_start, col_end, idx)

    data_end = data_start + len(staff_data) - 1

    # Conditional formatting on Target Achievement %
    ws.conditional_formatting.add(
        f"J{data_start}:J{data_end}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                   fill=PatternFill(bgColor="E8F5E9"),
                   font=Font(color=ACCENT_POSITIVE, bold=True))
    )
    ws.conditional_formatting.add(
        f"J{data_start}:J{data_end}",
        CellIsRule(operator="lessThan", formula=["0.8"],
                   fill=PatternFill(bgColor="FDEDEC"),
                   font=Font(color=ACCENT_NEGATIVE))
    )
    ws.conditional_formatting.add(
        f"J{data_start}:J{data_end}",
        CellIsRule(operator="between", formula=["0.8", "0.9999"],
                   fill=PatternFill(bgColor="FEF9E7"),
                   font=Font(color=ACCENT_WARNING))
    )

    # Data bars on Total Sales
    ws.conditional_formatting.add(
        f"C{data_start}:C{data_end}",
        DataBarRule(start_type="min", end_type="max",
                    color=PRIMARY)
    )

    # Totals row
    tr = data_end + 1
    ws.cell(row=tr, column=col_start, value="TOTALS")
    for c_idx in [1, 2, 6, 7]:
        col_letter = get_column_letter(col_start + c_idx)
        ws.cell(row=tr, column=col_start + c_idx).value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        if c_idx == 2:
            ws.cell(row=tr, column=col_start + c_idx).number_format = INT_FMT
        else:
            ws.cell(row=tr, column=col_start + c_idx).number_format = CURRENCY_FMT
    ws.cell(row=tr, column=col_start+3).value = f"=IFERROR(C{tr}/D{tr},0)"
    ws.cell(row=tr, column=col_start+3).number_format = CURRENCY_FMT
    ws.cell(row=tr, column=col_start+8).value = f"=IFERROR(C{tr}/I{tr},0)"
    ws.cell(row=tr, column=col_start+8).number_format = PCT_FMT
    style_total_row(ws, tr, col_start, col_end)

    widths = {"B": 14, "C": 14, "D": 14, "E": 16, "F": 14, "G": 12, "H": 16, "I": 14, "J": 20}
    set_col_widths(ws, widths)
    add_print_area(ws, last_col, tr)

    return ws, data_start, data_end


# ══════════════════════════════════════════════════════════════
# SHEET 8: Expenses
# ══════════════════════════════════════════════════════════════
def create_expenses():
    ws = wb.create_sheet("Expenses")
    apply_tab_color(ws)

    headers = ["Date", "Expense Category", "Description", "Amount",
               "Payment Method", "Vendor", "Receipt #", "Notes"]

    col_start = 2
    col_end = col_start + len(headers) - 1
    last_col = col_end

    setup_sheet(ws, title="Expenses — March 2025", last_col=last_col)

    hr = 4
    for i, h in enumerate(headers):
        ws.cell(row=hr, column=col_start + i, value=h)
    style_header_row(ws, hr, col_start, col_end)

    expense_entries = [
        (date(2025, 3, 1), "Rent", "Shop monthly rent - Dubai Mall", 25000, "Bank Transfer", "Emaar Properties", "RCP-3001", "Fixed monthly"),
        (date(2025, 3, 1), "Salaries", "Staff salaries - March", 32000, "Bank Transfer", "HR Payroll", "RCP-3002", "6 staff members"),
        (date(2025, 3, 1), "Utilities", "DEWA electricity bill", 3500, "Bank Transfer", "DEWA", "RCP-3003", "Feb consumption"),
        (date(2025, 3, 2), "Utilities", "Water bill", 450, "Bank Transfer", "DEWA", "RCP-3004", ""),
        (date(2025, 3, 3), "Supplies", "Thermal receipt paper rolls", 280, "Card", "Office Depot UAE", "RCP-3005", "20 rolls"),
        (date(2025, 3, 3), "Supplies", "Shopping bags & packaging", 650, "Card", "PackPro UAE", "RCP-3006", "Branded bags"),
        (date(2025, 3, 4), "Maintenance", "POS system maintenance", 500, "Bank Transfer", "TechServ ME", "RCP-3007", "Quarterly service"),
        (date(2025, 3, 5), "Marketing", "Google Ads campaign", 2000, "Card", "Google Ireland", "RCP-3008", "2-week campaign"),
        (date(2025, 3, 5), "Marketing", "Instagram sponsored posts", 800, "Card", "Meta Platforms", "RCP-3009", ""),
        (date(2025, 3, 6), "Transport", "Delivery vehicle fuel", 350, "Cash", "ADNOC", "RCP-3010", ""),
        (date(2025, 3, 7), "Supplies", "Display shelves & hooks", 1200, "Card", "ShopFittings UAE", "RCP-3011", "New display units"),
        (date(2025, 3, 8), "Insurance", "Shop insurance premium", 1800, "Bank Transfer", "Oman Insurance", "RCP-3012", "Quarterly premium"),
        (date(2025, 3, 9), "Maintenance", "AC maintenance service", 600, "Cash", "CoolTech Services", "RCP-3013", "Filter replacement"),
        (date(2025, 3, 10), "Transport", "Courier service - customer deliveries", 450, "Card", "Aramex", "RCP-3014", "March deliveries"),
        (date(2025, 3, 11), "Marketing", "Radio ad - Dubai Eye", 1500, "Bank Transfer", "Arabian Radio Network", "RCP-3015", "1-week spot"),
        (date(2025, 3, 12), "Supplies", "Cleaning supplies", 180, "Cash", "Carrefour", "RCP-3016", ""),
        (date(2025, 3, 13), "Miscellaneous", "Staff lunch - team meeting", 350, "Card", "Tim Hortons", "RCP-3017", "Monthly meeting"),
        (date(2025, 3, 14), "Utilities", "Internet & phone - Etisalat", 900, "Bank Transfer", "Etisalat", "RCP-3018", "Business plan"),
        (date(2025, 3, 15), "Salaries", "Overtime pay - March", 2500, "Bank Transfer", "HR Payroll", "RCP-3019", "Weekend shifts"),
        (date(2025, 3, 16), "Maintenance", "Security system check", 400, "Card", "SecureTech", "RCP-3020", ""),
        (date(2025, 3, 17), "Marketing", "Ramadan promotion flyers", 750, "Cash", "PrintZone", "RCP-3021", "2000 flyers"),
        (date(2025, 3, 18), "Transport", "Fuel - delivery vehicle", 300, "Cash", "ENOC", "RCP-3022", ""),
        (date(2025, 3, 19), "Supplies", "Price tags & labels", 150, "Card", "LabelPro", "RCP-3023", ""),
        (date(2025, 3, 20), "Miscellaneous", "Office coffee & tea supplies", 200, "Cash", "Carrefour", "RCP-3024", ""),
        (date(2025, 3, 21), "Marketing", "Social media influencer", 3000, "Bank Transfer", "Influencer Hub", "RCP-3025", "Product placement"),
        (date(2025, 3, 22), "Maintenance", "Glass door repair", 700, "Cash", "GlassFix Dubai", "RCP-3026", "Cracked panel"),
        (date(2025, 3, 23), "Supplies", "Anti-theft security tags", 450, "Card", "SecureTag ME", "RCP-3027", "50 tags"),
        (date(2025, 3, 24), "Insurance", "Inventory insurance", 1200, "Bank Transfer", "AXA Gulf", "RCP-3028", "Quarterly"),
        (date(2025, 3, 25), "Utilities", "DEWA electricity - mid month", 2800, "Bank Transfer", "DEWA", "RCP-3029", "Adjusted reading"),
        (date(2025, 3, 26), "Transport", "Courier - express deliveries", 550, "Card", "DHL Express", "RCP-3030", "Priority orders"),
        (date(2025, 3, 27), "Marketing", "Website hosting & SSL", 400, "Card", "GoDaddy", "RCP-3031", "Annual renewal"),
        (date(2025, 3, 28), "Miscellaneous", "Bank charges & fees", 180, "Bank Transfer", "Emirates NBD", "RCP-3032", "Transaction fees"),
        (date(2025, 3, 28), "Supplies", "Stationery & printer ink", 320, "Card", "Office Depot UAE", "RCP-3033", ""),
        (date(2025, 3, 29), "Maintenance", "Fire extinguisher inspection", 250, "Cash", "SafeFire UAE", "RCP-3034", "Annual check"),
        (date(2025, 3, 30), "Miscellaneous", "Customer gift cards", 500, "Card", "GiftCardCo", "RCP-3035", "Loyalty program"),
        (date(2025, 3, 30), "Marketing", "Email marketing platform", 250, "Card", "Mailchimp", "RCP-3036", "Monthly subscription"),
        (date(2025, 3, 31), "Salaries", "End of month bonus", 3000, "Bank Transfer", "HR Payroll", "RCP-3037", "Performance bonus"),
        (date(2025, 3, 31), "Utilities", "District cooling - Empower", 1200, "Bank Transfer", "Empower", "RCP-3038", "March billing"),
        (date(2025, 3, 31), "Transport", "Monthly vehicle maintenance", 800, "Card", "AutoServ Dubai", "RCP-3039", "Oil change & tires"),
        (date(2025, 3, 31), "Miscellaneous", "Waste management fee", 300, "Bank Transfer", "Averda", "RCP-3040", "Monthly collection"),
    ]

    data_start = 5
    for idx, row in enumerate(expense_entries):
        r = data_start + idx
        for c, val in enumerate(row):
            ws.cell(row=r, column=col_start + c, value=val)
        ws.cell(row=r, column=col_start + 0).number_format = DATE_FMT
        ws.cell(row=r, column=col_start + 0).alignment = align_date()
        ws.cell(row=r, column=col_start + 3).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start + 3).alignment = align_number()
        style_data_row(ws, r, col_start, col_end, idx)

    data_end = data_start + len(expense_entries) - 1

    # Totals row
    tr = data_end + 1
    ws.cell(row=tr, column=col_start, value="TOTALS")
    ws.cell(row=tr, column=col_start+3).value = f"=SUM(E{data_start}:E{data_end})"
    ws.cell(row=tr, column=col_start+3).number_format = CURRENCY_FMT
    style_total_row(ws, tr, col_start, col_end)

    # Category summary below
    sr = tr + 2
    ws.cell(row=sr, column=col_start, value="Summary by Category").font = font_subheader()
    sr += 1
    cats = ["Rent", "Salaries", "Utilities", "Supplies", "Maintenance",
            "Marketing", "Transport", "Insurance", "Miscellaneous"]
    ws.cell(row=sr, column=col_start, value="Category")
    ws.cell(row=sr, column=col_start+1, value="Total Amount")
    style_header_row(ws, sr, col_start, col_start+1)

    for i, cat in enumerate(cats):
        r = sr + 1 + i
        ws.cell(row=r, column=col_start, value=cat)
        ws.cell(row=r, column=col_start+1).value = (
            f"=SUMPRODUCT((C{data_start}:C{data_end}=B{r})*"
            f"(E{data_start}:E{data_end}))"
        )
        ws.cell(row=r, column=col_start+1).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+1).alignment = align_number()
        style_data_row(ws, r, col_start, col_start+1, i)

    widths = {"B": 14, "C": 18, "D": 32, "E": 14, "F": 16, "G": 22, "H": 14, "I": 22}
    set_col_widths(ws, widths)
    add_print_area(ws, last_col, tr)

    return ws, data_start, data_end


# ══════════════════════════════════════════════════════════════
# SHEET 9: Profit Analysis
# ══════════════════════════════════════════════════════════════
def create_profit_analysis():
    ws = wb.create_sheet("Profit Analysis")
    apply_tab_color(ws)

    col_start = 2
    col_end = 5
    last_col = col_end

    setup_sheet(ws, title="Profit & Loss Analysis — March 2025", last_col=last_col)

    # Header
    hr = 4
    ws.cell(row=hr, column=col_start, value="Item")
    ws.cell(row=hr, column=col_start+1, value="Amount (AED)")
    ws.cell(row=hr, column=col_start+2, value="% of Revenue")
    ws.cell(row=hr, column=col_start+3, value="Notes")
    style_header_row(ws, hr, col_start, col_end)

    ms_sheet = "'Monthly Summary'"
    ex_sheet = "'Expenses'"

    rows_data = [
        ("Revenue", f"={ms_sheet}!C5", "", "Total net sales for March"),
        ("Cost of Goods Sold (60%)", f"=-C5*0.6", "", "Estimated at 60% of revenue"),
        ("Gross Profit", "=C5+C6", "", "Revenue minus COGS"),
        ("", "", "", ""),
        ("Operating Expenses", f"={ms_sheet}!G5", "", "Total from Expenses sheet"),
        ("", "", "", ""),
        ("Net Profit", "=C7-C9", "", "Gross Profit minus Expenses"),
        ("Profit Margin %", "=IFERROR(C11/C5,0)", "", ""),
    ]

    data_start = 5
    for idx, (item, amount, pct, notes) in enumerate(rows_data):
        r = data_start + idx
        ws.cell(row=r, column=col_start, value=item)
        if amount:
            ws.cell(row=r, column=col_start+1, value=amount)
        ws.cell(row=r, column=col_start+1).number_format = CURRENCY_FMT
        ws.cell(row=r, column=col_start+1).alignment = align_number()
        if pct:
            ws.cell(row=r, column=col_start+2, value=pct)
        ws.cell(row=r, column=col_start+3, value=notes)
        ws.cell(row=r, column=col_start+3).font = font_caption()

        # Apply special styles for key rows
        if item == "Revenue":
            style_data_row(ws, r, col_start, col_end, 0)
            ws.cell(row=r, column=col_start).font = font_subheader()
        elif item == "Gross Profit":
            style_total_row(ws, r, col_start, col_end)
        elif item == "Net Profit":
            style_total_row(ws, r, col_start, col_end)
            ws.cell(row=r, column=col_start).font = Font(name="Calibri", size=12, bold=True, color=PRIMARY)
        elif item == "Profit Margin %":
            ws.cell(row=r, column=col_start+1).number_format = PCT_FMT
            style_data_row(ws, r, col_start, col_end, 4)
        elif item:
            style_data_row(ws, r, col_start, col_end, idx)

    # Conditional formatting for Profit Margin
    # Net Profit: green if positive, red if negative
    ws.conditional_formatting.add(
        f"C11",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill(bgColor="E8F5E9"),
                   font=Font(color=ACCENT_POSITIVE, bold=True, size=12))
    )
    ws.conditional_formatting.add(
        f"C11",
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill(bgColor="FDEDEC"),
                   font=Font(color=ACCENT_NEGATIVE, bold=True, size=12))
    )

    # Also for Gross Profit
    ws.conditional_formatting.add(
        f"C7",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill(bgColor="E8F5E9"),
                   font=Font(color=ACCENT_POSITIVE, bold=True))
    )

    # COGS row: red coloring
    ws.conditional_formatting.add(
        f"C6",
        CellIsRule(operator="lessThan", formula=["0"],
                   font=Font(color=ACCENT_NEGATIVE))
    )

    widths = {"B": 28, "C": 18, "D": 16, "E": 28}
    set_col_widths(ws, widths)

    return ws


# ══════════════════════════════════════════════════════════════
# SHEET 10: Dashboard
# ══════════════════════════════════════════════════════════════
def create_dashboard(se_data_end_ref=114):
    ws = wb.create_sheet("Dashboard")
    apply_tab_color(ws)
    setup_sheet(ws, last_col=16)

    ws.merge_cells("B2:P2")
    cell = ws.cell(row=2, column=2, value="TechZone Electronics — POS Dashboard")
    cell.font = Font(name="Calibri", size=20, bold=True, color=PRIMARY)
    cell.alignment = align_title()
    ws.row_dimensions[2].height = 40

    ws.merge_cells("B3:P3")
    ws.cell(row=3, column=2, value="March 2025 | Premium POS Report").font = font_caption()

    ms_sheet = "'Monthly Summary'"
    ds_sheet = "'Daily Summary'"
    sp_sheet = "'Staff Performance'"
    ex_sheet = "'Expenses'"
    pa_sheet = "'Profit Analysis'"

    # ── KPI CARDS ──────────────────────────────────────────
    # Row 5-7: KPI Cards (3 rows per card: label, value, spacer)
    kpi_row = 5
    kpi_data = [
        ("B", "C", "Total Revenue", f"={ms_sheet}!C5", CURRENCY_FMT),
        ("D", "E", "Total Expenses", f"={ms_sheet}!G5", CURRENCY_FMT),
        ("F", "G", "Net Profit", f"={pa_sheet}!C11", CURRENCY_FMT),
        ("H", "I", "Profit Margin %", f"={pa_sheet}!C12", PCT_FMT),
        ("J", "K", "# Transactions", f"={ms_sheet}!D5", INT_FMT),
        ("L", "M", "Avg Transaction", f"={ms_sheet}!E5", CURRENCY_FMT),
    ]

    ws.row_dimensions[kpi_row].height = 22      # label
    ws.row_dimensions[kpi_row + 1].height = 36   # value
    ws.row_dimensions[kpi_row + 2].height = 8    # spacer

    for start_col, end_col, label, formula, fmt in kpi_data:
        # Merge cells for the card
        ws.merge_cells(f"{start_col}{kpi_row}:{end_col}{kpi_row}")
        ws.merge_cells(f"{start_col}{kpi_row+1}:{end_col}{kpi_row+1}")

        # Label
        label_cell = ws[f"{start_col}{kpi_row}"]
        label_cell.value = label
        label_cell.font = font_kpi_label()
        label_cell.alignment = Alignment(horizontal="center", vertical="bottom")

        # Value
        val_cell = ws[f"{start_col}{kpi_row+1}"]
        val_cell.value = formula
        val_cell.font = font_kpi()
        val_cell.number_format = fmt
        val_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Card background - light fill
        for row in [kpi_row, kpi_row + 1]:
            for col_letter in [start_col, end_col]:
                ws[f"{col_letter}{row}"].fill = PatternFill("solid", fgColor=NEUTRAL_50)

    # ── CHART DATA AREAS ──────────────────────────────────
    # We'll create helper data areas below the charts, hidden

    # CHART 1: Daily Sales Trend (line chart) — placed at B9
    chart1_data_start = 50  # row 50 onwards for helper data
    ws.cell(row=chart1_data_start, column=2, value="Date")
    ws.cell(row=chart1_data_start, column=3, value="Daily Sales")
    for day in range(1, 32):
        r = chart1_data_start + day
        d = date(2025, 3, day)
        ws.cell(row=r, column=2, value=d)
        ws.cell(row=r, column=2).number_format = "DD-MMM"
        ws.cell(row=r, column=3).value = f"={ds_sheet}!C{4 + day}"

    chart1 = create_line_chart(width=28, height=13)
    data1 = Reference(ws, min_col=3, min_row=chart1_data_start, max_row=chart1_data_start + 31)
    cats1 = Reference(ws, min_col=2, min_row=chart1_data_start + 1, max_row=chart1_data_start + 31)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    setup_chart_titles(chart1, title="Daily Sales Trend — March 2025", y_title="Sales (AED)")
    apply_chart_colors(chart1)
    # Style the line
    s = chart1.series[0]
    s.graphicalProperties.line.width = 22000  # ~2pt
    s.smooth = True
    ws.add_chart(chart1, "B9")

    # CHART 2: Expense Breakdown (pie chart) — placed at J9
    # Helper data
    chart2_data_start = 85
    cats_expense = ["Rent", "Salaries", "Utilities", "Supplies", "Maintenance",
                    "Marketing", "Transport", "Insurance", "Miscellaneous"]
    ws.cell(row=chart2_data_start, column=2, value="Category")
    ws.cell(row=chart2_data_start, column=3, value="Amount")
    for i, cat in enumerate(cats_expense):
        r = chart2_data_start + 1 + i
        ws.cell(row=r, column=2, value=cat)
        # SUMPRODUCT referencing Expenses sheet
        ws.cell(row=r, column=3).value = f"=SUMPRODUCT(({ex_sheet}!C5:C44=B{r})*({ex_sheet}!E5:E44))"

    chart2 = create_pie_chart(width=16, height=13)
    data2 = Reference(ws, min_col=3, min_row=chart2_data_start, max_row=chart2_data_start + len(cats_expense))
    cats2 = Reference(ws, min_col=2, min_row=chart2_data_start + 1, max_row=chart2_data_start + len(cats_expense))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    setup_chart_titles(chart2, title="Expense Breakdown — March 2025")
    apply_pie_colors(chart2, len(cats_expense))
    ws.add_chart(chart2, "J9")

    # CHART 3: Staff Performance (horizontal bar chart) — placed at B26
    chart3_data_start = 100
    ws.cell(row=chart3_data_start, column=2, value="Staff")
    ws.cell(row=chart3_data_start, column=3, value="Total Sales")
    ws.cell(row=chart3_data_start, column=4, value="Target")
    for i, name in enumerate(STAFF):
        r = chart3_data_start + 1 + i
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3).value = f"={sp_sheet}!C{5 + i}"
        ws.cell(row=r, column=4).value = f"={sp_sheet}!I{5 + i}"

    chart3 = create_bar_chart(chart_type="bar", width=28, height=11)
    data3 = Reference(ws, min_col=3, min_row=chart3_data_start, max_col=4, max_row=chart3_data_start + len(STAFF))
    cats3 = Reference(ws, min_col=2, min_row=chart3_data_start + 1, max_row=chart3_data_start + len(STAFF))
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    setup_chart_titles(chart3, title="Staff Performance — Sales vs Target", x_title="Amount (AED)")
    apply_chart_colors(chart3)
    ws.add_chart(chart3, "B26")

    # CHART 4: Category Sales (bar chart) — placed at J26
    chart4_data_start = 115
    se_sheet = "'Sales Entry'"
    se_data_end = se_data_end_ref  # passed from create_sales_entry()
    # Categories
    categories = ["Accessories", "Audio", "Cables", "Devices", "Peripherals", "Storage", "Wearables"]
    ws.cell(row=chart4_data_start, column=2, value="Category")
    ws.cell(row=chart4_data_start, column=3, value="Total Sales")
    for i, cat in enumerate(categories):
        r = chart4_data_start + 1 + i
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3).value = (
            f"=SUMPRODUCT(({se_sheet}!E5:E{se_data_end}=B{r})*"
            f"({se_sheet}!M5:M{se_data_end}))"
        )

    chart4 = create_bar_chart(width=16, height=11)
    data4 = Reference(ws, min_col=3, min_row=chart4_data_start, max_row=chart4_data_start + len(categories))
    cats4 = Reference(ws, min_col=2, min_row=chart4_data_start + 1, max_row=chart4_data_start + len(categories))
    chart4.add_data(data4, titles_from_data=True)
    chart4.set_categories(cats4)
    setup_chart_titles(chart4, title="Sales by Category — March 2025", y_title="Sales (AED)")
    apply_chart_colors(chart4)
    ws.add_chart(chart4, "J26")

    # Column widths for dashboard
    for c in range(2, 17):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws.column_dimensions["A"].width = 3

    # Hide helper rows
    # We can't easily hide non-contiguous rows, so we'll just make them very small
    # Actually let's hide the helper data rows
    for r in range(50, 130):
        ws.row_dimensions[r].hidden = True

    return ws


# ══════════════════════════════════════════════════════════════
# BUILD ALL SHEETS
# ══════════════════════════════════════════════════════════════
print("Creating Sales Entry...")
_, _, _, se_data_end_actual = create_sales_entry()
print(f"  → Sales Entry data_end = {se_data_end_actual}")

print("Creating Daily Summary...")
create_daily_summary(se_data_end_ref=se_data_end_actual)

print("Creating Weekly Summary...")
create_weekly_summary()

print("Creating Monthly Summary...")
create_monthly_summary()

print("Creating Invoice...")
create_invoice()

print("Creating Inventory...")
create_inventory()

print("Creating Staff Performance...")
create_staff_performance(se_data_end_ref=se_data_end_actual)

print("Creating Expenses...")
create_expenses()

print("Creating Profit Analysis...")
create_profit_analysis()

print("Creating Dashboard...")
create_dashboard(se_data_end_ref=se_data_end_actual)

# Move Dashboard to be the first visible sheet
wb.move_sheet("Dashboard", offset=-9)

# Save
output_path = "/home/z/my-project/download/POS_Premium_Sample.xlsx"
wb.save(output_path)
print(f"\n✅ Saved to: {output_path}")
print(f"   Sheets: {wb.sheetnames}")
