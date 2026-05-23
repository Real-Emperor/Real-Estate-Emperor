#!/usr/bin/env python3
"""
Generate Premium POS Empty Template — 10 sheets, Bloomberg palette.
Ready for actual use and sale. No sample data — headers, formulas, structure only.
"""

import sys, os
sys.path.insert(0, "/home/z/my-project/skills/xlsx")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

# ── Import base template ──────────────────────────────────────
from templates.base import (
    use_palette_explicit, setup_sheet, style_header_row, style_data_row,
    style_total_row, font_title, font_header, font_subheader, font_body,
    font_caption, font_kpi, font_kpi_label, fill_header, fill_total, fill_data_row,
    border_header, border_total, align_title, align_header, align_number,
    align_text, align_date, COLUMN_WIDTHS, ROW_HEIGHTS,
    PRIMARY, PRIMARY_LIGHT, SECONDARY, HEADER_TEXT,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_50, NEUTRAL_0,
    CHART_COLORS, CF_POSITIVE_FILL, CF_POSITIVE_FONT,
    CF_NEGATIVE_FILL, CF_NEGATIVE_FONT, CF_WARNING_FILL, CF_WARNING_FONT,
    create_bar_chart, create_line_chart, create_pie_chart,
    setup_chart_titles, apply_chart_colors, apply_pie_colors,
    make_chart_title, auto_fit_columns, FORMATS,
    FONT_NAME, HEADER_BOLD,
)

# ── Activate bloomberg palette BEFORE creating styles ─────────
use_palette_explicit("bloomberg")

# ── Re-import tokens after palette activation ─────────────────
from templates.base import (
    PRIMARY, PRIMARY_LIGHT, SECONDARY, HEADER_TEXT,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_50, NEUTRAL_0,
    CHART_COLORS, CF_POSITIVE_FILL, CF_POSITIVE_FONT,
    CF_NEGATIVE_FILL, CF_NEGATIVE_FONT, CF_WARNING_FILL, CF_WARNING_FONT,
    FONT_NAME, HEADER_BOLD,
)

# ── Number Formats ────────────────────────────────────────────
FMT_CURRENCY = "#,##0.00"
FMT_INT = "#,##0"
FMT_PCT = "0.0%"
FMT_DATE = "YYYY-MM-DD"

# ── Standard constants ────────────────────────────────────────
COL_START = 2  # Start at column B
HEADER_ROW = 4
DATA_START = 5

# ── Workbook ──────────────────────────────────────────────────
wb = Workbook()
wb.properties.creator = "Z.ai"


def apply_col_widths(ws, widths, start_col=COL_START):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w


# ══════════════════════════════════════════════════════════════
# SHEET 1: Sales Entry
# ══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Sales Entry"
ws1.sheet_properties.tabColor = PRIMARY

HEADERS_S1 = ["Date", "Invoice #", "Item/Service", "Category", "Qty",
              "Unit Price", "Total", "Payment Method", "Staff Name",
              "Customer Name", "Discount", "Net Total", "Notes"]
COL_END_S1 = COL_START + len(HEADERS_S1) - 1
N_SALES = 500

setup_sheet(ws1, "Sales Entry — Premium POS", COL_END_S1)
apply_col_widths(ws1, [14, 14, 22, 14, 8, 14, 14, 16, 16, 18, 14, 14, 20])

for i, h in enumerate(HEADERS_S1):
    ws1.cell(row=HEADER_ROW, column=COL_START + i, value=h)
style_header_row(ws1, HEADER_ROW, COL_START, COL_END_S1)

# Write 500 empty data rows with formulas
for idx in range(N_SALES):
    r = DATA_START + idx
    # Total = Qty * Unit Price
    ws1.cell(row=r, column=COL_START + 6).value = "=IFERROR(F{0}*G{0},0)".format(r)
    ws1.cell(row=r, column=COL_START + 6).number_format = FMT_CURRENCY
    # Net Total = Total - Discount
    ws1.cell(row=r, column=COL_START + 11).value = "=H{0}-IFERROR(L{0},0)".format(r)
    ws1.cell(row=r, column=COL_START + 11).number_format = FMT_CURRENCY

    style_data_row(ws1, r, COL_START, COL_END_S1, idx)

    ws1.cell(row=r, column=COL_START).alignment = align_date()
    ws1.cell(row=r, column=COL_START).number_format = FMT_DATE
    ws1.cell(row=r, column=COL_START + 1).alignment = align_text()
    ws1.cell(row=r, column=COL_START + 2).alignment = align_text()
    ws1.cell(row=r, column=COL_START + 3).alignment = align_text()
    ws1.cell(row=r, column=COL_START + 4).alignment = align_number()
    ws1.cell(row=r, column=COL_START + 4).number_format = FMT_INT
    ws1.cell(row=r, column=COL_START + 5).alignment = align_number()
    ws1.cell(row=r, column=COL_START + 5).number_format = FMT_CURRENCY
    ws1.cell(row=r, column=COL_START + 6).alignment = align_number()
    ws1.cell(row=r, column=COL_START + 6).number_format = FMT_CURRENCY
    ws1.cell(row=r, column=COL_START + 7).alignment = align_text()
    ws1.cell(row=r, column=COL_START + 8).alignment = align_text()
    ws1.cell(row=r, column=COL_START + 9).alignment = align_text()
    ws1.cell(row=r, column=COL_START + 10).alignment = align_number()
    ws1.cell(row=r, column=COL_START + 10).number_format = FMT_CURRENCY
    ws1.cell(row=r, column=COL_START + 11).alignment = align_number()
    ws1.cell(row=r, column=COL_START + 11).number_format = FMT_CURRENCY
    ws1.cell(row=r, column=COL_START + 12).alignment = align_text()

# Payment Method data validation
dv_payment = DataValidation(type="list", formula1='"Cash,Card,Online,Cheque,Bank Transfer"', allow_blank=True)
dv_payment.error = "Please select a valid payment method"
dv_payment.errorTitle = "Invalid Payment Method"
dv_payment.prompt = "Select payment method"
dv_payment.promptTitle = "Payment Method"
ws1.add_data_validation(dv_payment)
pay_col = get_column_letter(COL_START + 7)
dv_payment.add("{0}{1}:{0}{2}".format(pay_col, DATA_START, DATA_START + N_SALES - 1))

# Totals row
TOTAL_ROW_S1 = DATA_START + N_SALES
ws1.cell(row=TOTAL_ROW_S1, column=COL_START, value="TOTAL")
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 4).value = "=SUM(F{0}:F{1})".format(DATA_START, TOTAL_ROW_S1 - 1)
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 4).number_format = FMT_INT
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 6).value = "=SUM(H{0}:H{1})".format(DATA_START, TOTAL_ROW_S1 - 1)
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 6).number_format = FMT_CURRENCY
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 10).value = "=SUM(L{0}:L{1})".format(DATA_START, TOTAL_ROW_S1 - 1)
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 10).number_format = FMT_CURRENCY
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 11).value = "=SUM(M{0}:M{1})".format(DATA_START, TOTAL_ROW_S1 - 1)
ws1.cell(row=TOTAL_ROW_S1, column=COL_START + 11).number_format = FMT_CURRENCY
style_total_row(ws1, TOTAL_ROW_S1, COL_START, COL_END_S1)
ws1.cell(row=TOTAL_ROW_S1, column=COL_START).alignment = align_text()

# Freeze panes at C5
ws1.freeze_panes = "C5"
ws1.print_area = "A1:{0}{1}".format(get_column_letter(COL_END_S1), TOTAL_ROW_S1)
ws1.print_title_rows = "{0}:{0}".format(HEADER_ROW)

# Sales Entry column references for other sheets
SE_DATE = get_column_letter(COL_START)        # B
SE_TOTAL = get_column_letter(COL_START + 6)   # H
SE_PAY = get_column_letter(COL_START + 7)     # I
SE_QTY = get_column_letter(COL_START + 4)     # F
SE_DISC = get_column_letter(COL_START + 10)   # L
SE_NET = get_column_letter(COL_START + 11)    # M
SE_STAFF = get_column_letter(COL_START + 8)   # J
SE_CAT = get_column_letter(COL_START + 3)     # E
SE_LAST = TOTAL_ROW_S1 - 1


# ══════════════════════════════════════════════════════════════
# SHEET 2: Daily Summary
# ══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Daily Summary")
ws2.sheet_properties.tabColor = PRIMARY

HEADERS_S2 = ["Date", "Total Sales", "# Transactions", "Avg Transaction",
              "Cash", "Card", "Online", "Bank Transfer", "# Items Sold", "Total Discount Given"]
COL_END_S2 = COL_START + len(HEADERS_S2) - 1
N_DAYS = 31

setup_sheet(ws2, "Daily Summary — Auto-Calculated", COL_END_S2)
apply_col_widths(ws2, [14, 16, 16, 16, 16, 16, 16, 16, 14, 18])

for i, h in enumerate(HEADERS_S2):
    ws2.cell(row=HEADER_ROW, column=COL_START + i, value=h)
style_header_row(ws2, HEADER_ROW, COL_START, COL_END_S2)

DS_DATE = get_column_letter(COL_START)        # B
DS_SALES = get_column_letter(COL_START + 1)   # C
DS_TXNS = get_column_letter(COL_START + 2)    # D

for day in range(N_DAYS):
    r = DATA_START + day
    # Total Sales
    ws2.cell(row=r, column=COL_START + 1).value = "=SUMPRODUCT(('Sales Entry'!{0}{2}:{0}{3}=B{1})*('Sales Entry'!{4}{2}:{4}{3}))".format(SE_DATE, r, DATA_START, SE_LAST, SE_NET)
    ws2.cell(row=r, column=COL_START + 1).number_format = FMT_CURRENCY
    # # Transactions
    ws2.cell(row=r, column=COL_START + 2).value = "=SUMPRODUCT(('Sales Entry'!{0}{2}:{0}{3}=B{1})*1)".format(SE_DATE, r, DATA_START, SE_LAST)
    ws2.cell(row=r, column=COL_START + 2).number_format = FMT_INT
    # Avg Transaction
    ws2.cell(row=r, column=COL_START + 3).value = "=IFERROR(C{0}/D{0},0)".format(r)
    ws2.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    # Cash
    ws2.cell(row=r, column=COL_START + 4).value = "=SUMPRODUCT(('Sales Entry'!{0}{2}:{0}{3}=B{1})*('Sales Entry'!{4}{2}:{4}{3}=\"Cash\")*('Sales Entry'!{5}{2}:{5}{3}))".format(SE_DATE, r, DATA_START, SE_LAST, SE_PAY, SE_NET)
    ws2.cell(row=r, column=COL_START + 4).number_format = FMT_CURRENCY
    # Card
    ws2.cell(row=r, column=COL_START + 5).value = "=SUMPRODUCT(('Sales Entry'!{0}{2}:{0}{3}=B{1})*('Sales Entry'!{4}{2}:{4}{3}=\"Card\")*('Sales Entry'!{5}{2}:{5}{3}))".format(SE_DATE, r, DATA_START, SE_LAST, SE_PAY, SE_NET)
    ws2.cell(row=r, column=COL_START + 5).number_format = FMT_CURRENCY
    # Online
    ws2.cell(row=r, column=COL_START + 6).value = "=SUMPRODUCT(('Sales Entry'!{0}{2}:{0}{3}=B{1})*('Sales Entry'!{4}{2}:{4}{3}=\"Online\")*('Sales Entry'!{5}{2}:{5}{3}))".format(SE_DATE, r, DATA_START, SE_LAST, SE_PAY, SE_NET)
    ws2.cell(row=r, column=COL_START + 6).number_format = FMT_CURRENCY
    # Bank Transfer
    ws2.cell(row=r, column=COL_START + 7).value = "=SUMPRODUCT(('Sales Entry'!{0}{2}:{0}{3}=B{1})*('Sales Entry'!{4}{2}:{4}{3}=\"Bank Transfer\")*('Sales Entry'!{5}{2}:{5}{3}))".format(SE_DATE, r, DATA_START, SE_LAST, SE_PAY, SE_NET)
    ws2.cell(row=r, column=COL_START + 7).number_format = FMT_CURRENCY
    # # Items Sold
    ws2.cell(row=r, column=COL_START + 8).value = "=SUMPRODUCT(('Sales Entry'!{0}{2}:{0}{3}=B{1})*('Sales Entry'!{4}{2}:{4}{3}))".format(SE_DATE, r, DATA_START, SE_LAST, SE_QTY)
    ws2.cell(row=r, column=COL_START + 8).number_format = FMT_INT
    # Total Discount Given
    ws2.cell(row=r, column=COL_START + 9).value = "=SUMPRODUCT(('Sales Entry'!{0}{2}:{0}{3}=B{1})*('Sales Entry'!{4}{2}:{4}{3}))".format(SE_DATE, r, DATA_START, SE_LAST, SE_DISC)
    ws2.cell(row=r, column=COL_START + 9).number_format = FMT_CURRENCY

    style_data_row(ws2, r, COL_START, COL_END_S2, day)
    ws2.cell(row=r, column=COL_START).alignment = align_date()
    ws2.cell(row=r, column=COL_START).number_format = FMT_DATE
    for cc in range(1, len(HEADERS_S2)):
        if cc in [2, 8]:
            ws2.cell(row=r, column=COL_START + cc).alignment = align_number()
            ws2.cell(row=r, column=COL_START + cc).number_format = FMT_INT
        else:
            ws2.cell(row=r, column=COL_START + cc).alignment = align_number()

# Totals row
TR_S2 = DATA_START + N_DAYS
ws2.cell(row=TR_S2, column=COL_START, value="TOTAL")
for cc in [1, 4, 5, 6, 7, 9]:
    col_l = get_column_letter(COL_START + cc)
    ws2.cell(row=TR_S2, column=COL_START + cc).value = "=SUM({0}{1}:{0}{2})".format(col_l, DATA_START, TR_S2 - 1)
    ws2.cell(row=TR_S2, column=COL_START + cc).number_format = FMT_CURRENCY
for cc in [2, 8]:
    col_l = get_column_letter(COL_START + cc)
    ws2.cell(row=TR_S2, column=COL_START + cc).value = "=SUM({0}{1}:{0}{2})".format(col_l, DATA_START, TR_S2 - 1)
    ws2.cell(row=TR_S2, column=COL_START + cc).number_format = FMT_INT
ws2.cell(row=TR_S2, column=COL_START + 3).value = "=IFERROR(C{0}/D{0},0)".format(TR_S2)
ws2.cell(row=TR_S2, column=COL_START + 3).number_format = FMT_CURRENCY
style_total_row(ws2, TR_S2, COL_START, COL_END_S2)
ws2.cell(row=TR_S2, column=COL_START).alignment = align_text()

ws2.print_area = "A1:{0}{1}".format(get_column_letter(COL_END_S2), TR_S2)
ws2.print_title_rows = "{0}:{0}".format(HEADER_ROW)


# ══════════════════════════════════════════════════════════════
# SHEET 3: Weekly Summary
# ══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Weekly Summary")
ws3.sheet_properties.tabColor = PRIMARY

HEADERS_S3 = ["Week #", "Week Start", "Week End", "Total Sales",
              "# Transactions", "Avg Transaction", "Top Staff", "Profit Estimate"]
COL_END_S3 = COL_START + len(HEADERS_S3) - 1
N_WEEKS = 52

setup_sheet(ws3, "Weekly Summary — 52 Weeks", COL_END_S3)
apply_col_widths(ws3, [10, 14, 14, 16, 16, 16, 16, 16])

for i, h in enumerate(HEADERS_S3):
    ws3.cell(row=HEADER_ROW, column=COL_START + i, value=h)
style_header_row(ws3, HEADER_ROW, COL_START, COL_END_S3)

for idx in range(N_WEEKS):
    r = DATA_START + idx
    ws3.cell(row=r, column=COL_START, value=idx + 1)
    # Total Sales from Daily Summary
    ws3.cell(row=r, column=COL_START + 3).value = "=SUMPRODUCT(('Daily Summary'!{0}{2}:{0}{3}>=C{1})*('Daily Summary'!{0}{2}:{0}{3}<=D{1})*('Daily Summary'!{4}{2}:{4}{3}))".format(DS_DATE, r, DATA_START, TR_S2 - 1, DS_SALES)
    ws3.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    # # Transactions
    ws3.cell(row=r, column=COL_START + 4).value = "=SUMPRODUCT(('Daily Summary'!{0}{2}:{0}{3}>=C{1})*('Daily Summary'!{0}{2}:{0}{3}<=D{1})*('Daily Summary'!{4}{2}:{4}{3}))".format(DS_DATE, r, DATA_START, TR_S2 - 1, DS_TXNS)
    ws3.cell(row=r, column=COL_START + 4).number_format = FMT_INT
    # Avg Transaction
    ws3.cell(row=r, column=COL_START + 5).value = "=IFERROR(E{0}/F{0},0)".format(r)
    ws3.cell(row=r, column=COL_START + 5).number_format = FMT_CURRENCY
    # Top Staff - leave as user reference (complex nested IF)
    # Profit Estimate (30% margin)
    ws3.cell(row=r, column=COL_START + 7).value = "=E{0}*0.3".format(r)
    ws3.cell(row=r, column=COL_START + 7).number_format = FMT_CURRENCY

    style_data_row(ws3, r, COL_START, COL_END_S3, idx)
    ws3.cell(row=r, column=COL_START).alignment = align_number()
    ws3.cell(row=r, column=COL_START + 1).alignment = align_date()
    ws3.cell(row=r, column=COL_START + 1).number_format = FMT_DATE
    ws3.cell(row=r, column=COL_START + 2).alignment = align_date()
    ws3.cell(row=r, column=COL_START + 2).number_format = FMT_DATE
    ws3.cell(row=r, column=COL_START + 3).alignment = align_number()
    ws3.cell(row=r, column=COL_START + 4).alignment = align_number()
    ws3.cell(row=r, column=COL_START + 5).alignment = align_number()
    ws3.cell(row=r, column=COL_START + 6).alignment = align_text()
    ws3.cell(row=r, column=COL_START + 7).alignment = align_number()

# Totals row
TR_S3 = DATA_START + N_WEEKS
ws3.cell(row=TR_S3, column=COL_START, value="TOTAL")
for cc in [3, 7]:
    col_l = get_column_letter(COL_START + cc)
    ws3.cell(row=TR_S3, column=COL_START + cc).value = "=SUM({0}{1}:{0}{2})".format(col_l, DATA_START, TR_S3 - 1)
    ws3.cell(row=TR_S3, column=COL_START + cc).number_format = FMT_CURRENCY
col_l4 = get_column_letter(COL_START + 4)
ws3.cell(row=TR_S3, column=COL_START + 4).value = "=SUM({0}{1}:{0}{2})".format(col_l4, DATA_START, TR_S3 - 1)
ws3.cell(row=TR_S3, column=COL_START + 4).number_format = FMT_INT
ws3.cell(row=TR_S3, column=COL_START + 5).value = "=IFERROR(E{0}/F{0},0)".format(TR_S3)
ws3.cell(row=TR_S3, column=COL_START + 5).number_format = FMT_CURRENCY
style_total_row(ws3, TR_S3, COL_START, COL_END_S3)
ws3.cell(row=TR_S3, column=COL_START).alignment = align_text()

ws3.print_area = "A1:{0}{1}".format(get_column_letter(COL_END_S3), TR_S3)
ws3.print_title_rows = "{0}:{0}".format(HEADER_ROW)


# ══════════════════════════════════════════════════════════════
# SHEET 4: Monthly Summary
# ══════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Monthly Summary")
ws4.sheet_properties.tabColor = PRIMARY

HEADERS_S4 = ["Month", "Total Sales", "# Transactions", "Avg Transaction",
              "Best Day Sales", "Best Day Date", "Total Expenses", "Net Profit"]
COL_END_S4 = COL_START + len(HEADERS_S4) - 1
N_MONTHS = 12

setup_sheet(ws4, "Monthly Summary — 12 Months", COL_END_S4)
apply_col_widths(ws4, [16, 16, 16, 16, 16, 16, 16, 16])

for i, h in enumerate(HEADERS_S4):
    ws4.cell(row=HEADER_ROW, column=COL_START + i, value=h)
style_header_row(ws4, HEADER_ROW, COL_START, COL_END_S4)

MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]

# Expenses references (Expenses sheet has same COL_START/HEADER_ROW/DATA_START pattern)
EXP_DATE_COL = get_column_letter(COL_START)      # B in Expenses
EXP_CAT_COL = get_column_letter(COL_START + 1)   # C in Expenses
EXP_AMT_COL = get_column_letter(COL_START + 3)   # E in Expenses
N_EXPENSES = 300
EXP_LAST = DATA_START + N_EXPENSES - 1

for idx in range(N_MONTHS):
    r = DATA_START + idx
    ws4.cell(row=r, column=COL_START, value=MONTH_NAMES[idx])

    # Total Sales from Daily Summary
    ws4.cell(row=r, column=COL_START + 1).value = "=SUMPRODUCT((MONTH('Daily Summary'!{0}{2}:{0}{3})={1})*('Daily Summary'!{4}{2}:{4}{3}))".format(DS_DATE, idx + 1, DATA_START, TR_S2 - 1, DS_SALES)
    ws4.cell(row=r, column=COL_START + 1).number_format = FMT_CURRENCY
    # # Transactions
    ws4.cell(row=r, column=COL_START + 2).value = "=SUMPRODUCT((MONTH('Daily Summary'!{0}{2}:{0}{3})={1})*('Daily Summary'!{4}{2}:{4}{3}))".format(DS_DATE, idx + 1, DATA_START, TR_S2 - 1, DS_TXNS)
    ws4.cell(row=r, column=COL_START + 2).number_format = FMT_INT
    # Avg Transaction
    ws4.cell(row=r, column=COL_START + 3).value = "=IFERROR(C{0}/D{0},0)".format(r)
    ws4.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    # Best Day Sales
    ws4.cell(row=r, column=COL_START + 4).value = "=IFERROR(MAX(IF(MONTH('Daily Summary'!{0}{2}:{0}{3})={1},'Daily Summary'!{4}{2}:{4}{3})),0)".format(DS_DATE, idx + 1, DATA_START, TR_S2 - 1, DS_SALES)
    ws4.cell(row=r, column=COL_START + 4).number_format = FMT_CURRENCY
    # Best Day Date
    ws4.cell(row=r, column=COL_START + 5).value = "=IFERROR(INDEX('Daily Summary'!{0}{2}:{0}{3},MATCH(F{1},IF(MONTH('Daily Summary'!{0}{2}:{0}{3})={4},'Daily Summary'!{5}{2}:{5}{3}),0)),\"\")".format(DS_DATE, r, DATA_START, TR_S2 - 1, idx + 1, DS_SALES)
    ws4.cell(row=r, column=COL_START + 5).number_format = FMT_DATE
    # Total Expenses from Expenses sheet
    ws4.cell(row=r, column=COL_START + 6).value = "=SUMPRODUCT((MONTH('Expenses'!{0}{2}:{0}{3})={1})*('Expenses'!{4}{2}:{4}{3}))".format(EXP_DATE_COL, idx + 1, DATA_START, EXP_LAST, EXP_AMT_COL)
    ws4.cell(row=r, column=COL_START + 6).number_format = FMT_CURRENCY
    # Net Profit
    ws4.cell(row=r, column=COL_START + 7).value = "=C{0}-H{0}".format(r)
    ws4.cell(row=r, column=COL_START + 7).number_format = FMT_CURRENCY

    style_data_row(ws4, r, COL_START, COL_END_S4, idx)
    ws4.cell(row=r, column=COL_START).alignment = align_text()
    for cc in [1, 3, 4, 6, 7]:
        ws4.cell(row=r, column=COL_START + cc).alignment = align_number()
    ws4.cell(row=r, column=COL_START + 2).alignment = align_number()
    ws4.cell(row=r, column=COL_START + 5).alignment = align_date()

# Totals row
TR_S4 = DATA_START + N_MONTHS
ws4.cell(row=TR_S4, column=COL_START, value="TOTAL")
for cc in [1, 6, 7]:
    col_l = get_column_letter(COL_START + cc)
    ws4.cell(row=TR_S4, column=COL_START + cc).value = "=SUM({0}{1}:{0}{2})".format(col_l, DATA_START, TR_S4 - 1)
    ws4.cell(row=TR_S4, column=COL_START + cc).number_format = FMT_CURRENCY
col_l2 = get_column_letter(COL_START + 2)
ws4.cell(row=TR_S4, column=COL_START + 2).value = "=SUM({0}{1}:{0}{2})".format(col_l2, DATA_START, TR_S4 - 1)
ws4.cell(row=TR_S4, column=COL_START + 2).number_format = FMT_INT
ws4.cell(row=TR_S4, column=COL_START + 3).value = "=IFERROR(C{0}/D{0},0)".format(TR_S4)
ws4.cell(row=TR_S4, column=COL_START + 3).number_format = FMT_CURRENCY
style_total_row(ws4, TR_S4, COL_START, COL_END_S4)
ws4.cell(row=TR_S4, column=COL_START).alignment = align_text()

# Bar chart: Monthly Sales vs Expenses
chart4 = create_bar_chart(chart_type="col", grouping="clustered", gap_width=80, width=22, height=12)
sales_ref = Reference(ws4, min_col=COL_START + 1, min_row=HEADER_ROW, max_row=DATA_START + N_MONTHS - 1)
exp_ref = Reference(ws4, min_col=COL_START + 6, min_row=HEADER_ROW, max_row=DATA_START + N_MONTHS - 1)
cats_ref = Reference(ws4, min_col=COL_START, min_row=DATA_START, max_row=DATA_START + N_MONTHS - 1)
chart4.add_data(sales_ref, titles_from_data=True)
chart4.add_data(exp_ref, titles_from_data=True)
chart4.set_categories(cats_ref)
chart4.shape = 4
setup_chart_titles(chart4, title="Monthly Sales vs Expenses", y_title="Amount", x_title="Month")
apply_chart_colors(chart4, [CHART_COLORS[0], CHART_COLORS[3]])
ws4.add_chart(chart4, "B{0}".format(TR_S4 + 2))

ws4.print_area = "A1:{0}{1}".format(get_column_letter(COL_END_S4), TR_S4)
ws4.print_title_rows = "{0}:{0}".format(HEADER_ROW)


# ══════════════════════════════════════════════════════════════
# SHEET 5: Invoice
# ══════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Invoice")
ws5.sheet_properties.tabColor = PRIMARY

setup_sheet(ws5, None, 8)

# Company header placeholder
ws5.merge_cells("B2:H2")
cell = ws5.cell(row=2, column=2, value="[Your Company Name]")
cell.font = Font(name=FONT_NAME, size=20, bold=HEADER_BOLD, color=PRIMARY)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[2].height = 40

ws5.merge_cells("B3:H3")
cell = ws5.cell(row=3, column=2, value="[Address Line 1 | City, Country]")
cell.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_600)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[3].height = 20

ws5.merge_cells("B4:H4")
cell = ws5.cell(row=4, column=2, value="Tel: [Phone]  |  VAT: [VAT Number]")
cell.font = Font(name=FONT_NAME, size=9, color=NEUTRAL_600)
cell.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[4].height = 18

# Divider
ws5.row_dimensions[5].height = 6
for c in range(2, 9):
    ws5.cell(row=5, column=c).border = Border(bottom=Side(style="medium", color=PRIMARY))

# Invoice details
ws5.cell(row=7, column=2, value="Invoice #:").font = font_subheader()
ws5.cell(row=7, column=6, value="Date:").font = font_subheader()
ws5.cell(row=8, column=2, value="Customer:").font = font_subheader()
ws5.cell(row=8, column=6, value="Staff:").font = font_subheader()

# Invoice items header
INV_HDR_ROW = 10
inv_headers = ["Item", "Qty", "Price", "Discount", "Total"]
for i, h in enumerate(inv_headers):
    ws5.cell(row=INV_HDR_ROW, column=2 + i, value=h)
style_header_row(ws5, INV_HDR_ROW, 2, 6)

# 15 empty rows with formulas
N_INV_ROWS = 15
for idx in range(N_INV_ROWS):
    r = INV_HDR_ROW + 1 + idx
    # Total = Qty * Price - Discount
    ws5.cell(row=r, column=6).value = "=IFERROR(C{0}*D{0}-IFERROR(E{0},0),0)".format(r)
    ws5.cell(row=r, column=6).number_format = FMT_CURRENCY
    style_data_row(ws5, r, 2, 6, idx)
    ws5.cell(row=r, column=2).alignment = align_text()
    ws5.cell(row=r, column=3).alignment = align_number()
    ws5.cell(row=r, column=3).number_format = FMT_INT
    ws5.cell(row=r, column=4).alignment = align_number()
    ws5.cell(row=r, column=4).number_format = FMT_CURRENCY
    ws5.cell(row=r, column=5).alignment = align_number()
    ws5.cell(row=r, column=5).number_format = FMT_CURRENCY
    ws5.cell(row=r, column=6).alignment = align_number()

# Totals section
INV_TOT_START = INV_HDR_ROW + 1 + N_INV_ROWS
ws5.row_dimensions[INV_TOT_START].height = 8

r = INV_TOT_START + 1
ws5.cell(row=r, column=4, value="Subtotal:").font = font_subheader()
ws5.cell(row=r, column=4).alignment = Alignment(horizontal="right", vertical="center")
ws5.cell(row=r, column=6).value = "=SUM(F{0}:F{1})".format(INV_HDR_ROW + 1, INV_HDR_ROW + N_INV_ROWS)
ws5.cell(row=r, column=6).number_format = FMT_CURRENCY
ws5.cell(row=r, column=6).font = font_body()
ws5.cell(row=r, column=6).alignment = align_number()
SUBTOTAL_ROW = r

r += 1
ws5.cell(row=r, column=4, value="Discount:").font = font_subheader()
ws5.cell(row=r, column=4).alignment = Alignment(horizontal="right", vertical="center")
ws5.cell(row=r, column=6).number_format = FMT_CURRENCY
ws5.cell(row=r, column=6).font = font_body()
ws5.cell(row=r, column=6).alignment = align_number()
DISCOUNT_ROW = r

r += 1
ws5.cell(row=r, column=4, value="VAT (5%):").font = font_subheader()
ws5.cell(row=r, column=4).alignment = Alignment(horizontal="right", vertical="center")
ws5.cell(row=r, column=6).value = "=IFERROR((F{0}-IFERROR(F{1},0))*0.05,0)".format(SUBTOTAL_ROW, DISCOUNT_ROW)
ws5.cell(row=r, column=6).number_format = FMT_CURRENCY
ws5.cell(row=r, column=6).font = font_body()
ws5.cell(row=r, column=6).alignment = align_number()
VAT_ROW = r

r += 1
for c in range(4, 7):
    ws5.cell(row=r, column=c).border = Border(top=Side(style="medium", color=PRIMARY))
ws5.cell(row=r, column=4, value="Grand Total:").font = Font(name=FONT_NAME, size=13, bold=True, color=PRIMARY)
ws5.cell(row=r, column=4).alignment = Alignment(horizontal="right", vertical="center")
ws5.cell(row=r, column=6).value = "=F{0}-IFERROR(F{1},0)+F{2}".format(SUBTOTAL_ROW, DISCOUNT_ROW, VAT_ROW)
ws5.cell(row=r, column=6).number_format = FMT_CURRENCY
ws5.cell(row=r, column=6).font = Font(name=FONT_NAME, size=13, bold=True, color=PRIMARY)
ws5.cell(row=r, column=6).alignment = align_number()
ws5.row_dimensions[r].height = 30
GRAND_TOTAL_ROW = r

r += 2
ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
cell = ws5.cell(row=r, column=2, value="Thank you for your business!")
cell.font = Font(name=FONT_NAME, size=11, italic=True, color=NEUTRAL_600)
cell.alignment = Alignment(horizontal="center", vertical="center")

# Column widths for invoice
apply_col_widths(ws5, [22, 8, 14, 14, 14], start_col=2)
ws5.print_area = "A1:H{0}".format(GRAND_TOTAL_ROW + 3)


# ══════════════════════════════════════════════════════════════
# SHEET 6: Inventory
# ══════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("Inventory")
ws6.sheet_properties.tabColor = PRIMARY

HEADERS_S6 = ["Product ID", "Item Name", "Category", "Unit Price", "Cost Price",
              "Margin", "Stock Qty", "Reorder Level", "Status", "Supplier",
              "Last Restocked", "Days Since Restock"]
COL_END_S6 = COL_START + len(HEADERS_S6) - 1
N_INV = 200

setup_sheet(ws6, "Inventory — Stock Tracking", COL_END_S6)
apply_col_widths(ws6, [12, 22, 14, 14, 14, 12, 12, 14, 14, 18, 16, 16])

for i, h in enumerate(HEADERS_S6):
    ws6.cell(row=HEADER_ROW, column=COL_START + i, value=h)
style_header_row(ws6, HEADER_ROW, COL_START, COL_END_S6)

INVENTORY_CATEGORIES = ["Grains", "Dairy", "Bakery", "Meat", "Beverages",
                         "Produce", "Snacks", "Cleaning", "Electronics", "Other"]

for idx in range(N_INV):
    r = DATA_START + idx
    # Margin = (Unit Price - Cost Price) / Unit Price
    ws6.cell(row=r, column=COL_START + 5).value = "=IFERROR((E{0}-F{0})/E{0},0)".format(r)
    ws6.cell(row=r, column=COL_START + 5).number_format = FMT_PCT
    # Status = IF(Stock<=Reorder,"LOW STOCK","In Stock")
    ws6.cell(row=r, column=COL_START + 8).value = '=IF(H{0}<=I{0},"LOW STOCK","In Stock")'.format(r)
    # Days Since Restock
    ws6.cell(row=r, column=COL_START + 11).value = "=IFERROR(TODAY()-L{0},\"\")".format(r)
    ws6.cell(row=r, column=COL_START + 11).number_format = FMT_INT

    style_data_row(ws6, r, COL_START, COL_END_S6, idx)
    ws6.cell(row=r, column=COL_START).alignment = align_text()
    ws6.cell(row=r, column=COL_START + 1).alignment = align_text()
    ws6.cell(row=r, column=COL_START + 2).alignment = align_text()
    ws6.cell(row=r, column=COL_START + 3).alignment = align_number()
    ws6.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    ws6.cell(row=r, column=COL_START + 4).alignment = align_number()
    ws6.cell(row=r, column=COL_START + 4).number_format = FMT_CURRENCY
    ws6.cell(row=r, column=COL_START + 5).alignment = align_number()
    ws6.cell(row=r, column=COL_START + 5).number_format = FMT_PCT
    ws6.cell(row=r, column=COL_START + 6).alignment = align_number()
    ws6.cell(row=r, column=COL_START + 6).number_format = FMT_INT
    ws6.cell(row=r, column=COL_START + 7).alignment = align_number()
    ws6.cell(row=r, column=COL_START + 7).number_format = FMT_INT
    ws6.cell(row=r, column=COL_START + 8).alignment = align_text()
    ws6.cell(row=r, column=COL_START + 9).alignment = align_text()
    ws6.cell(row=r, column=COL_START + 10).alignment = align_date()
    ws6.cell(row=r, column=COL_START + 10).number_format = FMT_DATE
    ws6.cell(row=r, column=COL_START + 11).alignment = align_number()

TR_S6 = DATA_START + N_INV

# Category dropdown validation
dv_cat = DataValidation(type="list", formula1='"' + ','.join(INVENTORY_CATEGORIES) + '"', allow_blank=True)
dv_cat.error = "Please select a valid category"
dv_cat.errorTitle = "Invalid Category"
ws6.add_data_validation(dv_cat)
cat_col = get_column_letter(COL_START + 2)
dv_cat.add("{0}{1}:{0}{2}".format(cat_col, DATA_START, TR_S6 - 1))

# Conditional formatting: Status column
status_col_letter = get_column_letter(COL_START + 8)
status_range = "{0}{1}:{0}{2}".format(status_col_letter, DATA_START, TR_S6 - 1)
ws6.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"LOW STOCK"'], fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))
ws6.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=['"In Stock"'], fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT))

ws6.print_area = "A1:{0}{1}".format(get_column_letter(COL_END_S6), TR_S6 - 1)
ws6.print_title_rows = "{0}:{0}".format(HEADER_ROW)


# ══════════════════════════════════════════════════════════════
# SHEET 7: Staff Performance
# ══════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Staff Performance")
ws7.sheet_properties.tabColor = PRIMARY

HEADERS_S7 = ["Staff Name", "Total Sales", "# Transactions", "Avg Transaction",
              "Top Category", "Shift", "Commission (5%)", "Sales Target",
              "Target Achievement %"]
COL_END_S7 = COL_START + len(HEADERS_S7) - 1
N_STAFF = 25

setup_sheet(ws7, "Staff Performance — Employee Tracking", COL_END_S7)
apply_col_widths(ws7, [18, 16, 16, 16, 16, 12, 16, 16, 18])

for i, h in enumerate(HEADERS_S7):
    ws7.cell(row=HEADER_ROW, column=COL_START + i, value=h)
style_header_row(ws7, HEADER_ROW, COL_START, COL_END_S7)

for idx in range(N_STAFF):
    r = DATA_START + idx
    # Total Sales = SUMPRODUCT
    ws7.cell(row=r, column=COL_START + 1).value = "=SUMPRODUCT(('Sales Entry'!{0}{2}:{0}{3}=B{1})*('Sales Entry'!{4}{2}:{4}{3}))".format(SE_STAFF, r, DATA_START, SE_LAST, SE_NET)
    ws7.cell(row=r, column=COL_START + 1).number_format = FMT_CURRENCY
    # # Transactions
    ws7.cell(row=r, column=COL_START + 2).value = "=SUMPRODUCT(('Sales Entry'!{0}{2}:{0}{3}=B{1})*1)".format(SE_STAFF, r, DATA_START, SE_LAST)
    ws7.cell(row=r, column=COL_START + 2).number_format = FMT_INT
    # Avg Transaction
    ws7.cell(row=r, column=COL_START + 3).value = "=IFERROR(C{0}/D{0},0)".format(r)
    ws7.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    # Top Category - SUMPRODUCT per category, find max
    cat_formulas_parts = []
    for ci, cat in enumerate(INVENTORY_CATEGORIES):
        part = "SUMPRODUCT(('Sales Entry'!{0}{2}:{0}{3}=B{1})*('Sales Entry'!{4}{2}:{4}{3}=\"{5}\")*('Sales Entry'!{6}{2}:{6}{3}))".format(SE_STAFF, r, DATA_START, SE_LAST, SE_CAT, cat, SE_NET)
        cat_formulas_parts.append(part)
    
    # Build nested IF for top category (simplified - just top 3 for formula length)
    top_cat_formula = '=IF({0}>={1},{2},IF({1}>={3},{4},{5}))'.format(
        cat_formulas_parts[0], cat_formulas_parts[1],
        '"{0}"'.format(INVENTORY_CATEGORIES[0]),
        cat_formulas_parts[2],
        '"{0}"'.format(INVENTORY_CATEGORIES[1]),
        '"{0}"'.format(INVENTORY_CATEGORIES[2])
    )
    ws7.cell(row=r, column=COL_START + 4, value=top_cat_formula)
    # Commission (5%)
    ws7.cell(row=r, column=COL_START + 6).value = "=C{0}*0.05".format(r)
    ws7.cell(row=r, column=COL_START + 6).number_format = FMT_CURRENCY
    # Target Achievement %
    ws7.cell(row=r, column=COL_START + 8).value = "=IFERROR(C{0}/I{0},0)".format(r)
    ws7.cell(row=r, column=COL_START + 8).number_format = FMT_PCT

    style_data_row(ws7, r, COL_START, COL_END_S7, idx)
    ws7.cell(row=r, column=COL_START).alignment = align_text()
    ws7.cell(row=r, column=COL_START + 1).alignment = align_number()
    ws7.cell(row=r, column=COL_START + 1).number_format = FMT_CURRENCY
    ws7.cell(row=r, column=COL_START + 2).alignment = align_number()
    ws7.cell(row=r, column=COL_START + 2).number_format = FMT_INT
    ws7.cell(row=r, column=COL_START + 3).alignment = align_number()
    ws7.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    ws7.cell(row=r, column=COL_START + 4).alignment = align_text()
    ws7.cell(row=r, column=COL_START + 5).alignment = align_text()
    ws7.cell(row=r, column=COL_START + 6).alignment = align_number()
    ws7.cell(row=r, column=COL_START + 6).number_format = FMT_CURRENCY
    ws7.cell(row=r, column=COL_START + 7).alignment = align_number()
    ws7.cell(row=r, column=COL_START + 7).number_format = FMT_CURRENCY
    ws7.cell(row=r, column=COL_START + 8).alignment = align_number()
    ws7.cell(row=r, column=COL_START + 8).number_format = FMT_PCT

TR_S7 = DATA_START + N_STAFF

# Totals row
ws7.cell(row=TR_S7, column=COL_START, value="TOTAL")
for cc in [1, 6]:
    col_l = get_column_letter(COL_START + cc)
    ws7.cell(row=TR_S7, column=COL_START + cc).value = "=SUM({0}{1}:{0}{2})".format(col_l, DATA_START, TR_S7 - 1)
    ws7.cell(row=TR_S7, column=COL_START + cc).number_format = FMT_CURRENCY
col_l2 = get_column_letter(COL_START + 2)
ws7.cell(row=TR_S7, column=COL_START + 2).value = "=SUM({0}{1}:{0}{2})".format(col_l2, DATA_START, TR_S7 - 1)
ws7.cell(row=TR_S7, column=COL_START + 2).number_format = FMT_INT
ws7.cell(row=TR_S7, column=COL_START + 3).value = "=IFERROR(C{0}/D{0},0)".format(TR_S7)
ws7.cell(row=TR_S7, column=COL_START + 3).number_format = FMT_CURRENCY
style_total_row(ws7, TR_S7, COL_START, COL_END_S7)
ws7.cell(row=TR_S7, column=COL_START).alignment = align_text()

# Conditional formatting: Target Achievement %
ach_col = get_column_letter(COL_START + 8)
ach_range = "{0}{1}:{0}{2}".format(ach_col, DATA_START, TR_S7 - 1)
ws7.conditional_formatting.add(ach_range, CellIsRule(operator="greaterThan", formula=["1"], fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT))
ws7.conditional_formatting.add(ach_range, CellIsRule(operator="between", formula=["0.8", "1"], fill=CF_WARNING_FILL, font=CF_WARNING_FONT))
ws7.conditional_formatting.add(ach_range, CellIsRule(operator="lessThan", formula=["0.8"], fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))

# Data bars on Total Sales
sales_col = get_column_letter(COL_START + 1)
sales_range = "{0}{1}:{0}{2}".format(sales_col, DATA_START, TR_S7 - 1)
ws7.conditional_formatting.add(sales_range, DataBarRule(start_type="min", end_type="max", color=PRIMARY_LIGHT))

ws7.print_area = "A1:{0}{1}".format(get_column_letter(COL_END_S7), TR_S7)
ws7.print_title_rows = "{0}:{0}".format(HEADER_ROW)


# ══════════════════════════════════════════════════════════════
# SHEET 8: Expenses
# ══════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Expenses")
ws8.sheet_properties.tabColor = PRIMARY

HEADERS_S8 = ["Date", "Expense Category", "Description", "Amount",
              "Payment Method", "Vendor", "Receipt #", "Notes"]
COL_END_S8 = COL_START + len(HEADERS_S8) - 1

setup_sheet(ws8, "Expenses — Expense Tracking", COL_END_S8)
apply_col_widths(ws8, [14, 18, 24, 14, 16, 18, 14, 20])

for i, h in enumerate(HEADERS_S8):
    ws8.cell(row=HEADER_ROW, column=COL_START + i, value=h)
style_header_row(ws8, HEADER_ROW, COL_START, COL_END_S8)

EXPENSE_CATEGORIES = ["Rent", "Salaries", "Utilities", "Supplies",
                       "Maintenance", "Marketing", "Transport",
                       "Insurance", "Miscellaneous"]

for idx in range(N_EXPENSES):
    r = DATA_START + idx
    style_data_row(ws8, r, COL_START, COL_END_S8, idx)
    ws8.cell(row=r, column=COL_START).alignment = align_date()
    ws8.cell(row=r, column=COL_START).number_format = FMT_DATE
    ws8.cell(row=r, column=COL_START + 1).alignment = align_text()
    ws8.cell(row=r, column=COL_START + 2).alignment = align_text()
    ws8.cell(row=r, column=COL_START + 3).alignment = align_number()
    ws8.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    ws8.cell(row=r, column=COL_START + 4).alignment = align_text()
    ws8.cell(row=r, column=COL_START + 5).alignment = align_text()
    ws8.cell(row=r, column=COL_START + 6).alignment = align_text()
    ws8.cell(row=r, column=COL_START + 7).alignment = align_text()

TR_S8 = DATA_START + N_EXPENSES

# Category dropdown
dv_expcat = DataValidation(type="list", formula1='"' + ','.join(EXPENSE_CATEGORIES) + '"', allow_blank=True)
dv_expcat.error = "Please select a valid expense category"
dv_expcat.errorTitle = "Invalid Category"
ws8.add_data_validation(dv_expcat)
expcat_col = get_column_letter(COL_START + 1)
dv_expcat.add("{0}{1}:{0}{2}".format(expcat_col, DATA_START, TR_S8 - 1))

# Totals row
ws8.cell(row=TR_S8, column=COL_START, value="TOTAL")
ws8.cell(row=TR_S8, column=COL_START + 3).value = "=SUM(E{0}:E{1})".format(DATA_START, TR_S8 - 1)
ws8.cell(row=TR_S8, column=COL_START + 3).number_format = FMT_CURRENCY
style_total_row(ws8, TR_S8, COL_START, COL_END_S8)
ws8.cell(row=TR_S8, column=COL_START).alignment = align_text()

# Category totals section
CAT_TOTAL_START = TR_S8 + 3
ws8.cell(row=CAT_TOTAL_START, column=COL_START, value="Totals by Category")
ws8.cell(row=CAT_TOTAL_START, column=COL_START).font = font_subheader()

ws8.cell(row=CAT_TOTAL_START + 1, column=COL_START, value="Category")
ws8.cell(row=CAT_TOTAL_START + 1, column=COL_START + 1, value="Total Amount")
style_header_row(ws8, CAT_TOTAL_START + 1, COL_START, COL_START + 1)

for idx, cat in enumerate(EXPENSE_CATEGORIES):
    r = CAT_TOTAL_START + 2 + idx
    ws8.cell(row=r, column=COL_START, value=cat)
    ws8.cell(row=r, column=COL_START + 1).value = "=SUMPRODUCT(({0}{1}:{0}{2}=\"{3}\")*(E{1}:E{2}))".format(expcat_col, DATA_START, TR_S8 - 1, cat)
    ws8.cell(row=r, column=COL_START + 1).number_format = FMT_CURRENCY
    style_data_row(ws8, r, COL_START, COL_START + 1, idx)
    ws8.cell(row=r, column=COL_START).alignment = align_text()
    ws8.cell(row=r, column=COL_START + 1).alignment = align_number()

CAT_TOTAL_END = CAT_TOTAL_START + 2 + len(EXPENSE_CATEGORIES)
ws8.cell(row=CAT_TOTAL_END, column=COL_START, value="GRAND TOTAL")
ws8.cell(row=CAT_TOTAL_END, column=COL_START + 1).value = "=SUM(C{0}:C{1})".format(CAT_TOTAL_START + 2, CAT_TOTAL_END - 1)
ws8.cell(row=CAT_TOTAL_END, column=COL_START + 1).number_format = FMT_CURRENCY
style_total_row(ws8, CAT_TOTAL_END, COL_START, COL_START + 1)

ws8.print_area = "A1:{0}{1}".format(get_column_letter(COL_END_S8), TR_S8)
ws8.print_title_rows = "{0}:{0}".format(HEADER_ROW)


# ══════════════════════════════════════════════════════════════
# SHEET 9: Profit Analysis
# ══════════════════════════════════════════════════════════════
ws9 = wb.create_sheet("Profit Analysis")
ws9.sheet_properties.tabColor = PRIMARY

setup_sheet(ws9, "Profit Analysis — P&L Statement", 8)

# Section: Revenue
SEC_ROW = 4
ws9.cell(row=SEC_ROW, column=COL_START, value="REVENUE")
ws9.cell(row=SEC_ROW, column=COL_START).font = Font(name=FONT_NAME, size=13, bold=HEADER_BOLD, color=PRIMARY)
ws9.merge_cells(start_row=SEC_ROW, start_column=COL_START, end_row=SEC_ROW, end_column=8)
ws9.row_dimensions[SEC_ROW].height = 28

REV_HDR = SEC_ROW + 1
rev_headers = ["Month", "Revenue", "COGS (est. 70%)", "Gross Profit",
               "Operating Expenses", "Net Profit", "Profit Margin %"]
for i, h in enumerate(rev_headers):
    ws9.cell(row=REV_HDR, column=COL_START + i, value=h)
style_header_row(ws9, REV_HDR, COL_START, COL_START + len(rev_headers) - 1)

for idx in range(N_MONTHS):
    r = REV_HDR + 1 + idx
    ws9.cell(row=r, column=COL_START, value=MONTH_NAMES[idx])
    # Revenue = from Monthly Summary
    ws9.cell(row=r, column=COL_START + 1).value = "='Monthly Summary'!C{0}".format(DATA_START + idx)
    ws9.cell(row=r, column=COL_START + 1).number_format = FMT_CURRENCY
    # COGS (70%)
    ws9.cell(row=r, column=COL_START + 2).value = "=C{0}*0.7".format(r)
    ws9.cell(row=r, column=COL_START + 2).number_format = FMT_CURRENCY
    # Gross Profit
    ws9.cell(row=r, column=COL_START + 3).value = "=C{0}-D{0}".format(r)
    ws9.cell(row=r, column=COL_START + 3).number_format = FMT_CURRENCY
    # Operating Expenses from Monthly Summary
    ws9.cell(row=r, column=COL_START + 4).value = "='Monthly Summary'!H{0}".format(DATA_START + idx)
    ws9.cell(row=r, column=COL_START + 4).number_format = FMT_CURRENCY
    # Net Profit
    ws9.cell(row=r, column=COL_START + 5).value = "=E{0}-F{0}".format(r)
    ws9.cell(row=r, column=COL_START + 5).number_format = FMT_CURRENCY
    # Profit Margin %
    ws9.cell(row=r, column=COL_START + 6).value = "=IFERROR(G{0}/C{0},0)".format(r)
    ws9.cell(row=r, column=COL_START + 6).number_format = FMT_PCT

    style_data_row(ws9, r, COL_START, COL_START + len(rev_headers) - 1, idx)
    ws9.cell(row=r, column=COL_START).alignment = align_text()
    for cc in range(1, 6):
        ws9.cell(row=r, column=COL_START + cc).alignment = align_number()
    ws9.cell(row=r, column=COL_START + 6).alignment = align_number()

PA_TOTAL = REV_HDR + 1 + N_MONTHS
ws9.cell(row=PA_TOTAL, column=COL_START, value="TOTAL")
for cc in [1, 2, 3, 4, 5]:
    col_l = get_column_letter(COL_START + cc)
    ws9.cell(row=PA_TOTAL, column=COL_START + cc).value = "=SUM({0}{1}:{0}{2})".format(col_l, REV_HDR + 1, PA_TOTAL - 1)
    ws9.cell(row=PA_TOTAL, column=COL_START + cc).number_format = FMT_CURRENCY
ws9.cell(row=PA_TOTAL, column=COL_START + 6).value = "=IFERROR(G{0}/C{0},0)".format(PA_TOTAL)
ws9.cell(row=PA_TOTAL, column=COL_START + 6).number_format = FMT_PCT
style_total_row(ws9, PA_TOTAL, COL_START, COL_START + len(rev_headers) - 1)
ws9.cell(row=PA_TOTAL, column=COL_START).alignment = align_text()

# Conditional formatting: Net Profit positive/negative
np_col = get_column_letter(COL_START + 5)
np_range = "{0}{1}:{0}{2}".format(np_col, REV_HDR + 1, PA_TOTAL)
ws9.conditional_formatting.add(np_range, CellIsRule(operator="greaterThan", formula=["0"], fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT))
ws9.conditional_formatting.add(np_range, CellIsRule(operator="lessThan", formula=["0"], fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))

# Profit Margin %
pm_col = get_column_letter(COL_START + 6)
pm_range = "{0}{1}:{0}{2}".format(pm_col, REV_HDR + 1, PA_TOTAL)
ws9.conditional_formatting.add(pm_range, CellIsRule(operator="greaterThan", formula=["0"], fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT))
ws9.conditional_formatting.add(pm_range, CellIsRule(operator="lessThan", formula=["0"], fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))

apply_col_widths(ws9, [16, 16, 18, 16, 18, 16, 16])

ws9.print_area = "A1:{0}{1}".format(get_column_letter(COL_START + len(rev_headers) - 1), PA_TOTAL)
ws9.print_title_rows = "{0}:{0}".format(REV_HDR)


# ══════════════════════════════════════════════════════════════
# SHEET 10: Dashboard
# ══════════════════════════════════════════════════════════════
ws10 = wb.create_sheet("Dashboard")
ws10.sheet_properties.tabColor = PRIMARY

setup_sheet(ws10, None, 14)

# Title
ws10.merge_cells("B2:N2")
cell = ws10.cell(row=2, column=2, value="Premium POS Dashboard")
cell.font = Font(name=FONT_NAME, size=22, bold=HEADER_BOLD, color=PRIMARY)
cell.alignment = Alignment(horizontal="left", vertical="center")
ws10.row_dimensions[2].height = 40

# Subtitle
ws10.merge_cells("B3:N3")
cell = ws10.cell(row=3, column=2, value="Real-time KPI overview with cross-sheet analytics")
cell.font = font_caption()
cell.alignment = Alignment(horizontal="left", vertical="center")

# ── KPI Cards ─────────────────────────────────────────────────
KPI_ROW = 5
ws10.row_dimensions[KPI_ROW].height = 18
ws10.row_dimensions[KPI_ROW + 1].height = 42
ws10.row_dimensions[KPI_ROW + 2].height = 6

kpi_data = [
    ("Total Revenue", "='Monthly Summary'!C{0}".format(TR_S4), FMT_CURRENCY),
    ("Total Expenses", "='Monthly Summary'!H{0}".format(TR_S4), FMT_CURRENCY),
    ("Net Profit", "='Profit Analysis'!G{0}".format(PA_TOTAL), FMT_CURRENCY),
    ("Profit Margin %", "='Profit Analysis'!H{0}".format(PA_TOTAL), FMT_PCT),
    ("# Transactions", "='Monthly Summary'!D{0}".format(TR_S4), FMT_INT),
    ("Avg Transaction", "='Monthly Summary'!E{0}".format(TR_S4), FMT_CURRENCY),
]

kpi_start_col = 2
kpi_width = 2

for i, (label, formula, fmt) in enumerate(kpi_data):
    col = kpi_start_col + i * kpi_width
    # KPI Card background
    for row_offset in range(3):
        for col_offset in range(kpi_width):
            c = col + col_offset
            ws10.cell(row=KPI_ROW + row_offset, column=c).fill = PatternFill("solid", fgColor=NEUTRAL_100)

    # Label
    ws10.merge_cells(start_row=KPI_ROW, start_column=col, end_row=KPI_ROW, end_column=col + kpi_width - 1)
    label_cell = ws10.cell(row=KPI_ROW, column=col, value=label)
    label_cell.font = font_kpi_label()
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.fill = PatternFill("solid", fgColor=NEUTRAL_100)

    # Value
    ws10.merge_cells(start_row=KPI_ROW + 1, start_column=col, end_row=KPI_ROW + 1, end_column=col + kpi_width - 1)
    value_cell = ws10.cell(row=KPI_ROW + 1, column=col, value=formula)
    value_cell.font = font_kpi()
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.number_format = fmt
    value_cell.fill = PatternFill("solid", fgColor=NEUTRAL_100)

# KPI column widths
for i in range(12):
    ws10.column_dimensions[get_column_letter(kpi_start_col + i)].width = 10

# ── Charts ────────────────────────────────────────────────────
CHART_START_ROW = KPI_ROW + 4

# Chart a: Monthly Revenue bar chart
chart_a = create_bar_chart(chart_type="col", grouping="clustered", gap_width=80, width=18, height=11)
data_a = Reference(ws4, min_col=COL_START + 1, min_row=HEADER_ROW, max_row=DATA_START + N_MONTHS - 1)
cats_a = Reference(ws4, min_col=COL_START, min_row=DATA_START, max_row=DATA_START + N_MONTHS - 1)
chart_a.add_data(data_a, titles_from_data=True)
chart_a.set_categories(cats_a)
chart_a.shape = 4
setup_chart_titles(chart_a, title="Monthly Revenue", y_title="Revenue")
apply_chart_colors(chart_a, [CHART_COLORS[0]])
ws10.add_chart(chart_a, "B{0}".format(CHART_START_ROW))

# Chart b: Expense breakdown pie chart
chart_b = create_pie_chart(width=14, height=11)
data_b = Reference(ws8, min_col=COL_START + 1, min_row=CAT_TOTAL_START + 1, max_row=CAT_TOTAL_START + 1 + len(EXPENSE_CATEGORIES))
cats_b = Reference(ws8, min_col=COL_START, min_row=CAT_TOTAL_START + 2, max_row=CAT_TOTAL_START + 1 + len(EXPENSE_CATEGORIES))
chart_b.add_data(data_b, titles_from_data=True)
chart_b.set_categories(cats_b)
setup_chart_titles(chart_b, title="Expense Breakdown")
apply_pie_colors(chart_b, len(EXPENSE_CATEGORIES))
ws10.add_chart(chart_b, "J{0}".format(CHART_START_ROW))

# Chart c: Sales trend line chart
CHART_START_ROW2 = CHART_START_ROW + 13
chart_c = create_line_chart(width=18, height=11)
data_c = Reference(ws4, min_col=COL_START + 1, min_row=HEADER_ROW, max_row=DATA_START + N_MONTHS - 1)
cats_c = Reference(ws4, min_col=COL_START, min_row=DATA_START, max_row=DATA_START + N_MONTHS - 1)
chart_c.add_data(data_c, titles_from_data=True)
chart_c.set_categories(cats_c)
setup_chart_titles(chart_c, title="Sales Trend", y_title="Revenue")
apply_chart_colors(chart_c, [CHART_COLORS[0]])
ws10.add_chart(chart_c, "B{0}".format(CHART_START_ROW2))

# Chart d: Top 5 Staff bar chart
chart_d = create_bar_chart(chart_type="bar", grouping="clustered", gap_width=80, width=14, height=11)
data_d = Reference(ws7, min_col=COL_START + 1, min_row=HEADER_ROW, max_row=DATA_START + min(4, N_STAFF - 1))
cats_d = Reference(ws7, min_col=COL_START, min_row=DATA_START, max_row=DATA_START + min(4, N_STAFF - 1))
chart_d.add_data(data_d, titles_from_data=True)
chart_d.set_categories(cats_d)
setup_chart_titles(chart_d, title="Top 5 Staff Sales", x_title="Sales")
apply_chart_colors(chart_d, [CHART_COLORS[0]])
ws10.add_chart(chart_d, "J{0}".format(CHART_START_ROW2))


# ══════════════════════════════════════════════════════════════
# Final: Save workbook
# ══════════════════════════════════════════════════════════════
OUTPUT = "/home/z/my-project/download/POS_Premium_Empty.xlsx"
wb.save(OUTPUT)
print("Saved: {0}".format(OUTPUT))
print("Sheets: {0} — {1}".format(len(wb.sheetnames), ', '.join(wb.sheetnames)))
