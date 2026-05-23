"""
Generate 3 Professional Excel Data Analysis Reports
=====================================================
1. Al Matar Restaurant (warm palette)
2. Al Madina Grocery (professional palette)
3. Luxe Beauty Salon (elegant palette)

Author: Ahmed Ali — Data Analysis Services
Website: ahmed-ali-ops.vercel.app
"""

import sys, os
sys.path.insert(0, '/home/z/my-project/skills/xlsx')
sys.path.insert(0, '/home/z/my-project/skills/xlsx/templates')
from base import *

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from copy import copy
import datetime


# ============================================================
# UTILITY FUNCTIONS
# ============================================================

FOOTER_TEXT = "Prepared by Ahmed Ali | Data Analysis Services | ahmed-ali-ops.vercel.app"

def add_footer(ws, row, col_start, col_end):
    """Add footer row to a sheet."""
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=FOOTER_TEXT)
    cell.font = font_caption()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20


def write_kpi_row(ws, row, col_start, kpis):
    """
    Write a row of KPIs: each KPI is (label, value, number_format).
    KPIs are spaced across columns with label above value.
    """
    num_kpis = len(kpis)
    # Each KPI takes 2 columns: label, value — with 1 spacer col between
    for i, (label, value, fmt) in enumerate(kpis):
        label_col = col_start + i * 2
        value_col = label_col + 1

        # Label
        lc = ws.cell(row=row, column=label_col, value=label)
        lc.font = font_kpi_label()
        lc.alignment = Alignment(horizontal="center", vertical="bottom")

        # Value
        vc = ws.cell(row=row + 1, column=value_col, value=value)
        vc.font = font_kpi()
        vc.alignment = Alignment(horizontal="center", vertical="top")
        if fmt:
            vc.number_format = fmt

    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 36
    return row + 2


def write_table(ws, start_row, col_start, headers, data, total_row=None,
                num_fmts=None, col_widths=None):
    """
    Write a styled table with headers, data rows, optional total row.
    Returns the next available row.
    """
    num_cols = len(headers)
    col_end = col_start + num_cols - 1

    # Header row
    for j, h in enumerate(headers):
        ws.cell(row=start_row, column=col_start + j, value=h)
    style_header_row(ws, start_row, col_start, col_end)

    # Data rows
    for i, row_data in enumerate(data):
        r = start_row + 1 + i
        for j, val in enumerate(row_data):
            cell = ws.cell(row=r, column=col_start + j, value=val)
            if num_fmts and j < len(num_fmts) and num_fmts[j]:
                cell.number_format = num_fmts[j]
            # Right-align numbers
            if isinstance(val, (int, float)):
                cell.alignment = align_number()
            else:
                cell.alignment = align_text()
        style_data_row(ws, r, col_start, col_end, i)

    next_row = start_row + 1 + len(data)

    # Total row
    if total_row is not None:
        for j, val in enumerate(total_row):
            cell = ws.cell(row=next_row, column=col_start + j, value=val)
            if num_fmts and j < len(num_fmts) and num_fmts[j]:
                cell.number_format = num_fmts[j]
            if isinstance(val, (int, float)):
                cell.alignment = align_number()
            else:
                cell.alignment = align_text()
        style_total_row(ws, next_row, col_start, col_end)
        next_row += 1

    # Set column widths
    if col_widths:
        for j, w in enumerate(col_widths):
            if w:
                ws.column_dimensions[get_column_letter(col_start + j)].width = w

    return next_row


def add_conditional_formatting(ws, cell_range, conditions):
    """
    Add conditional formatting rules.
    conditions: list of (operator, value, fill, font)
    """
    for op, val, fill, font in conditions:
        rule = CellIsRule(operator=op, formula=[str(val)], fill=fill, font=font)
        ws.conditional_formatting.add(cell_range, rule)


def add_bar_chart(ws, title, categories_ref, data_refs, data_labels,
                  chart_row, chart_col, width=18, height=10,
                  y_title=None, x_title=None, chart_type="col",
                  grouping="clustered", gap_width=80, overlap=100):
    """Add a bar chart to the sheet."""
    chart = create_bar_chart(chart_type=chart_type, grouping=grouping,
                             gap_width=gap_width, overlap=overlap,
                             width=width, height=height)
    for ref, label in zip(data_refs, data_labels):
        chart.add_data(ref, titles_from_data=False)
    chart.set_categories(categories_ref)

    for i, label in enumerate(data_labels):
        if i < len(chart.series):
            chart.series[i].tx = SeriesLabel(v=label)

    setup_chart_titles(chart, title=title, y_title=y_title, x_title=x_title)
    apply_chart_colors(chart)
    ws.add_chart(chart, f"{get_column_letter(chart_col)}{chart_row}")
    return chart


def add_line_chart(ws, title, categories_ref, data_refs, data_labels,
                   chart_row, chart_col, width=18, height=10,
                   y_title=None, x_title=None):
    """Add a line chart to the sheet."""
    chart = create_line_chart(width=width, height=height)
    for ref, label in zip(data_refs, data_labels):
        chart.add_data(ref, titles_from_data=False)
    chart.set_categories(categories_ref)

    for i, label in enumerate(data_labels):
        if i < len(chart.series):
            chart.series[i].tx = SeriesLabel(v=label)

    setup_chart_titles(chart, title=title, y_title=y_title, x_title=x_title)
    apply_chart_colors(chart)
    ws.add_chart(chart, f"{get_column_letter(chart_col)}{chart_row}")
    return chart


def add_pie_chart(ws, title, categories_ref, data_ref, count,
                  chart_row, chart_col, width=14, height=10):
    """Add a pie chart to the sheet."""
    chart = create_pie_chart(width=width, height=height)
    chart.add_data(data_ref, titles_from_data=False)
    chart.set_categories(categories_ref)
    setup_chart_titles(chart, title=title)
    apply_pie_colors(chart, count)
    ws.add_chart(chart, f"{get_column_letter(chart_col)}{chart_row}")
    return chart


# ============================================================
# FILE 1: AL MATAR RESTAURANT
# ============================================================

def generate_restaurant_report():
    use_palette_explicit("warm")
    wb = Workbook()
    wb.properties.creator = "Z.ai"

    # ---- Sheet 1: Executive Summary ----
    ws = wb.active
    ws.title = "Executive Summary"
    setup_sheet(ws, "Al Matar Restaurant — Executive Dashboard", 13)

    # KPIs
    kpis = [
        ("Total Revenue", 1847500, "#,##0"),
        ("Total Expenses", 1218350, "#,##0"),
        ("Net Profit", 629150, "#,##0"),
        ("Profit Margin", 0.3406, "0.0%"),
        ("Total Orders", 34280, "#,##0"),
        ("Avg Order Value", 53.88, "#,##0.00"),
    ]
    row = write_kpi_row(ws, 4, 2, kpis)

    # Spacer
    row += 1

    # Monthly trend table
    headers = ["Month", "Revenue (AED)", "Expenses (AED)", "Net Profit (AED)", "Profit Margin", "Orders", "Avg Order (AED)"]
    data = [
        ["Aug 2025", 278500, 189200, 89300, 0.3207, 5120, 54.39],
        ["Sep 2025", 296800, 198500, 98300, 0.3312, 5480, 54.16],
        ["Oct 2025", 312400, 205600, 106800, 0.3419, 5790, 53.95],
        ["Nov 2025", 325600, 213400, 112200, 0.3446, 5960, 54.63],
        ["Dec 2025", 308900, 202650, 106250, 0.3439, 5720, 54.00],
        ["Jan 2026", 325300, 209000, 116300, 0.3575, 6210, 52.38],
    ]
    num_fmts = [None, "#,##0", "#,##0", "#,##0", "0.0%", "#,##0", "#,##0.00"]
    total_row = ["TOTAL", 1847500, 1218350, 629150, 0.3406, 34280, 53.88]
    row = write_table(ws, row, 2, headers, data, total_row, num_fmts)

    # Chart data references
    cat_ref = Reference(ws, min_col=2, min_row=8, max_row=13)  # Months

    # Bar chart: Revenue vs Expenses
    rev_ref = Reference(ws, min_col=3, min_row=7, max_row=13)
    exp_ref = Reference(ws, min_col=4, min_row=7, max_row=13)
    add_bar_chart(ws, "Revenue vs Expenses", cat_ref, [rev_ref, exp_ref],
                  ["Revenue", "Expenses"], row, 2, width=18, height=10,
                  y_title="AED")

    # Line chart: Profit Margin trend
    margin_ref = Reference(ws, min_col=5, min_row=7, max_row=13)
    add_line_chart(ws, "Profit Margin Trend", cat_ref, [margin_ref],
                   ["Profit Margin"], row, 9, width=12, height=10,
                   y_title="Margin")

    row += 1
    # Line chart: Orders trend
    orders_ref = Reference(ws, min_col=7, min_row=7, max_row=13)
    add_line_chart(ws, "Monthly Orders Trend", cat_ref, [orders_ref],
                   ["Orders"], row + 16, 2, width=12, height=10,
                   y_title="Orders")

    # Footer
    add_footer(ws, ws.max_row + 18, 2, 13)
    auto_fit_columns(ws)

    # ---- Sheet 2: Sales Analysis ----
    ws2 = wb.create_sheet("Sales Analysis")
    setup_sheet(ws2, "Sales Performance — January 2026", 9)

    row = 4

    # Category table
    headers = ["Category", "Orders", "Revenue (AED)", "% of Revenue", "Avg Order (AED)", "YoY Growth", "Trend"]
    cat_data = [
        ["Grills & BBQ", 1420, 128500, 0.3949, 90.49, 0.12, "▲ Up"],
        ["Rice & Biryani", 1380, 95600, 0.2938, 69.28, 0.08, "▲ Up"],
        ["Shawarma", 1650, 72600, 0.2232, 44.00, 0.15, "▲ Up"],
        ["Beverages", 5800, 40600, 0.1248, 7.00, 0.05, "→ Stable"],
        ["Desserts", 920, 28400, 0.0873, 30.87, -0.03, "▼ Down"],
        ["Appetizers", 1100, 24200, 0.0744, 22.00, 0.06, "→ Stable"],
        ["Seafood", 420, 32500, 0.0999, 77.38, 0.18, "▲ Up"],
        ["Salads", 680, 18500, 0.0569, 27.21, 0.02, "→ Stable"],
    ]
    num_fmts = [None, "#,##0", "#,##0", "0.0%", "#,##0.00", "0.0%", None]
    total_row = ["TOTAL", 13370, 325300, 1.0, 36.23, 0.09, "—"]

    # Write header + data first for chart references
    row = write_table(ws2, row, 2, headers, cat_data, total_row, num_fmts)

    # Charts for category table
    cat_names = Reference(ws2, min_col=2, min_row=5, max_row=12)
    cat_revenue = Reference(ws2, min_col=4, min_row=4, max_row=12)
    cat_orders = Reference(ws2, min_col=3, min_row=4, max_row=12)

    add_pie_chart(ws2, "Revenue by Category", cat_names, cat_revenue, 8, row + 1, 2, width=14, height=10)
    add_bar_chart(ws2, "Orders by Category", cat_names, [cat_orders],
                  ["Orders"], row + 1, 8, width=14, height=10, y_title="Orders")

    row += 18  # space for charts

    # Daily sales table
    row += 1
    headers2 = ["Date", "Revenue (AED)", "Orders", "Top Category"]
    top_cats = ["Grills & BBQ", "Shawarma", "Rice & Biryani", "Grills & BBQ", "Seafood",
                "Shawarma", "Rice & Biryani", "Grills & BBQ", "Beverages", "Shawarma",
                "Grills & BBQ", "Rice & Biryani", "Shawarma", "Seafood", "Grills & BBQ",
                "Rice & Biryani", "Shawarma", "Grills & BBQ", "Beverages", "Rice & Biryani",
                "Grills & BBQ", "Shawarma", "Seafood", "Grills & BBQ", "Rice & Biryani",
                "Shawarma", "Grills & BBQ", "Beverages", "Grills & BBQ", "Rice & Biryani",
                "Shawarma"]
    daily_rev = [9800, 10200, 11500, 10800, 11200, 10600, 10100, 9900, 10400, 11000,
                 10800, 11200, 10600, 11500, 10900, 11100, 10400, 9800, 10700, 10300,
                 11200, 10900, 11600, 10100, 10800, 11300, 10500, 10200, 11000, 11400, 10900]
    daily_orders = [188, 196, 220, 208, 215, 204, 194, 190, 200, 212,
                    208, 215, 204, 221, 210, 214, 200, 188, 206, 198,
                    215, 210, 223, 194, 208, 218, 202, 196, 212, 219, 210]

    daily_data = []
    for i in range(31):
        date_str = f"2026-01-{i+1:02d}"
        daily_data.append([date_str, daily_rev[i], daily_orders[i], top_cats[i]])

    num_fmts2 = ["YYYY-MM-DD", "#,##0", "#,##0", None]
    total_rev = sum(daily_rev)
    total_ord = sum(daily_orders)
    total_row2 = ["TOTAL", total_rev, total_ord, "—"]

    table_start = row
    row = write_table(ws2, row, 2, headers2, daily_data, total_row2, num_fmts2)

    # Line chart: Daily Revenue
    daily_dates = Reference(ws2, min_col=2, min_row=table_start + 1, max_row=table_start + 31)
    daily_rev_ref = Reference(ws2, min_col=3, min_row=table_start, max_row=table_start + 31)
    add_line_chart(ws2, "Daily Revenue — January 2026", daily_dates, [daily_rev_ref],
                   ["Revenue"], row + 1, 2, width=22, height=10, y_title="AED")

    add_footer(ws2, ws.max_row + 20, 2, 9)
    auto_fit_columns(ws2)

    # ---- Sheet 3: Expense Analysis ----
    ws3 = wb.create_sheet("Expense Analysis")
    setup_sheet(ws3, "Expense Breakdown — January 2026", 9)

    row = 4
    headers = ["Category", "Budget (AED)", "Actual (AED)", "Variance (AED)", "Variance %", "Status", "Notes"]
    exp_data = [
        ["Rent", 45000, 45000, 0, 0.0, "On Budget", "Fixed lease"],
        ["Staff Salaries", 68000, 69500, -1500, -0.0221, "Over Budget", "Overtime pay"],
        ["Food Supplies", 42000, 39800, 2200, 0.0524, "Under Budget", "Bulk discount"],
        ["Utilities", 12000, 13200, -1200, -0.1, "Over Budget", "AC usage high"],
        ["Marketing", 15000, 14800, 200, 0.0133, "On Budget", "Social media ads"],
        ["Maintenance", 8000, 9200, -1200, -0.15, "Over Budget", "Kitchen repair"],
        ["Delivery Commissions", 10000, 11500, -1500, -0.15, "Over Budget", "Talabat surge"],
        ["Insurance", 5000, 5000, 0, 0.0, "On Budget", "Annual policy"],
        ["POS Fees", 3000, 3100, -100, -0.0333, "On Budget", "Card transactions"],
        ["Cleaning", 4000, 3800, 200, 0.05, "Under Budget", "Negotiated rate"],
        ["Miscellaneous", 6000, 5400, 600, 0.1, "Under Budget", "Lower than expected"],
        ["Licenses", 2000, 2000, 0, 0.0, "On Budget", "Annual renewal"],
    ]
    num_fmts = [None, "#,##0", "#,##0", "#,##0", "0.0%", None, None]
    budget_total = sum(r[1] for r in exp_data)
    actual_total = sum(r[2] for r in exp_data)
    var_total = actual_total - budget_total
    var_pct_total = var_total / budget_total if budget_total else 0
    total_row = ["TOTAL", budget_total, actual_total, var_total, var_pct_total, "", ""]

    table_start = row
    row = write_table(ws3, row, 2, headers, exp_data, total_row, num_fmts)

    # Conditional formatting on Variance % column (col 6 = F)
    cf_range = f"F{table_start+1}:F{table_start+len(exp_data)}"
    ws3.conditional_formatting.add(cf_range,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))
    ws3.conditional_formatting.add(cf_range,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT))

    # Conditional formatting on Status column (col 7 = G)
    status_range = f"G{table_start+1}:G{table_start+len(exp_data)}"
    ws3.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"Over Budget"'],
                   fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))
    ws3.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"Under Budget"'],
                   fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT))

    # Charts
    cat_names = Reference(ws3, min_col=2, min_row=table_start+1, max_row=table_start+len(exp_data))
    budget_ref = Reference(ws3, min_col=3, min_row=table_start, max_row=table_start+len(exp_data))
    actual_ref = Reference(ws3, min_col=4, min_row=table_start, max_row=table_start+len(exp_data))
    add_bar_chart(ws3, "Budget vs Actual Expenses", cat_names, [budget_ref, actual_ref],
                  ["Budget", "Actual"], row + 1, 2, width=20, height=10, y_title="AED")

    # Monthly expense comparison
    row += 18
    row += 1
    headers3 = ["Category", "Oct 2025", "Nov 2025", "Dec 2025", "Jan 2026"]
    monthly_exp = [
        ["Rent", 45000, 45000, 45000, 45000],
        ["Staff Salaries", 67200, 68000, 68500, 69500],
        ["Food Supplies", 41500, 40200, 40800, 39800],
        ["Utilities", 11500, 11800, 12400, 13200],
        ["Marketing", 14000, 14500, 15200, 14800],
        ["Maintenance", 7500, 8200, 7800, 9200],
        ["Delivery Commissions", 9500, 10200, 10800, 11500],
        ["Insurance", 5000, 5000, 5000, 5000],
        ["POS Fees", 2800, 2900, 2950, 3100],
        ["Cleaning", 4200, 4000, 3900, 3800],
        ["Miscellaneous", 5800, 5500, 5700, 5400],
        ["Licenses", 2000, 2000, 2000, 2000],
    ]
    num_fmts3 = [None, "#,##0", "#,##0", "#,##0", "#,##0"]
    row = write_table(ws3, row, 2, headers3, monthly_exp, num_fmts=num_fmts3)

    # Line chart: Monthly expense trend
    months_ref = Reference(ws3, min_col=3, min_row=row - len(monthly_exp) - 1, max_col=6)
    # For simplicity, chart total expenses per month
    total_exp_row = row
    ws3.cell(row=total_exp_row, column=2, value="TOTAL")
    for mi in range(3, 7):
        idx = mi - 2  # map col 3→1, 4→2, 5→3, 6→4
        total_val = sum(r[idx] for r in monthly_exp)
        ws3.cell(row=total_exp_row, column=mi, value=total_val)
        ws3.cell(row=total_exp_row, column=mi).number_format = "#,##0"
    style_total_row(ws3, total_exp_row, 2, 6)

    month_labels = Reference(ws3, min_col=3, max_col=6, min_row=total_exp_row - len(monthly_exp) - 1)
    total_vals = Reference(ws3, min_col=3, max_col=6, min_row=total_exp_row, max_row=total_exp_row)
    # Use a simpler approach: reference individual months
    # Let's write a small helper data area for the chart
    chart_data_row = total_exp_row + 2
    chart_labels = ["Oct 2025", "Nov 2025", "Dec 2025", "Jan 2026"]
    totals_by_month = [sum(r[i] for r in monthly_exp) for i in range(1, 5)]
    for ci, (lbl, val) in enumerate(zip(chart_labels, totals_by_month)):
        ws3.cell(row=chart_data_row, column=2 + ci, value=lbl)
        ws3.cell(row=chart_data_row + 1, column=2 + ci, value=val)
        ws3.cell(row=chart_data_row + 1, column=2 + ci).number_format = "#,##0"

    m_cats = Reference(ws3, min_col=2, max_col=5, min_row=chart_data_row)
    m_data = Reference(ws3, min_col=2, max_col=5, min_row=chart_data_row + 1)
    add_line_chart(ws3, "Monthly Total Expense Trend", m_cats, [m_data],
                   ["Total Expenses"], total_exp_row + 4, 2, width=14, height=10, y_title="AED")

    add_footer(ws3, ws3.max_row + 5, 2, 9)
    auto_fit_columns(ws3)

    # ---- Sheet 4: Staff & Operations ----
    ws4 = wb.create_sheet("Staff & Operations")
    setup_sheet(ws4, "Staff Performance & Operations — January 2026", 9)

    row = 4
    headers = ["Name", "Role", "Hours Worked", "Clients Served", "Revenue Generated (AED)", "Efficiency Score", "Rating"]
    staff_data = [
        ["Mohammed Al Suwaidi", "Head Chef", 220, 890, 142500, 0.95, 4.8],
        ["Fatima Al Zaabi", "Floor Manager", 200, 1200, 98000, 0.88, 4.6],
        ["Ahmed Hassan", "Grill Master", 210, 720, 86500, 0.92, 4.5],
        ["Sara Ibrahim", "Cashier", 190, 2400, 52300, 0.85, 4.3],
        ["Khalid Omar", "Waiter", 180, 1650, 48200, 0.80, 4.1],
        ["Nour Ali", "Pastry Chef", 200, 580, 38900, 0.90, 4.7],
        ["Youssef Majeed", "Delivery Lead", 195, 980, 42100, 0.82, 4.0],
        ["Mariam Saeed", "Hostess", 175, 1800, 35400, 0.78, 4.2],
    ]
    num_fmts = [None, None, "#,##0", "#,##0", "#,##0", "0%", "0.0"]
    total_row = ["TOTAL", "", 1570, 10220, 543900, 0.86, 4.4]
    row = write_table(ws4, row, 2, headers, staff_data, total_row, num_fmts)

    # Bar chart: Staff Revenue
    staff_names = Reference(ws4, min_col=2, min_row=5, max_row=12)
    staff_rev = Reference(ws4, min_col=6, min_row=4, max_row=12)
    add_bar_chart(ws4, "Revenue by Staff Member", staff_names, [staff_rev],
                  ["Revenue (AED)"], row + 1, 2, width=18, height=10, y_title="AED")

    # Peak hours table
    row += 18
    row += 1
    headers2 = ["Time Slot", "Avg Orders", "Avg Revenue (AED)"]
    peak_data = [
        ["10:00 AM – 12:00 PM", 45, 2430],
        ["12:00 PM – 2:00 PM", 128, 6920],
        ["2:00 PM – 5:00 PM", 62, 3350],
        ["5:00 PM – 8:00 PM", 145, 7830],
        ["8:00 PM – 10:00 PM", 98, 5290],
        ["10:00 PM – 12:00 AM", 42, 2270],
    ]
    num_fmts2 = [None, "#,##0", "#,##0"]
    row = write_table(ws4, row, 2, headers2, peak_data, num_fmts=num_fmts2)

    # Line chart: Peak Hours Revenue
    slot_ref = Reference(ws4, min_col=2, min_row=row - len(peak_data), max_row=row - 1)
    peak_rev_ref = Reference(ws4, min_col=4, min_row=row - len(peak_data) - 1, max_row=row - 1)
    add_line_chart(ws4, "Peak Hours Revenue", slot_ref, [peak_rev_ref],
                   ["Avg Revenue"], row + 1, 2, width=16, height=10, y_title="AED")

    add_footer(ws4, ws4.max_row + 14, 2, 9)
    auto_fit_columns(ws4)

    # ---- Sheet 5: Recommendations ----
    ws5 = wb.create_sheet("Recommendations")
    setup_sheet(ws5, "Key Insights & Strategic Recommendations", 8)

    row = 4
    headers = ["#", "Category", "Insight", "Impact", "Action Item", "Priority"]
    rec_data = [
        [1, "Revenue", "Seafood category shows highest YoY growth (18%) with strong margins", "High",
         "Expand seafood menu with 3-4 new premium items", "High"],
        [2, "Operations", "Evening peak (5-8 PM) generates 29% of daily revenue", "High",
         "Add 2 staff members during evening shift", "High"],
        [3, "Cost Control", "Delivery commissions exceed budget by 15% due to Talabat surge pricing", "High",
         "Negotiate flat-rate commission or add direct ordering channel", "High"],
        [4, "Marketing", "Beverage category has low avg order value (AED 7) but high volume", "Medium",
         "Introduce combo deals pairing beverages with main dishes", "Medium"],
        [5, "Staff", "Mohammed Al Suwaidi generates 26% of total revenue as Head Chef", "Medium",
         "Develop succession plan and document signature recipes", "Medium"],
        [6, "Customer", "Dessert category declined 3% YoY despite overall growth", "Medium",
         "Introduce seasonal dessert specials and loyalty discounts", "Medium"],
        [7, "Efficiency", "Kitchen maintenance costs spiked 15% due to aging equipment", "Low",
         "Schedule preventive maintenance and budget for equipment upgrade", "Low"],
        [8, "Growth", "Profit margin improved to 35.7% in Jan — highest in 6 months", "High",
         "Reinvest margin gains into marketing for customer acquisition", "Medium"],
    ]
    num_fmts = [None, None, None, None, None, None]
    row = write_table(ws5, row, 2, headers, rec_data, num_fmts=num_fmts)

    # Conditional formatting on Priority column (col 7 = G)
    priority_range = f"G5:G{4 + len(rec_data)}"
    ws5.conditional_formatting.add(priority_range,
        CellIsRule(operator='equal', formula=['"High"'],
                   fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))
    ws5.conditional_formatting.add(priority_range,
        CellIsRule(operator='equal', formula=['"Medium"'],
                   fill=CF_WARNING_FILL, font=CF_WARNING_FONT))
    ws5.conditional_formatting.add(priority_range,
        CellIsRule(operator='equal', formula=['"Low"'],
                   fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT))

    # Conditional formatting on Impact column (col 5 = E)
    impact_range = f"E5:E{4 + len(rec_data)}"
    ws5.conditional_formatting.add(impact_range,
        CellIsRule(operator='equal', formula=['"High"'],
                   fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))
    ws5.conditional_formatting.add(impact_range,
        CellIsRule(operator='equal', formula=['"Medium"'],
                   fill=CF_WARNING_FILL, font=CF_WARNING_FONT))

    add_footer(ws5, row + 2, 2, 8)
    auto_fit_columns(ws5)

    filepath = "/home/z/my-project/download/sample_report_restaurant.xlsx"
    wb.save(filepath)
    print(f"✅ Saved: {filepath}")
    return filepath


# ============================================================
# FILE 2: AL MADINA GROCERY
# ============================================================

def generate_grocery_report():
    use_palette_explicit("professional")
    wb = Workbook()
    wb.properties.creator = "Z.ai"

    # ---- Sheet 1: Executive Summary ----
    ws = wb.active
    ws.title = "Executive Summary"
    setup_sheet(ws, "Al Madina Grocery — Executive Dashboard", 13)

    kpis = [
        ("Total Revenue", 2634500, "#,##0"),
        ("Total Expenses", 1928600, "#,##0"),
        ("Net Profit", 705900, "#,##0"),
        ("Profit Margin", 0.2679, "0.0%"),
        ("Total Transactions", 48750, "#,##0"),
        ("Avg Basket Size", 54.04, "#,##0.00"),
    ]
    row = write_kpi_row(ws, 4, 2, kpis)
    row += 1

    headers = ["Month", "Revenue (AED)", "Expenses (AED)", "Net Profit (AED)", "Profit Margin", "Transactions", "Avg Basket (AED)"]
    data = [
        ["Aug 2025", 398200, 296500, 101700, 0.2554, 7200, 55.31],
        ["Sep 2025", 412800, 305200, 107600, 0.2607, 7450, 55.41],
        ["Oct 2025", 435600, 318400, 117200, 0.2691, 7820, 55.70],
        ["Nov 2025", 448900, 327800, 121100, 0.2698, 8100, 55.42],
        ["Dec 2025", 467200, 338500, 128700, 0.2754, 8500, 54.96],
        ["Jan 2026", 472000, 342200, 129800, 0.2750, 9680, 48.76],
    ]
    num_fmts = [None, "#,##0", "#,##0", "#,##0", "0.0%", "#,##0", "#,##0.00"]
    total_row = ["TOTAL", 2634500, 1928600, 705900, 0.2679, 48750, 54.04]
    row = write_table(ws, row, 2, headers, data, total_row, num_fmts)

    # Charts
    cat_ref = Reference(ws, min_col=2, min_row=8, max_row=13)
    rev_ref = Reference(ws, min_col=3, min_row=7, max_row=13)
    exp_ref = Reference(ws, min_col=4, min_row=7, max_row=13)
    add_bar_chart(ws, "Revenue vs Expenses", cat_ref, [rev_ref, exp_ref],
                  ["Revenue", "Expenses"], row, 2, width=18, height=10, y_title="AED")

    margin_ref = Reference(ws, min_col=5, min_row=7, max_row=13)
    add_line_chart(ws, "Profit Margin Trend", cat_ref, [margin_ref],
                   ["Profit Margin"], row, 9, width=12, height=10, y_title="Margin")

    trans_ref = Reference(ws, min_col=7, min_row=7, max_row=13)
    add_line_chart(ws, "Monthly Transactions", cat_ref, [trans_ref],
                   ["Transactions"], row + 16, 2, width=12, height=10, y_title="Count")

    add_footer(ws, ws.max_row + 18, 2, 13)
    auto_fit_columns(ws)

    # ---- Sheet 2: Product Analysis ----
    ws2 = wb.create_sheet("Product Analysis")
    setup_sheet(ws2, "Product Performance — January 2026", 10)

    row = 4
    headers = ["Product", "Category", "Units Sold", "Revenue (AED)", "Unit Price (AED)", "Cost Price (AED)", "Profit Margin", "Stock Status"]
    prod_data = [
        ["Al Ain Water (24pk)", "Beverages", 2850, 57000, 20.00, 14.00, 0.30, "Adequate"],
        ["Basmati Rice (5kg)", "Grains & Pulses", 1920, 76800, 40.00, 28.00, 0.30, "Adequate"],
        ["Fresh Chicken (1kg)", "Meat & Poultry", 2340, 70200, 30.00, 22.00, 0.267, "Low Stock"],
        ["Olive Oil (1L)", "Cooking Essentials", 1680, 67200, 40.00, 26.00, 0.35, "Adequate"],
        ["Labneh (500g)", "Dairy", 2100, 37800, 18.00, 11.00, 0.389, "Adequate"],
        ["Arabic Bread (10pk)", "Bakery", 3200, 32000, 10.00, 6.50, 0.35, "Surplus"],
        ["Dates (1kg)", "Dry Fruits", 1450, 43500, 30.00, 18.00, 0.40, "Adequate"],
        ["Milk (2L)", "Dairy", 2800, 42000, 15.00, 10.00, 0.333, "Low Stock"],
        ["Coca-Cola (12pk)", "Beverages", 1950, 39000, 20.00, 14.00, 0.30, "Adequate"],
        ["Saffron (1g)", "Spices", 680, 34000, 50.00, 30.00, 0.40, "Critical"],
        ["Frozen Falafel (1kg)", "Frozen Foods", 1560, 23400, 15.00, 9.00, 0.40, "Adequate"],
        ["Honey (500g)", "Dry Fruits", 920, 36800, 40.00, 24.00, 0.40, "Low Stock"],
        ["Laundry Detergent", "Household", 1380, 27600, 20.00, 13.00, 0.35, "Adequate"],
        ["Diapers (Large)", "Baby Care", 1100, 33000, 30.00, 20.00, 0.333, "Adequate"],
        ["Tea Bags (100pk)", "Beverages", 2050, 41000, 20.00, 12.00, 0.40, "Surplus"],
    ]
    num_fmts = [None, None, "#,##0", "#,##0", "#,##0.00", "#,##0.00", "0.0%", None]
    row = write_table(ws2, row, 2, headers, prod_data, num_fmts=num_fmts)

    # Category Summary
    row += 1
    row += 1
    ws2.cell(row=row, column=2, value="Category Summary").font = font_subheader()
    row += 1
    cat_headers = ["Category", "Revenue (AED)", "% of Revenue", "Avg Margin"]
    cat_summary = [
        ["Beverages", 137000, 0.2195, 0.333],
        ["Grains & Pulses", 76800, 0.1231, 0.30],
        ["Meat & Poultry", 70200, 0.1125, 0.267],
        ["Cooking Essentials", 67200, 0.1077, 0.35],
        ["Dairy", 79800, 0.1279, 0.361],
        ["Bakery", 32000, 0.0513, 0.35],
        ["Dry Fruits", 80300, 0.1287, 0.40],
        ["Spices", 34000, 0.0545, 0.40],
        ["Frozen Foods", 23400, 0.0375, 0.40],
        ["Household", 27600, 0.0442, 0.35],
        ["Baby Care", 33000, 0.0529, 0.333],
    ]
    num_fmts2 = [None, "#,##0", "0.0%", "0.0%"]
    cat_table_start = row
    row = write_table(ws2, row, 2, cat_headers, cat_summary, num_fmts=num_fmts2)

    # Charts
    prod_names = Reference(ws2, min_col=2, min_row=5, max_row=19)
    prod_rev = Reference(ws2, min_col=5, min_row=4, max_row=19)
    add_bar_chart(ws2, "Top Products by Revenue", prod_names, [prod_rev],
                  ["Revenue (AED)"], row + 1, 2, width=18, height=10,
                  y_title="AED", chart_type="bar")  # horizontal bar

    cat_names_ref = Reference(ws2, min_col=2, min_row=cat_table_start+1, max_row=cat_table_start+len(cat_summary))
    cat_rev_ref = Reference(ws2, min_col=3, min_row=cat_table_start, max_row=cat_table_start+len(cat_summary))
    add_pie_chart(ws2, "Revenue by Category", cat_names_ref, cat_rev_ref,
                  len(cat_summary), row + 1, 8, width=14, height=10)

    cat_margin_ref = Reference(ws2, min_col=5, min_row=cat_table_start, max_row=cat_table_start+len(cat_summary))
    add_bar_chart(ws2, "Margin by Category", cat_names_ref, [cat_margin_ref],
                  ["Avg Margin"], row + 16, 2, width=14, height=10, y_title="Margin")

    add_footer(ws2, ws2.max_row + 18, 2, 10)
    auto_fit_columns(ws2)

    # ---- Sheet 3: Inventory Management ----
    ws3 = wb.create_sheet("Inventory Management")
    setup_sheet(ws3, "Inventory & Supply Chain — January 2026", 11)

    row = 4
    headers = ["Product", "Current Stock", "Reorder Level", "Status", "Days of Supply", "Last Order Date", "Supplier", "Lead Time (days)", "Action"]
    inv_data = [
        ["Al Ain Water (24pk)", 450, 200, "Adequate", 18, "2026-01-25", "Al Ain Foods", 3, "Monitor"],
        ["Basmati Rice (5kg)", 180, 150, "Adequate", 14, "2026-01-22", "Al Khaleej Rice", 5, "Monitor"],
        ["Fresh Chicken (1kg)", 65, 100, "Low Stock", 5, "2026-01-28", "Al Rawdah", 1, "Reorder Now"],
        ["Olive Oil (1L)", 120, 100, "Adequate", 12, "2026-01-20", "Filipppo Berio", 7, "Monitor"],
        ["Labneh (500g)", 200, 150, "Adequate", 16, "2026-01-24", "Almarai", 2, "Monitor"],
        ["Arabic Bread (10pk)", 520, 200, "Surplus", 28, "2026-01-29", "Modern Bakery", 1, "Reduce Order"],
        ["Dates (1kg)", 95, 100, "Low Stock", 8, "2026-01-23", "Al Foah", 5, "Reorder Now"],
        ["Milk (2L)", 80, 120, "Low Stock", 6, "2026-01-29", "Almarai", 1, "Reorder Now"],
        ["Coca-Cola (12pk)", 160, 150, "Adequate", 15, "2026-01-21", "Coca-Cola MEA", 4, "Monitor"],
        ["Saffron (1g)", 15, 30, "Critical", 3, "2026-01-15", "Iranian Spice Co", 10, "Urgent Reorder"],
        ["Frozen Falafel (1kg)", 130, 100, "Adequate", 11, "2026-01-26", "Al Areesh", 3, "Monitor"],
        ["Honey (500g)", 45, 60, "Low Stock", 7, "2026-01-19", "Al Shifa", 7, "Reorder Now"],
        ["Laundry Detergent", 140, 100, "Adequate", 13, "2026-01-22", "Unilever MEA", 5, "Monitor"],
        ["Diapers (Large)", 110, 100, "Adequate", 10, "2026-01-24", "P&G MEA", 4, "Monitor"],
        ["Tea Bags (100pk)", 280, 150, "Surplus", 22, "2026-01-27", "Lipton MEA", 5, "Reduce Order"],
    ]
    num_fmts = [None, "#,##0", "#,##0", None, "#,##0", "YYYY-MM-DD", None, "#,##0", None]
    row = write_table(ws3, row, 2, headers, inv_data, num_fmts=num_fmts)

    # Conditional formatting on Status column (col 5 = E)
    status_range = f"E5:E{4 + len(inv_data)}"
    crit_fill = PatternFill(bgColor="FDEDEC")
    crit_font = Font(color=ACCENT_NEGATIVE)
    low_fill = PatternFill(bgColor="FEF9E7")
    low_font = Font(color=ACCENT_WARNING)
    ok_fill = PatternFill(bgColor="E8F5E9")
    ok_font = Font(color=ACCENT_POSITIVE)
    sur_fill = PatternFill(bgColor="D6E4F0")
    sur_font = Font(color=PRIMARY)

    ws3.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"Critical"'], fill=crit_fill, font=crit_font))
    ws3.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"Low Stock"'], fill=low_fill, font=low_font))
    ws3.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"Adequate"'], fill=ok_fill, font=ok_font))
    ws3.conditional_formatting.add(status_range,
        CellIsRule(operator='equal', formula=['"Surplus"'], fill=sur_fill, font=sur_font))

    # Action column conditional formatting (col 10 = J)
    action_range = f"J5:J{4 + len(inv_data)}"
    ws3.conditional_formatting.add(action_range,
        CellIsRule(operator='equal', formula=['"Urgent Reorder"'], fill=crit_fill, font=crit_font))
    ws3.conditional_formatting.add(action_range,
        CellIsRule(operator='equal', formula=['"Reorder Now"'], fill=low_fill, font=low_font))
    ws3.conditional_formatting.add(action_range,
        CellIsRule(operator='equal', formula=['"Reduce Order"'], fill=sur_fill, font=sur_font))

    # Supplier Summary
    row += 1
    row += 1
    ws3.cell(row=row, column=2, value="Supplier Summary").font = font_subheader()
    row += 1
    sup_headers = ["Supplier", "Products", "Total Orders", "On-Time Delivery %", "Rating"]
    sup_data = [
        ["Almarai", 2, 28, 0.96, 4.8],
        ["Al Ain Foods", 1, 22, 0.98, 4.9],
        ["Al Khaleej Rice", 1, 18, 0.94, 4.5],
        ["Al Rawdah", 1, 30, 0.92, 4.3],
        ["Filipppo Berio", 1, 12, 0.88, 4.0],
        ["Al Foah", 1, 15, 0.93, 4.4],
        ["Coca-Cola MEA", 1, 16, 0.97, 4.7],
        ["Iranian Spice Co", 1, 8, 0.85, 3.8],
        ["Al Areesh", 1, 20, 0.95, 4.6],
        ["Al Shifa", 1, 10, 0.90, 4.2],
        ["Unilever MEA", 1, 14, 0.96, 4.5],
        ["P&G MEA", 1, 16, 0.97, 4.6],
        ["Lipton MEA", 1, 14, 0.94, 4.4],
        ["Modern Bakery", 1, 30, 0.99, 4.9],
    ]
    num_fmts2 = [None, "#,##0", "#,##0", "0.0%", "0.0"]
    row = write_table(ws3, row, 2, sup_headers, sup_data, num_fmts=num_fmts2)

    add_footer(ws3, row + 2, 2, 11)
    auto_fit_columns(ws3)

    # ---- Sheet 4: Customer Analysis ----
    ws4 = wb.create_sheet("Customer Analysis")
    setup_sheet(ws4, "Customer Insights — January 2026", 9)

    row = 4
    headers = ["Segment", "Customers", "Avg Spend (AED)", "Visits/Month", "Revenue (AED)", "% of Revenue"]
    seg_data = [
        ["Daily Shoppers", 3200, 28.50, 24, 2188800, 0.4635],
        ["Weekly Families", 1800, 85.00, 4, 612000, 0.1297],
        ["Monthly Bulk", 650, 210.00, 1, 136500, 0.0289],
        ["Occasional", 2200, 42.00, 2, 184800, 0.0392],
        ["Wholesale", 85, 750.00, 3, 191250, 0.0405],
    ]
    num_fmts = [None, "#,##0", "#,##0.00", "#,##0", "#,##0", "0.0%"]
    row = write_table(ws4, row, 2, headers, seg_data, num_fmts=num_fmts)

    # Charts
    seg_names = Reference(ws4, min_col=2, min_row=5, max_row=9)
    seg_rev = Reference(ws4, min_col=6, min_row=4, max_row=9)
    add_pie_chart(ws4, "Revenue by Customer Segment", seg_names, seg_rev,
                  len(seg_data), row + 1, 2, width=14, height=10)

    # Day-of-week analysis
    row += 18
    row += 1
    headers2 = ["Day", "Revenue (AED)", "Transactions", "Avg Basket (AED)"]
    dow_data = [
        ["Saturday", 82000, 1680, 48.81],
        ["Sunday", 68000, 1420, 47.89],
        ["Monday", 58000, 1180, 49.15],
        ["Tuesday", 55000, 1120, 49.11],
        ["Wednesday", 60000, 1240, 48.39],
        ["Thursday", 65000, 1350, 48.15],
        ["Friday", 84500, 1780, 47.47],
    ]
    num_fmts2 = [None, "#,##0", "#,##0", "#,##0.00"]
    dow_table_start = row
    row = write_table(ws4, row, 2, headers2, dow_data, num_fmts=num_fmts2)

    # Bar chart: Revenue by Day
    day_names = Reference(ws4, min_col=2, min_row=dow_table_start+1, max_row=dow_table_start+len(dow_data))
    day_rev = Reference(ws4, min_col=3, min_row=dow_table_start, max_row=dow_table_start+len(dow_data))
    add_bar_chart(ws4, "Revenue by Day of Week", day_names, [day_rev],
                  ["Revenue (AED)"], row + 1, 2, width=14, height=10, y_title="AED")

    add_footer(ws4, ws4.max_row + 14, 2, 9)
    auto_fit_columns(ws4)

    # ---- Sheet 5: Recommendations ----
    ws5 = wb.create_sheet("Recommendations")
    setup_sheet(ws5, "Key Insights & Strategic Recommendations", 8)

    row = 4
    headers = ["#", "Category", "Insight", "Impact", "Action Item", "Priority"]
    rec_data = [
        [1, "Inventory", "Saffron stock is critical (3 days of supply) with 10-day lead time", "High",
         "Place urgent reorder with Iranian Spice Co; identify backup supplier", "High"],
        [2, "Revenue", "Daily Shoppers contribute 46% of revenue but have low avg spend (AED 28.50)", "High",
         "Implement cross-sell promotions at checkout to increase basket size", "High"],
        [3, "Supply Chain", "Arabic Bread surplus (28 days) with high waste risk", "Medium",
         "Reduce order quantity by 40% and negotiate smaller daily deliveries", "Medium"],
        [4, "Customer", "Wholesale segment has highest avg spend (AED 750) but only 85 customers", "High",
         "Launch B2B loyalty program with volume discounts to grow this segment", "High"],
        [5, "Operations", "Weekend revenue (Fri+Sat) is 35% of weekly total", "Medium",
         "Extend weekend hours and add promotional displays at entrance", "Medium"],
        [6, "Pricing", "Fresh Chicken margin (26.7%) is lowest across all products", "Medium",
         "Renegotiate supplier terms or increase retail price by AED 2/kg", "Medium"],
        [7, "Inventory", "Milk and Dates are below reorder level with high turnover", "High",
         "Set up automated reorder triggers when stock hits 120% of reorder level", "High"],
        [8, "Growth", "Monthly bulk buyers visit only once — potential for repeat purchase", "Low",
         "Introduce subscription boxes with home delivery for bulk items", "Low"],
    ]
    row = write_table(ws5, row, 2, headers, rec_data, num_fmts=None)

    # Conditional formatting
    priority_range = f"G5:G{4 + len(rec_data)}"
    ws5.conditional_formatting.add(priority_range,
        CellIsRule(operator='equal', formula=['"High"'], fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))
    ws5.conditional_formatting.add(priority_range,
        CellIsRule(operator='equal', formula=['"Medium"'], fill=CF_WARNING_FILL, font=CF_WARNING_FONT))
    ws5.conditional_formatting.add(priority_range,
        CellIsRule(operator='equal', formula=['"Low"'], fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT))

    impact_range = f"E5:E{4 + len(rec_data)}"
    ws5.conditional_formatting.add(impact_range,
        CellIsRule(operator='equal', formula=['"High"'], fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))
    ws5.conditional_formatting.add(impact_range,
        CellIsRule(operator='equal', formula=['"Medium"'], fill=CF_WARNING_FILL, font=CF_WARNING_FONT))

    add_footer(ws5, row + 2, 2, 8)
    auto_fit_columns(ws5)

    filepath = "/home/z/my-project/download/sample_report_grocery.xlsx"
    wb.save(filepath)
    print(f"✅ Saved: {filepath}")
    return filepath


# ============================================================
# FILE 3: LUXE BEAUTY SALON
# ============================================================

def generate_salon_report():
    use_palette_explicit("elegant")
    wb = Workbook()
    wb.properties.creator = "Z.ai"

    # ---- Sheet 1: Executive Summary ----
    ws = wb.active
    ws.title = "Executive Summary"
    setup_sheet(ws, "Luxe Beauty Salon — Executive Dashboard", 13)

    kpis = [
        ("Total Revenue", 982500, "#,##0"),
        ("Total Expenses", 614800, "#,##0"),
        ("Net Profit", 367700, "#,##0"),
        ("Profit Margin", 0.3742, "0.0%"),
        ("Total Bookings", 4280, "#,##0"),
        ("Avg Booking Value", 229.56, "#,##0.00"),
    ]
    row = write_kpi_row(ws, 4, 2, kpis)
    row += 1

    headers = ["Month", "Revenue (AED)", "Expenses (AED)", "Net Profit (AED)", "Profit Margin", "Bookings", "Avg Booking (AED)"]
    data = [
        ["Aug 2025", 142800, 92500, 50300, 0.3522, 620, 230.32],
        ["Sep 2025", 156200, 99800, 56400, 0.3611, 680, 229.71],
        ["Oct 2025", 168400, 105200, 63200, 0.3753, 730, 230.68],
        ["Nov 2025", 175600, 108900, 66700, 0.3798, 760, 231.05],
        ["Dec 2025", 172800, 107400, 65400, 0.3781, 845, 204.50],
        ["Jan 2026", 166700, 101000, 65700, 0.3941, 645, 258.45],
    ]
    num_fmts = [None, "#,##0", "#,##0", "#,##0", "0.0%", "#,##0", "#,##0.00"]
    total_row = ["TOTAL", 982500, 614800, 367700, 0.3742, 4280, 229.56]
    row = write_table(ws, row, 2, headers, data, total_row, num_fmts)

    # Charts
    cat_ref = Reference(ws, min_col=2, min_row=8, max_row=13)
    rev_ref = Reference(ws, min_col=3, min_row=7, max_row=13)
    exp_ref = Reference(ws, min_col=4, min_row=7, max_row=13)
    add_bar_chart(ws, "Revenue vs Expenses", cat_ref, [rev_ref, exp_ref],
                  ["Revenue", "Expenses"], row, 2, width=18, height=10, y_title="AED")

    margin_ref = Reference(ws, min_col=5, min_row=7, max_row=13)
    add_line_chart(ws, "Profit Margin Trend", cat_ref, [margin_ref],
                   ["Profit Margin"], row, 9, width=12, height=10, y_title="Margin")

    bookings_ref = Reference(ws, min_col=7, min_row=7, max_row=13)
    add_line_chart(ws, "Monthly Bookings Trend", cat_ref, [bookings_ref],
                   ["Bookings"], row + 16, 2, width=12, height=10, y_title="Bookings")

    add_footer(ws, ws.max_row + 18, 2, 13)
    auto_fit_columns(ws)

    # ---- Sheet 2: Service Analysis ----
    ws2 = wb.create_sheet("Service Analysis")
    setup_sheet(ws2, "Service Performance — January 2026", 10)

    row = 4
    headers = ["Service", "Category", "Bookings", "Revenue (AED)", "Avg Price (AED)", "Duration (min)", "Utilization %", "% of Revenue"]
    svc_data = [
        ["Hair Coloring", "Hair", 380, 114000, 300.00, 120, 0.82, 0.1368],
        ["Haircut & Styling", "Hair", 520, 78000, 150.00, 45, 0.91, 0.0936],
        ["Bridal Package", "Bridal", 85, 127500, 1500.00, 360, 0.95, 0.1530],
        ["Facial Treatment", "Skin Care", 290, 58000, 200.00, 60, 0.78, 0.0696],
        ["Manicure & Pedicure", "Nails", 640, 51200, 80.00, 45, 0.88, 0.0614],
        ["Keratin Treatment", "Hair", 195, 87750, 450.00, 180, 0.72, 0.1053],
        ["Henna Design", "Bridal", 280, 33600, 120.00, 90, 0.85, 0.0403],
        ["Hair Extensions", "Hair", 110, 55000, 500.00, 150, 0.65, 0.0660],
        ["Scalp Treatment", "Hair", 160, 32000, 200.00, 60, 0.70, 0.0384],
        ["Makeup", "Bridal", 420, 84000, 200.00, 60, 0.90, 0.1008],
    ]
    num_fmts = [None, None, "#,##0", "#,##0", "#,##0.00", "#,##0", "0.0%", "0.0%"]
    row = write_table(ws2, row, 2, headers, svc_data, num_fmts=num_fmts)

    # Service Category Summary
    row += 1
    row += 1
    ws2.cell(row=row, column=2, value="Service Category Summary").font = font_subheader()
    row += 1
    cat_headers = ["Category", "Bookings", "Revenue (AED)", "% of Revenue", "Avg Utilization"]
    cat_summary = [
        ["Hair", 1365, 366750, 0.4398, 0.775],
        ["Bridal", 785, 245100, 0.2941, 0.90],
        ["Skin Care", 290, 58000, 0.0696, 0.78],
        ["Nails", 640, 51200, 0.0614, 0.88],
    ]
    num_fmts2 = [None, "#,##0", "#,##0", "0.0%", "0.0%"]
    row = write_table(ws2, row, 2, cat_headers, cat_summary, num_fmts=num_fmts2)

    # Charts
    svc_names = Reference(ws2, min_col=2, min_row=5, max_row=14)
    svc_rev = Reference(ws2, min_col=5, min_row=4, max_row=14)
    add_pie_chart(ws2, "Revenue by Service", svc_names, svc_rev,
                  len(svc_data), row + 1, 2, width=14, height=10)

    svc_bookings = Reference(ws2, min_col=4, min_row=4, max_row=14)
    add_bar_chart(ws2, "Bookings by Service", svc_names, [svc_bookings],
                  ["Bookings"], row + 1, 8, width=14, height=10, y_title="Bookings")

    svc_util = Reference(ws2, min_col=8, min_row=4, max_row=14)
    add_bar_chart(ws2, "Utilization by Service", svc_names, [svc_util],
                  ["Utilization %"], row + 16, 2, width=14, height=10, y_title="Utilization")

    add_footer(ws2, ws2.max_row + 18, 2, 10)
    auto_fit_columns(ws2)

    # ---- Sheet 3: Staff Performance ----
    ws3 = wb.create_sheet("Staff Performance")
    setup_sheet(ws3, "Staff Performance & Productivity — January 2026", 11)

    row = 4
    headers = ["Name", "Role", "Clients", "Revenue (AED)", "Avg Ticket (AED)", "Utilization", "Rating", "Retention Rate", "Upsell Rate"]
    staff_data = [
        ["Layla Al Maktoum", "Senior Stylist", 145, 52200, 360.00, 0.92, 4.9, 0.88, 0.35],
        ["Noura Al Ketbi", "Color Specialist", 120, 43200, 360.00, 0.88, 4.8, 0.85, 0.30],
        ["Sara Ahmed", "Bridal Expert", 65, 39000, 600.00, 0.95, 5.0, 0.92, 0.45],
        ["Huda Hassan", "Nail Technician", 180, 14400, 80.00, 0.85, 4.5, 0.80, 0.20],
        ["Mariam Al Suwaidi", "Facial Therapist", 110, 22000, 200.00, 0.78, 4.6, 0.82, 0.25],
        ["Aisha Mohammed", "Junior Stylist", 95, 14250, 150.00, 0.72, 4.2, 0.75, 0.15],
        ["Fatima Al Zaabi", "Henna Artist", 160, 19200, 120.00, 0.82, 4.4, 0.78, 0.22],
        ["Khadija Omar", "Makeup Artist", 175, 35000, 200.00, 0.90, 4.7, 0.86, 0.28],
    ]
    num_fmts = [None, None, "#,##0", "#,##0", "#,##0.00", "0.0%", "0.0", "0.0%", "0.0%"]
    row = write_table(ws3, row, 2, headers, staff_data, num_fmts=num_fmts)

    # Staff Revenue Breakdown by Service (simplified)
    row += 1
    row += 1
    ws3.cell(row=row, column=2, value="Staff Revenue by Service Category").font = font_subheader()
    row += 1
    svc_headers = ["Staff Member", "Hair Services", "Bridal Services", "Skin Care", "Nails", "Other"]
    svc_staff_data = [
        ["Layla Al Maktoum", 31200, 10500, 4500, 0, 6000],
        ["Noura Al Ketbi", 36000, 0, 0, 0, 7200],
        ["Sara Ahmed", 0, 39000, 0, 0, 0],
        ["Huda Hassan", 0, 0, 0, 14400, 0],
        ["Mariam Al Suwaidi", 0, 0, 22000, 0, 0],
        ["Aisha Mohammed", 9500, 0, 0, 0, 4750],
        ["Fatima Al Zaabi", 0, 8400, 0, 0, 10800],
        ["Khadija Omar", 0, 35000, 0, 0, 0],
    ]
    num_fmts2 = [None, "#,##0", "#,##0", "#,##0", "#,##0", "#,##0"]
    row = write_table(ws3, row, 2, svc_headers, svc_staff_data, num_fmts=num_fmts2)

    # Charts
    staff_names = Reference(ws3, min_col=2, min_row=5, max_row=12)
    staff_rev = Reference(ws3, min_col=5, min_row=4, max_row=12)
    add_bar_chart(ws3, "Revenue by Staff", staff_names, [staff_rev],
                  ["Revenue (AED)"], row + 1, 2, width=18, height=10,
                  y_title="AED", chart_type="bar")

    staff_rating = Reference(ws3, min_col=8, min_row=4, max_row=12)
    add_bar_chart(ws3, "Staff Rating Comparison", staff_names, [staff_rating],
                  ["Rating"], row + 1, 9, width=14, height=10, y_title="Rating (out of 5)")

    add_footer(ws3, ws3.max_row + 14, 2, 11)
    auto_fit_columns(ws3)

    # ---- Sheet 4: Client Insights ----
    ws4 = wb.create_sheet("Client Insights")
    setup_sheet(ws4, "Client Analytics — January 2026", 9)

    row = 4
    headers = ["Segment", "Clients", "Avg Spend (AED)", "Visits/Month", "Revenue (AED)", "Retention Rate"]
    seg_data = [
        ["Regular Monthly", 1850, 185.00, 3, 1027875, 0.82],
        ["Bridal", 280, 650.00, 1, 182000, 0.45],
        ["Occasional", 950, 120.00, 2, 228000, 0.58],
        ["Tourist", 420, 280.00, 1, 117600, 0.20],
        ["VIP", 65, 950.00, 4, 247000, 0.92],
    ]
    num_fmts = [None, "#,##0", "#,##0.00", "#,##0", "#,##0", "0.0%"]
    row = write_table(ws4, row, 2, headers, seg_data, num_fmts=num_fmts)

    # Pie chart
    seg_names = Reference(ws4, min_col=2, min_row=5, max_row=9)
    seg_rev = Reference(ws4, min_col=6, min_row=4, max_row=9)
    add_pie_chart(ws4, "Revenue by Client Segment", seg_names, seg_rev,
                  len(seg_data), row + 1, 2, width=14, height=10)

    # New vs Returning breakdown
    row += 18
    row += 1
    ws4.cell(row=row, column=2, value="New vs Returning Clients (6-Month Trend)").font = font_subheader()
    row += 1
    nr_headers = ["Month", "New Clients", "Returning Clients", "Total", "Retention Rate"]
    nr_data = [
        ["Aug 2025", 185, 435, 620, 0.702],
        ["Sep 2025", 195, 485, 680, 0.713],
        ["Oct 2025", 210, 520, 730, 0.712],
        ["Nov 2025", 225, 535, 760, 0.704],
        ["Dec 2025", 260, 585, 845, 0.692],
        ["Jan 2026", 190, 455, 645, 0.706],
    ]
    num_fmts2 = [None, "#,##0", "#,##0", "#,##0", "0.0%"]
    nr_start = row
    row = write_table(ws4, row, 2, nr_headers, nr_data, num_fmts=num_fmts2)

    # Line chart: New vs Returning
    months_ref = Reference(ws4, min_col=2, min_row=nr_start+1, max_row=nr_start+len(nr_data))
    new_ref = Reference(ws4, min_col=3, min_row=nr_start, max_row=nr_start+len(nr_data))
    ret_ref = Reference(ws4, min_col=4, min_row=nr_start, max_row=nr_start+len(nr_data))
    add_line_chart(ws4, "New vs Returning Clients", months_ref, [new_ref, ret_ref],
                   ["New Clients", "Returning Clients"], row + 1, 2, width=16, height=10,
                   y_title="Clients")

    # Day-of-week and time-slot analysis
    row += 18
    row += 1
    ws4.cell(row=row, column=2, value="Day & Time Slot Analysis").font = font_subheader()
    row += 1
    dt_headers = ["Day", "Time Slot", "Bookings", "Revenue (AED)", "Avg Booking (AED)"]
    dt_data = [
        ["Saturday", "9 AM – 12 PM", 42, 10920, 260.00],
        ["Saturday", "12 PM – 4 PM", 58, 15080, 260.00],
        ["Saturday", "4 PM – 9 PM", 72, 19440, 270.00],
        ["Sunday", "9 AM – 12 PM", 28, 7000, 250.00],
        ["Sunday", "12 PM – 4 PM", 45, 11700, 260.00],
        ["Sunday", "4 PM – 9 PM", 55, 14300, 260.00],
        ["Monday", "10 AM – 5 PM", 48, 12000, 250.00],
        ["Tuesday", "10 AM – 5 PM", 52, 13000, 250.00],
        ["Wednesday", "10 AM – 5 PM", 50, 12500, 250.00],
        ["Thursday", "10 AM – 9 PM", 78, 21060, 270.00],
        ["Friday", "9 AM – 9 PM", 125, 35000, 280.00],
    ]
    num_fmts3 = [None, None, "#,##0", "#,##0", "#,##0.00"]
    row = write_table(ws4, row, 2, dt_headers, dt_data, num_fmts=num_fmts3)

    add_footer(ws4, row + 2, 2, 9)
    auto_fit_columns(ws4)

    # ---- Sheet 5: Recommendations ----
    ws5 = wb.create_sheet("Recommendations")
    setup_sheet(ws5, "Key Insights & Strategic Recommendations", 8)

    row = 4
    headers = ["#", "Category", "Insight", "Impact", "Action Item", "Priority"]
    rec_data = [
        [1, "Revenue", "Bridal Package has highest avg price (AED 1,500) and 95% utilization", "High",
         "Increase bridal package price by 10% and add premium add-ons", "High"],
        [2, "Staff", "Sara Ahmed has 5.0 rating and highest retention rate (92%) but serves only 65 clients", "High",
         "Increase Sara's capacity by assigning junior assistant for bridal prep", "High"],
        [3, "Client", "Tourist segment has 20% retention — opportunity for re-engagement", "Medium",
         "Launch digital follow-up campaign with discount for next visit", "Medium"],
        [4, "Service", "Hair Extensions utilization is only 65% — lowest performing service", "Medium",
         "Promote extensions with free consultation and before/after showcase", "Medium"],
        [5, "Operations", "Friday and Saturday generate 40% of weekly bookings", "High",
         "Add premium pricing for weekend peak slots to maximize revenue per slot", "High"],
        [6, "Client", "VIP segment (65 clients) generates AED 247K — AED 3,808 avg per client", "High",
         "Create exclusive VIP membership with priority booking and complimentary services", "High"],
        [7, "Service", "Scalp Treatment has 70% utilization — growth potential in hair wellness", "Low",
         "Partner with dermatologists for referrals and add treatment packages", "Medium"],
        [8, "Marketing", "New client acquisition dropped 27% in Jan vs Dec", "Medium",
         "Launch referral program: existing clients get AED 50 credit for each referral", "High"],
    ]
    row = write_table(ws5, row, 2, headers, rec_data, num_fmts=None)

    # Conditional formatting
    priority_range = f"G5:G{4 + len(rec_data)}"
    ws5.conditional_formatting.add(priority_range,
        CellIsRule(operator='equal', formula=['"High"'], fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))
    ws5.conditional_formatting.add(priority_range,
        CellIsRule(operator='equal', formula=['"Medium"'], fill=CF_WARNING_FILL, font=CF_WARNING_FONT))
    ws5.conditional_formatting.add(priority_range,
        CellIsRule(operator='equal', formula=['"Low"'], fill=CF_POSITIVE_FILL, font=CF_POSITIVE_FONT))

    impact_range = f"E5:E{4 + len(rec_data)}"
    ws5.conditional_formatting.add(impact_range,
        CellIsRule(operator='equal', formula=['"High"'], fill=CF_NEGATIVE_FILL, font=CF_NEGATIVE_FONT))
    ws5.conditional_formatting.add(impact_range,
        CellIsRule(operator='equal', formula=['"Medium"'], fill=CF_WARNING_FILL, font=CF_WARNING_FONT))

    add_footer(ws5, row + 2, 2, 8)
    auto_fit_columns(ws5)

    filepath = "/home/z/my-project/download/sample_report_salon.xlsx"
    wb.save(filepath)
    print(f"✅ Saved: {filepath}")
    return filepath


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    print("=" * 60)
    print("Generating 3 Professional Excel Reports")
    print("=" * 60)

    f1 = generate_restaurant_report()
    f2 = generate_grocery_report()
    f3 = generate_salon_report()

    print("\n" + "=" * 60)
    print("All reports generated successfully!")
    print(f"  1. {f1}")
    print(f"  2. {f2}")
    print(f"  3. {f3}")
    print("=" * 60)
