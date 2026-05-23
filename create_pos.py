#!/usr/bin/env python3
"""
Professional POS (Point of Sale) — Empty Template
===================================================
8-sheet workbook with formulas, validation, conditional formatting, charts.
Uses base.py styling system with professional palette.
"""

import sys, os
sys.path.insert(0, "/home/z/my-project/skills/xlsx")

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.chart import Reference

from templates.base import (
    # Palette
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_50, NEUTRAL_0,
    HEADER_TEXT, CHART_COLORS,
    CF_POSITIVE_FILL, CF_POSITIVE_FONT,
    CF_NEGATIVE_FILL, CF_NEGATIVE_FONT,
    CF_WARNING_FILL, CF_WARNING_FONT,
    COLUMN_WIDTHS, FORMATS, ROW_HEIGHTS,
    # Style factories
    font_title, font_header, font_subheader, font_body, font_caption,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    # Sheet helpers
    setup_sheet, style_header_row, style_data_row, style_total_row,
    # Chart helpers
    create_bar_chart, setup_chart_titles, apply_chart_colors,
)

# ============================================================
# Workbook init
# ============================================================
wb = Workbook()
wb.properties.creator = "Z.ai"

TAB_COLOR = PRIMARY  # deep blue

# ============================================================
# HELPER: set column widths from a list of (width_key, custom_width)
# ============================================================
def set_col_widths(ws, widths):
    """widths: list of int/str values. Position 0 = col B."""
    for i, w in enumerate(widths):
        col = i + 2  # B=2
        if isinstance(w, str) and w in COLUMN_WIDTHS:
            ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS[w]
        else:
            ws.column_dimensions[get_column_letter(col)].width = w


def apply_body_style(ws, row, col_start, col_end, row_idx, formats=None):
    """Style a data row: alternating fill + body font + optional number formats."""
    fill = fill_data_row(row_idx)
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font_body()
    ws.row_dimensions[row].height = ROW_HEIGHTS["data"]
    if formats:
        for col_off, fmt in formats:
            ws.cell(row=row, column=col_start + col_off).number_format = fmt


# ============================================================
# 1. SALES ENTRY
# ============================================================
ws1 = wb.active
ws1.title = "Sales Entry"
ws1.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws1, title="Sales Entry", last_col=12)

headers_se = ["Date", "Invoice #", "Item / Service", "Category",
              "Qty", "Unit Price", "Total", "Payment Method",
              "Staff Name", "Customer Name", "Notes"]

col_widths_se = [14, 14, 22, 16, 10, 14, 14, 16, 16, 20, 24]
set_col_widths(ws1, col_widths_se)

# Header row = 4
HR = 4
for i, h in enumerate(headers_se):
    ws1.cell(row=HR, column=2 + i, value=h)
style_header_row(ws1, HR, 2, 12)

# Data rows 5..304 (300 rows)
DATA_ROWS_SE = 300
for r in range(HR + 1, HR + 1 + DATA_ROWS_SE):
    row_idx = r - HR  # 1-based alternating index
    # Total formula: =IFERROR(F5*G5,0)  — Qty (col F=6) * Unit Price (col G=7) → Total (col H=8)
    ws1.cell(row=r, column=8).value = f'=IFERROR({get_column_letter(6)}{r}*{get_column_letter(7)}{r},0)'
    # Apply styles
    apply_body_style(ws1, r, 2, 12, row_idx,
                     formats=[(4, FORMATS["decimal_2"]), (5, FORMATS["decimal_2"]),
                              (6, FORMATS["decimal_2"])])

# Total row
TR = HR + 1 + DATA_ROWS_SE
ws1.cell(row=TR, column=2, value="TOTALS")
ws1.cell(row=TR, column=6).value = f'=SUMPRODUCT(({get_column_letter(6)}{HR+1}:{get_column_letter(6)}{TR-1}))'
ws1.cell(row=TR, column=7).value = f'=SUMPRODUCT(({get_column_letter(7)}{HR+1}:{get_column_letter(7)}{TR-1}))'
ws1.cell(row=TR, column=8).value = f'=SUMPRODUCT(({get_column_letter(8)}{HR+1}:{get_column_letter(8)}{TR-1}))'
style_total_row(ws1, TR, 2, 12)
ws1.cell(row=TR, column=6).number_format = FORMATS["decimal_2"]
ws1.cell(row=TR, column=7).number_format = FORMATS["decimal_2"]
ws1.cell(row=TR, column=8).number_format = FORMATS["decimal_2"]

# Data validation: Payment Method
dv_pm = DataValidation(type="list", formula1='"Cash,Card,Online,Cheque"', allow_blank=True)
dv_pm.error = "Please select a valid payment method"
dv_pm.errorTitle = "Invalid Payment Method"
ws1.add_data_validation(dv_pm)
dv_pm.add(f'{get_column_letter(9)}{HR+1}:{get_column_letter(9)}{TR-1}')

# Data validation: Staff Name (linked to Staff sheet)
dv_staff = DataValidation(type="list", formula1="=Staff!$B$5:$B$24", allow_blank=True)
dv_staff.error = "Please select a staff member"
dv_staff.errorTitle = "Invalid Staff Name"
ws1.add_data_validation(dv_staff)
dv_staff.add(f'{get_column_letter(10)}{HR+1}:{get_column_letter(10)}{TR-1}')

# Freeze panes at C5
ws1.freeze_panes = "C5"

# Print area & title rows
ws1.print_area = f'B2:L{TR}'
ws1.print_title_rows = f'{HR}:{HR}'


# ============================================================
# 2. DAILY SUMMARY
# ============================================================
ws2 = wb.create_sheet("Daily Summary")
ws2.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws2, title="Daily Summary", last_col=9)

headers_ds = ["Date", "Total Sales", "# Transactions", "Avg Transaction",
              "Cash Total", "Card Total", "Online Total", "# Items Sold"]
col_widths_ds = [14, 16, 16, 16, 14, 14, 14, 16]
set_col_widths(ws2, col_widths_ds)

HR2 = 4
for i, h in enumerate(headers_ds):
    ws2.cell(row=HR2, column=2 + i, value=h)
style_header_row(ws2, HR2, 2, 9)

# 31 data rows
DAYS = 31
se_range = f"'Sales Entry'!$B${HR+1}:$B${HR+DATA_ROWS_SE}"
for r in range(HR2 + 1, HR2 + 1 + DAYS):
    row_idx = r - HR2
    dr = r  # current row
    date_col = get_column_letter(2)  # B = Date

    # Total Sales: =SUMPRODUCT((Sales Entry!$B$5:$B$304=B5)*(Sales Entry!$H$5:$H$304))
    ws2.cell(row=dr, column=3).value = (
        f"=SUMPRODUCT(({se_range}={date_col}{dr})*('Sales Entry'!$H${HR+1}:$H${HR+DATA_ROWS_SE}))"
    )
    # # Transactions: =SUMPRODUCT((Sales Entry!$B$5:$B$304=B5)*1)
    ws2.cell(row=dr, column=4).value = (
        f"=SUMPRODUCT(({se_range}={date_col}{dr})*1)"
    )
    # Avg Transaction: =IFERROR(C5/D5,0)
    ws2.cell(row=dr, column=5).value = f"=IFERROR({get_column_letter(3)}{dr}/{get_column_letter(4)}{dr},0)"
    # Cash Total
    ws2.cell(row=dr, column=6).value = (
        f"=SUMPRODUCT(({se_range}={date_col}{dr})*('Sales Entry'!$I${HR+1}:$I${HR+DATA_ROWS_SE}=\"Cash\")*('Sales Entry'!$H${HR+1}:$H${HR+DATA_ROWS_SE}))"
    )
    # Card Total
    ws2.cell(row=dr, column=7).value = (
        f"=SUMPRODUCT(({se_range}={date_col}{dr})*('Sales Entry'!$I${HR+1}:$I${HR+DATA_ROWS_SE}=\"Card\")*('Sales Entry'!$H${HR+1}:$H${HR+DATA_ROWS_SE}))"
    )
    # Online Total
    ws2.cell(row=dr, column=8).value = (
        f"=SUMPRODUCT(({se_range}={date_col}{dr})*('Sales Entry'!$I${HR+1}:$I${HR+DATA_ROWS_SE}=\"Online\")*('Sales Entry'!$H${HR+1}:$H${HR+DATA_ROWS_SE}))"
    )
    # # Items Sold
    ws2.cell(row=dr, column=9).value = (
        f"=SUMPRODUCT(({se_range}={date_col}{dr})*('Sales Entry'!$F${HR+1}:$F${HR+DATA_ROWS_SE}))"
    )

    apply_body_style(ws2, dr, 2, 9, row_idx,
                     formats=[(1, FORMATS["decimal_2"]), (2, FORMATS["integer"]),
                              (3, FORMATS["decimal_2"]), (4, FORMATS["decimal_2"]),
                              (5, FORMATS["decimal_2"]), (6, FORMATS["decimal_2"]),
                              (7, FORMATS["integer"])])

# Total row
TR2 = HR2 + 1 + DAYS
ws2.cell(row=TR2, column=2, value="TOTALS")
ws2.cell(row=TR2, column=3).value = f"=SUM({get_column_letter(3)}{HR2+1}:{get_column_letter(3)}{TR2-1})"
ws2.cell(row=TR2, column=4).value = f"=SUM({get_column_letter(4)}{HR2+1}:{get_column_letter(4)}{TR2-1})"
ws2.cell(row=TR2, column=5).value = f"=IFERROR({get_column_letter(3)}{TR2}/{get_column_letter(4)}{TR2},0)"
ws2.cell(row=TR2, column=6).value = f"=SUM({get_column_letter(6)}{HR2+1}:{get_column_letter(6)}{TR2-1})"
ws2.cell(row=TR2, column=7).value = f"=SUM({get_column_letter(7)}{HR2+1}:{get_column_letter(7)}{TR2-1})"
ws2.cell(row=TR2, column=8).value = f"=SUM({get_column_letter(8)}{HR2+1}:{get_column_letter(8)}{TR2-1})"
ws2.cell(row=TR2, column=9).value = f"=SUM({get_column_letter(9)}{HR2+1}:{get_column_letter(9)}{TR2-1})"
style_total_row(ws2, TR2, 2, 9)
for c in [3, 5, 6, 7, 8]:
    ws2.cell(row=TR2, column=c).number_format = FORMATS["decimal_2"]
ws2.cell(row=TR2, column=4).number_format = FORMATS["integer"]
ws2.cell(row=TR2, column=9).number_format = FORMATS["integer"]

ws2.freeze_panes = "C5"
ws2.print_area = f'B2:I{TR2}'
ws2.print_title_rows = f'{HR2}:{HR2}'


# ============================================================
# 3. WEEKLY SUMMARY
# ============================================================
ws3 = wb.create_sheet("Weekly Summary")
ws3.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws3, title="Weekly Summary", last_col=8)

headers_ws = ["Week #", "Week Start", "Week End", "Total Sales",
              "# Transactions", "Avg Transaction", "Top Staff"]
col_widths_ws = [10, 14, 14, 16, 16, 16, 18]
set_col_widths(ws3, col_widths_ws)

HR3 = 4
for i, h in enumerate(headers_ws):
    ws3.cell(row=HR3, column=2 + i, value=h)
style_header_row(ws3, HR3, 2, 8)

WEEKS = 52
for r in range(HR3 + 1, HR3 + 1 + WEEKS):
    row_idx = r - HR3
    dr = r
    # Week #
    ws3.cell(row=dr, column=2, value=row_idx)
    # Week Start & Week End left blank for user input
    # Total Sales: =SUMPRODUCT(('Daily Summary'!$B$5:$B$35>=C5)*('Daily Summary'!$B$5:$B$35<=D5)*('Daily Summary'!$C$5:$C$35))
    ws3.cell(row=dr, column=5).value = (
        f"=SUMPRODUCT(('Daily Summary'!$B${HR2+1}:$B${HR2+DAYS}>={get_column_letter(3)}{dr})*"
        f"('Daily Summary'!$B${HR2+1}:$B${HR2+DAYS}<={get_column_letter(4)}{dr})*"
        f"('Daily Summary'!$C${HR2+1}:$C${HR2+DAYS}))"
    )
    # # Transactions
    ws3.cell(row=dr, column=6).value = (
        f"=SUMPRODUCT(('Daily Summary'!$B${HR2+1}:$B${HR2+DAYS}>={get_column_letter(3)}{dr})*"
        f"('Daily Summary'!$B${HR2+1}:$B${HR2+DAYS}<={get_column_letter(4)}{dr})*"
        f"('Daily Summary'!$D${HR2+1}:$D${HR2+DAYS}))"
    )
    # Avg Transaction
    ws3.cell(row=dr, column=7).value = f"=IFERROR({get_column_letter(5)}{dr}/{get_column_letter(6)}{dr},0)"
    # Top Staff - left as placeholder text
    ws3.cell(row=dr, column=8, value="")

    apply_body_style(ws3, dr, 2, 8, row_idx,
                     formats=[(3, FORMATS["decimal_2"]), (4, FORMATS["integer"]),
                              (5, FORMATS["decimal_2"])])

# Total row
TR3 = HR3 + 1 + WEEKS
ws3.cell(row=TR3, column=2, value="TOTALS")
ws3.cell(row=TR3, column=5).value = f"=SUM({get_column_letter(5)}{HR3+1}:{get_column_letter(5)}{TR3-1})"
ws3.cell(row=TR3, column=6).value = f"=SUM({get_column_letter(6)}{HR3+1}:{get_column_letter(6)}{TR3-1})"
ws3.cell(row=TR3, column=7).value = f"=IFERROR({get_column_letter(5)}{TR3}/{get_column_letter(6)}{TR3},0)"
style_total_row(ws3, TR3, 2, 8)
ws3.cell(row=TR3, column=5).number_format = FORMATS["decimal_2"]
ws3.cell(row=TR3, column=6).number_format = FORMATS["integer"]
ws3.cell(row=TR3, column=7).number_format = FORMATS["decimal_2"]

ws3.freeze_panes = "C5"
ws3.print_area = f'B2:H{TR3}'
ws3.print_title_rows = f'{HR3}:{HR3}'


# ============================================================
# 4. MONTHLY SUMMARY
# ============================================================
ws4 = wb.create_sheet("Monthly Summary")
ws4.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws4, title="Monthly Summary", last_col=7)

headers_ms = ["Month", "Total Sales", "# Transactions", "Avg Transaction",
              "Best Day Sales", "Best Day Date"]
col_widths_ms = [16, 16, 16, 16, 16, 16]
set_col_widths(ws4, col_widths_ms)

HR4 = 4
for i, h in enumerate(headers_ms):
    ws4.cell(row=HR4, column=2 + i, value=h)
style_header_row(ws4, HR4, 2, 7)

MONTHS = 12
MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]
for r in range(HR4 + 1, HR4 + 1 + MONTHS):
    row_idx = r - HR4
    dr = r
    ws4.cell(row=dr, column=2, value=MONTH_NAMES[row_idx - 1])
    # Total Sales: =SUMPRODUCT((MONTH('Daily Summary'!$B$5:$B$35)=row_idx)*('Daily Summary'!$C$5:$C$35))
    ws4.cell(row=dr, column=3).value = (
        f"=SUMPRODUCT((MONTH('Daily Summary'!$B${HR2+1}:$B${HR2+DAYS})={row_idx})*"
        f"('Daily Summary'!$C${HR2+1}:$C${HR2+DAYS}))"
    )
    # # Transactions
    ws4.cell(row=dr, column=4).value = (
        f"=SUMPRODUCT((MONTH('Daily Summary'!$B${HR2+1}:$B${HR2+DAYS})={row_idx})*"
        f"('Daily Summary'!$D${HR2+1}:$D${HR2+DAYS}))"
    )
    # Avg Transaction
    ws4.cell(row=dr, column=5).value = f"=IFERROR({get_column_letter(3)}{dr}/{get_column_letter(4)}{dr},0)"
    # Best Day Sales: =MAX(IF(MONTH('Daily Summary'!$B$5:$B$35)=row_idx,'Daily Summary'!$C$5:$C$35))
    # Using SUMPRODUCT approach (no dynamic arrays)
    ws4.cell(row=dr, column=6).value = (
        f"=MAX(IF(MONTH('Daily Summary'!$B${HR2+1}:$B${HR2+DAYS})={row_idx},"
        f"'Daily Summary'!$C${HR2+1}:$C${HR2+DAYS}))"
    )
    # Best Day Date — leave blank for user; complex to do without dynamic arrays
    ws4.cell(row=dr, column=7, value="")

    apply_body_style(ws4, dr, 2, 7, row_idx,
                     formats=[(1, FORMATS["decimal_2"]), (2, FORMATS["integer"]),
                              (3, FORMATS["decimal_2"]), (4, FORMATS["decimal_2"])])

# Total row
TR4 = HR4 + 1 + MONTHS
ws4.cell(row=TR4, column=2, value="TOTALS")
ws4.cell(row=TR4, column=3).value = f"=SUM({get_column_letter(3)}{HR4+1}:{get_column_letter(3)}{TR4-1})"
ws4.cell(row=TR4, column=4).value = f"=SUM({get_column_letter(4)}{HR4+1}:{get_column_letter(4)}{TR4-1})"
ws4.cell(row=TR4, column=5).value = f"=IFERROR({get_column_letter(3)}{TR4}/{get_column_letter(4)}{TR4},0)"
style_total_row(ws4, TR4, 2, 7)
ws4.cell(row=TR4, column=3).number_format = FORMATS["decimal_2"]
ws4.cell(row=TR4, column=4).number_format = FORMATS["integer"]
ws4.cell(row=TR4, column=5).number_format = FORMATS["decimal_2"]

# Bar chart of monthly sales
chart = create_bar_chart(chart_type="col", grouping="clustered", width=18, height=10)
data_ref = Reference(ws4, min_col=3, min_row=HR4, max_row=HR4 + MONTHS, max_col=3)
cats_ref = Reference(ws4, min_col=2, min_row=HR4 + 1, max_row=HR4 + MONTHS)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
setup_chart_titles(chart, title="Monthly Sales", y_title="Sales Amount")
apply_chart_colors(chart, [PRIMARY])
ws4.add_chart(chart, "B18")

ws4.freeze_panes = "C5"
ws4.print_area = f'B2:G{TR4}'
ws4.print_title_rows = f'{HR4}:{HR4}'


# ============================================================
# 5. INVOICE
# ============================================================
ws5 = wb.create_sheet("Invoice")
ws5.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws5, title="Invoice", last_col=8)

# Company header placeholder (row 2-3, merged)
ws5.merge_cells("B2:H2")
ws5.cell(row=2, column=2, value="YOUR COMPANY NAME")
ws5.cell(row=2, column=2).font = Font(name="Calibri", size=18, bold=True, color=PRIMARY)
ws5.cell(row=2, column=2).alignment = Alignment(horizontal="center", vertical="center")

ws5.merge_cells("B3:H3")
ws5.cell(row=3, column=2, value="Address Line | Phone | Email | Website")
ws5.cell(row=3, column=2).font = font_caption()
ws5.cell(row=3, column=2).alignment = Alignment(horizontal="center", vertical="center")

# Invoice details section
ws5.cell(row=5, column=2, value="Invoice #:")
ws5.cell(row=5, column=2).font = font_subheader()
ws5.cell(row=5, column=3, value="")  # User fills in

ws5.cell(row=5, column=5, value="Date:")
ws5.cell(row=5, column=5).font = font_subheader()
ws5.cell(row=5, column=6, value="")  # User fills in

ws5.cell(row=6, column=2, value="Customer:")
ws5.cell(row=6, column=2).font = font_subheader()
ws5.cell(row=6, column=3, value="")

ws5.cell(row=6, column=5, value="Staff:")
ws5.cell(row=6, column=5).font = font_subheader()
ws5.cell(row=6, column=6, value="")

# Items table
items_header_row = 8
headers_inv = ["Item", "Qty", "Price", "Total"]
col_widths_inv_items = [30, 10, 14, 14]
# Position items at B-E
for i, h in enumerate(headers_inv):
    ws5.cell(row=items_header_row, column=2 + i, value=h)
style_header_row(ws5, items_header_row, 2, 5)

# 10 item rows
for r in range(items_header_row + 1, items_header_row + 11):
    row_idx = r - items_header_row
    # Total = Qty * Price
    ws5.cell(row=r, column=5).value = f'=IFERROR({get_column_letter(3)}{r}*{get_column_letter(4)}{r},0)'
    apply_body_style(ws5, r, 2, 5, row_idx,
                     formats=[(2, FORMATS["decimal_2"]), (3, FORMATS["decimal_2"])])

# Subtotal
sub_row = items_header_row + 11
ws5.cell(row=sub_row, column=2, value="Subtotal")
ws5.cell(row=sub_row, column=5).value = f'=SUM({get_column_letter(5)}{items_header_row+1}:{get_column_letter(5)}{items_header_row+10})'
style_total_row(ws5, sub_row, 2, 5)
ws5.cell(row=sub_row, column=5).number_format = FORMATS["decimal_2"]

# VAT 5%
vat_row = sub_row + 1
ws5.cell(row=vat_row, column=2, value="VAT 5%")
ws5.cell(row=vat_row, column=5).value = f'=IFERROR({get_column_letter(5)}{sub_row}*0.05,0)'
apply_body_style(ws5, vat_row, 2, 5, 2)
ws5.cell(row=vat_row, column=5).number_format = FORMATS["decimal_2"]

# Discount
disc_row = vat_row + 1
ws5.cell(row=disc_row, column=2, value="Discount")
ws5.cell(row=disc_row, column=5, value="")  # User enters discount amount
apply_body_style(ws5, disc_row, 2, 5, 3)
ws5.cell(row=disc_row, column=5).number_format = FORMATS["decimal_2"]

# Grand Total
gt_row = disc_row + 1
ws5.cell(row=gt_row, column=2, value="GRAND TOTAL")
ws5.cell(row=gt_row, column=5).value = f'=IFERROR({get_column_letter(5)}{sub_row}+{get_column_letter(5)}{vat_row}-{get_column_letter(5)}{disc_row},0)'
style_total_row(ws5, gt_row, 2, 5)
ws5.cell(row=gt_row, column=5).number_format = FORMATS["decimal_2"]
ws5.cell(row=gt_row, column=2).font = Font(name="Calibri", size=12, bold=True, color=PRIMARY)

# Thank you note
ty_row = gt_row + 2
ws5.cell(row=ty_row, column=2, value="Thank you for your business!")
ws5.cell(row=ty_row, column=2).font = font_caption()
ws5.cell(row=ty_row, column=2).alignment = Alignment(horizontal="center")

# Column widths
set_col_widths(ws5, [30, 10, 14, 14, 10, 14, 14, 14])

# Print area — receipt style
ws5.print_area = f'B2:E{ty_row}'
ws5.page_setup.orientation = 'portrait'
ws5.page_setup.paperSize = ws5.PAPERSIZE_A5


# ============================================================
# 6. INVENTORY
# ============================================================
ws6 = wb.create_sheet("Inventory")
ws6.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws6, title="Inventory", last_col=10)

headers_inv = ["Product ID", "Item Name", "Category", "Unit Price",
               "Stock Qty", "Reorder Level", "Status", "Supplier", "Last Restocked"]
col_widths_inv = [14, 22, 16, 14, 12, 14, 14, 18, 16]
set_col_widths(ws6, col_widths_inv)

HR6 = 4
for i, h in enumerate(headers_inv):
    ws6.cell(row=HR6, column=2 + i, value=h)
style_header_row(ws6, HR6, 2, 10)

INV_ROWS = 150
for r in range(HR6 + 1, HR6 + 1 + INV_ROWS):
    row_idx = r - HR6
    dr = r
    # Status: =IF(F5<=G5,"LOW STOCK","In Stock") — Stock Qty(F=6) vs Reorder Level(G=7)
    ws6.cell(row=dr, column=8).value = f'=IF({get_column_letter(6)}{dr}<={get_column_letter(7)}{dr},"LOW STOCK","In Stock")'
    apply_body_style(ws6, dr, 2, 10, row_idx,
                     formats=[(3, FORMATS["decimal_2"])])

# Total row
TR6 = HR6 + 1 + INV_ROWS
ws6.cell(row=TR6, column=2, value="SUMMARY")
ws6.cell(row=TR6, column=5).value = f"=SUM({get_column_letter(5)}{HR6+1}:{get_column_letter(5)}{TR6-1})"
ws6.cell(row=TR6, column=7).value = f"=SUM({get_column_letter(7)}{HR6+1}:{get_column_letter(7)}{TR6-1})"
style_total_row(ws6, TR6, 2, 10)
ws6.cell(row=TR6, column=5).number_format = FORMATS["decimal_2"]
ws6.cell(row=TR6, column=7).number_format = FORMATS["integer"]

# Conditional formatting: Status column (H)
status_range = f'{get_column_letter(8)}{HR6+1}:{get_column_letter(8)}{HR6+INV_ROWS}'

# LOW STOCK = red fill + red font
ws6.conditional_formatting.add(
    status_range,
    CellIsRule(
        operator="equal",
        formula=['"LOW STOCK"'],
        fill=PatternFill("solid", fgColor="FDEDEC"),
        font=Font(color=ACCENT_NEGATIVE, bold=True)
    )
)

# In Stock = green fill + green font
ws6.conditional_formatting.add(
    status_range,
    CellIsRule(
        operator="equal",
        formula=['"In Stock"'],
        fill=PatternFill("solid", fgColor="E8F5E9"),
        font=Font(color=ACCENT_POSITIVE)
    )
)

# Category dropdown
dv_cat = DataValidation(type="list", formula1='"Electronics,Clothing,Food & Beverage,Stationery,Health & Beauty,Home & Garden,Sports,Toys,Other"', allow_blank=True)
dv_cat.error = "Please select a valid category"
ws6.add_data_validation(dv_cat)
dv_cat.add(f'{get_column_letter(4)}{HR6+1}:{get_column_letter(4)}{TR6-1}')

ws6.freeze_panes = "C5"
ws6.print_area = f'B2:J{TR6}'
ws6.print_title_rows = f'{HR6}:{HR6}'


# ============================================================
# 7. STAFF PERFORMANCE
# ============================================================
ws7 = wb.create_sheet("Staff Performance")
ws7.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws7, title="Staff Performance", last_col=9)

headers_sp = ["Staff Name", "Total Sales", "# Transactions", "Avg Transaction",
              "Top Category", "Shift", "Commission (5%)"]
col_widths_sp = [20, 16, 16, 16, 18, 14, 16]
set_col_widths(ws7, col_widths_sp)

HR7 = 4
for i, h in enumerate(headers_sp):
    ws7.cell(row=HR7, column=2 + i, value=h)
style_header_row(ws7, HR7, 2, 8)

STAFF_ROWS = 20
se_staff_range = f"'Sales Entry'!$J${HR+1}:$J${HR+DATA_ROWS_SE}"
for r in range(HR7 + 1, HR7 + 1 + STAFF_ROWS):
    row_idx = r - HR7
    dr = r
    staff_col = get_column_letter(2)  # B = Staff Name

    # Total Sales: =SUMPRODUCT((Sales Entry!$J$5:$J$304=B5)*(Sales Entry!$H$5:$H$304))
    ws7.cell(row=dr, column=3).value = (
        f"=SUMPRODUCT(({se_staff_range}={staff_col}{dr})*('Sales Entry'!$H${HR+1}:$H${HR+DATA_ROWS_SE}))"
    )
    # # Transactions
    ws7.cell(row=dr, column=4).value = (
        f"=SUMPRODUCT(({se_staff_range}={staff_col}{dr})*1)"
    )
    # Avg Transaction
    ws7.cell(row=dr, column=5).value = f"=IFERROR({get_column_letter(3)}{dr}/{get_column_letter(4)}{dr},0)"
    # Top Category — left blank
    ws7.cell(row=dr, column=6, value="")
    # Commission (5%)
    ws7.cell(row=dr, column=8).value = f"=IFERROR({get_column_letter(3)}{dr}*0.05,0)"

    apply_body_style(ws7, dr, 2, 8, row_idx,
                     formats=[(1, FORMATS["decimal_2"]), (2, FORMATS["integer"]),
                              (3, FORMATS["decimal_2"]), (6, FORMATS["decimal_2"])])

# Total row
TR7 = HR7 + 1 + STAFF_ROWS
ws7.cell(row=TR7, column=2, value="TOTALS")
ws7.cell(row=TR7, column=3).value = f"=SUM({get_column_letter(3)}{HR7+1}:{get_column_letter(3)}{TR7-1})"
ws7.cell(row=TR7, column=4).value = f"=SUM({get_column_letter(4)}{HR7+1}:{get_column_letter(4)}{TR7-1})"
ws7.cell(row=TR7, column=5).value = f"=IFERROR({get_column_letter(3)}{TR7}/{get_column_letter(4)}{TR7},0)"
ws7.cell(row=TR7, column=8).value = f"=SUM({get_column_letter(8)}{HR7+1}:{get_column_letter(8)}{TR7-1})"
style_total_row(ws7, TR7, 2, 8)
ws7.cell(row=TR7, column=3).number_format = FORMATS["decimal_2"]
ws7.cell(row=TR7, column=4).number_format = FORMATS["integer"]
ws7.cell(row=TR7, column=5).number_format = FORMATS["decimal_2"]
ws7.cell(row=TR7, column=8).number_format = FORMATS["decimal_2"]

# Shift dropdown
dv_shift = DataValidation(type="list", formula1='"Morning,Evening,Night"', allow_blank=True)
ws7.add_data_validation(dv_shift)
dv_shift.add(f'{get_column_letter(7)}{HR7+1}:{get_column_letter(7)}{TR7-1}')

# Data bars on Total Sales column (C)
data_bar_rule = DataBarRule(
    start_type="min", end_type="max",
    color=PRIMARY,
    showValue=True,
    minLength=None, maxLength=None
)
ws7.conditional_formatting.add(
    f'{get_column_letter(3)}{HR7+1}:{get_column_letter(3)}{TR7-1}',
    data_bar_rule
)

ws7.freeze_panes = "C5"
ws7.print_area = f'B2:H{TR7}'
ws7.print_title_rows = f'{HR7}:{HR7}'


# ============================================================
# 8. PRODUCT LIST
# ============================================================
ws8 = wb.create_sheet("Product List")
ws8.sheet_properties.tabColor = TAB_COLOR

setup_sheet(ws8, title="Product List", last_col=9)

headers_pl = ["Product ID", "Item Name", "Category", "Unit Price",
              "Cost Price", "Margin %", "Description", "Supplier"]
col_widths_pl = [14, 22, 16, 14, 14, 12, 32, 18]
set_col_widths(ws8, col_widths_pl)

HR8 = 4
for i, h in enumerate(headers_pl):
    ws8.cell(row=HR8, column=2 + i, value=h)
style_header_row(ws8, HR8, 2, 9)

PROD_ROWS = 200
for r in range(HR8 + 1, HR8 + 1 + PROD_ROWS):
    row_idx = r - HR8
    dr = r
    # Margin %: =IFERROR((E5-D5)/E5,0) — Unit Price (E) - Cost Price (D) / Unit Price (E)
    # Col mapping: B=Product ID, C=Item Name, D=Category, E=Unit Price, F=Cost Price, G=Margin%
    ws8.cell(row=dr, column=7).value = f"=IFERROR(({get_column_letter(5)}{dr}-{get_column_letter(6)}{dr})/{get_column_letter(5)}{dr},0)"
    apply_body_style(ws8, dr, 2, 9, row_idx,
                     formats=[(3, FORMATS["decimal_2"]), (4, FORMATS["decimal_2"]),
                              (5, FORMATS["percentage"])])

# Total row
TR8 = HR8 + 1 + PROD_ROWS
ws8.cell(row=TR8, column=2, value="SUMMARY")
ws8.cell(row=TR8, column=5).value = f"=IFERROR(SUM({get_column_letter(5)}{HR8+1}:{get_column_letter(5)}{TR8-1}),0)"
ws8.cell(row=TR8, column=6).value = f"=IFERROR(SUM({get_column_letter(6)}{HR8+1}:{get_column_letter(6)}{TR8-1}),0)"
ws8.cell(row=TR8, column=7).value = f"=IFERROR(({get_column_letter(5)}{TR8}-{get_column_letter(6)}{TR8})/{get_column_letter(5)}{TR8},0)"
style_total_row(ws8, TR8, 2, 9)
ws8.cell(row=TR8, column=5).number_format = FORMATS["decimal_2"]
ws8.cell(row=TR8, column=6).number_format = FORMATS["decimal_2"]
ws8.cell(row=TR8, column=7).number_format = FORMATS["percentage"]

# Category dropdown
dv_cat2 = DataValidation(type="list", formula1='"Electronics,Clothing,Food & Beverage,Stationery,Health & Beauty,Home & Garden,Sports,Toys,Other"', allow_blank=True)
ws8.add_data_validation(dv_cat2)
dv_cat2.add(f'{get_column_letter(4)}{HR8+1}:{get_column_letter(4)}{TR8-1}')

ws8.freeze_panes = "C5"
ws8.print_area = f'B2:I{TR8}'
ws8.print_title_rows = f'{HR8}:{HR8}'


# ============================================================
# Reorder sheets: Sales Entry, Daily Summary, Weekly Summary,
# Monthly Summary, Invoice, Inventory, Staff Performance, Product List
# ============================================================
# Already in correct order since we used wb.active for Sales Entry
# and created sheets sequentially.

# ============================================================
# Save
# ============================================================
OUTPUT = "/home/z/my-project/download/POS_Professional_Empty.xlsx"
os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
wb.save(OUTPUT)
print(f"✅  Saved: {OUTPUT}")
